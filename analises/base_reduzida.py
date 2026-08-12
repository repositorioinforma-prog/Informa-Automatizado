"""
Processador de Base Reduzida.

Remove, de cada tabela do arquivo, as colunas cuja "Base reduzida" seja
menor que o limite escolhido — preservando a formatação original E as
mesclagens de cabeçalho (ex.: um título "Regiões" mesclado sobre várias
colunas de região encolhe pra cobrir só as colunas que sobraram).

Reaproveita o mesmo motor usado pelo código 05 do Relatório Automatizado
(core/planilha_utils.py::excluir_colunas_base_reduzida) — mesma lógica
testada nos dois lugares, em vez de duas implementações separadas.

Diferente das análises de correspondência, não depende do `dados`
(pandas) carregado no início do app: trabalha direto com o Workbook do
openpyxl (para preservar formatação original), então tem upload de
arquivo próprio.
"""
import io
import zipfile

import openpyxl
import streamlit as st

from core.planilha_utils import (
    encontrar_inicio_bloco,
    encontrar_fim_bloco,
    linha_vazia_ate_coluna,
    normalizar_texto_maiusculo,
    excluir_colunas_base_reduzida,
)


def _dividir_em_blocos_de_tabela(ws):
    """
    Varre a aba de cima a baixo identificando blocos de tabela separados
    por pelo menos uma linha totalmente em branco. Devolve uma lista de
    (linha_inicio, linha_fim) — 1-based, inclusive nos dois extremos.
    """
    max_col = ws.max_column
    max_row = ws.max_row
    blocos = []

    r = 1
    while r <= max_row:
        if linha_vazia_ate_coluna(ws, r, max_col):
            r += 1
            continue
        inicio = encontrar_inicio_bloco(ws, r, max_col)
        fim = encontrar_fim_bloco(ws, r, max_col)
        blocos.append((inicio, fim))
        r = fim + 1

    return blocos


def _copiar_bloco_para(ws_origem, linha_ini, linha_fim, ws_destino, linha_destino_ini):
    """Copia um intervalo de linhas (valores, estilo e mesclagens internas
    ao bloco) de uma aba pra outra, começando em `linha_destino_ini`."""
    from copy import copy as _copy_style

    max_col = ws_origem.max_column
    offset = linha_destino_ini - linha_ini

    for r in range(linha_ini, linha_fim + 1):
        for c in range(1, max_col + 1):
            origem = ws_origem.cell(row=r, column=c)
            destino = ws_destino.cell(row=r + offset, column=c)
            destino.value = origem.value
            if origem.has_style:
                destino.font = _copy_style(origem.font)
                destino.border = _copy_style(origem.border)
                destino.fill = _copy_style(origem.fill)
                destino.number_format = origem.number_format
                destino.protection = _copy_style(origem.protection)
                destino.alignment = _copy_style(origem.alignment)
        altura = ws_origem.row_dimensions.get(r)
        if altura and altura.height:
            ws_destino.row_dimensions[r + offset].height = altura.height

    for rng in ws_origem.merged_cells.ranges:
        if rng.min_row >= linha_ini and rng.max_row <= linha_fim:
            ws_destino.merge_cells(
                start_row=rng.min_row + offset, start_column=rng.min_col,
                end_row=rng.max_row + offset, end_column=rng.max_col,
            )

    for c in range(1, max_col + 1):
        largura = ws_origem.column_dimensions.get(
            openpyxl.utils.get_column_letter(c)
        )
        if largura and largura.width:
            ws_destino.column_dimensions[openpyxl.utils.get_column_letter(c)].width = largura.width

    return linha_fim - linha_ini + 1  # quantas linhas ocupou no destino


def _processar_tabelas_base_reduzida(wb, aba_original, limite=20):
    """
    Localiza cada bloco de tabela na aba selecionada, exclui as colunas
    de "Base reduzida" abaixo do limite (com mesclagem preservada) e
    devolve:
        - um workbook consolidado com todas as tabelas já filtradas;
        - uma lista de (nome_arquivo, BytesIO) com cada tabela isolada.
    """
    ws = wb[aba_original]
    blocos = _dividir_em_blocos_de_tabela(ws)

    n_blocos_com_base_reduzida = 0
    for inicio, fim in blocos:
        tem_base_reduzida = any(
            normalizar_texto_maiusculo(ws.cell(row=r, column=1).value) == "BASE REDUZIDA"
            for r in range(inicio, fim + 1)
        )
        if tem_base_reduzida:
            n_blocos_com_base_reduzida += 1

    # a exclusão já detecta os blocos e o limite sozinha, olhando a aba inteira
    excluir_colunas_base_reduzida(ws, limite=limite)

    wb_consolidado = openpyxl.Workbook()
    ws_consolidado = wb_consolidado.active
    ws_consolidado.title = "Consolidado"

    arquivos_tabelas = []
    linha_destino = 1
    for inicio, fim in blocos:
        n_linhas = _copiar_bloco_para(ws, inicio, fim, ws_consolidado, linha_destino)
        linha_destino += n_linhas + 1  # +1 = linha em branco entre tabelas

        wb_tabela = openpyxl.Workbook()
        ws_tabela = wb_tabela.active
        _copiar_bloco_para(ws, inicio, fim, ws_tabela, 1)
        tabela_bytes = io.BytesIO()
        wb_tabela.save(tabela_bytes)
        tabela_bytes.seek(0)
        arquivos_tabelas.append((f"tabela_{inicio}.xlsx", tabela_bytes))

    consolidado_bytes = io.BytesIO()
    wb_consolidado.save(consolidado_bytes)
    consolidado_bytes.seek(0)

    return consolidado_bytes, arquivos_tabelas, n_blocos_com_base_reduzida


def processador_base_reduzida():
    """
    Tela do Processador de Base Reduzida: upload de um Excel com várias
    tabelas empilhadas numa aba, remoção das colunas com base abaixo do
    limite escolhido (preservando mesclagem de cabeçalho) e download da
    planilha consolidada e/ou das tabelas separadas em .zip.
    """
    st.header("Processador de Base Reduzida")
    st.caption(
        "Remove, de cada tabela do arquivo, as colunas cuja 'Base reduzida' "
        "seja menor que o limite escolhido, preservando a formatação e as "
        "mesclagens de cabeçalho originais."
    )

    uploaded_file = st.file_uploader(
        "Faça upload do arquivo Excel", type=["xlsx"], key="upload_base_reduzida"
    )

    if not uploaded_file:
        st.info("Envie um arquivo .xlsx para começar.")
        return

    wb = openpyxl.load_workbook(uploaded_file, rich_text=True)
    abas = wb.sheetnames
    col1, col2 = st.columns([2, 1])
    with col1:
        aba = st.selectbox("Selecione a aba para processar:", abas, key="aba_base_reduzida")
    with col2:
        limite = st.number_input(
            "Limite da base", min_value=1, value=20, step=1, key="limite_base_reduzida",
            help="Colunas com 'Base reduzida' menor que este valor são excluídas.",
        )

    if st.button("Processar", key="processar_base_reduzida"):
        with st.spinner("Processando tabelas..."):
            consolidado_bytes, arquivos_tabelas, n_com_base = _processar_tabelas_base_reduzida(
                wb, aba, limite=int(limite)
            )

        st.session_state["br_consolidado_bytes"] = consolidado_bytes.getvalue()

        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, "w") as zip_file:
            for nome, tabela_bytes in arquivos_tabelas:
                zip_file.writestr(nome, tabela_bytes.read())
        zip_buffer.seek(0)
        st.session_state["br_zip_bytes"] = zip_buffer.getvalue()

        st.success(
            f"Processamento concluído! {len(arquivos_tabelas)} tabela(s) encontrada(s), "
            f"{n_com_base} com linha 'Base reduzida' (colunas < {int(limite)} excluídas nelas)."
        )

    if "br_consolidado_bytes" in st.session_state:
        st.download_button(
            label="Baixar planilha consolidada",
            data=st.session_state["br_consolidado_bytes"],
            file_name="Bases reduzidas - Consolidado.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="download_br_consolidado"
        )

    if "br_zip_bytes" in st.session_state:
        st.download_button(
            label="Baixar todas as tabelas processadas (ZIP)",
            data=st.session_state["br_zip_bytes"],
            file_name="Tabelas_processadas.zip",
            mime="application/zip",
            key="download_br_zip"
        )
