"""
Códigos Individuais.

Roda um único código do "Relatório Automatizado" isoladamente — mesma
lógica exata, sem passar pelo assistente completo. Útil pra testar/
depurar um código específico direto num arquivo real, sem precisar
rodar os 15 passos pra descobrir se um deles está funcionando.

Não duplica nenhuma lógica: chama as mesmas funções de
core/relatorio_automatizado_math.py que o assistente usa.

O código "Legendas" tem sua própria tela dedicada (upload de dois
arquivos) e não está listado aqui.

Não depende do `dados` carregado no início do app: tem upload próprio.
"""
import io
import os

import openpyxl
import streamlit as st

from core.planilha_utils import worksheet_para_html
from core.cabecalho_imagem import inserir_imagem_cabecalho
from core.cabecalho_correcao_math import parsear_blocos_cabecalho_referencia, aplicar_codigo_13
from core.capas_resultados_math import aplicar_codigo_15
from core.relatorio_automatizado_math import (
    aplicar_codigo_01,
    aplicar_codigo_02,
    aplicar_codigo_03,
    aplicar_codigo_04,
    aplicar_codigo_05,
    aplicar_codigo_06,
    aplicar_codigo_06_autofit,
    aplicar_codigo_07,
    aplicar_codigo_08,
    aplicar_codigo_09,
    aplicar_codigo_10,
    aplicar_codigo_11,
    aplicar_codigo_12,
    aplicar_codigo_14,
    aplicar_codigo_16,
    TERMOS_EXCLUSAO_CODIGO_03,
)

# Registro de todos os códigos: número, nome, onde a função atua
# ("ws" = planilha ativa, "ws_param" = planilha + 1 parâmetro extra,
# "wb" = workbook inteiro) e a função em si. Códigos ainda não portados
# entram com tipo "pendente".
CODIGOS = [
    {"num": "01", "nome": "Preenche células do SPSS pirata", "tipo": "ws", "fn": aplicar_codigo_01},
    {"num": "02", "nome": "Muda Bordas (espessa → média)", "tipo": "ws", "fn": aplicar_codigo_02},
    {
        "num": "03", "nome": "Ordenar tabelas (lista de exclusão padrão + termos extras)",
        "tipo": "ws_termos", "fn": aplicar_codigo_03,
    },
    {"num": "04", "nome": "Formatação de rótulos específicos", "tipo": "ws", "fn": aplicar_codigo_04},
    {
        "num": "05", "nome": "Formata layout / Base reduzida", "tipo": "ws_param", "fn": aplicar_codigo_05,
        "param": {"chave": "limite", "label": "Limite mínimo de Base reduzida", "default": 25, "min": 1},
    },
    {"num": "06", "nome": "Ajuste DIN 9 na linha de Pergunta", "tipo": "ws", "fn": aplicar_codigo_06},
    {"num": "06b", "nome": "Autoajuste geral de altura (AutoFit aproximado)", "tipo": "ws", "fn": aplicar_codigo_06_autofit},
    {"num": "07", "nome": "Adiciona linhas e quebra de página", "tipo": "ws", "fn": aplicar_codigo_07},
    {"num": "08", "nome": "Base pequena", "tipo": "ws", "fn": aplicar_codigo_08},
    {"num": "09", "nome": "Ajusta a altura das labels (AutoFit aproximado)", "tipo": "ws", "fn": aplicar_codigo_09},
    {"num": "10", "nome": "Ajusta o título da Renda", "tipo": "ws", "fn": aplicar_codigo_10},
    {"num": "11", "nome": "Ajusta a altura dos Títulos (AutoFit aproximado)", "tipo": "ws", "fn": aplicar_codigo_11},
    {"num": "12", "nome": "Ajuste da altura das Perguntas (AutoFit aproximado)", "tipo": "ws", "fn": aplicar_codigo_12},
    {
        "num": "13", "nome": "Corrigir cabeçalhos repetidos (precisa de arquivo de referência)",
        "tipo": "ws_ref_arquivo", "fn": aplicar_codigo_13,
    },
    {"num": "14", "nome": "Cabeçalho/rodapé de impressão fixos", "tipo": "wb", "fn": aplicar_codigo_14},
    {"num": "15", "nome": "InserirCapasResultados", "tipo": "ws", "fn": aplicar_codigo_15},
    {
        "num": "16", "nome": "Índice / Sumário (por quebras de página)", "tipo": "ws", "fn": aplicar_codigo_16,
    },
]


def modulo_codigos_individuais():
    st.header("Códigos Individuais")
    st.caption(
        "Roda um único código do Relatório Automatizado isoladamente — "
        "mesma função exata que o assistente completo usa, sem lógica "
        "duplicada. Útil pra testar ou depurar um código específico "
        "direto num arquivo real."
    )
    st.info(
        "O código \"Legendas\" tem tela própria no menu (upload de dois "
        "arquivos) — não está listado aqui."
    )

    opcoes = [f"Código {c['num']} — {c['nome']}" for c in CODIGOS]
    escolha = st.selectbox("Qual código você quer rodar?", opcoes, key="ci_escolha")
    codigo = CODIGOS[opcoes.index(escolha)]

    if codigo["tipo"] == "pendente":
        st.warning(
            f"Código {codigo['num']} ainda não foi portado para Python — "
            "nada a rodar por aqui ainda."
        )
        return

    arquivo = st.file_uploader("Arquivo .xlsx", type=["xlsx"], key="ci_upload")
    if not arquivo:
        st.info("Envie um arquivo .xlsx para começar.")
        return

    arquivo_referencia = None
    blocos_referencia_cache = None
    if codigo["tipo"] == "ws_ref_arquivo":
        st.caption(
            "Arquivo com o cabeçalho \"correto\" de cada tipo de "
            "segmentação — um bloco \"Titulo: ...\" por tipo (Sexo, "
            "Escolaridade, Religião etc.), cada um com a linha de grupo "
            "e a linha de subcategoria certas."
        )
        arquivo_referencia = st.file_uploader(
            "Arquivo de referência dos cabeçalhos (.xlsx)", type=["xlsx"], key="ci_ref_13"
        )
        if arquivo_referencia is None:
            st.info("Envie também o arquivo de referência dos cabeçalhos para continuar.")
            return

    valor_param = None
    if codigo["tipo"] == "ws_param":
        p = codigo["param"]
        valor_param = st.number_input(
            p["label"], min_value=p["min"], value=p["default"], step=1,
            key=f"ci_param_{codigo['num']}"
        )

    lista_termos = None
    if codigo["tipo"] == "ws_termos":
        st.caption(
            f"A lista padrão já tem {len(TERMOS_EXCLUSAO_CODIGO_03)} termos "
            "herdados do VBA original (ex.: 'Base', 'Não sabe', 'Branco/Nulo')."
        )
        termos_extra_txt = st.text_area(
            "Termos extras a incluir na lista de exclusão (um por linha) — "
            "essas linhas ficam fixas na posição original e não entram na "
            "reordenação, além dos termos padrão:",
            key="ci_termos_extra_03",
            height=120,
        )
        termos_extra = [t.strip() for t in termos_extra_txt.splitlines() if t.strip()]
        lista_termos = TERMOS_EXCLUSAO_CODIGO_03 + termos_extra
        if termos_extra:
            st.caption(f"{len(termos_extra)} termo(s) extra(s) adicionado(s) para esta execução.")

    incluir_logo = False
    if codigo["num"] == "14":
        caminho_logo = os.path.join(
            os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "assets", "logo.jpg"
        )
        incluir_logo = st.checkbox(
            "⚠️ [Experimental] Incluir o logo no canto superior direito de toda "
            "página (5 cm de largura) — não confirmado que funciona no Excel de "
            "verdade",
            value=False, key="ci_incluir_logo_14",
        )

    if st.button("Aplicar", key="ci_aplicar"):
        with st.spinner("Aplicando..."):
            wb = openpyxl.load_workbook(arquivo, rich_text=True)
            alvo = wb if codigo["tipo"] == "wb" else wb.active

            if codigo["tipo"] == "ws_param":
                resultado = codigo["fn"](alvo, valor_param)
            elif codigo["tipo"] == "ws_termos":
                resultado = codigo["fn"](alvo, lista_termos)
            elif codigo["tipo"] == "ws_ref_arquivo":
                blocos_ref = parsear_blocos_cabecalho_referencia(arquivo_referencia)
                resultado = codigo["fn"](alvo, blocos_ref)
            else:
                resultado = codigo["fn"](alvo)

            saida = io.BytesIO()
            wb.save(saida)
            bytes_finais = saida.getvalue()

            if codigo["num"] == "14" and incluir_logo:
                bytes_finais = inserir_imagem_cabecalho(
                    bytes_finais, caminho_logo, largura_cm=5.0, posicao="R",
                    aplicar_em_todas_abas=True,
                )

        if isinstance(resultado, tuple):
            resultado_txt = ", ".join(str(x) for x in resultado) + " alteração(ões)"
        elif isinstance(resultado, dict):
            if resultado.get("status") == "ok":
                resultado_txt = (
                    f"capa do Total na linha {resultado['linha_capa_total']}"
                    + (
                        f", capa dos Segmentos na linha {resultado['linha_capa_segmentos']}"
                        if resultado.get("linha_capa_segmentos")
                        else " (nenhuma tabela segmentada encontrada)"
                    )
                )
            else:
                resultado_txt = resultado.get("mensagem", "erro desconhecido")
        else:
            resultado_txt = f"{resultado} alteração(ões)"
        st.session_state["ci_resultado_bytes"] = bytes_finais
        st.session_state["ci_resultado_nome"] = arquivo.name
        st.session_state["ci_resultado_msg"] = f"Código {codigo['num']} aplicado: {resultado_txt}."

    if "ci_resultado_bytes" in st.session_state:
        st.success(st.session_state["ci_resultado_msg"])
        st.download_button(
            "Baixar resultado",
            data=st.session_state["ci_resultado_bytes"],
            file_name=f"codigo_{codigo['num']}_{st.session_state['ci_resultado_nome']}",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="ci_download"
        )

        st.subheader("Pré-visualização do resultado")
        wb_preview = openpyxl.load_workbook(io.BytesIO(st.session_state["ci_resultado_bytes"]))
        aba_escolhida = st.selectbox(
            "Aba", wb_preview.sheetnames, key="ci_preview_aba"
        ) if len(wb_preview.sheetnames) > 1 else wb_preview.sheetnames[0]
        ws_preview = wb_preview[aba_escolhida]

        html_preview, truncou_linhas, truncou_colunas = worksheet_para_html(ws_preview)
        st.components.v1.html(html_preview, height=620, scrolling=True)

        avisos = []
        if truncou_linhas:
            avisos.append(f"mostrando as primeiras 150 linhas de {ws_preview.max_row}")
        if truncou_colunas:
            avisos.append(f"mostrando as primeiras 30 colunas de {ws_preview.max_column}")
        aviso_txt = " e ".join(avisos)
        st.caption(
            (f"⚠️ Pré-visualização parcial ({aviso_txt}) — " if aviso_txt else "")
            + "reproduz mesclagem, cor de fundo, negrito/itálico, alinhamento e bordas, "
            "mas é uma aproximação; o arquivo baixado é a fonte da verdade."
        )
