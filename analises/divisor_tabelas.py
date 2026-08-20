"""Tela do Divisor de Tabelas de Regiões (v2)."""
from __future__ import annotations

import io

import openpyxl
import pandas as pd
import streamlit as st

from core.divisor_tabelas_math import processar_workbook, ALTURA_UTIL, ALTURA_LINHA_PADRAO
from core.relatorio_automatizado_math import aplicar_codigo_16, remover_termo_do_sumario


_ROTULO_ACAO = {
    "dividida": "Dividida em partes",
    "legenda_movida": "Legenda movida pra página própria",
    "nao_corrigida": "Não corrigida automaticamente",
}


def _result_rows(resumo):
    rows = []
    for a in resumo["acoes"]:
        rows.append({
            "Aba": a["table"].sheet,
            "Tabela": a["titulo"],
            "Linhas físicas": a["tamanho"],
            "Altura total": round(a["altura"], 1),
            "Tem legenda": "Sim" if a["tem_legenda"] else "Não",
            "Protegida": "Sim" if a["protegida"] else "Não",
            "Ação": _ROTULO_ACAO.get(a["acao"], a["acao"]),
            "Partes": a.get("parts", "-"),
        })
    return rows


def modulo_divisor_tabelas():
    st.header("Divisor de Tabelas de Regiões")
    linhas_uteis_equivalentes = int(ALTURA_UTIL / ALTURA_LINHA_PADRAO)
    st.write(
        f"Envie um relatório Excel finalizado. Cada tabela vive num slot de página "
        f"de 32 linhas de referência (altura {ALTURA_LINHA_PADRAO:g} cada) — 2 acima "
        f"+ {linhas_uteis_equivalentes} linhas úteis + 2 abaixo. O orçamento é por "
        f"ALTURA real (não contagem de linha), já que nem toda linha tem a mesma "
        f"altura — soma a altura de título, cabeçalho, labels, Base, Pergunta e "
        f"legenda (quando houver) do bloco inteiro contra {ALTURA_UTIL:g} de altura "
        f"útil. Tabela que passar disso é candidata a ajuste."
    )
    st.info(
        "Tabelas normais que ultrapassam o limite têm os labels divididos em "
        "partes, repetindo cabeçalho/Base/Pergunta/legenda em cada uma, com "
        "'Continuação' a partir da 2ª. Tabelas protegidas (ex.: IPV, Avaliação "
        "com Aprovação/Regular/Reprovação) nunca têm os labels divididos — se "
        "tiverem legenda, ela é movida pra uma página própria; sem legenda, a "
        "tabela só é reportada, sem alteração automática."
    )

    uploaded = st.file_uploader(
        "Carregar relatório Excel",
        type=["xlsx", "xlsm"],
        key="divisor_tabelas_upload",
    )
    if uploaded is None:
        return

    if st.button("Analisar e ajustar tabelas", key="divisor_tabelas_processar"):
        try:
            raw = uploaded.getvalue()
            keep_vba = uploaded.name.lower().endswith(".xlsm")
            wb = openpyxl.load_workbook(io.BytesIO(raw), keep_vba=keep_vba, rich_text=True)

            with st.status("Analisando tabelas...", expanded=True) as status:
                resumo = processar_workbook(wb)
                status.update(label="Análise concluída", state="complete")

            out = io.BytesIO()
            wb.save(out)
            out_bytes = out.getvalue()

            st.session_state["divisor_tabelas_resultado"] = out_bytes
            st.session_state["divisor_tabelas_resumo"] = resumo
            st.session_state["divisor_tabelas_nome"] = uploaded.name
            st.session_state.pop("divisor_tabelas_resultado_com_indice", None)
        except Exception as exc:
            st.exception(exc)
            return

    resumo = st.session_state.get("divisor_tabelas_resumo")
    out_bytes = st.session_state.get("divisor_tabelas_resultado")
    if not resumo or out_bytes is None:
        return

    c1, c2, c3, c4, c5, c6 = st.columns(6)
    c1.metric("Abas analisadas", resumo["sheets_analyzed"])
    c2.metric("Tabelas candidatas", resumo["tables_candidatas"])
    c3.metric("Divididas", resumo["tables_divididas"])
    c4.metric("Legenda movida", resumo["legenda_movida"])
    c5.metric("Não corrigidas", resumo["nao_corrigidas"])
    c6.metric("Partes geradas", resumo["partes_geradas"])

    if resumo["nao_corrigidas"]:
        titulos_nao_corrigidas = [
            a["titulo"] for a in resumo["acoes"] if a["acao"] == "nao_corrigida"
        ]
        lista = "\n".join(f"- {t}" for t in titulos_nao_corrigidas)
        st.warning(
            "Tabela(s) protegida(s) sem legenda (ou onde nem sem legenda cabe "
            "nas linhas úteis) — passam do limite mas não têm como ser "
            "corrigidas automaticamente sem dividir os labels no meio. "
            "Precisam de ajuste manual:\n\n" + lista
        )

    if resumo["tables_candidatas"]:
        st.success(f"{resumo['tables_candidatas']} tabela(s) precisaram de ajuste.")
        rows = _result_rows(resumo)
        if rows:
            st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)
    else:
        st.success("Nenhuma tabela ultrapassa o limite de linhas úteis segundo o critério atual.")

    original_name = st.session_state.get("divisor_tabelas_nome", "relatorio.xlsx")
    if "." in original_name:
        stem, ext = original_name.rsplit(".", 1)
        output_name = f"{stem}_tabelas_divididas.{ext}"
    else:
        output_name = original_name + "_tabelas_divididas.xlsx"

    st.download_button(
        "Baixar relatório processado",
        data=out_bytes,
        file_name=output_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="divisor_tabelas_download",
    )

    st.divider()
    st.subheader("Índice / Sumário (código 16)")
    st.caption(
        "O Índice lista os títulos junto com o número de página REAL, "
        "calculado a partir das quebras de página do arquivo — por isso "
        "só faz sentido rodar depois de dividir as tabelas, nunca antes "
        "(dividir tabela desloca tudo que vem depois e muda a "
        "paginação; se o índice for gerado antes, os números de página "
        "ficam errados). Rode aqui, direto em cima do arquivo já "
        "dividido acima, sem precisar baixar e reenviar em outra tela."
    )
    if st.button("Rodar Índice sobre o arquivo já dividido", key="divisor_tabelas_rodar_indice"):
        with st.spinner("Gerando o Índice..."):
            wb_com_indice = openpyxl.load_workbook(io.BytesIO(out_bytes), rich_text=True)
            n_titulos = aplicar_codigo_16(wb_com_indice.active)
            saida_indice = io.BytesIO()
            wb_com_indice.save(saida_indice)
            st.session_state["divisor_tabelas_resultado_com_indice"] = saida_indice.getvalue()
        st.success(f"Índice gerado com {n_titulos} título(s) listado(s).")

    resultado_com_indice = st.session_state.get("divisor_tabelas_resultado_com_indice")
    if resultado_com_indice is not None:
        if "." in original_name:
            stem, ext = original_name.rsplit(".", 1)
            output_name_indice = f"{stem}_tabelas_divididas_com_indice.{ext}"
        else:
            output_name_indice = original_name + "_tabelas_divididas_com_indice.xlsx"
        st.download_button(
            "Baixar relatório dividido + com Índice",
            data=resultado_com_indice,
            file_name=output_name_indice,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="divisor_tabelas_download_com_indice",
        )

        st.markdown("**Prévia do Sumário**")
        wb_preview = openpyxl.load_workbook(io.BytesIO(resultado_com_indice), rich_text=True)
        sumario_ws = wb_preview["Sumário"]
        preview_rows = [
            {"Título": sumario_ws.cell(row=r, column=1).value, "Página": sumario_ws.cell(row=r, column=2).value}
            for r in range(1, sumario_ws.max_row + 1)
            if sumario_ws.cell(row=r, column=1).value is not None
        ]
        st.dataframe(pd.DataFrame(preview_rows), use_container_width=True, hide_index=True)

        st.caption("Remover termo de todos os títulos do Sumário:")
        col_a, col_b = st.columns(2)
        with col_a:
            if st.button("Remover \"(Estimulada e Única)\"", key="divisor_tabelas_remover_unica"):
                wb_edit = openpyxl.load_workbook(io.BytesIO(resultado_com_indice), rich_text=True)
                n = remover_termo_do_sumario(wb_edit, "(Estimulada e Única)")
                saida = io.BytesIO()
                wb_edit.save(saida)
                st.session_state["divisor_tabelas_resultado_com_indice"] = saida.getvalue()
                st.success(f"Termo removido de {n} título(s).")
                st.rerun()
        with col_b:
            if st.button("Remover \"(Estimulada e Múltipla)\"", key="divisor_tabelas_remover_multipla"):
                wb_edit = openpyxl.load_workbook(io.BytesIO(resultado_com_indice), rich_text=True)
                n = remover_termo_do_sumario(wb_edit, "(Estimulada e Múltipla)")
                saida = io.BytesIO()
                wb_edit.save(saida)
                st.session_state["divisor_tabelas_resultado_com_indice"] = saida.getvalue()
                st.success(f"Termo removido de {n} título(s).")
                st.rerun()

