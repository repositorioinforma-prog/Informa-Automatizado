"""
Exclusões — Equilíbrio de Cotas por Exclusão.

Diferente das análises de correspondência, não depende do `dados`
carregado no início do app: tem upload próprio (Excel ou SPSS .sav).

Fluxo: carrega o banco, detecta/seleciona variáveis de cota (padrão P1,
P2, P3, P7), define a amostra esperada e a meta (% ou N bruto) de cada
categoria, e calcula quais entrevistas excluir para equilibrar a amostra
dentro de uma tolerância — sem nunca deixar nenhuma categoria abaixo da
base mínima, e sem excluir ninguém de categorias já abaixo da meta.

Não modifica e não depende dos módulos de correspondência.
"""
import os
import tempfile
import unicodedata

import openpyxl
import pandas as pd
import pyreadstat
import streamlit as st

from core.config import ROTULOS_VARIAVEIS, VARIAVEIS_EXCLUSAO_PADRAO, _rotulo_amigavel
from core.exclusoes_math import calcular_exclusoes

# Nomes alternativos (aliases) usados para reconhecer cada variável de cota
# num arquivo de metas externo, além do próprio código e do rótulo amigável
# de ROTULOS_VARIAVEIS. Comparação ignora acento/caixa/espaço nas pontas.
ALIASES_VARIAVEIS_COTA = {
    "P1": ["sexo", "genero", "gênero"],
    "P2": ["idade", "faixa etaria", "faixa etária"],
    "P3": ["renda", "renda familiar", "classe", "classe social"],
    "P7": ["regiao", "região", "regioes", "regiões", "sub-regiao", "sub-região"],
}


def _normalizar_texto(valor):
    """Minúsculas, sem acento, sem espaços nas pontas — para comparação tolerante."""
    s = str(valor).strip().lower()
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")
    return s


def _carregar_dados_exclusao(uploaded_file):
    """
    Carrega Excel (.xlsx) ou SPSS (.sav) para o módulo de Exclusões.
    Mais enxuto que `core.dados.carregar_dados` (sem tabelas de preview),
    já que aqui o foco é ir direto para a configuração das cotas.
    """
    nome = uploaded_file.name.lower()

    if nome.endswith(".xlsx"):
        dados = pd.read_excel(uploaded_file)
        return dados

    elif nome.endswith(".sav"):
        with tempfile.NamedTemporaryFile(delete=False, suffix=".sav") as tmp:
            tmp.write(uploaded_file.getvalue())
            tmp_path = tmp.name
        try:
            dados, meta = pyreadstat.read_sav(tmp_path)
        finally:
            os.unlink(tmp_path)
        return dados

    else:
        st.error("Formato não suportado. Use um arquivo .xlsx ou .sav.")
        return None


def _ler_arquivo_metas(uploaded_file):
    """
    Lê um arquivo de metas no formato de blocos empilhados (como
    Amostra_RM.xlsx): cada bloco começa numa linha onde a 2ª coluna é
    literalmente '%' — a 1ª coluna dessa linha é o nome (amigável) da
    variável; as linhas seguintes, até o próximo cabeçalho, são as
    categorias com seu percentual (2ª coluna) e valor bruto (3ª coluna).

    Retorna uma lista de dicts: [{"rotulo": str, "categorias": {cat: {"pct": float|None, "bruto": int|None}}}]
    """
    wb = openpyxl.load_workbook(uploaded_file, data_only=True)
    ws = wb[wb.sheetnames[0]]

    blocos = []
    bloco_atual = None

    for linha in ws.iter_rows(values_only=True):
        if linha is None or all(v is None for v in linha):
            continue
        valores = (list(linha) + [None, None, None])[:3]
        col1, col2, col3 = valores

        eh_cabecalho = isinstance(col2, str) and col2.strip() == "%"
        if eh_cabecalho:
            if bloco_atual is not None:
                blocos.append(bloco_atual)
            bloco_atual = {"rotulo": str(col1).strip() if col1 is not None else "", "categorias": {}}
            continue

        if bloco_atual is not None and col1 is not None:
            try:
                pct = float(col2) if col2 is not None else None
            except (TypeError, ValueError):
                pct = None
            try:
                bruto = int(col3) if col3 is not None else None
            except (TypeError, ValueError):
                bruto = None
            if pct is not None or bruto is not None:
                bloco_atual["categorias"][str(col1).strip()] = {"pct": pct, "bruto": bruto}

    if bloco_atual is not None:
        blocos.append(bloco_atual)

    return blocos


def _casar_variavel(rotulo, variaveis_disponiveis):
    """Tenta reconhecer a qual variável de cota (P1, P2...) um rótulo do arquivo de metas se refere."""
    alvo = _normalizar_texto(rotulo)

    for v in variaveis_disponiveis:
        if _normalizar_texto(v) == alvo:
            return v

    for v in variaveis_disponiveis:
        if _normalizar_texto(_rotulo_amigavel(v)) == alvo:
            return v

    for v in variaveis_disponiveis:
        aliases = [_normalizar_texto(a) for a in ALIASES_VARIAVEIS_COTA.get(v, [])]
        if alvo in aliases:
            return v

    return None


def _casar_categoria(categoria_arquivo, categorias_dados):
    """Casa o texto de uma categoria do arquivo de metas com a categoria real nos dados."""
    alvo = _normalizar_texto(categoria_arquivo)
    for c in categorias_dados:
        if _normalizar_texto(c) == alvo:
            return c
    return None


def _importar_metas(uploaded_file, variaveis_cota, dados):
    """
    Lê e casa o arquivo de metas com as variáveis/categorias selecionadas.
    Retorna (metas_importadas, avisos), onde metas_importadas é
    {variavel: {categoria_nos_dados: {"pct": float|None, "bruto": int|None}}}.
    """
    blocos = _ler_arquivo_metas(uploaded_file)
    metas_importadas = {}
    variaveis_nao_reconhecidas = []
    categorias_nao_reconhecidas = []  # (rotulo_variavel, categoria_arquivo, categorias_disponiveis)

    for bloco in blocos:
        v = _casar_variavel(bloco["rotulo"], variaveis_cota)
        if v is None:
            variaveis_nao_reconhecidas.append(bloco["rotulo"])
            continue

        categorias_dados = dados[v].dropna().unique().tolist()
        metas_importadas.setdefault(v, {})

        for cat_arquivo, valores in bloco["categorias"].items():
            cat_dados = _casar_categoria(cat_arquivo, categorias_dados)
            if cat_dados is None:
                categorias_nao_reconhecidas.append((bloco["rotulo"], cat_arquivo, categorias_dados))
                continue
            metas_importadas[v][cat_dados] = valores

    avisos = []
    if variaveis_nao_reconhecidas:
        avisos.append(
            "Variáveis do arquivo de metas não reconhecidas: "
            f"{', '.join(variaveis_nao_reconhecidas)}. "
            f"Variáveis de cota selecionadas atualmente: {', '.join(variaveis_cota) or '(nenhuma)'}. "
            "Confirme se a variável correspondente está marcada em "
            "\"Variáveis de cota a usar no equilíbrio\" — o casamento só "
            "procura dentro do que está selecionado ali."
        )
    if categorias_nao_reconhecidas:
        detalhes = "; ".join(
            f"'{rotulo} / {cat}' (categorias disponíveis nos dados: {', '.join(str(c) for c in disponiveis)})"
            for rotulo, cat, disponiveis in categorias_nao_reconhecidas
        )
        avisos.append(
            "Categorias do arquivo de metas sem correspondência exata nos dados: "
            f"{detalhes}."
        )

    return metas_importadas, avisos


def _gerar_sintaxe_spss(id_var, ids_excluir, tratar_como_texto=None, estilo="any"):
    """
    Gera sintaxe SPSS que exclui os IDs informados.

    'tratar_como_texto': None = detecta pelo tipo Python dos IDs (padrão);
    True = força aspas (variável string no SPSS); False = força sem aspas
    (variável numérica no SPSS). Use o parâmetro explícito quando o tipo
    detectado automaticamente não bater com o tipo real da variável no
    arquivo .sav onde a sintaxe vai ser rodada — é comum IDs longos que
    parecem número serem, na verdade, variáveis de texto no SPSS.

    'estilo': "any" usa SELECT IF + ANY() (mais compacto); "or" usa
    SELECT IF com condições OR encadeadas (mais verboso, porém evita
    qualquer peculiaridade específica da função ANY() em certas versões/
    configurações de SPSS — útil se o erro 4325 persistir mesmo depois de
    ajustar o tipo do ID).
    """
    if not ids_excluir:
        return "* Nenhuma exclusao necessaria: a amostra ja esta dentro da tolerancia. *."

    if tratar_como_texto is None:
        ids_sao_texto = any(isinstance(i, str) for i in ids_excluir)
    else:
        ids_sao_texto = tratar_como_texto

    def _fmt(i):
        if ids_sao_texto:
            return f'"{i}"'
        # evita "1005.0" quando o ID veio como float sem casas decimais
        try:
            if float(i).is_integer():
                return str(int(i))
        except (TypeError, ValueError):
            pass
        return str(i)

    valores_fmt = [_fmt(i) for i in ids_excluir]

    if estilo == "or":
        condicoes = [f"{id_var} = {v}" for v in valores_fmt]
        por_linha = 4
        grupos = [condicoes[i:i + por_linha] for i in range(0, len(condicoes), por_linha)]
        corpo = " OR\n    ".join(" OR ".join(g) for g in grupos)
        linha_select = [
            "SELECT IF (NOT (",
            f"    {corpo}",
            ")).",
        ]
    else:
        por_linha = 10
        grupos = [valores_fmt[i:i + por_linha] for i in range(0, len(valores_fmt), por_linha)]
        corpo = ",\n    ".join(", ".join(g) for g in grupos)
        linha_select = [
            f"SELECT IF (NOT ANY({id_var},",
            f"    {corpo})).",
        ]

    linhas = [
        "* Exclusao de casos para equilibrio de cotas - gerado automaticamente. *.",
        f"* Total de casos a excluir: {len(ids_excluir)}. *.",
        "FILTER OFF.",
        "USE ALL.",
        *linha_select,
        "EXECUTE.",
    ]
    return "\n".join(linhas)


def modulo_exclusoes():
    st.header("Exclusões — Equilíbrio de Cotas")
    st.caption(
        "Carregue o banco, defina a amostra esperada e a meta de cada "
        "categoria das variáveis de cota, e o app calcula quais entrevistas "
        "excluir para equilibrar a amostra — sem furar a base mínima e sem "
        "excluir ninguém de categorias já abaixo da meta."
    )

    uploaded_file = st.file_uploader(
        "Carregar banco (Excel ou SPSS)", type=["xlsx", "sav"], key="upload_exclusoes"
    )

    if not uploaded_file:
        st.info("Envie um arquivo .xlsx ou .sav para começar.")
        return

    dados = _carregar_dados_exclusao(uploaded_file)
    if dados is None or dados.empty:
        return

    st.success(f"Banco carregado: {dados.shape[0]} entrevistas, {dados.shape[1]} variáveis.")

    colunas = dados.columns.tolist()

    # -------------------------------------------------------------
    # Configuração geral
    # -------------------------------------------------------------
    col_a, col_b, col_c = st.columns(3)
    with col_a:
        amostra_esperada = st.number_input(
            "Amostra esperada (n total do projeto)",
            min_value=1, value=int(dados.shape[0]), step=1, key="excl_amostra_esperada"
        )
    with col_b:
        tolerancia_pct = st.number_input(
            "Tolerância (± pontos percentuais)",
            min_value=0.0, value=2.0, step=0.5, format="%.1f", key="excl_tolerancia_pct"
        )
    with col_c:
        base_minima = st.number_input(
            "Base mínima por categoria",
            min_value=1, value=30, step=1, key="excl_base_minima"
        )

    id_var = st.selectbox(
        "Qual é a variável de ID da entrevista?",
        colunas,
        index=0,
        key="excl_id_var"
    )

    # -------------------------------------------------------------
    # Seleção das variáveis de cota
    # -------------------------------------------------------------
    disponiveis_padrao = [v for v in VARIAVEIS_EXCLUSAO_PADRAO if v in colunas and v != id_var]
    if disponiveis_padrao:
        st.caption(
            f"Detectadas automaticamente: {', '.join(disponiveis_padrao)}. "
            "Você pode adicionar ou remover variáveis abaixo."
        )

    colunas_restantes = [c for c in colunas if c != id_var]
    variaveis_cota = st.multiselect(
        "Variáveis de cota a usar no equilíbrio:",
        colunas_restantes,
        default=disponiveis_padrao,
        key="excl_variaveis_cota"
    )

    if not variaveis_cota:
        st.warning("Selecione pelo menos uma variável de cota.")
        return

    # -------------------------------------------------------------
    # Importar metas de um arquivo (opcional)
    # -------------------------------------------------------------
    st.subheader("Importar metas (opcional)")
    st.caption(
        "Arquivo Excel com blocos por variável — cabeçalho com o nome da "
        "variável e as colunas '%' e 'Valor Bruto', seguido das categorias "
        "(igual ao modelo). O app tenta casar automaticamente o nome da "
        "variável e de cada categoria com o que está selecionado abaixo."
    )
    uploaded_metas = st.file_uploader(
        "Importar arquivo de metas (.xlsx)", type=["xlsx"], key="upload_metas_exclusoes"
    )

    if uploaded_metas is not None:
        fingerprint = f"{uploaded_metas.name}:{uploaded_metas.size}"
        if st.session_state.get("excl_metas_fingerprint") != fingerprint:
            metas_importadas, avisos = _importar_metas(uploaded_metas, variaveis_cota, dados)
            st.session_state["excl_metas_importadas"] = metas_importadas
            st.session_state["excl_metas_fingerprint"] = fingerprint
            # muda a versão -> força os campos abaixo a nascerem com os
            # novos valores (widgets já criados não atualizam sozinhos)
            st.session_state["excl_metas_versao"] = st.session_state.get("excl_metas_versao", 0) + 1

            total_categorias = sum(len(cs) for cs in metas_importadas.values())
            st.success(
                f"Metas importadas: {len(metas_importadas)} variável(is), "
                f"{total_categorias} categoria(s) reconhecida(s)."
            )
            for aviso in avisos:
                st.warning(aviso)

    metas_importadas = st.session_state.get("excl_metas_importadas", {})
    metas_versao = st.session_state.get("excl_metas_versao", 0)

    # -------------------------------------------------------------
    # Metas por categoria (% e/ou N bruto)
    # -------------------------------------------------------------
    st.subheader("Metas por categoria")
    st.caption(
        "Preencha o % esperado (calcula o N bruto automaticamente a partir "
        "da amostra esperada) — ou informe o N bruto diretamente, se "
        "preferir; nesse caso ele tem prioridade sobre o %."
    )

    with st.form(key=f"form_metas_exclusoes__v{metas_versao}"):
        campos_pct = {}
        campos_n = {}

        for v in variaveis_cota:
            st.markdown(f"**{v}**")
            categorias = sorted(
                dados[v].dropna().unique().tolist(), key=lambda x: str(x)
            )
            campos_pct[v] = {}
            campos_n[v] = {}

            for c in categorias:
                col1, col2, col3 = st.columns([3, 1, 1])
                atual = int((dados[v] == c).sum())
                importado = metas_importadas.get(v, {}).get(c, {})
                pct_padrao = importado.get("pct") or 0.0
                n_padrao = importado.get("bruto") or 0

                with col1:
                    st.write(f"'{c}' — atual: {atual}")
                with col2:
                    campos_pct[v][c] = st.number_input(
                        "% esperado", min_value=0.0, max_value=100.0,
                        value=float(pct_padrao), step=0.01, format="%.2f",
                        key=f"excl_pct__{v}__{c}__v{metas_versao}"
                    )
                with col3:
                    campos_n[v][c] = st.number_input(
                        "N bruto (opcional)", min_value=0, value=int(n_padrao), step=1,
                        key=f"excl_n__{v}__{c}__v{metas_versao}"
                    )

        calcular = st.form_submit_button("Calcular exclusões")

    if calcular:
        # monta metas finais: N bruto tem prioridade sobre %
        metas = {}
        for v in variaveis_cota:
            metas[v] = {}
            for c in campos_pct[v]:
                n_manual = campos_n[v][c]
                if n_manual > 0:
                    metas[v][c] = n_manual
                else:
                    metas[v][c] = round((campos_pct[v][c] / 100) * amostra_esperada)

        tolerancia_n_abs = round((tolerancia_pct / 100) * amostra_esperada)
        tolerancia_n = {
            v: {c: tolerancia_n_abs for c in metas[v]} for v in metas
        }

        with st.spinner("Calculando exclusões..."):
            resultado = calcular_exclusoes(
                dados, id_var, metas, tolerancia_n, base_minima=int(base_minima)
            )

        st.session_state["excl_resultado"] = {
            "ids_excluir": resultado["ids_excluir"],
            "resumo": resultado["resumo"],
            "id_var": id_var,
        }

    _exibir_resultado_exclusoes()


def _exibir_resultado_exclusoes():
    """Renderiza o resultado guardado em session_state (fora do form, para
    permitir baixar arquivos sem que o clique do botão de download reinicie
    o cálculo)."""
    resultado = st.session_state.get("excl_resultado")
    if not resultado:
        return

    ids_excluir = resultado["ids_excluir"]
    resumo = resultado["resumo"]
    id_var = resultado["id_var"]

    st.markdown("---")
    st.subheader("Resultado")
    st.write(f"**{len(ids_excluir)}** entrevista(s) marcada(s) para exclusão.")
    colunas_pct = [c for c in resumo.columns if c.endswith("(%)")]
    st.dataframe(
        resumo.style.format({c: "{:.2f}" for c in colunas_pct}),
        use_container_width=True
    )

    if ids_excluir:
        df_ids = pd.DataFrame({id_var: ids_excluir})
        st.download_button(
            "Baixar lista de IDs a excluir (CSV)",
            data=df_ids.to_csv(index=False).encode("utf-8"),
            file_name="ids_excluir.csv",
            mime="text/csv",
            key="download_excl_csv"
        )

        tipo_id = st.radio(
            f"No SPSS, a variável '{id_var}' é numérica ou texto?",
            ["Detectar automaticamente", "Numérica", "Texto"],
            index=0,
            horizontal=True,
            key="excl_tipo_id_spss",
            help=(
                "Se o SPSS acusar erro 4325 (\"arguments to ANY... must be "
                "either all strings or all numeric\") ao rodar a sintaxe, "
                "troque essa opção — o tipo real da variável no seu banco "
                "SPSS pode ser diferente do que o app detectou no arquivo "
                "que você carregou aqui (é comum IDs longos serem texto no "
                "SPSS mesmo parecendo só números)."
            )
        )
        tratar_como_texto = {
            "Detectar automaticamente": None,
            "Numérica": False,
            "Texto": True,
        }[tipo_id]

        estilo_label = st.radio(
            "Estilo da sintaxe",
            ["ANY() — mais compacto", "OR encadeado — mais compatível"],
            index=0,
            horizontal=True,
            key="excl_estilo_sintaxe",
            help=(
                "Se o erro 4325 persistir mesmo depois de ajustar o tipo do "
                "ID acima, tente o estilo 'OR encadeado' — evita a função "
                "ANY() por completo, usando comparações simples (var = "
                "valor) encadeadas com OR, que é a forma mais básica e "
                "universalmente compatível de fazer isso no SPSS."
            )
        )
        estilo = "any" if estilo_label.startswith("ANY") else "or"

        sintaxe = _gerar_sintaxe_spss(
            id_var, ids_excluir, tratar_como_texto=tratar_como_texto, estilo=estilo
        )
        st.download_button(
            "Baixar sintaxe SPSS (.sps)",
            data=sintaxe.encode("utf-8"),
            file_name="exclusoes.sps",
            mime="text/plain",
            key="download_excl_sps"
        )
        with st.expander("Ver sintaxe SPSS gerada"):
            st.code(sintaxe, language="sql")
    else:
        st.success("Nenhuma exclusão necessária — a amostra já está dentro da tolerância.")
