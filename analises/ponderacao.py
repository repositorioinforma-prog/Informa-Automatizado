"""Interface Streamlit para ponderação de bases SPSS."""
from __future__ import annotations

import hashlib
import io
import os
import re
import tempfile
import zipfile
from collections import OrderedDict
from copy import copy
from datetime import datetime, timezone
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd
import pyreadstat
import streamlit as st
from openpyxl.comments import Comment
from openpyxl.formatting.formatting import ConditionalFormattingList
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from core.ponderacao_math import (
    calcular_razao_simples,
    calcular_raking,
    categorias_validas,
    frequencia_percentual,
    gerar_syntax_spss,
    resumo_pesos,
)
from core.ponderacao_exclusoes_math import (
    gerar_syntax_exclusao_ids,
    sugerir_exclusoes_ponderacao,
)


# Regra apenas visual do registro histórico: diferenças exibidas entre
# -2,10 e +2,10 p.p. permanecem sem preenchimento.
LIMITE_DESTAQUE_VISUAL_PP = 2.10


def _ler_sav(uploaded_file):
    with tempfile.NamedTemporaryFile(delete=False, suffix=".sav") as tmp:
        tmp.write(uploaded_file.getvalue())
        caminho = tmp.name

    try:
        dados, meta = pyreadstat.read_sav(caminho, apply_value_formats=False)
    finally:
        os.unlink(caminho)

    labels = getattr(meta, "variable_value_labels", {}) or {}
    return dados, meta, labels



def _nome_exibicao_variavel(meta, variavel: str) -> str:
    """Usa o rótulo da variável no SPSS; se não houver, preserva o nome técnico."""
    nomes_para_labels = getattr(meta, "column_names_to_labels", {}) or {}
    label = nomes_para_labels.get(variavel)
    if label is not None and str(label).strip():
        return str(label).strip()
    return str(variavel)


def _descricao_variavel(meta, variavel: str) -> str:
    label = _nome_exibicao_variavel(meta, variavel)
    return f"{variavel} — {label}" if label != str(variavel) else str(variavel)


def _sugerir_variaveis(dados: pd.DataFrame, meta) -> list[str]:
    """Sugere até cinco margens sem depender de um nome fixo para região."""
    colunas = list(dados.columns)
    escolhidas: list[str] = []

    grupos = [
        (("sexo", "genero", "gênero"), ("P1",)),
        (("faixa etaria", "faixa etária", "idade"), ("P2",)),
        (("renda",), ("P3",)),
        (("regiao", "região", "mesorregiao", "mesorregião", "rpa"), ("P7_C",)),
        (("voto", "presidente", "2022", "nacional"), ("VN",)),
    ]

    for palavras, legados in grupos:
        candidato = None
        for legado in legados:
            if legado in colunas and legado not in escolhidas:
                candidato = legado
                break
        if candidato is None:
            for variavel in colunas:
                if variavel in escolhidas:
                    continue
                texto = f"{variavel} {_nome_exibicao_variavel(meta, variavel)}".casefold()
                if any(p.casefold() in texto for p in palavras):
                    candidato = variavel
                    break
        if candidato is not None:
            escolhidas.append(candidato)

    return escolhidas[:5]


def _nome_arquivo_seguro(texto: str) -> str:
    nome = re.sub(r"[^0-9A-Za-zÀ-ÖØ-öø-ÿ._ -]+", "_", str(texto)).strip(" ._")
    nome = re.sub(r"\s+", "_", nome)
    return nome[:100] or "ponderacao"


def _sugerir_variavel_id(dados: pd.DataFrame) -> str | None:
    """Sugere um identificador único para a syntax de exclusão."""
    if dados.empty:
        return None

    nomes_preferidos = (
        "id",
        "id_entrevista",
        "identificador",
        "codigo",
        "código",
        "nquest",
        "nr",
        "respondentid",
        "respondent_id",
    )
    normalizadas = {str(c).strip().casefold(): c for c in dados.columns}
    for nome in nomes_preferidos:
        coluna = normalizadas.get(nome.casefold())
        if coluna is not None and dados[coluna].notna().all() and dados[coluna].is_unique:
            return coluna

    for coluna in dados.columns:
        serie = dados[coluna]
        texto = str(coluna).strip().casefold()
        if "id" in texto and serie.notna().all() and serie.is_unique:
            return coluna

    for coluna in dados.columns:
        serie = dados[coluna]
        if serie.notna().all() and serie.is_unique:
            return coluna
    return None

def _calcular_ponderacao_para_teste_exclusao(
    dados: pd.DataFrame,
    alvos,
    labels_por_variavel: dict,
    metodo_ui: str,
    tolerancia: float,
    max_iter: int,
):
    if metodo_ui.startswith("Método atual"):
        return calcular_razao_simples(
            dados,
            alvos,
            labels=labels_por_variavel,
            tolerancia_pp=float(tolerancia),
        )
    return calcular_raking(
        dados,
        alvos,
        labels=labels_por_variavel,
        tolerancia_pp=float(tolerancia),
        max_iteracoes=int(max_iter),
    )


def _refinar_prefixo_exclusoes_pelo_peso(
    dados: pd.DataFrame,
    sugestao,
    resultado_atual,
    alvos,
    labels_por_variavel: dict,
    metodo_ui: str,
    tolerancia: float,
    max_iter: int,
):
    """Procura o melhor N dentro da faixa permitida, após reponderar cada cenário avaliado.

    O N mínimo funciona como piso, não como obrigação. Para bases grandes, a busca
    usa uma varredura ampla e depois refina a vizinhança do melhor ponto encontrado.
    Em empates, preserva mais entrevistas (menor número de exclusões).
    """
    total = len(sugestao.indices_excluir)
    melhor_k = 0
    melhor_resultado = resultado_atual
    melhor_valor = float(resultado_atual.maior_diferenca_pp)

    if total == 0:
        return melhor_k, melhor_resultado

    avaliados: dict[int, tuple[float, object]] = {
        0: (melhor_valor, resultado_atual)
    }

    def avaliar(k: int) -> None:
        k = max(0, min(int(k), total))
        if k in avaliados:
            return
        dados_teste = dados.drop(index=sugestao.indices_excluir[:k])
        try:
            teste = _calcular_ponderacao_para_teste_exclusao(
                dados_teste,
                alvos,
                labels_por_variavel,
                metodo_ui,
                tolerancia,
                max_iter,
            )
        except Exception:
            return
        valor = float(teste.maior_diferenca_pp)
        if np.isfinite(valor):
            avaliados[k] = (valor, teste)

    def melhor_avaliado() -> tuple[int, float, object]:
        candidatos = [
            (k, valor_resultado[0], valor_resultado[1])
            for k, valor_resultado in avaliados.items()
            if np.isfinite(valor_resultado[0])
        ]
        # Primeiro menor diferença; em empate, menor k para preservar mais casos.
        return min(candidatos, key=lambda item: (item[1], item[0]))

    metodo_simples = metodo_ui.startswith("Método atual")
    limite_varredura_completa = 120 if metodo_simples else 24

    if total <= limite_varredura_completa:
        for k in range(1, total + 1):
            avaliar(k)
    else:
        # Etapa 1: cobre toda a faixa do N atual até o piso informado.
        quantidade = 60 if metodo_simples else 12
        pontos = sorted(
            set(
                max(1, min(total, int(round(x))))
                for x in np.linspace(1, total, num=quantidade, endpoint=True)
            )
        )
        for k in pontos:
            avaliar(k)

        # Etapa 2: refina a vizinhança do melhor ponto da varredura ampla.
        if metodo_simples:
            melhor_k_tmp, _, _ = melhor_avaliado()
            grade = sorted(set([0, total, *pontos]))
            pos = grade.index(melhor_k_tmp) if melhor_k_tmp in grade else 0
            esquerda = grade[max(0, pos - 1)]
            direita = grade[min(len(grade) - 1, pos + 1)]
            for k in range(esquerda + 1, direita):
                avaliar(k)
        else:
            # Raking é mais caro; estreita a faixa em duas rodadas antes da busca local.
            for _ in range(2):
                melhor_k_tmp, _, _ = melhor_avaliado()
                grade = sorted(set([0, total, *avaliados.keys()]))
                pos = grade.index(melhor_k_tmp)
                esquerda = grade[max(0, pos - 1)]
                direita = grade[min(len(grade) - 1, pos + 1)]
                if direita - esquerda <= 8:
                    break
                extras = sorted(
                    set(
                        int(round(x))
                        for x in np.linspace(esquerda, direita, num=7, endpoint=True)
                    )
                )
                for k in extras:
                    avaliar(k)
            melhor_k_tmp, _, _ = melhor_avaliado()
            grade = sorted(set([0, total, *avaliados.keys()]))
            pos = grade.index(melhor_k_tmp)
            esquerda = grade[max(0, pos - 1)]
            direita = grade[min(len(grade) - 1, pos + 1)]
            if direita - esquerda <= 10:
                for k in range(esquerda + 1, direita):
                    avaliar(k)

    melhor_k, melhor_valor, melhor_resultado = melhor_avaliado()
    return melhor_k, melhor_resultado


def _label_categoria(labels_var, codigo):
    if codigo in labels_var:
        return str(labels_var[codigo])
    # Em SAV, códigos inteiros frequentemente chegam como float.
    if isinstance(codigo, (float, np.floating)) and float(codigo).is_integer():
        inteiro = int(codigo)
        if inteiro in labels_var:
            return str(labels_var[inteiro])
    return str(codigo)


def _chave_ordenacao_codigo(codigo):
    """Ordena códigos numéricos antes dos textuais, sempre em ordem crescente."""
    try:
        numero = float(codigo)
        if np.isfinite(numero):
            return (0, numero, "")
    except (TypeError, ValueError):
        pass
    return (1, 0.0, str(codigo).casefold())


def _tabela_editor(dados, variavel, labels_var):
    categorias = sorted(
        categorias_validas(dados[variavel]),
        key=_chave_ordenacao_codigo,
    )
    freq = frequencia_percentual(dados[variavel], categorias=categorias)
    return pd.DataFrame(
        {
            "Código": freq["codigo"],
            "Categoria": [_label_categoria(labels_var, c) for c in freq["codigo"]],
            "Frequência": freq["frequencia"].astype(int),
            "% Amostra": freq["percentual"],
            "% Universo": [np.nan] * len(freq),
        }
    )


def _parse_percentuais_colados(texto: str) -> list[float]:
    """Lê percentuais copiados de uma coluna/linha do Excel.

    Aceita quebra de linha, TAB, ponto e vírgula ou barra vertical como separadores,
    além de decimal com vírgula ou ponto e o símbolo opcional de porcentagem.
    """
    tokens = [
        token.strip()
        for token in re.split(r"[\r\n\t;|]+", str(texto).strip())
        if token.strip()
    ]

    valores = []
    for token in tokens:
        limpo = token.replace("%", "").strip().replace(" ", "")
        if not limpo:
            continue

        # Trata tanto 12,62 quanto 12.62. Se os dois separadores aparecerem,
        # considera o último deles como separador decimal.
        if "," in limpo and "." in limpo:
            if limpo.rfind(",") > limpo.rfind("."):
                limpo = limpo.replace(".", "").replace(",", ".")
            else:
                limpo = limpo.replace(",", "")
        else:
            limpo = limpo.replace(",", ".")

        try:
            valor = float(limpo)
        except ValueError as exc:
            raise ValueError(f"Valor de universo inválido: '{token}'.") from exc

        if not 0 <= valor <= 100:
            raise ValueError(
                f"O percentual {valor:g} está fora do intervalo de 0 a 100."
            )
        valores.append(valor)

    if not valores:
        raise ValueError("Cole pelo menos um percentual do universo.")
    return valores


def _assinatura_editor(base_editor: pd.DataFrame):
    """Identifica a estrutura e a frequência da base para detectar mudanças na amostra."""
    return tuple(
        (
            repr(linha["Código"]),
            str(linha["Categoria"]),
            int(linha["Frequência"]),
            round(float(linha["% Amostra"]), 8),
        )
        for _, linha in base_editor.iterrows()
    )


def _assinatura_calculo_ponderacao(
    uploaded,
    variaveis,
    editores,
    metodo_ui: str,
    tolerancia: float,
    max_iter: int,
    identificacao: str,
):
    """Identifica exatamente a configuração que originou um resultado.

    A assinatura permite preservar o cálculo no session_state durante reruns do
    Streamlit (inclusive os causados por downloads), sem reaproveitar resultado
    antigo quando a base, as metas, o método ou a identificação forem alterados.
    """
    arquivo_hash = hashlib.sha256(uploaded.getvalue()).hexdigest()
    metas = []

    for variavel in variaveis:
        tabela = editores[variavel]
        for _, linha in tabela.iterrows():
            valor = pd.to_numeric(pd.Series([linha["% Universo"]]), errors="coerce").iloc[0]
            valor_assinatura = None if pd.isna(valor) else round(float(valor), 10)
            metas.append(
                (
                    str(variavel),
                    _chave_codigo_perfil(linha["Código"]),
                    valor_assinatura,
                )
            )

    return (
        arquivo_hash,
        tuple(str(v) for v in variaveis),
        tuple(metas),
        str(metodo_ui),
        round(float(tolerancia), 10),
        int(max_iter),
        str(identificacao).strip(),
    )


def _chave_codigo_perfil(codigo) -> str:
    """Cria uma chave estável para casar códigos SPSS entre SAV, sessão e Excel."""
    if codigo is None or (isinstance(codigo, float) and np.isnan(codigo)):
        return "NULL"

    if isinstance(codigo, (int, float, np.integer, np.floating)) and not isinstance(codigo, bool):
        numero = float(codigo)
        if np.isfinite(numero):
            if numero.is_integer():
                return f"N:{int(numero)}"
            return f"N:{format(numero, '.15g')}"

    return f"S:{str(codigo).strip()}"


def _perfil_vazio(origem: str = "Sessão atual") -> dict:
    return {"origem": origem, "variaveis": OrderedDict()}


def _perfil_sessao() -> dict:
    perfil = st.session_state.get("ponderacao_perfil_universo")
    if not isinstance(perfil, dict) or "variaveis" not in perfil:
        perfil = _perfil_vazio()
        st.session_state["ponderacao_perfil_universo"] = perfil
    return perfil


def _importar_perfil_universo_xlsx(conteudo: bytes, nome_arquivo: str) -> dict:
    """Importa metas de Universo de um modelo preenchido ou registro histórico."""
    try:
        wb = openpyxl.load_workbook(io.BytesIO(conteudo), data_only=True, read_only=True)
    except Exception as exc:
        raise ValueError(f"Não foi possível abrir o Excel informado: {exc}") from exc

    if "Universo" in wb.sheetnames:
        ws = wb["Universo"]
        nome_aba = "Universo"
    elif "Registro" in wb.sheetnames:
        ws = wb["Registro"]
        nome_aba = "Registro"
    else:
        raise ValueError(
            "O arquivo precisa possuir a aba 'Universo' (modelo preenchido) ou "
            "'Registro' (ponderação histórica gerada pelo módulo)."
        )

    cabecalho = None
    mapa_colunas = {}
    obrigatorias = {"Variável SPSS", "Código SPSS", "% Universo"}

    for r in range(1, min(ws.max_row, 100) + 1):
        valores = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        mapa = {str(v).strip(): idx + 1 for idx, v in enumerate(valores) if v is not None}
        if obrigatorias.issubset(mapa):
            cabecalho = r
            mapa_colunas = mapa
            break

    if cabecalho is None:
        raise ValueError(
            f"Não encontrei na aba '{nome_aba}' as colunas Variável SPSS, Código SPSS e % Universo."
        )

    perfil = _perfil_vazio(f"Importado de {nome_arquivo}")
    col_var = mapa_colunas["Variável SPSS"]
    col_codigo = mapa_colunas["Código SPSS"]
    col_universo = mapa_colunas["% Universo"]
    col_rotulo = mapa_colunas.get("Rótulo da variável")
    col_categoria = mapa_colunas.get("Categoria")

    total_alvos = 0
    for r in range(cabecalho + 1, ws.max_row + 1):
        variavel = ws.cell(r, col_var).value
        codigo = ws.cell(r, col_codigo).value
        universo = ws.cell(r, col_universo).value
        if variavel is None and codigo is None and universo is None:
            continue
        if variavel is None or codigo is None or universo in (None, ""):
            continue

        try:
            percentual = float(universo)
        except (TypeError, ValueError) as exc:
            raise ValueError(
                f"% Universo inválido na linha {r} da aba {nome_aba}: {universo!r}."
            ) from exc
        if not np.isfinite(percentual) or not 0 <= percentual <= 100:
            raise ValueError(
                f"% Universo fora de 0 a 100 na linha {r} da aba {nome_aba}: {percentual:g}."
            )

        nome_var = str(variavel).strip()
        rotulo = ws.cell(r, col_rotulo).value if col_rotulo else None
        categoria = ws.cell(r, col_categoria).value if col_categoria else None
        chave = _chave_codigo_perfil(codigo)

        entrada_var = perfil["variaveis"].setdefault(
            nome_var,
            {
                "rotulo": str(rotulo).strip() if rotulo not in (None, "") else nome_var,
                "alvos": OrderedDict(),
            },
        )
        anterior = entrada_var["alvos"].get(chave)
        if anterior is not None and not np.isclose(anterior["percentual"], percentual):
            raise ValueError(
                f"O arquivo contém dois Universos diferentes para {nome_var}, código {codigo}."
            )
        entrada_var["alvos"][chave] = {
            "codigo": codigo,
            "categoria": str(categoria).strip() if categoria not in (None, "") else str(codigo),
            "percentual": percentual,
        }
        total_alvos += 1

    if not total_alvos:
        raise ValueError(f"A aba {nome_aba} não contém nenhum percentual de Universo preenchido.")

    return perfil


def _gerar_modelo_universo_excel(
    dados: pd.DataFrame,
    meta,
    value_labels: dict,
    variaveis: list[str],
    arquivo_origem: str,
) -> bytes:
    """Gera um modelo reutilizável do Universo a partir da estrutura do SAV atual."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Universo"
    ws.sheet_view.showGridLines = False

    azul = "1F4E78"
    azul_claro = "D9EAF7"
    amarelo = "FFF2CC"
    verde = "E2F0D9"
    cinza = "E7E6E6"
    borda_fina = Side(style="thin", color="B7B7B7")
    borda = Border(left=borda_fina, right=borda_fina, top=borda_fina, bottom=borda_fina)

    ws.merge_cells("A1:E1")
    ws["A1"] = "MODELO DE UNIVERSO - PONDERAÇÃO"
    ws["A1"].font = Font(name="Arial", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor=azul)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    ws["A2"] = "Arquivo SPSS de origem"
    ws["B2"] = arquivo_origem
    ws["A3"] = "Orientação"
    ws["B3"] = (
        "Preencha somente a coluna % Universo. Não altere Variável SPSS nem Código SPSS, "
        "pois o Analítico usa esses dois campos para reaplicar o Universo com segurança."
    )
    ws.merge_cells("B2:E2")
    ws.merge_cells("B3:E3")
    for celula in (ws["A2"], ws["A3"]):
        celula.font = Font(name="Arial", size=10, bold=True)
        celula.fill = PatternFill("solid", fgColor=cinza)
        celula.border = borda
    for celula in (ws["B2"], ws["B3"]):
        celula.font = Font(name="Arial", size=10)
        celula.alignment = Alignment(vertical="center", wrap_text=True)
        celula.border = borda
    ws.row_dimensions[3].height = 34

    cabecalho = 5
    colunas = [
        "Variável SPSS",
        "Rótulo da variável",
        "Código SPSS",
        "Categoria",
        "% Universo",
    ]
    for c, texto in enumerate(colunas, 1):
        celula = ws.cell(cabecalho, c, texto)
        celula.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        celula.fill = PatternFill("solid", fgColor=azul)
        celula.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        celula.border = borda
    ws.row_dimensions[cabecalho].height = 30

    linha = cabecalho + 1
    faixas_variaveis = []
    for variavel in variaveis:
        tabela = _tabela_editor(
            dados,
            variavel,
            (value_labels or {}).get(variavel, {}) or {},
        )
        rotulo = _nome_exibicao_variavel(meta, variavel)
        inicio = linha
        for _, item in tabela.iterrows():
            valores = [
                variavel,
                rotulo,
                item["Código"],
                item["Categoria"],
                None,
            ]
            for c, valor in enumerate(valores, 1):
                celula = ws.cell(linha, c, valor)
                celula.font = Font(name="Arial", size=10)
                celula.border = borda
                celula.alignment = Alignment(vertical="center", wrap_text=True)
                if c <= 4:
                    celula.fill = PatternFill("solid", fgColor=verde)
                else:
                    celula.fill = PatternFill("solid", fgColor=amarelo)
                    celula.font = Font(name="Arial", size=10, color="0000FF")
                    celula.number_format = "0.00"
            linha += 1
        fim = linha - 1
        if fim >= inicio:
            faixas_variaveis.append((variavel, rotulo, inicio, fim))

    if linha > cabecalho + 1:
        validacao = DataValidation(
            type="decimal",
            operator="between",
            formula1="0",
            formula2="100",
            allow_blank=True,
        )
        validacao.error = "Informe um percentual entre 0 e 100."
        validacao.errorTitle = "Percentual inválido"
        validacao.prompt = "Preencha o percentual do Universo desta categoria."
        validacao.promptTitle = "% Universo"
        ws.add_data_validation(validacao)
        validacao.add(f"E{cabecalho + 1}:E{linha - 1}")

    ws.freeze_panes = f"A{cabecalho + 1}"
    ws.auto_filter.ref = f"A{cabecalho}:E{max(cabecalho, linha - 1)}"
    larguras = {"A": 19, "B": 38, "C": 14, "D": 34, "E": 16}
    for coluna, largura in larguras.items():
        ws.column_dimensions[coluna].width = largura
    for r in range(cabecalho + 1, linha):
        _ajustar_altura_linha(
            ws,
            r,
            [
                (ws.cell(r, 2).value, "B", 1.05),
                (ws.cell(r, 4).value, "D", 1.05),
            ],
            minima=18.0,
        )

    conf = wb.create_sheet("Conferência")
    conf.sheet_view.showGridLines = False
    conf.merge_cells("A1:D1")
    conf["A1"] = "CONFERÊNCIA DO UNIVERSO"
    conf["A1"].font = Font(name="Arial", size=14, bold=True, color="FFFFFF")
    conf["A1"].fill = PatternFill("solid", fgColor=azul)
    conf["A1"].alignment = Alignment(horizontal="center", vertical="center")
    headers_conf = ["Variável SPSS", "Rótulo da variável", "Total Universo (%)", "Status"]
    for c, texto in enumerate(headers_conf, 1):
        celula = conf.cell(3, c, texto)
        celula.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        celula.fill = PatternFill("solid", fgColor=azul)
        celula.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        celula.border = borda
    for idx, (variavel, rotulo, _inicio, _fim) in enumerate(faixas_variaveis, start=4):
        conf.cell(idx, 1, variavel)
        conf.cell(idx, 2, rotulo)
        conf.cell(idx, 3, f'=SUMIF(Universo!$A:$A,A{idx},Universo!$E:$E)')
        conf.cell(idx, 4, f'=IF(ABS(C{idx}-100)<=0.1,"OK","REVISAR")')
        for c in range(1, 5):
            celula = conf.cell(idx, c)
            celula.font = Font(name="Arial", size=10)
            celula.border = borda
            celula.alignment = Alignment(vertical="center", wrap_text=True)
        conf.cell(idx, 3).number_format = "0.00"
    for coluna, largura in {"A": 19, "B": 38, "C": 20, "D": 14}.items():
        conf.column_dimensions[coluna].width = largura
    conf.freeze_panes = None

    try:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
        wb.calculation.calcMode = "auto"
    except Exception:
        pass

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _gerar_relatorio_sugestao_exclusoes_excel(
    sugestao,
    id_var: str,
    identificacao: str,
    metodo: str,
    tolerancia_pp: float,
    maior_ponderada_antes_pp: float,
    maior_ponderada_depois_pp: float | None,
    objetivo_exclusao: str = "Encontrar a melhor ponderação até o N mínimo",
) -> bytes:
    """Documenta a sugestão de exclusões sem alterar a base original."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resumo"
    ws.sheet_view.showGridLines = False

    azul = "1F4E78"
    amarelo = "FFF2CC"
    cinza = "E7E6E6"
    vermelho = "F4CCCC"
    verde = "D9EAD3"
    borda = Border(
        left=Side(style="thin", color="B7B7B7"),
        right=Side(style="thin", color="B7B7B7"),
        top=Side(style="thin", color="B7B7B7"),
        bottom=Side(style="thin", color="B7B7B7"),
    )

    ws.merge_cells("A1:D1")
    ws["A1"] = "Sugestão de exclusões para teste de ponderação"
    ws["A1"].font = Font(name="Arial", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor=azul)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 25

    modo_n_exato = objetivo_exclusao.startswith("Chegar exatamente")
    resumo_itens = [
        ("Projeto", identificacao),
        ("Variável de ID", id_var),
        ("Método de ponderação", metodo),
        ("Objetivo das exclusões", objetivo_exclusao),
        ("Tolerância", float(tolerancia_pp)),
        ("N inicial", int(sugestao.n_inicial)),
    ]
    if modo_n_exato:
        resumo_itens.extend([
            ("N-alvo exato informado", int(sugestao.n_minimo)),
            ("Máximo de exclusões até o N-alvo", int(sugestao.limite_exclusoes)),
            ("N-alvo atingido", "SIM" if getattr(sugestao, "atingiu_n_alvo", False) else "NÃO"),
            ("Exclusões que faltariam para o N-alvo", int(getattr(sugestao, "faltam_exclusoes_para_alvo", 0))),
        ])
    else:
        resumo_itens.extend([
            ("N mínimo permitido", int(sugestao.n_minimo)),
            ("Faixa de N permitida", f"{int(sugestao.n_inicial)} até {int(sugestao.n_minimo)}"),
            ("Máximo de exclusões permitidas", int(sugestao.limite_exclusoes)),
            ("N recomendado pelo assistente", int(sugestao.n_final)),
        ])
    resumo_itens.extend([
        ("Base mínima protegida por categoria", int(getattr(sugestao, "base_minima_categoria", 0))),
        ("Exclusões sugeridas", len(sugestao.ids_excluir)),
        ("N após sugestão", int(sugestao.n_final)),
        ("Maior diferença bruta antes (p.p.)", float(sugestao.maior_diferenca_bruta_antes_pp)),
        ("Maior diferença bruta após (p.p.)", float(sugestao.maior_diferenca_bruta_depois_pp)),
        ("Maior diferença ponderada antes (p.p.)", float(maior_ponderada_antes_pp)),
        (
            "Maior diferença ponderada estimada após (p.p.)",
            None if maior_ponderada_depois_pp is None else float(maior_ponderada_depois_pp),
        ),
        ("Motivo da parada", sugestao.motivo_parada),
    ])
    for r, (rotulo, valor) in enumerate(resumo_itens, start=3):
        ws.cell(r, 1, rotulo)
        ws.cell(r, 2, valor)
        ws.cell(r, 1).font = Font(name="Arial", size=10, bold=True)
        ws.cell(r, 1).fill = PatternFill("solid", fgColor=cinza)
        for c in (1, 2):
            ws.cell(r, c).border = borda
            ws.cell(r, c).alignment = Alignment(vertical="center", wrap_text=True)
            ws.cell(r, c).font = Font(name="Arial", size=10, bold=(c == 1))
    for r in range(3, 3 + len(resumo_itens)):
        if isinstance(ws.cell(r, 2).value, float):
            ws.cell(r, 2).number_format = "0.00"

    nota_row = len(resumo_itens) + 5
    ws.merge_cells(start_row=nota_row, start_column=1, end_row=nota_row + 2, end_column=4)
    ws.cell(nota_row, 1,
        "ATENÇÃO: esta é uma sugestão heurística para apoiar testes. O Analítico não exclui casos automaticamente. "
        "Revise os IDs, salve uma cópia do banco e recalcule a ponderação depois de aplicar qualquer exclusão."
    )
    ws.cell(nota_row, 1).fill = PatternFill("solid", fgColor=amarelo)
    ws.cell(nota_row, 1).alignment = Alignment(wrap_text=True, vertical="center")
    ws.cell(nota_row, 1).font = Font(name="Arial", size=10, bold=True)
    ws.cell(nota_row, 1).border = borda

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.freeze_panes = None

    margens = wb.create_sheet("Margens")
    margens.sheet_view.showGridLines = False
    colunas = [
        "variavel", "codigo", "categoria", "n_antes", "pct_antes",
        "pct_universo", "diferenca_antes_pp", "n_depois", "pct_depois",
        "diferenca_depois_pp", "excesso_n_aprox", "status",
    ]
    for c, nome in enumerate(colunas, start=1):
        cel = margens.cell(1, c, nome)
        cel.fill = PatternFill("solid", fgColor=azul)
        cel.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        cel.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cel.border = borda
    for r, (_, linha) in enumerate(sugestao.resumo_margens.iterrows(), start=2):
        for c, nome in enumerate(colunas, start=1):
            valor = linha.get(nome)
            cel = margens.cell(r, c, valor)
            cel.font = Font(name="Arial", size=10)
            cel.border = borda
            cel.alignment = Alignment(vertical="center", wrap_text=True)
            if nome in {"pct_antes", "pct_universo", "diferenca_antes_pp", "pct_depois", "diferenca_depois_pp", "excesso_n_aprox"}:
                cel.number_format = "0.00"
        if str(linha.get("status")) == "REVISAR":
            margens.cell(r, 12).fill = PatternFill("solid", fgColor=vermelho)
        else:
            margens.cell(r, 12).fill = PatternFill("solid", fgColor=verde)
    larguras = [16, 12, 34, 12, 14, 14, 18, 12, 14, 18, 16, 12]
    for idx, largura in enumerate(larguras, start=1):
        margens.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = largura
    margens.freeze_panes = "A2"

    ids = wb.create_sheet("IDs sugeridos")
    ids.sheet_view.showGridLines = False
    ids.cell(1, 1, id_var)
    ids.cell(1, 1).fill = PatternFill("solid", fgColor=azul)
    ids.cell(1, 1).font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    ids.cell(1, 1).alignment = Alignment(horizontal="center")
    ids.cell(1, 1).border = borda
    for r, valor in enumerate(sugestao.ids_excluir, start=2):
        ids.cell(r, 1, valor)
        ids.cell(r, 1).font = Font(name="Arial", size=10)
        ids.cell(r, 1).border = borda
    ids.column_dimensions["A"].width = 26
    ids.freeze_panes = "A2"

    sobras = wb.create_sheet("Sobras remanescentes")
    sobras.sheet_view.showGridLines = False
    if sugestao.sobras_remanescentes.empty:
        sobras["A1"] = "Nenhuma categoria acima da tolerância bruta após a sugestão."
    else:
        sob_cols = [
            "variavel", "codigo", "categoria", "n_depois", "pct_depois",
            "pct_universo", "diferenca_depois_pp", "excesso_n_aprox",
        ]
        for c, nome in enumerate(sob_cols, start=1):
            cel = sobras.cell(1, c, nome)
            cel.fill = PatternFill("solid", fgColor=azul)
            cel.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
            cel.border = borda
        for r, (_, linha) in enumerate(sugestao.sobras_remanescentes.iterrows(), start=2):
            for c, nome in enumerate(sob_cols, start=1):
                cel = sobras.cell(r, c, linha.get(nome))
                cel.font = Font(name="Arial", size=10)
                cel.border = borda
                cel.alignment = Alignment(vertical="center", wrap_text=True)
                if nome in {"pct_depois", "pct_universo", "diferenca_depois_pp", "excesso_n_aprox"}:
                    cel.number_format = "0.00"
        for idx, largura in enumerate([16, 12, 34, 12, 14, 14, 18, 16], start=1):
            sobras.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = largura
    sobras.freeze_panes = None

    protecoes = wb.create_sheet("Proteções")
    protecoes.sheet_view.showGridLines = False
    restricoes_df = getattr(sugestao, "restricoes_ativas", pd.DataFrame())
    if restricoes_df is None or restricoes_df.empty:
        protecoes["A1"] = "Nenhuma proteção ficou ativa no ponto final da simulação."
    else:
        prot_cols = ["variavel", "codigo", "categoria", "n_final", "base_minima", "motivo"]
        for c, nome in enumerate(prot_cols, start=1):
            cel = protecoes.cell(1, c, nome)
            cel.fill = PatternFill("solid", fgColor=azul)
            cel.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
            cel.border = borda
            cel.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for r, (_, linha) in enumerate(restricoes_df.iterrows(), start=2):
            for c, nome in enumerate(prot_cols, start=1):
                cel = protecoes.cell(r, c, linha.get(nome))
                cel.font = Font(name="Arial", size=10)
                cel.border = borda
                cel.alignment = Alignment(vertical="center", wrap_text=True)
        for idx, largura in enumerate([16, 12, 34, 12, 14, 48], start=1):
            protecoes.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = largura
    protecoes.freeze_panes = None

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _gerar_pacote_ponderacao_zip(
    nome_registro: str,
    excel_historico: bytes | None,
    excel_calculo: bytes | None,
    syntax: str,
    excel_tecnico: bytes | None,
    modelo_universo: bytes | None = None,
    relatorio_exclusoes: bytes | None = None,
    ids_exclusoes_csv: bytes | None = None,
    syntax_exclusoes: str | None = None,
) -> bytes:
    """Agrupa os entregáveis em um ZIP que, ao extrair, cria uma pasta do projeto."""
    pasta = f"Ponderacao_{nome_registro}/"
    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
        if excel_historico is not None:
            zf.writestr(f"{pasta}Ponderacao_{nome_registro}.xlsx", excel_historico)
        if excel_calculo is not None:
            zf.writestr(
                f"{pasta}Calculo_da_Ponderacao_{nome_registro}.xlsx",
                excel_calculo,
            )
        zf.writestr(f"{pasta}Pesos_{nome_registro}.sps", syntax.encode("utf-8"))
        if excel_tecnico is not None:
            zf.writestr(
                f"{pasta}Relatorio_tecnico_{nome_registro}.xlsx",
                excel_tecnico,
            )
        if modelo_universo is not None:
            zf.writestr(
                f"{pasta}Modelo_Universo_{nome_registro}.xlsx",
                modelo_universo,
            )
        if relatorio_exclusoes is not None:
            zf.writestr(
                f"{pasta}Sugestao_Exclusoes_{nome_registro}.xlsx",
                relatorio_exclusoes,
            )
        if ids_exclusoes_csv is not None:
            zf.writestr(
                f"{pasta}IDs_Exclusao_{nome_registro}.csv",
                ids_exclusoes_csv,
            )
        if syntax_exclusoes:
            zf.writestr(
                f"{pasta}Exclusoes_{nome_registro}.sps",
                syntax_exclusoes.encode("utf-8"),
            )
    return buffer.getvalue()


def _compatibilidade_perfil(base_editor: pd.DataFrame, perfil: dict, variavel: str) -> dict:
    entrada = (perfil.get("variaveis") or {}).get(variavel)
    alvos = entrada.get("alvos", {}) if entrada else {}
    atuais = OrderedDict(
        (
            _chave_codigo_perfil(linha["Código"]),
            {"codigo": linha["Código"], "categoria": str(linha["Categoria"])},
        )
        for _, linha in base_editor.iterrows()
    )

    casados = [chave for chave in atuais if chave in alvos]
    sem_universo = [atuais[chave] for chave in atuais if chave not in alvos]
    ausentes_na_amostra = [alvos[chave] for chave in alvos if chave not in atuais]
    labels_divergentes = []
    for chave in casados:
        categoria_atual = atuais[chave]["categoria"].strip().casefold()
        categoria_salva = str(alvos[chave].get("categoria", "")).strip().casefold()
        if categoria_salva and categoria_atual != categoria_salva:
            labels_divergentes.append(
                {
                    "codigo": atuais[chave]["codigo"],
                    "atual": atuais[chave]["categoria"],
                    "salva": alvos[chave].get("categoria", ""),
                }
            )

    return {
        "casados": len(casados),
        "total_atual": len(atuais),
        "sem_universo": sem_universo,
        "ausentes_na_amostra": ausentes_na_amostra,
        "labels_divergentes": labels_divergentes,
    }


def _aplicar_perfil_no_editor(base_editor: pd.DataFrame, perfil: dict, variavel: str):
    atualizado = base_editor.copy()
    entrada = (perfil.get("variaveis") or {}).get(variavel)
    alvos = entrada.get("alvos", {}) if entrada else {}
    for idx, linha in atualizado.iterrows():
        alvo = alvos.get(_chave_codigo_perfil(linha["Código"]))
        if alvo is not None:
            atualizado.at[idx, "% Universo"] = float(alvo["percentual"])
    return atualizado, _compatibilidade_perfil(base_editor, perfil, variavel)


def _sincronizar_editor_com_perfil(
    perfil: dict, variavel: str, rotulo: str, editor: pd.DataFrame
) -> None:
    """Atualiza os códigos visíveis sem apagar alvos de categorias ausentes nesta amostra."""
    variaveis = perfil.setdefault("variaveis", OrderedDict())
    entrada = variaveis.setdefault(
        variavel, {"rotulo": rotulo, "alvos": OrderedDict()}
    )
    entrada["rotulo"] = rotulo
    alvos = entrada.setdefault("alvos", OrderedDict())

    for _, linha in editor.iterrows():
        chave = _chave_codigo_perfil(linha["Código"])
        valor = pd.to_numeric(pd.Series([linha["% Universo"]]), errors="coerce").iloc[0]
        if pd.isna(valor):
            alvos.pop(chave, None)
            continue
        alvos[chave] = {
            "codigo": linha["Código"],
            "categoria": str(linha["Categoria"]),
            "percentual": float(valor),
        }

    if not alvos:
        variaveis.pop(variavel, None)


def _remover_variavel_do_perfil(perfil: dict, variavel: str) -> None:
    (perfil.get("variaveis") or {}).pop(variavel, None)


def _contar_alvos_perfil(perfil: dict) -> int:
    return sum(
        len(entrada.get("alvos", {}))
        for entrada in (perfil.get("variaveis") or {}).values()
    )


def _alvos_do_editor(editor: pd.DataFrame, variavel: str) -> OrderedDict:
    if editor["% Universo"].isna().any():
        faltantes = editor.loc[editor["% Universo"].isna(), "Categoria"].astype(str).tolist()
        raise ValueError(
            f"Preencha o % Universo de todas as categorias de '{variavel}'. "
            f"Faltando: {', '.join(faltantes)}"
        )

    alvos = OrderedDict()
    for _, linha in editor.iterrows():
        alvos[linha["Código"]] = float(linha["% Universo"])
    return alvos


def _nome_aba_seguro(nome: str, usados: set[str]) -> str:
    base = re.sub(r"[\\/*?:\[\]]", "_", str(nome)).strip() or "Variavel"
    base = base[:31]
    candidato = base
    contador = 2
    while candidato in usados:
        sufixo = f"_{contador}"
        candidato = f"{base[:31-len(sufixo)]}{sufixo}"
        contador += 1
    usados.add(candidato)
    return candidato


def _gerar_relatorio_excel(resultado, tolerancia_pp: float, total_casos: int) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resumo"
    resumo = resumo_pesos(resultado.pesos)

    linhas = [
        ("Método", resultado.metodo),
        ("Tolerância máxima (p.p.)", tolerancia_pp),
        ("Dentro da tolerância", "SIM" if resultado.convergiu else "NÃO"),
        ("Iterações", resultado.iteracoes),
        ("Maior diferença absoluta (p.p.)", resultado.maior_diferenca_pp),
        ("Casos na base", total_casos),
        ("Casos com peso válido", resumo["casos_com_peso"]),
        ("Soma dos pesos", resumo["soma_pesos"]),
        ("Peso médio", resumo["media"]),
        ("Peso mínimo", resumo["minimo"]),
        ("Peso máximo", resumo["maximo"]),
        ("CV dos pesos", resumo["cv"]),
        ("Efeito de desenho aprox. (1 + CV²)", resumo["efeito_desenho_aprox"]),
        ("N efetivo aprox.", resumo["n_efetivo_aprox"]),
    ]
    for r, (chave, valor) in enumerate(linhas, start=1):
        ws.cell(r, 1, chave)
        ws.cell(r, 2, valor)
        ws.cell(r, 1).font = Font(bold=True)
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 24

    usados = {"Resumo"}
    colunas = [
        ("codigo", "Código"),
        ("categoria", "Categoria"),
        ("frequencia_sem_peso", "Frequência sem peso"),
        ("percentual_sem_peso", "% sem peso"),
        ("percentual_universo", "% universo"),
        ("fator", "Fator de peso"),
        ("frequencia_com_peso", "Frequência com peso"),
        ("percentual_com_peso", "% com peso"),
        ("diferenca_pp", "Diferença (p.p.)"),
        ("diferenca_abs_pp", "|Diferença| (p.p.)"),
        ("status", "Status"),
    ]

    for variavel, bloco in resultado.diagnostico.groupby("variavel", sort=False):
        nome_aba = _nome_aba_seguro(str(variavel), usados)
        aba = wb.create_sheet(nome_aba)
        aba.append([titulo for _, titulo in colunas])
        for c in aba[1]:
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="2563EB")

        for _, linha in bloco.iterrows():
            aba.append([linha[chave] for chave, _ in colunas])

        for row in aba.iter_rows(min_row=2, max_row=aba.max_row):
            for idx in [4, 5, 8, 9, 10]:
                row[idx - 1].number_format = "0.00"
            row[5].number_format = "0.000000"
            status = row[10].value
            if status == "REVISAR":
                row[10].fill = PatternFill("solid", fgColor="FECACA")
            else:
                row[10].fill = PatternFill("solid", fgColor="DCFCE7")

        larguras = [14, 28, 20, 14, 14, 16, 22, 14, 18, 20, 12]
        for idx, largura in enumerate(larguras, start=1):
            aba.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = largura
        aba.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()



def _frequencia_registro(
    serie: pd.Series,
    categorias: list,
    pesos: pd.Series | None = None,
) -> pd.DataFrame:
    """Frequência no formato usado no arquivo histórico (Percent/Valid Percent/Cumulative)."""
    if pesos is None:
        valores = pd.Series(1.0, index=serie.index, dtype="float64")
    else:
        valores = pd.to_numeric(pesos, errors="coerce").reindex(serie.index)

    mascara_peso = valores.notna()
    mascara_valida = mascara_peso & serie.notna()
    total_geral = float(valores.loc[mascara_peso].sum())
    total_valido = float(valores.loc[mascara_valida].sum())

    linhas = []
    acumulado = 0.0
    for categoria in categorias:
        mascara_cat = mascara_valida & (serie == categoria)
        frequencia = float(valores.loc[mascara_cat].sum())
        percentual = (frequencia / total_geral * 100.0) if total_geral else np.nan
        valido = (frequencia / total_valido * 100.0) if total_valido else np.nan
        if np.isfinite(valido):
            acumulado += valido
        linhas.append(
            {
                "codigo": categoria,
                "frequencia": frequencia,
                "percentual": percentual,
                "percentual_valido": valido,
                "percentual_acumulado": acumulado if np.isfinite(valido) else np.nan,
            }
        )

    return pd.DataFrame(linhas), total_geral, total_valido


def _copiar_estilo(origem, destino):
    destino._style = copy(origem._style)
    if origem.has_style:
        destino.font = copy(origem.font)
        destino.fill = copy(origem.fill)
        destino.border = copy(origem.border)
        destino.alignment = copy(origem.alignment)
        destino.number_format = origem.number_format
        destino.protection = copy(origem.protection)


def _estilo_com_borda_inferior(celula, estilo_origem, espessura="medium"):
    _copiar_estilo(estilo_origem, celula)
    lado = Side(style=espessura, color="000000")
    celula.border = Border(
        left=copy(celula.border.left),
        right=copy(celula.border.right),
        top=copy(celula.border.top),
        bottom=lado,
    )


def _contar_linhas_quebradas(texto, largura_coluna: float, fator_largura: float = 1.0) -> int:
    """Estima a quantidade de linhas que o Excel usará com quebra automática.

    O cálculo é propositalmente conservador para títulos longos. Em células de
    categoria usamos um pequeno ganho de largura porque a fonte normal ocupa
    menos espaço que os cabeçalhos em negrito.
    """
    if texto is None:
        return 1
    texto = str(texto).strip()
    if not texto:
        return 1

    capacidade = max(4, int(float(largura_coluna or 8.43) * fator_largura))
    total = 0
    for paragrafo in texto.splitlines() or [texto]:
        palavras = paragrafo.split()
        if not palavras:
            total += 1
            continue
        linha = 0
        linhas_paragrafo = 1
        for palavra in palavras:
            tamanho = len(palavra)
            if linha == 0:
                linha = tamanho
            elif linha + 1 + tamanho <= capacidade:
                linha += 1 + tamanho
            else:
                linhas_paragrafo += 1
                linha = tamanho
        total += linhas_paragrafo
    return max(1, total)


def _altura_por_linhas(qtd_linhas: int, minima: float = 16.2) -> float:
    # Aproxima o AutoFit do Excel no modelo visual usado pela equipe.
    return max(minima, min(120.0, 13.2 * max(1, qtd_linhas) + 0.6))


def _ajustar_altura_linha(
    ws,
    linha: int,
    textos_colunas: list[tuple[object, str, float]],
    minima: float = 16.2,
) -> None:
    """Ajusta a altura da linha considerando texto, coluna e wrap text.

    Cada item é (texto, coluna, fator_largura).
    """
    maior = 1
    for texto, coluna, fator in textos_colunas:
        largura = ws.column_dimensions[coluna].width or 8.43
        maior = max(maior, _contar_linhas_quebradas(texto, largura, fator))
    ws.row_dimensions[linha].height = _altura_por_linhas(maior, minima=minima)


def _forcar_quebra_texto(celula) -> None:
    alinhamento = copy(celula.alignment)
    celula.alignment = Alignment(
        horizontal=alinhamento.horizontal or "center",
        vertical=alinhamento.vertical or "center",
        text_rotation=alinhamento.textRotation or 0,
        wrap_text=True,
        shrink_to_fit=alinhamento.shrinkToFit,
        indent=alinhamento.indent or 0,
    )


def _materializar_bordas_mescladas(ws, linha_inicial: int, linha_final: int) -> None:
    """Propaga as bordas do canto superior esquerdo para toda célula mesclada.

    O openpyxl não redistribui automaticamente as bordas quando o estilo é
    aplicado depois do merge. Sem isso, partes do contorno podem desaparecer.
    """
    for intervalo in list(ws.merged_cells.ranges):
        if intervalo.max_row < linha_inicial or intervalo.min_row > linha_final:
            continue
        intervalo.format()


def _gerar_arquivo_historico_excel(
    resultado,
    tolerancia_pp: float,
    dados: pd.DataFrame,
    meta,
    variaveis: list[str],
    labels_por_variavel: dict,
    arquivo_origem: str,
    identificacao: str,
) -> bytes:
    """Preenche o modelo histórico enviado pelo usuário, adaptando blocos às variáveis do SAV."""
    modelo = Path(__file__).resolve().parents[1] / "assets" / "modelo_ponderacao.xlsx"
    if not modelo.exists():
        raise FileNotFoundError("O modelo histórico de ponderação não foi encontrado no projeto.")

    wb = openpyxl.load_workbook(modelo)
    ws = wb["Ponderação"] if "Ponderação" in wb.sheetnames else wb.active

    # O modelo antigo possuía uma regra de formatação condicional na coluna K.
    # Como os blocos são reposicionados dinamicamente, a regra podia atingir cabeçalhos
    # textuais como "Pontos". O destaque passa a ser aplicado apenas por código,
    # exclusivamente nas células numéricas de diferença.
    ws.conditional_formatting = ConditionalFormattingList()

    # Captura os estilos do modelo antes de limpar os blocos de exemplo.
    refs = {
        "titulo_esq": copy(ws["B5"]),
        "titulo_dir": copy(ws["M5"]),
        "cab_valid_esq": copy(ws["B6"]),
        "cab_freq_esq": copy(ws["D6"]),
        "cab_percent_esq": copy(ws["E6"]),
        "cab_validpct_esq": copy(ws["F6"]),
        "cab_cum_esq": copy(ws["G6"]),
        "cab_universo": copy(ws["I6"]),
        "cab_diff": copy(ws["K6"]),
        "cab_valid_dir": copy(ws["M6"]),
        "cab_freq_dir": copy(ws["O6"]),
        "cab_percent_dir": copy(ws["P6"]),
        "cab_validpct_dir": copy(ws["Q6"]),
        "cab_cum_dir": copy(ws["R6"]),
        "valid_esq": copy(ws["B7"]),
        "cat_primeira_esq": copy(ws["C7"]),
        "cat_meio_esq": copy(ws["C8"]),
        "freq_primeira_esq": copy(ws["D7"]),
        "freq_meio_esq": copy(ws["D8"]),
        "pct_primeira_esq": copy(ws["E7"]),
        "pct_meio_esq": copy(ws["E8"]),
        "validpct_primeira_esq": copy(ws["F7"]),
        "validpct_meio_esq": copy(ws["F8"]),
        "cum_primeira_esq": copy(ws["G7"]),
        "cum_meio_esq": copy(ws["G8"]),
        "universo_primeira": copy(ws["I7"]),
        "universo_meio": copy(ws["I14"]),
        "universo_ultima": copy(ws["I16"]),
        "diff_primeira": copy(ws["K7"]),
        "diff_meio": copy(ws["K14"]),
        "diff_ultima": copy(ws["K16"]),
        "valid_dir": copy(ws["M7"]),
        "cat_primeira_dir": copy(ws["N7"]),
        "cat_meio_dir": copy(ws["N8"]),
        "freq_primeira_dir": copy(ws["O7"]),
        "freq_meio_dir": copy(ws["O8"]),
        "pct_primeira_dir": copy(ws["P7"]),
        "pct_meio_dir": copy(ws["P8"]),
        "validpct_primeira_dir": copy(ws["Q7"]),
        "validpct_meio_dir": copy(ws["Q8"]),
        "cum_primeira_dir": copy(ws["R7"]),
        "cum_meio_dir": copy(ws["R8"]),
        "total_cat_esq": copy(ws["C9"]),
        "total_freq_esq": copy(ws["D9"]),
        "total_pct_esq": copy(ws["E9"]),
        "total_validpct_esq": copy(ws["F9"]),
        "total_cum_esq": copy(ws["G9"]),
        "total_cat_dir": copy(ws["N9"]),
        "total_freq_dir": copy(ws["O9"]),
        "total_pct_dir": copy(ws["P9"]),
        "total_validpct_dir": copy(ws["Q9"]),
        "total_cum_dir": copy(ws["R9"]),
    }

    # Mantém cabeçalhos, larguras e identidade visual; remove apenas os exemplos fixos.
    for intervalo in list(ws.merged_cells.ranges):
        if intervalo.min_row >= 5:
            ws.unmerge_cells(str(intervalo))
    if ws.max_row >= 5:
        ws.delete_rows(5, ws.max_row - 4)

    ws["B3"] = "AMOSTRA S/ PONDERAÇÃO"
    ws["M3"] = "AMOSTRA C/ PONDERAÇÃO"
    ws["I1"] = "% Universo"
    ws["K1"] = "Diferença p/ universo"
    for coord in ["B3", "M3", "I1", "K1"]:
        _forcar_quebra_texto(ws[coord])

    diagnostico = resultado.diagnostico.copy()
    linha_atual = 5

    for indice_var, variavel in enumerate(variaveis):
        bloco = diagnostico.loc[diagnostico["variavel"] == variavel].copy()
        categorias = bloco["codigo"].tolist()
        if not categorias:
            continue

        titulo = _nome_exibicao_variavel(meta, variavel)
        labels_var = labels_por_variavel.get(variavel, {})
        nomes_categorias = [labels_var.get(c, str(c)) for c in categorias]
        freq_sem, total_geral_sem, total_valido_sem = _frequencia_registro(
            dados[variavel], categorias
        )
        freq_com, total_geral_com, total_valido_com = _frequencia_registro(
            dados[variavel], categorias, resultado.pesos
        )
        universo = dict(zip(bloco["codigo"], bloco["percentual_universo"]))

        # Título do bloco.
        ws.merge_cells(start_row=linha_atual, start_column=2, end_row=linha_atual, end_column=7)
        ws.merge_cells(start_row=linha_atual, start_column=13, end_row=linha_atual, end_column=18)
        ws.cell(linha_atual, 2, titulo)
        ws.cell(linha_atual, 13, titulo)
        _copiar_estilo(refs["titulo_esq"], ws.cell(linha_atual, 2))
        _copiar_estilo(refs["titulo_dir"], ws.cell(linha_atual, 13))
        ws.cell(linha_atual, 11, f"Limite: ±{tolerancia_pp:.2f} p.p.")
        _copiar_estilo(refs["titulo_esq"], ws.cell(linha_atual, 11))
        for coluna in [2, 11, 13]:
            _forcar_quebra_texto(ws.cell(linha_atual, coluna))
        # Os títulos laterais estão mesclados em seis colunas e normalmente cabem em uma linha.
        ws.row_dimensions[linha_atual].height = 18.0

        cab = linha_atual + 1
        ws.merge_cells(start_row=cab, start_column=2, end_row=cab, end_column=3)
        ws.merge_cells(start_row=cab, start_column=13, end_row=cab, end_column=14)
        ws.cell(cab, 2, " ")
        ws.cell(cab, 4, "Frequency")
        ws.cell(cab, 5, "Percent")
        ws.cell(cab, 6, "Valid Percent")
        ws.cell(cab, 7, "Cumulative Percent")
        ws.cell(cab, 9, titulo)
        ws.cell(cab, 11, "Pontos")
        ws.cell(cab, 13, " ")
        ws.cell(cab, 15, "Frequency")
        ws.cell(cab, 16, "Percent")
        ws.cell(cab, 17, "Valid Percent")
        ws.cell(cab, 18, "Cumulative Percent")
        for chave, coord in [
            ("cab_valid_esq", (cab, 2)), ("cab_freq_esq", (cab, 4)),
            ("cab_percent_esq", (cab, 5)), ("cab_validpct_esq", (cab, 6)),
            ("cab_cum_esq", (cab, 7)), ("cab_universo", (cab, 9)),
            ("cab_diff", (cab, 11)), ("cab_valid_dir", (cab, 13)),
            ("cab_freq_dir", (cab, 15)), ("cab_percent_dir", (cab, 16)),
            ("cab_validpct_dir", (cab, 17)), ("cab_cum_dir", (cab, 18)),
        ]:
            _copiar_estilo(refs[chave], ws.cell(*coord))
            _forcar_quebra_texto(ws.cell(*coord))

        _ajustar_altura_linha(
            ws,
            cab,
            [
                ("Frequency", "D", 1.0), ("Percent", "E", 1.0),
                ("Valid Percent", "F", 1.0), ("Cumulative Percent", "G", 1.0),
                (titulo, "I", 1.0), ("Pontos", "K", 1.0),
                ("Frequency", "O", 1.0), ("Percent", "P", 1.0),
                ("Valid Percent", "Q", 1.0), ("Cumulative Percent", "R", 1.0),
            ],
            minima=27.0,
        )

        primeira = cab + 1
        ultima_cat = primeira + len(categorias) - 1
        total_row = ultima_cat + 1
        ws.merge_cells(start_row=primeira, start_column=2, end_row=total_row, end_column=2)
        ws.merge_cells(start_row=primeira, start_column=13, end_row=total_row, end_column=13)
        ws.cell(primeira, 2, "Valid")
        ws.cell(primeira, 13, "Valid")
        _copiar_estilo(refs["valid_esq"], ws.cell(primeira, 2))
        _copiar_estilo(refs["valid_dir"], ws.cell(primeira, 13))

        for pos, (codigo, categoria) in enumerate(zip(categorias, nomes_categorias)):
            r = primeira + pos
            primeiro = pos == 0
            ultimo = pos == len(categorias) - 1
            sem = freq_sem.iloc[pos]
            com = freq_com.iloc[pos]

            ws.cell(r, 3, categoria)
            ws.cell(r, 4, sem["frequencia"])
            ws.cell(r, 5, sem["percentual"])
            ws.cell(r, 6, sem["percentual_valido"])
            ws.cell(r, 7, sem["percentual_acumulado"])
            ws.cell(r, 9, float(universo[codigo]))
            # A diferença é gravada como valor numérico, não como fórmula. Isso evita
            # que regras do Excel alcancem acidentalmente cabeçalhos textuais quando
            # os blocos mudam de posição.
            ws.cell(r, 11, round(float(bloco.iloc[pos]["diferenca_pp"]), 2))
            ws.cell(r, 14, categoria)
            ws.cell(r, 15, com["frequencia"])
            ws.cell(r, 16, com["percentual"])
            ws.cell(r, 17, com["percentual_valido"])
            ws.cell(r, 18, com["percentual_acumulado"])

            estilo_cat_esq = refs["cat_primeira_esq"] if primeiro else refs["cat_meio_esq"]
            estilo_freq_esq = refs["freq_primeira_esq"] if primeiro else refs["freq_meio_esq"]
            estilo_pct_esq = refs["pct_primeira_esq"] if primeiro else refs["pct_meio_esq"]
            estilo_validpct_esq = refs["validpct_primeira_esq"] if primeiro else refs["validpct_meio_esq"]
            estilo_cum_esq = refs["cum_primeira_esq"] if primeiro else refs["cum_meio_esq"]
            estilo_cat_dir = refs["cat_primeira_dir"] if primeiro else refs["cat_meio_dir"]
            estilo_freq_dir = refs["freq_primeira_dir"] if primeiro else refs["freq_meio_dir"]
            estilo_pct_dir = refs["pct_primeira_dir"] if primeiro else refs["pct_meio_dir"]
            estilo_validpct_dir = refs["validpct_primeira_dir"] if primeiro else refs["validpct_meio_dir"]
            estilo_cum_dir = refs["cum_primeira_dir"] if primeiro else refs["cum_meio_dir"]

            for estilo, coluna in [
                (estilo_cat_esq, 3), (estilo_freq_esq, 4), (estilo_pct_esq, 5),
                (estilo_validpct_esq, 6), (estilo_cum_esq, 7),
                (estilo_cat_dir, 14), (estilo_freq_dir, 15), (estilo_pct_dir, 16),
                (estilo_validpct_dir, 17), (estilo_cum_dir, 18),
            ]:
                _copiar_estilo(estilo, ws.cell(r, coluna))

            _forcar_quebra_texto(ws.cell(r, 3))
            _forcar_quebra_texto(ws.cell(r, 14))
            _ajustar_altura_linha(
                ws, r, [(categoria, "C", 1.10), (categoria, "N", 1.10)], minima=16.2
            )

            if len(categorias) == 1:
                _estilo_com_borda_inferior(ws.cell(r, 9), refs["universo_primeira"])
                _estilo_com_borda_inferior(ws.cell(r, 11), refs["diff_primeira"])
            elif primeiro:
                _copiar_estilo(refs["universo_primeira"], ws.cell(r, 9))
                _copiar_estilo(refs["diff_primeira"], ws.cell(r, 11))
            elif ultimo:
                _copiar_estilo(refs["universo_ultima"], ws.cell(r, 9))
                _copiar_estilo(refs["diff_ultima"], ws.cell(r, 11))
            else:
                _copiar_estilo(refs["universo_meio"], ws.cell(r, 9))
                _copiar_estilo(refs["diff_meio"], ws.cell(r, 11))

            _forcar_quebra_texto(ws.cell(r, 9))
            _forcar_quebra_texto(ws.cell(r, 11))
            ws.cell(r, 3).comment = Comment(f"Código SPSS: {codigo}", "Analítico")
            ws.cell(r, 14).comment = Comment(f"Código SPSS: {codigo}", "Analítico")

            # O modelo visual pode conter preenchimentos de exemplos antigos.
            # Limpa sempre a célula de diferença antes de aplicar a regra atual.
            ws.cell(r, 11).fill = PatternFill(fill_type=None)

            # Usa o valor arredondado como ele aparece no Excel. Assim, +2,10 e
            # -2,10 ficam brancos; somente > +2,10 ou < -2,10 ficam vermelhos.
            diff_assinado = float(bloco.iloc[pos]["diferenca_pp"])
            diff_exibido = round(diff_assinado, 2)
            if (
                np.isfinite(diff_exibido)
                and abs(diff_exibido) > LIMITE_DESTAQUE_VISUAL_PP
            ):
                ws.cell(r, 11).fill = PatternFill("solid", fgColor="FECACA")

        # Total válido em cada lado.
        ws.cell(total_row, 3, "Total")
        ws.cell(total_row, 4, total_valido_sem)
        ws.cell(total_row, 5, (total_valido_sem / total_geral_sem * 100.0) if total_geral_sem else np.nan)
        ws.cell(total_row, 6, 100.0 if total_valido_sem else np.nan)
        ws.cell(total_row, 14, "Total")
        ws.cell(total_row, 15, total_valido_com)
        ws.cell(total_row, 16, (total_valido_com / total_geral_com * 100.0) if total_geral_com else np.nan)
        ws.cell(total_row, 17, 100.0 if total_valido_com else np.nan)
        for chave, coord in [
            ("total_cat_esq", (total_row, 3)), ("total_freq_esq", (total_row, 4)),
            ("total_pct_esq", (total_row, 5)), ("total_validpct_esq", (total_row, 6)),
            ("total_cum_esq", (total_row, 7)), ("total_cat_dir", (total_row, 14)),
            ("total_freq_dir", (total_row, 15)), ("total_pct_dir", (total_row, 16)),
            ("total_validpct_dir", (total_row, 17)), ("total_cum_dir", (total_row, 18)),
        ]:
            _copiar_estilo(refs[chave], ws.cell(*coord))

        # Na tabela visual, a frequência ponderada (coluna O) é exibida como
        # número inteiro. O valor interno continua preservado para os cálculos.
        for r in range(primeira, total_row + 1):
            ws.cell(r, 15).number_format = "0"
            for c in [5, 6, 7, 9, 11, 16, 17, 18]:
                ws.cell(r, c).number_format = "0.00"

        # Propaga as bordas dos merges (Valid e cabeçalhos B:C / M:N).
        # Isso evita contornos incompletos depois de gerar o arquivo dinamicamente.
        _materializar_bordas_mescladas(ws, linha_atual, total_row)
        ws.row_dimensions[total_row].height = 16.2

        linha_atual = total_row + 2

    # Metadados tornam o arquivo realmente útil como registro histórico.
    if "Registro" in wb.sheetnames:
        del wb["Registro"]
    reg = wb.create_sheet("Registro")
    resumo = resumo_pesos(resultado.pesos)
    meta_linhas = [
        ("Identificação", identificacao),
        ("Arquivo SPSS de origem", arquivo_origem),
        ("Gerado em", datetime.now(timezone.utc).strftime("%d/%m/%Y %H:%M UTC")),
        ("Método", resultado.metodo),
        ("Tolerância máxima (p.p.)", tolerancia_pp),
        ("Dentro da tolerância", "SIM" if resultado.convergiu else "NÃO"),
        ("Maior diferença absoluta (p.p.)", resultado.maior_diferenca_pp),
        ("Iterações", resultado.iteracoes),
        ("Casos na base", len(dados)),
        ("Casos com peso válido", resumo["casos_com_peso"]),
        ("Peso mínimo", resumo["minimo"]),
        ("Peso máximo", resumo["maximo"]),
        ("Peso médio", resumo["media"]),
    ]
    reg.append(["REGISTRO HISTÓRICO DA PONDERAÇÃO", None])
    reg.merge_cells("A1:B1")
    reg["A1"].font = Font(bold=True, size=14)
    for chave, valor in meta_linhas:
        reg.append([chave, valor])
    reg.append([])
    reg.append([
        "Variável SPSS", "Rótulo da variável", "Código SPSS", "Categoria",
        "% Amostra", "% Universo", "% Ponderada", "Diferença (p.p.)", "Status"
    ])
    cab_reg = reg.max_row
    for c in reg[cab_reg]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F4E78")

    for variavel in variaveis:
        bloco = diagnostico.loc[diagnostico["variavel"] == variavel]
        titulo = _nome_exibicao_variavel(meta, variavel)
        for _, linha in bloco.iterrows():
            detalhe_row = reg.max_row + 1
            reg.append([
                variavel,
                titulo,
                linha["codigo"],
                linha["categoria"],
                linha["percentual_sem_peso"],
                linha["percentual_universo"],
                linha["percentual_com_peso"],
                f"=G{detalhe_row}-F{detalhe_row}",
                f'=IF(ABS(H{detalhe_row})<=$B$6,"OK","REVISAR")',
            ])
            for c in range(5, 9):
                reg.cell(reg.max_row, c).number_format = "0.00"
            if linha["status"] == "REVISAR":
                reg.cell(reg.max_row, 9).fill = PatternFill("solid", fgColor="FECACA")
            else:
                reg.cell(reg.max_row, 9).fill = PatternFill("solid", fgColor="DCFCE7")

    larguras = [20, 34, 14, 30, 14, 14, 14, 18, 12]
    for i, largura in enumerate(larguras, 1):
        reg.column_dimensions[openpyxl.utils.get_column_letter(i)].width = largura
    for row in reg.iter_rows():
        for celula in row:
            if celula.value is not None:
                _forcar_quebra_texto(celula)
    for r in range(1, reg.max_row + 1):
        textos = []
        for c in range(1, reg.max_column + 1):
            celula = reg.cell(r, c)
            if celula.value is not None:
                textos.append((celula.value, openpyxl.utils.get_column_letter(c), 1.10))
        if textos:
            _ajustar_altura_linha(reg, r, textos, minima=16.2)
    # A aba Registro deve abrir totalmente livre, sem linhas/colunas congeladas.
    reg.freeze_panes = None
    ws.sheet_view.showGridLines = False
    reg.sheet_view.showGridLines = False

    try:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
        wb.calculation.calcMode = "auto"
    except Exception:
        pass

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _literal_spss_documentacao(valor) -> str:
    """Formata um código como literal SPSS para o arquivo de documentação."""
    if isinstance(valor, (str, np.str_)):
        return "'" + str(valor).replace("'", "''") + "'"
    if isinstance(valor, (int, np.integer)):
        return str(int(valor))
    if isinstance(valor, (float, np.floating)):
        numero = float(valor)
        if np.isfinite(numero) and numero.is_integer():
            return str(int(numero))
        return format(numero, ".15g")
    return str(valor)


def _gerar_calculo_ponderacao_excel(
    resultado,
    dados: pd.DataFrame,
    meta,
    variaveis: list[str],
    arquivo_origem: str,
    identificacao: str,
) -> bytes:
    """Gera a memória de cálculo dos pesos em formato semelhante à planilha histórica da equipe.

    Cada variável recebe uma aba própria com amostra, universo, razão inicial,
    fator efetivamente utilizado e as linhas correspondentes da syntax SPSS.
    """
    wb = openpyxl.Workbook()
    resumo_ws = wb.active
    resumo_ws.title = "Resumo"
    resumo_ws.sheet_view.showGridLines = False

    # Paleta inspirada no arquivo Cálculo da Ponderação usado pela equipe.
    amarelo = PatternFill("solid", fgColor="FFF2CC")
    amarelo_forte = PatternFill("solid", fgColor="FFF200")
    verde = PatternFill("solid", fgColor="E2F0D9")
    verde_forte = PatternFill("solid", fgColor="92D050")
    azul = PatternFill("solid", fgColor="D9EAF7")
    cinza = PatternFill("solid", fgColor="E7E6E6")
    cinza_claro = PatternFill("solid", fgColor="F2F2F2")
    branco = PatternFill("solid", fgColor="FFFFFF")
    borda_fina = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )
    borda_total = Border(
        top=Side(style="thin", color="000000"),
        bottom=Side(style="medium", color="000000"),
    )

    def estilizar_cabecalho(celula, fill):
        celula.fill = fill
        celula.font = Font(name="Arial", size=10, bold=True)
        celula.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        celula.border = borda_fina

    def estilizar_dado(celula, *, fill=None, alinhamento="center", formato=None, fonte_cor=None):
        if fill is not None:
            celula.fill = fill
        celula.font = Font(name="Arial", size=10, color=fonte_cor or "000000")
        celula.alignment = Alignment(horizontal=alinhamento, vertical="center", wrap_text=True)
        celula.border = borda_fina
        if formato:
            celula.number_format = formato

    # Resumo / memória metodológica.
    resumo_ws.merge_cells("A1:F1")
    resumo_ws["A1"] = "MEMÓRIA DE CÁLCULO DA PONDERAÇÃO"
    resumo_ws["A1"].font = Font(name="Arial", size=14, bold=True)
    resumo_ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    resumo_ws["A1"].fill = verde_forte
    resumo_ws.row_dimensions[1].height = 26

    resumo_linhas = [
        ("Identificação", identificacao),
        ("Arquivo SPSS de origem", arquivo_origem),
        ("Método", resultado.metodo),
        ("Gerado em", datetime.now(timezone.utc).strftime("%d/%m/%Y %H:%M UTC")),
        ("Casos na base", len(dados)),
    ]
    for idx, (rotulo, valor) in enumerate(resumo_linhas, start=3):
        resumo_ws.cell(idx, 1, rotulo).font = Font(name="Arial", size=10, bold=True)
        resumo_ws.cell(idx, 1).fill = cinza
        resumo_ws.cell(idx, 1).border = borda_fina
        resumo_ws.cell(idx, 2, valor)
        resumo_ws.merge_cells(start_row=idx, start_column=2, end_row=idx, end_column=6)
        resumo_ws.cell(idx, 2).border = borda_fina
        resumo_ws.cell(idx, 2).alignment = Alignment(vertical="center", wrap_text=True)

    resumo_ws["A10"] = "Como o peso é obtido"
    resumo_ws["A10"].font = Font(name="Arial", size=11, bold=True)
    resumo_ws["A10"].fill = amarelo_forte
    resumo_ws.merge_cells("A10:F10")

    if resultado.metodo == "Razão simples":
        metodo_texto = (
            "Para cada categoria: fator = % Universo / % Amostra. "
            "O peso final de cada entrevista é o produto dos fatores das variáveis selecionadas."
        )
    else:
        metodo_texto = (
            "A coluna Razão inicial mostra % Universo / % Amostra. No ajuste iterativo, "
            "o Fator utilizado registra o fator acumulado após as iterações de calibração. "
            "O peso final continua sendo o produto dos componentes gerados para cada variável."
        )
    resumo_ws["A11"] = metodo_texto
    resumo_ws.merge_cells("A11:F12")
    resumo_ws["A11"].alignment = Alignment(vertical="top", wrap_text=True)
    resumo_ws["A11"].fill = amarelo
    resumo_ws["A11"].border = borda_fina

    componentes = [f"peso{i}" for i in range(1, len(variaveis) + 1)]
    produto = "*".join(componentes)
    resumo_ws["A14"] = "Fórmula final no SPSS"
    resumo_ws["A14"].font = Font(name="Arial", size=10, bold=True)
    resumo_ws["B14"] = f"COMPUTE peso={produto}."
    resumo_ws.merge_cells("B14:F14")
    resumo_ws["B14"].font = Font(name="Consolas", size=10)
    resumo_ws["B14"].fill = cinza_claro
    resumo_ws["B14"].alignment = Alignment(wrap_text=True)

    cab_resumo = 17
    headers = ["Ordem", "Variável SPSS", "Rótulo da variável", "Componente", "Categorias", "Aba"]
    for c, texto in enumerate(headers, start=1):
        cel = resumo_ws.cell(cab_resumo, c, texto)
        estilizar_cabecalho(cel, azul)

    usados = {"Resumo"}
    diagnostico = resultado.diagnostico.copy()
    nomes_abas = {}

    for indice, variavel in enumerate(variaveis, start=1):
        titulo = _nome_exibicao_variavel(meta, variavel)
        nome_aba = _nome_aba_seguro(titulo or variavel, usados)
        nomes_abas[variavel] = nome_aba
        bloco = diagnostico.loc[diagnostico["variavel"] == variavel].copy()

        r = cab_resumo + indice
        valores = [indice, variavel, titulo, f"peso{indice}", len(bloco), nome_aba]
        for c, valor in enumerate(valores, start=1):
            cel = resumo_ws.cell(r, c, valor)
            estilizar_dado(cel, fill=branco, alinhamento="left" if c in (2, 3, 6) else "center")

        ws = wb.create_sheet(nome_aba)
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = None
        ws.column_dimensions["A"].width = 36
        ws.column_dimensions["B"].width = 14
        ws.column_dimensions["C"].width = 16
        ws.column_dimensions["D"].width = 16
        ws.column_dimensions["E"].width = 18
        ws.column_dimensions["F"].width = 18

        ws["B1"] = "Memória de cálculo — valores de Universo são a referência do projeto"
        ws.merge_cells("B1:F1")
        ws["B1"].font = Font(name="Arial", size=11, bold=True)
        ws["B1"].fill = amarelo_forte
        ws["B1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[1].height = 30

        ws["A3"] = titulo
        ws["A3"].font = Font(name="Arial", size=12, bold=True)
        ws["A4"] = f"Variável SPSS: {variavel} | Componente: peso{indice}"
        ws["A4"].font = Font(name="Arial", size=10, italic=True)
        ws.merge_cells("A4:F4")

        headers_var = ["Categoria", "Código SPSS", "AMOSTRA (%)", "UNIVERSO (%)", "Razão inicial U/A", "Fator utilizado"]
        fills_var = [cinza, cinza, azul, amarelo_forte, verde, verde_forte]
        for c, (texto, fill) in enumerate(zip(headers_var, fills_var), start=1):
            estilizar_cabecalho(ws.cell(5, c, texto), fill)
        ws.row_dimensions[5].height = 32

        primeira = 6
        for pos, (_, linha) in enumerate(bloco.iterrows()):
            rr = primeira + pos
            codigo = linha["codigo"]
            categoria = linha["categoria"]
            ws.cell(rr, 1, categoria)
            ws.cell(rr, 2, codigo)
            ws.cell(rr, 3, float(linha["percentual_sem_peso"]))
            ws.cell(rr, 4, float(linha["percentual_universo"]))
            ws.cell(rr, 5, f'=IFERROR(D{rr}/C{rr},"")')

            if resultado.metodo == "Razão simples":
                ws.cell(rr, 6, f"=E{rr}")
                fator_fill = verde
                fator_cor = "000000"
            else:
                # No raking o fator é o produto acumulado das iterações; é um
                # resultado importado do motor de cálculo, não a razão inicial.
                ws.cell(rr, 6, float(linha["fator"]))
                fator_fill = verde
                fator_cor = "008000"

            estilizar_dado(ws.cell(rr, 1), alinhamento="left")
            estilizar_dado(ws.cell(rr, 2), fill=cinza_claro)
            estilizar_dado(ws.cell(rr, 3), fill=azul, formato="0.00")
            estilizar_dado(ws.cell(rr, 4), fill=amarelo, formato="0.00")
            estilizar_dado(ws.cell(rr, 5), fill=verde, formato="0.000000")
            estilizar_dado(ws.cell(rr, 6), fill=fator_fill, formato="0.000000", fonte_cor=fator_cor)
            _ajustar_altura_linha(ws, rr, [(categoria, "A", 1.15)], minima=18.0)

        ultima = primeira + len(bloco) - 1
        total = ultima + 1
        ws.cell(total, 1, "Total")
        ws.cell(total, 3, f"=SUM(C{primeira}:C{ultima})")
        ws.cell(total, 4, f"=SUM(D{primeira}:D{ultima})")
        for c in range(1, 7):
            cel = ws.cell(total, c)
            cel.font = Font(name="Arial", size=10, bold=True)
            cel.alignment = Alignment(horizontal="center", vertical="center")
            cel.border = borda_total
            if c == 3:
                cel.fill = azul
                cel.number_format = "0.00"
            elif c == 4:
                cel.fill = amarelo
                cel.number_format = "0.00"

        info = total + 3
        ws.cell(info, 1, "Número da Pergunta / Variável SPSS")
        ws.cell(info, 1).font = Font(name="Arial", size=10, bold=True)
        ws.cell(info, 1).fill = amarelo_forte
        ws.cell(info + 1, 1, variavel)
        ws.cell(info + 1, 1).font = Font(name="Arial", size=10, bold=True)
        ws.cell(info + 1, 1).fill = amarelo
        ws.cell(info, 3, "Componente")
        ws.cell(info, 3).font = Font(name="Arial", size=10, bold=True)
        ws.cell(info, 3).fill = verde_forte
        ws.cell(info + 1, 3, f"peso{indice}")
        ws.cell(info + 1, 3).fill = verde

        syntax_start = info + 4
        ws.cell(syntax_start, 1, "Linhas geradas para a syntax SPSS")
        ws.merge_cells(start_row=syntax_start, start_column=1, end_row=syntax_start, end_column=6)
        ws.cell(syntax_start, 1).font = Font(name="Arial", size=10, bold=True)
        ws.cell(syntax_start, 1).fill = cinza

        for pos, (_, linha) in enumerate(bloco.iterrows(), start=1):
            rr = syntax_start + pos
            codigo_lit = _literal_spss_documentacao(linha["codigo"])
            fator = float(linha["fator"])
            texto = f"IF ({variavel}={codigo_lit}) peso{indice}={format(fator, '.15g')}."
            ws.cell(rr, 1, texto)
            ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=6)
            ws.cell(rr, 1).font = Font(name="Consolas", size=9)
            ws.cell(rr, 1).fill = cinza_claro
            ws.cell(rr, 1).alignment = Alignment(vertical="center", wrap_text=True)
            ws.row_dimensions[rr].height = 20

        final_row = syntax_start + len(bloco) + 2
        ws.cell(final_row, 1, "Observação")
        ws.cell(final_row, 1).font = Font(name="Arial", size=10, bold=True)
        if resultado.metodo == "Razão simples":
            observacao = "Fator utilizado = Universo / Amostra."
        else:
            observacao = (
                "No ajuste iterativo, Razão inicial documenta o ponto de partida; "
                "Fator utilizado é o fator acumulado efetivamente aplicado pelo algoritmo."
            )
        ws.cell(final_row + 1, 1, observacao)
        ws.merge_cells(start_row=final_row + 1, start_column=1, end_row=final_row + 2, end_column=6)
        ws.cell(final_row + 1, 1).alignment = Alignment(vertical="top", wrap_text=True)
        ws.cell(final_row + 1, 1).fill = amarelo

    # Syntax completa no Resumo para documentar a composição final dos pesos.
    syntax = gerar_syntax_spss(resultado.fatores)
    inicio_syntax = cab_resumo + len(variaveis) + 3
    resumo_ws.cell(inicio_syntax, 1, "SYNTAX SPSS GERADA")
    resumo_ws.merge_cells(start_row=inicio_syntax, start_column=1, end_row=inicio_syntax, end_column=6)
    resumo_ws.cell(inicio_syntax, 1).font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    resumo_ws.cell(inicio_syntax, 1).fill = PatternFill("solid", fgColor="1F4E78")
    for offset, linha_syntax in enumerate(syntax.splitlines(), start=1):
        rr = inicio_syntax + offset
        resumo_ws.cell(rr, 1, linha_syntax)
        resumo_ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=6)
        resumo_ws.cell(rr, 1).font = Font(name="Consolas", size=9)
        resumo_ws.cell(rr, 1).fill = cinza_claro if linha_syntax else branco
        resumo_ws.cell(rr, 1).alignment = Alignment(vertical="center", wrap_text=True)

    resumo_ws.column_dimensions["A"].width = 22
    resumo_ws.column_dimensions["B"].width = 22
    resumo_ws.column_dimensions["C"].width = 38
    resumo_ws.column_dimensions["D"].width = 16
    resumo_ws.column_dimensions["E"].width = 14
    resumo_ws.column_dimensions["F"].width = 24
    for r in range(1, resumo_ws.max_row + 1):
        if resumo_ws.row_dimensions[r].height is None:
            resumo_ws.row_dimensions[r].height = 18

    try:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
        wb.calculation.calcMode = "auto"
    except Exception:
        pass

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _gerar_sav_com_peso(dados, meta, pesos: pd.Series) -> bytes:
    saida = dados.copy()
    saida["peso"] = pesos

    column_labels = dict(getattr(meta, "column_names_to_labels", {}) or {})
    column_labels["peso"] = "Peso de ponderação gerado pelo Analítico"
    value_labels = dict(getattr(meta, "variable_value_labels", {}) or {})

    with tempfile.NamedTemporaryFile(delete=False, suffix=".sav") as tmp:
        caminho = tmp.name

    try:
        pyreadstat.write_sav(
            saida,
            caminho,
            column_labels=column_labels,
            variable_value_labels=value_labels,
        )
        with open(caminho, "rb") as f:
            return f.read()
    finally:
        if os.path.exists(caminho):
            os.unlink(caminho)


def modulo_ponderacao():
    st.header("Ponderação")
    st.caption(
        "Gera frequências diretamente da base SPSS, recebe os percentuais do universo, "
        "calcula os pesos, confere a diferença final e gera os arquivos para uso no SPSS."
    )

    uploaded = st.file_uploader(
        "Carregue a base SPSS (.sav)", type=["sav"], key="ponderacao_upload_sav"
    )
    if not uploaded:
        st.info("Envie a base .sav para iniciar a ponderação.")
        return

    try:
        dados, meta, value_labels = _ler_sav(uploaded)
    except Exception as exc:
        st.error(f"Não foi possível ler o arquivo SPSS: {exc}")
        return

    st.success(f"Base carregada: {len(dados)} registros e {len(dados.columns)} variáveis.")

    st.subheader("Universo reutilizável")
    st.caption(
        "O Universo fica salvo durante esta sessão mesmo quando você trocar a base .sav. "
        "Também é possível importar um Modelo de Universo preenchido ou a aba Registro de uma ponderação anterior."
    )
    perfil = _perfil_sessao()
    total_alvos_perfil = _contar_alvos_perfil(perfil)
    if total_alvos_perfil:
        st.info(
            f"Universo disponível na sessão: {len(perfil.get('variaveis', {}))} variável(is), "
            f"{total_alvos_perfil} categoria(s). Origem: {perfil.get('origem', 'Sessão atual')}."
        )

    arquivo_universo = st.file_uploader(
        "Importar Universo preenchido (.xlsx)",
        type=["xlsx"],
        key="ponderacao_upload_universo_xlsx",
        help=(
            "Use um Modelo_Universo_*.xlsx preenchido ou um Ponderacao_*.xlsx histórico. "
            "O app lê somente Variável SPSS + Código SPSS + % Universo."
        ),
    )
    col_importar, col_esquecer = st.columns([3, 1])
    if col_importar.button(
        "Importar Universo",
        key="ponderacao_importar_universo",
        use_container_width=True,
        disabled=arquivo_universo is None,
    ):
        try:
            novo_perfil = _importar_perfil_universo_xlsx(
                arquivo_universo.getvalue(), arquivo_universo.name
            )
            st.session_state["ponderacao_perfil_universo"] = novo_perfil
            st.session_state["ponderacao_reaplicar_perfil"] = True
            variaveis_importadas = [
                v for v in novo_perfil["variaveis"] if v in dados.columns
            ][:5]
            if variaveis_importadas:
                st.session_state["ponderacao_variaveis"] = variaveis_importadas
            perfil = novo_perfil
            total_alvos_perfil = _contar_alvos_perfil(perfil)
            st.success(
                f"Universo importado: {len(perfil['variaveis'])} variável(is) e "
                f"{total_alvos_perfil} categoria(s)."
            )
        except ValueError as exc:
            st.error(str(exc))

    if col_esquecer.button(
        "Esquecer Universo",
        key="ponderacao_esquecer_universo",
        use_container_width=True,
        disabled=not bool(total_alvos_perfil),
        help="Remove apenas o Universo reutilizável salvo na sessão. Não altera arquivos no computador.",
    ):
        st.session_state["ponderacao_perfil_universo"] = _perfil_vazio()
        st.session_state["ponderacao_reaplicar_perfil"] = False
        perfil = st.session_state["ponderacao_perfil_universo"]
        total_alvos_perfil = 0
        st.success("Universo reutilizável removido da sessão.")

    identificacao_padrao = Path(uploaded.name).stem if getattr(uploaded, "name", None) else "Ponderação"
    identificacao = st.text_input(
        "Identificação do projeto / registro histórico",
        value=identificacao_padrao,
        key="ponderacao_identificacao",
        help="Esse nome será salvo no arquivo histórico e usado no nome do download.",
    )

    perfil = _perfil_sessao()
    preferencias_perfil = [
        v for v in perfil.get("variaveis", {}) if v in dados.columns
    ][:5]
    preferencias = preferencias_perfil or _sugerir_variaveis(dados, meta)

    if "ponderacao_variaveis" in st.session_state:
        selecionadas_validas = [
            v for v in st.session_state["ponderacao_variaveis"] if v in dados.columns
        ][:5]
        if selecionadas_validas != list(st.session_state["ponderacao_variaveis"]):
            st.session_state["ponderacao_variaveis"] = selecionadas_validas or preferencias

    variaveis = st.multiselect(
        "Variáveis usadas na ponderação (até 5)",
        options=list(dados.columns),
        default=preferencias,
        format_func=lambda v: _descricao_variavel(meta, v),
        key="ponderacao_variaveis",
        help=(
            "O app usa o nome e o rótulo que vierem no SPSS. Região não precisa ter um nome fixo: "
            "pode ser Região X, Mesorregião, RPA ou outro rótulo do projeto."
        ),
    )

    if not variaveis:
        st.info("Selecione as variáveis que serão aproximadas ao universo.")
        return
    if len(variaveis) > 5:
        st.error("Selecione no máximo 5 variáveis por ponderação.")
        return

    modelo_universo = None
    erro_modelo_universo = None
    try:
        modelo_universo = _gerar_modelo_universo_excel(
            dados=dados,
            meta=meta,
            value_labels=value_labels,
            variaveis=list(variaveis),
            arquivo_origem=getattr(uploaded, "name", "Base SPSS"),
        )
    except Exception as exc:
        erro_modelo_universo = str(exc)

    st.markdown("#### Modelo para preenchimento do Universo")
    st.caption(
        "O Analítico já montou um Excel com as variáveis selecionadas, códigos e labels "
        "da base atual. Preencha somente a coluna % Universo, salve e depois importe o "
        "mesmo arquivo acima quando precisar refazer a ponderação."
    )
    nome_modelo = _nome_arquivo_seguro(identificacao.strip() or identificacao_padrao)
    if modelo_universo is not None:
        st.download_button(
            "Baixar modelo do Universo (.xlsx)",
            data=modelo_universo,
            file_name=f"Modelo_Universo_{nome_modelo}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="ponderacao_download_modelo_universo",
            use_container_width=True,
            help=(
                "O modelo é criado a partir dos códigos e labels do SAV atual e pode ser "
                "reutilizado em novas versões da amostra do mesmo projeto."
            ),
        )
    else:
        st.warning(
            "Não foi possível gerar o modelo do Universo nesta execução. "
            f"Detalhe técnico: {erro_modelo_universo}"
        )

    perfil = _perfil_sessao()
    if _contar_alvos_perfil(perfil):
        total_categorias = 0
        total_casados = 0
        total_sem_universo = 0
        total_ausentes = 0
        total_labels_divergentes = 0
        for variavel in variaveis:
            base_tmp = _tabela_editor(
                dados, variavel, value_labels.get(variavel, {}) or {}
            )
            comp = _compatibilidade_perfil(base_tmp, perfil, variavel)
            total_categorias += comp["total_atual"]
            total_casados += comp["casados"]
            total_sem_universo += len(comp["sem_universo"])
            total_ausentes += len(comp["ausentes_na_amostra"])
            total_labels_divergentes += len(comp["labels_divergentes"])

        if total_casados:
            st.success(
                f"Compatibilidade do Universo: {total_casados}/{total_categorias} "
                "categorias da base atual encontradas por variável + código SPSS."
            )
        if total_sem_universo:
            st.warning(
                f"{total_sem_universo} categoria(s) da base atual ainda não possuem Universo salvo."
            )
        if total_ausentes:
            st.warning(
                f"{total_ausentes} categoria(s) do Universo salvo não aparecem na amostra atual. "
                "Isso pode acontecer após exclusões e será mostrado em cada variável."
            )
        if total_labels_divergentes:
            st.warning(
                f"{total_labels_divergentes} código(s) têm label diferente do registro anterior. "
                "O pareamento continua sendo feito pelo código SPSS, mas vale conferir."
            )

        if st.button(
            "Reaplicar Universo salvo nesta sessão",
            key="ponderacao_reaplicar_universo",
            use_container_width=True,
        ):
            st.session_state["ponderacao_reaplicar_perfil"] = True

    metodo_ui = st.radio(
        "Método",
        [
            "Método atual — razão simples e multiplicação",
            "Ajuste iterativo — tentar aproximar todas as margens",
        ],
        horizontal=False,
        key="ponderacao_metodo",
    )
    tolerancia = st.number_input(
        "Diferença máxima aceitável (pontos percentuais)",
        min_value=0.0,
        max_value=20.0,
        value=2.0,
        step=0.1,
        key="ponderacao_tolerancia",
    )
    max_iter = 100
    if metodo_ui.startswith("Ajuste iterativo"):
        max_iter = int(
            st.number_input(
                "Máximo de iterações",
                min_value=1,
                max_value=1000,
                value=100,
                step=10,
                key="ponderacao_max_iter",
            )
        )

    st.subheader("Amostra e universo")
    st.caption(
        "A frequência e o % da amostra são calculados da própria base. "
        "Preencha somente a coluna % Universo. Totais como 99,99 ou 100,01 são normalizados para 100 no cálculo."
    )

    editores = {}
    labels_por_variavel = {}
    reaplicar_perfil = bool(st.session_state.get("ponderacao_reaplicar_perfil", False))
    perfil = _perfil_sessao()
    for variavel in variaveis:
        labels_var = value_labels.get(variavel, {}) or {}
        base_editor = _tabela_editor(dados, variavel, labels_var)
        labels_por_variavel[variavel] = {
            linha["Código"]: str(linha["Categoria"])
            for _, linha in base_editor.iterrows()
        }

        estado_key = f"ponderacao_base_editor_{variavel}"
        assinatura_key = f"ponderacao_assinatura_editor_{variavel}"
        revisao_key = f"ponderacao_revisao_editor_{variavel}"
        assinatura = _assinatura_editor(base_editor)

        assinatura_mudou = st.session_state.get(assinatura_key) != assinatura
        if assinatura_mudou:
            novo_editor, _ = _aplicar_perfil_no_editor(base_editor, perfil, variavel)
            st.session_state[estado_key] = novo_editor
            st.session_state[assinatura_key] = assinatura
            st.session_state[revisao_key] = 0
        elif estado_key not in st.session_state:
            novo_editor, _ = _aplicar_perfil_no_editor(base_editor, perfil, variavel)
            st.session_state[estado_key] = novo_editor
        elif reaplicar_perfil:
            novo_editor, _ = _aplicar_perfil_no_editor(base_editor, perfil, variavel)
            st.session_state[estado_key] = novo_editor
            st.session_state[revisao_key] = (
                int(st.session_state.get(revisao_key, 0)) + 1
            )

        titulo_variavel = _nome_exibicao_variavel(meta, variavel)
        with st.expander(_descricao_variavel(meta, variavel), expanded=True):
            codigos_ordem = ", ".join(
                str(codigo) for codigo in st.session_state[estado_key]["Código"].tolist()
            )
            st.caption(
                f"Ordem dos códigos originais do SPSS: {codigos_ordem}. "
                f"O registro histórico usará o nome/rótulo '{titulo_variavel}'. "
                "Você pode preencher manualmente na tabela ou colar todos os percentuais de uma vez."
            )

            comp_perfil = _compatibilidade_perfil(base_editor, perfil, variavel)
            if variavel in perfil.get("variaveis", {}):
                st.caption(
                    f"Universo reutilizável: {comp_perfil['casados']}/{comp_perfil['total_atual']} "
                    "categorias atuais pareadas pelo código SPSS."
                )
                if comp_perfil["sem_universo"]:
                    codigos = ", ".join(str(item["codigo"]) for item in comp_perfil["sem_universo"])
                    st.warning(f"Sem Universo salvo nesta base: código(s) {codigos}.")
                if comp_perfil["ausentes_na_amostra"]:
                    detalhes = ", ".join(
                        f"{item['codigo']} ({item['categoria']} = {item['percentual']:.2f}%)"
                        for item in comp_perfil["ausentes_na_amostra"]
                    )
                    st.warning(
                        "Universo com categoria(s) sem casos nesta amostra: " + detalhes + ". "
                        "Se alguma meta for maior que 0%, o cálculo será bloqueado para evitar "
                        "normalizar o Universo incorretamente."
                    )
                if comp_perfil["labels_divergentes"]:
                    detalhes = "; ".join(
                        f"código {item['codigo']}: atual '{item['atual']}' / salvo '{item['salva']}'"
                        for item in comp_perfil["labels_divergentes"]
                    )
                    st.warning("Labels diferentes para o mesmo código SPSS: " + detalhes)

            texto_colado = st.text_area(
                "Colar % Universo em lote",
                placeholder=(
                    "Exemplo copiado de uma coluna do Excel:\n"
                    "12,62\n24,52\n25,32"
                ),
                height=105,
                key=f"ponderacao_colar_universo_{variavel}",
                help=(
                    "Aceita valores em linhas separadas, TAB, ponto e vírgula ou barra vertical. "
                    "A quantidade deve ser igual ao número de categorias exibidas."
                ),
            )
            col_aplicar, col_limpar = st.columns([3, 1])
            if col_aplicar.button(
                "Aplicar valores colados",
                key=f"ponderacao_aplicar_colados_{variavel}",
                use_container_width=True,
            ):
                try:
                    valores = _parse_percentuais_colados(texto_colado)
                    esperado = len(st.session_state[estado_key])
                    if len(valores) != esperado:
                        raise ValueError(
                            f"Foram encontrados {len(valores)} valores, mas '{variavel}' "
                            f"possui {esperado} categorias."
                        )
                    atualizado = st.session_state[estado_key].copy()
                    atualizado["% Universo"] = valores
                    st.session_state[estado_key] = atualizado
                    st.session_state[revisao_key] = (
                        int(st.session_state.get(revisao_key, 0)) + 1
                    )
                    st.success(
                        f"{len(valores)} percentuais aplicados em '{variavel}' na ordem dos códigos."
                    )
                except ValueError as exc:
                    st.error(str(exc))

            if col_limpar.button(
                "Limpar",
                key=f"ponderacao_limpar_universo_{variavel}",
                use_container_width=True,
            ):
                limpo = st.session_state[estado_key].copy()
                limpo["% Universo"] = np.nan
                st.session_state[estado_key] = limpo
                _remover_variavel_do_perfil(perfil, variavel)
                st.session_state[revisao_key] = (
                    int(st.session_state.get(revisao_key, 0)) + 1
                )

            revisao = int(st.session_state.get(revisao_key, 0))
            editado = st.data_editor(
                st.session_state[estado_key],
                hide_index=True,
                disabled=["Código", "Categoria", "Frequência", "% Amostra"],
                column_config={
                    "% Amostra": st.column_config.NumberColumn(format="%.2f"),
                    "% Universo": st.column_config.NumberColumn(
                        format="%.2f", min_value=0.0, max_value=100.0
                    ),
                },
                key=f"ponderacao_editor_{variavel}_{revisao}",
                use_container_width=True,
            )
            st.session_state[estado_key] = editado.copy()
            _sincronizar_editor_com_perfil(perfil, variavel, titulo_variavel, editado)
            st.session_state["ponderacao_perfil_universo"] = perfil
            editores[variavel] = editado

            total_universo = pd.to_numeric(
                editado["% Universo"], errors="coerce"
            ).sum(min_count=1)
            if pd.notna(total_universo):
                st.caption(f"Total informado do universo: {total_universo:.2f}%")

    if reaplicar_perfil:
        st.session_state["ponderacao_reaplicar_perfil"] = False

    identificacao_final = identificacao.strip() or identificacao_padrao
    assinatura_atual = _assinatura_calculo_ponderacao(
        uploaded=uploaded,
        variaveis=list(variaveis),
        editores=editores,
        metodo_ui=metodo_ui,
        tolerancia=float(tolerancia),
        max_iter=max_iter,
        identificacao=identificacao_final,
    )
    cache_key = "ponderacao_resultado_calculado"
    cache_resultado = st.session_state.get(cache_key)
    clicou_calcular = st.button(
        "Calcular ponderação",
        type="primary",
        key="ponderacao_calcular",
    )
    resultado_reutilizado = False

    if clicou_calcular:
        try:
            for variavel in variaveis:
                base_atual = _tabela_editor(
                    dados, variavel, value_labels.get(variavel, {}) or {}
                )
                comp = _compatibilidade_perfil(base_atual, perfil, variavel)
                ausentes_positivos = [
                    item for item in comp["ausentes_na_amostra"]
                    if float(item.get("percentual", 0.0)) > 0
                ]
                if ausentes_positivos:
                    detalhes = ", ".join(
                        f"código {item['codigo']} ({item['categoria']}: {item['percentual']:.2f}%)"
                        for item in ausentes_positivos
                    )
                    raise ValueError(
                        f"'{variavel}' possui categoria(s) do Universo sem nenhum caso na amostra atual: "
                        f"{detalhes}. Não é possível atingir essas metas após as exclusões."
                    )

            alvos = OrderedDict(
                (variavel, _alvos_do_editor(editores[variavel], variavel))
                for variavel in variaveis
            )

            if metodo_ui.startswith("Método atual"):
                resultado = calcular_razao_simples(
                    dados,
                    alvos,
                    labels=labels_por_variavel,
                    tolerancia_pp=float(tolerancia),
                )
            else:
                resultado = calcular_raking(
                    dados,
                    alvos,
                    labels=labels_por_variavel,
                    tolerancia_pp=float(tolerancia),
                    max_iteracoes=max_iter,
                )
        except Exception as exc:
            # Não mantém um resultado antigo visível para uma configuração que falhou.
            if isinstance(cache_resultado, dict) and cache_resultado.get("assinatura") == assinatura_atual:
                st.session_state.pop(cache_key, None)
            st.error(f"Não foi possível calcular a ponderação: {exc}")
            return
    elif (
        isinstance(cache_resultado, dict)
        and cache_resultado.get("assinatura") == assinatura_atual
        and cache_resultado.get("resultado") is not None
    ):
        resultado = cache_resultado["resultado"]
        resultado_reutilizado = True
    else:
        if isinstance(cache_resultado, dict):
            st.caption(
                "A configuração atual é diferente da última ponderação calculada. "
                "Clique em Calcular ponderação para gerar um novo resultado."
            )
        return

    if resultado_reutilizado:
        st.caption(
            "Resultado preservado na sessão. Você pode baixar todos os arquivos abaixo "
            "sem calcular a ponderação novamente."
        )

    resumo = resumo_pesos(resultado.pesos)
    st.subheader("Validação")
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Maior diferença", f"{resultado.maior_diferenca_pp:.2f} p.p.")
    c2.metric("Tolerância", f"{tolerancia:.2f} p.p.")
    c3.metric("Peso mínimo", f"{resumo['minimo']:.3f}")
    c4.metric("Peso máximo", f"{resumo['maximo']:.3f}")

    if resultado.convergiu:
        st.success(
            f"Todas as categorias ficaram dentro de {tolerancia:.2f} p.p. do universo."
        )
    else:
        st.warning(
            f"Ainda existem categorias acima de {tolerancia:.2f} p.p. de diferença. "
            "Revise a tabela abaixo e, se estiver no método atual, teste o ajuste iterativo."
        )

    if resultado.metodo == "Raking iterativo":
        st.caption(f"Iterações executadas: {resultado.iteracoes}.")

    diagnostico_exibicao = resultado.diagnostico.copy()
    for coluna in [
        "percentual_sem_peso",
        "percentual_universo",
        "percentual_com_peso",
        "diferenca_pp",
        "diferenca_abs_pp",
    ]:
        diagnostico_exibicao[coluna] = diagnostico_exibicao[coluna].round(2)
    diagnostico_exibicao["fator"] = diagnostico_exibicao["fator"].round(6)

    st.dataframe(
        diagnostico_exibicao[
            [
                "variavel",
                "codigo",
                "categoria",
                "frequencia_sem_peso",
                "percentual_sem_peso",
                "percentual_universo",
                "fator",
                "frequencia_com_peso",
                "percentual_com_peso",
                "diferenca_pp",
                "status",
            ]
        ],
        use_container_width=True,
        hide_index=True,
    )

    extremos = resumo["maximo"] / resumo["minimo"] if resumo["minimo"] and resumo["minimo"] > 0 else np.nan
    if np.isfinite(extremos) and extremos >= 10:
        st.warning(
            f"Há grande amplitude entre pesos (máx./mín. = {extremos:.1f}x). "
            "Vale revisar categorias com pouca base antes de usar o resultado."
        )

    # ------------------------------------------------------------------
    # Assistente de exclusões para testes de ponderação
    # ------------------------------------------------------------------
    assistente_relatorio = None
    assistente_ids_csv = None
    assistente_syntax = None

    st.markdown("---")
    st.subheader("Assistente de exclusões para ponderação")
    st.caption(
        "Ferramenta experimental de apoio a testes. O Analítico simula exclusões e recalcula a "
        "ponderação sem alterar o banco original. As travas de cota são tratadas como restrições "
        "rígidas: um caso só pode ser sugerido se todas as categorias dele continuarem acima da base mínima."
    )

    modo_objetivo_assistente = st.radio(
        "Objetivo da simulação",
        [
            "Chegar exatamente ao N-alvo informado",
            "Encontrar a melhor ponderação até o N mínimo",
        ],
        index=0,
        horizontal=True,
        key="ponderacao_assistente_objetivo",
        help=(
            "Use o primeiro modo quando o projeto precisa terminar exatamente com um N contratado, como 1.000. "
            "Use o segundo quando o N informado é apenas um piso: o assistente avalia cenários entre a base atual "
            "e esse mínimo e recomenda o N que produzir a menor diferença ponderada encontrada."
        ),
    )
    forcar_n_alvo = modo_objetivo_assistente.startswith("Chegar exatamente")

    if resultado.convergiu:
        if forcar_n_alvo:
            st.info(
                "A ponderação atual já está dentro da tolerância. Como o objetivo selecionado é chegar a um N-alvo, "
                "o assistente ainda pode procurar quais entrevistas retirar preservando o melhor equilíbrio possível."
            )
        else:
            st.info(
                "A ponderação atual já está dentro da tolerância configurada. Neste modo, o assistente ainda pode "
                "avaliar a faixa até o N mínimo e verificar se existe um ponto com ponderação ainda melhor; se a "
                "base atual continuar sendo o melhor cenário, ele recomendará zero exclusões."
            )

    id_padrao = _sugerir_variavel_id(dados)
    colunas_id = list(dados.columns)
    indice_id = colunas_id.index(id_padrao) if id_padrao in colunas_id else 0
    col_id, col_n = st.columns(2)
    with col_id:
        id_var_assistente = st.selectbox(
            "Variável de ID da entrevista",
            options=colunas_id,
            index=indice_id,
            key="ponderacao_assistente_id_var",
            help="O ID precisa ser único. Ele será usado na lista e na syntax SPSS de exclusão.",
        )
    with col_n:
        rotulo_n = (
            "Base-alvo exata após exclusões (N)"
            if forcar_n_alvo
            else "Base mínima permitida (N)"
        )
        n_minimo_assistente = int(
            st.number_input(
                rotulo_n,
                min_value=1,
                max_value=max(1, int(len(dados))),
                value=int(len(dados)),
                step=1,
                key="ponderacao_assistente_n_minimo",
                help=(
                    "No modo de N-alvo exato, o assistente tenta terminar neste total. "
                    "No modo de melhor ponderação, este valor é apenas o piso: por exemplo, com base 400 e mínimo "
                    "300, o resultado recomendado pode ser 352 se esse for o melhor cenário encontrado."
                ),
            )
        )

    codigos_protegidos_assistente = OrderedDict()
    with st.expander("Travas e proteções de exclusão", expanded=True):
        st.caption(
            "Uma entrevista é bloqueada para exclusão se pertencer a qualquer categoria protegida. "
            "Ex.: com mínimo 30, uma categoria com base 34 pode perder no máximo 4 casos; ao chegar a 30, "
            "nenhum outro respondente dessa categoria poderá ser sugerido."
        )
        proteger_minimo = st.checkbox(
            "Proteger automaticamente a base mínima de todas as categorias usadas na ponderação",
            value=True,
            key="ponderacao_assistente_proteger_minimo",
        )
        base_minima_categoria = int(
            st.number_input(
                "Base mínima por categoria",
                min_value=0,
                max_value=max(0, int(len(dados))),
                value=30,
                step=1,
                key="ponderacao_assistente_base_minima_categoria",
                disabled=not proteger_minimo,
                help=(
                    "Regra geral de segurança. Com 30, nenhuma categoria das variáveis de ponderação "
                    "poderá terminar com menos de 30 entrevistas por causa das sugestões do assistente."
                ),
            )
        ) if proteger_minimo else 0

        st.markdown("**Categorias que não podem ser excluídas de jeito algum**")
        for variavel in variaveis:
            editor_var = editores[variavel]
            opcoes = editor_var["Código"].tolist()
            labels_var_ui = {
                linha["Código"]: str(linha["Categoria"])
                for _, linha in editor_var.iterrows()
            }
            selecionados = st.multiselect(
                _descricao_variavel(meta, variavel),
                options=opcoes,
                default=[],
                format_func=lambda codigo, mapa=labels_var_ui: (
                    f"{codigo} — {mapa.get(codigo, str(codigo))}"
                ),
                key=f"ponderacao_assistente_protegidos_{variavel}",
                help=(
                    "Qualquer entrevista pertencente a uma destas categorias será totalmente bloqueada "
                    "para exclusão, mesmo que também pertença a outras categorias em excesso."
                ),
            )
            codigos_protegidos_assistente[variavel] = set(selecionados)

        linhas_capacidade = []
        for variavel in variaveis:
            for _, linha in editores[variavel].iterrows():
                codigo = linha["Código"]
                n_atual = int(linha["Frequência"])
                manual = codigo in codigos_protegidos_assistente.get(variavel, set())
                capacidade = 0 if manual else max(0, n_atual - base_minima_categoria)
                linhas_capacidade.append(
                    {
                        "Variável": variavel,
                        "Código": codigo,
                        "Categoria": str(linha["Categoria"]),
                        "N atual": n_atual,
                        "Mínimo": base_minima_categoria if proteger_minimo else 0,
                        "Máx. removível nesta cota": capacidade if proteger_minimo else n_atual,
                        "Nunca excluir": "SIM" if manual else "",
                    }
                )
        if linhas_capacidade:
            st.dataframe(
                pd.DataFrame(linhas_capacidade),
                use_container_width=True,
                hide_index=True,
            )

    max_exclusoes_assistente = max(0, int(len(dados)) - n_minimo_assistente)
    texto_exclusoes = (
        f"Exclusões necessárias para o N-alvo: {max_exclusoes_assistente}"
        if forcar_n_alvo
        else f"Máximo de exclusões que podem ser avaliadas: {max_exclusoes_assistente}"
    )
    st.caption(
        f"Base atual: {len(dados)} | {texto_exclusoes} | "
        f"Base mínima por categoria: {base_minima_categoria} | "
        f"Tolerância: ±{float(tolerancia):.2f} p.p."
    )

    alvos_assistente = OrderedDict(
        (variavel, _alvos_do_editor(editores[variavel], variavel))
        for variavel in variaveis
    )
    protegidos_assinatura = tuple(
        (variavel, tuple(sorted((repr(v) for v in valores))))
        for variavel, valores in codigos_protegidos_assistente.items()
    )
    assinatura_assistente = (
        assinatura_atual,
        str(id_var_assistente),
        int(n_minimo_assistente),
        modo_objetivo_assistente,
        int(base_minima_categoria),
        protegidos_assinatura,
    )
    assistente_key = "ponderacao_assistente_exclusoes_resultado"
    cache_assistente = st.session_state.get(assistente_key)

    if st.button(
        "Analisar possíveis exclusões",
        key="ponderacao_assistente_analisar",
        type="secondary",
        use_container_width=True,
        disabled=max_exclusoes_assistente <= 0,
    ):
        try:
            with st.spinner(
                "Analisando combinações de perfil, faixa de N e proteções de cota..."
            ):
                sugestao_limite = sugerir_exclusoes_ponderacao(
                    dados=dados,
                    id_var=id_var_assistente,
                    alvos=alvos_assistente,
                    n_minimo=n_minimo_assistente,
                    tolerancia_pp=float(tolerancia),
                    labels=labels_por_variavel,
                    max_perfis_avaliados=30,
                    parar_ao_atingir_tolerancia_bruta=False,
                    base_minima_categoria=base_minima_categoria,
                    codigos_protegidos=codigos_protegidos_assistente,
                    # Nos dois modos geramos uma trajetória até o piso permitido.
                    # No modo "melhor ponderação", depois escolhemos o melhor prefixo dessa trajetória.
                    forcar_atingir_n_minimo=True,
                )

                if forcar_n_alvo:
                    sugestao_final = sugestao_limite
                    resultado_pos = resultado
                    if sugestao_final.indices_excluir:
                        dados_teste = dados.drop(index=sugestao_final.indices_excluir)
                        resultado_pos = _calcular_ponderacao_para_teste_exclusao(
                            dados_teste,
                            alvos_assistente,
                            labels_por_variavel,
                            metodo_ui,
                            float(tolerancia),
                            max_iter,
                        )
                    resultado_no_limite = resultado_pos
                else:
                    melhor_k, resultado_pos = _refinar_prefixo_exclusoes_pelo_peso(
                        dados=dados,
                        sugestao=sugestao_limite,
                        resultado_atual=resultado,
                        alvos=alvos_assistente,
                        labels_por_variavel=labels_por_variavel,
                        metodo_ui=metodo_ui,
                        tolerancia=float(tolerancia),
                        max_iter=max_iter,
                    )

                    if melhor_k == len(sugestao_limite.ids_excluir):
                        sugestao_final = sugestao_limite
                    else:
                        sugestao_final = sugerir_exclusoes_ponderacao(
                            dados=dados,
                            id_var=id_var_assistente,
                            alvos=alvos_assistente,
                            n_minimo=int(len(dados)) - int(melhor_k),
                            tolerancia_pp=float(tolerancia),
                            labels=labels_por_variavel,
                            max_perfis_avaliados=30,
                            parar_ao_atingir_tolerancia_bruta=False,
                            base_minima_categoria=base_minima_categoria,
                            codigos_protegidos=codigos_protegidos_assistente,
                            forcar_atingir_n_minimo=True,
                        )
                        # O alvo temporário acima serve apenas para reconstruir exatamente o prefixo escolhido.
                        # Para documentação, preservamos o piso original informado pelo usuário.
                        sugestao_final.n_minimo = n_minimo_assistente
                        sugestao_final.atingiu_n_alvo = sugestao_final.n_final <= n_minimo_assistente
                        sugestao_final.faltam_exclusoes_para_alvo = max(
                            0, sugestao_final.n_final - n_minimo_assistente
                        )
                        sugestao_final.limite_exclusoes = max_exclusoes_assistente
                        sugestao_final.atingiu_limite_exclusoes = (
                            len(sugestao_final.ids_excluir) >= max_exclusoes_assistente
                            and not sugestao_final.atingiu_tolerancia_bruta
                        )
                    if melhor_k == 0 and max_exclusoes_assistente > 0:
                        sugestao_final.motivo_parada = (
                            "A base atual foi o melhor cenário ponderado entre os pontos avaliados; "
                            "nenhuma exclusão testada reduziu a maior diferença ponderada."
                        )
                    elif melhor_k > 0:
                        sugestao_final.motivo_parada = (
                            f"Melhor ponto ponderado encontrado na faixa permitida: N={sugestao_final.n_final} "
                            f"com {len(sugestao_final.ids_excluir)} exclusão(ões)."
                        )

                    resultado_no_limite = resultado
                    if sugestao_limite.indices_excluir:
                        try:
                            dados_limite = dados.drop(index=sugestao_limite.indices_excluir)
                            resultado_no_limite = _calcular_ponderacao_para_teste_exclusao(
                                dados_limite,
                                alvos_assistente,
                                labels_por_variavel,
                                metodo_ui,
                                float(tolerancia),
                                max_iter,
                            )
                        except Exception:
                            resultado_no_limite = resultado

            st.session_state[assistente_key] = {
                "assinatura": assinatura_assistente,
                "sugestao": sugestao_final,
                "sugestao_limite": sugestao_limite,
                "resultado_pos": resultado_pos,
                "resultado_no_limite": resultado_no_limite,
                "id_var": id_var_assistente,
                "objetivo": modo_objetivo_assistente,
            }
            cache_assistente = st.session_state[assistente_key]
        except Exception as exc:
            st.session_state.pop(assistente_key, None)
            cache_assistente = None
            st.error(f"Não foi possível analisar exclusões: {exc}")

    if (
        isinstance(cache_assistente, dict)
        and cache_assistente.get("assinatura") == assinatura_assistente
    ):
        sugestao = cache_assistente["sugestao"]
        sugestao_limite = cache_assistente["sugestao_limite"]
        resultado_pos = cache_assistente["resultado_pos"]
        resultado_no_limite = cache_assistente.get("resultado_no_limite", resultado_pos)

        a1, a2, a3, a4 = st.columns(4)
        a1.metric("Exclusões sugeridas", len(sugestao.ids_excluir))
        a2.metric("N estimado após exclusões", int(sugestao.n_final))
        a3.metric("Diferença ponderada atual", f"{resultado.maior_diferenca_pp:.2f} p.p.")
        a4.metric(
            "Diferença estimada após reponderar",
            f"{resultado_pos.maior_diferenca_pp:.2f} p.p.",
        )

        if forcar_n_alvo and not sugestao.atingiu_n_alvo:
            st.error(
                f"O N-alvo de {n_minimo_assistente} não pôde ser atingido sem violar as travas. "
                f"A simulação parou em N={sugestao.n_final}; ainda seriam necessárias "
                f"{sugestao.faltam_exclusoes_para_alvo} exclusão(ões)."
            )
        elif forcar_n_alvo and sugestao.atingiu_n_alvo:
            if float(resultado_pos.maior_diferenca_pp) <= float(tolerancia):
                st.success(
                    f"O N-alvo de {n_minimo_assistente} foi atingido respeitando as proteções. "
                    f"Após reponderar, a maior diferença estimada ficou em "
                    f"{resultado_pos.maior_diferenca_pp:.2f} p.p."
                )
            elif resultado_pos.maior_diferenca_pp < resultado.maior_diferenca_pp:
                st.warning(
                    f"O N-alvo foi atingido e a ponderação melhorou de "
                    f"{resultado.maior_diferenca_pp:.2f} para {resultado_pos.maior_diferenca_pp:.2f} p.p., "
                    "mas ainda há margens fora da tolerância configurada."
                )
            else:
                st.warning(
                    f"O N-alvo foi atingido respeitando as travas, porém a maior diferença estimada após "
                    f"reponderar ficou em {resultado_pos.maior_diferenca_pp:.2f} p.p. Revise as sobras antes de aplicar."
                )
        elif len(sugestao.ids_excluir) == 0:
            if max_exclusoes_assistente > 0:
                st.warning(
                    "Dentro do limite e das proteções informadas, a heurística não encontrou exclusões "
                    "que melhorassem a maior diferença ponderada."
                )
        elif resultado_pos.maior_diferenca_pp < resultado.maior_diferenca_pp:
            st.success(
                f"O melhor ponto ponderado encontrado na faixa de N={len(dados)} até N={n_minimo_assistente} "
                f"foi N={sugestao.n_final}, com {len(sugestao.ids_excluir)} exclusão(ões). "
                f"A maior diferença estimada caiu de {resultado.maior_diferenca_pp:.2f} para "
                f"{resultado_pos.maior_diferenca_pp:.2f} p.p. Recalcule a ponderação na base real depois de aplicar "
                "qualquer exclusão."
            )
        else:
            st.warning(
                "A aproximação das margens brutas não produziu melhora na maior diferença ponderada. "
                "As exclusões devem ser tratadas apenas como indicação para revisão manual."
            )

        if not forcar_n_alvo:
            if sugestao_limite.n_final > n_minimo_assistente:
                st.info(
                    f"As travas impediram que a trajetória fosse testada até N={n_minimo_assistente}. "
                    f"O menor N alcançável sem violar as proteções foi {sugestao_limite.n_final}."
                )
            elif sugestao.n_final > n_minimo_assistente:
                st.caption(
                    f"O N={n_minimo_assistente} era apenas o piso da busca. O assistente recomendou parar antes, "
                    f"em N={sugestao.n_final}, porque esse foi o melhor resultado ponderado encontrado."
                )

        if not sugestao.restricoes_ativas.empty:
            st.markdown("**Proteções ativas no ponto em que a simulação terminou**")
            st.dataframe(
                sugestao.restricoes_ativas,
                use_container_width=True,
                hide_index=True,
            )

        tabela_sobras = (
            sugestao_limite.sobras_remanescentes
            if not sugestao_limite.sobras_remanescentes.empty
            else sugestao.sobras_remanescentes
        )
        if not tabela_sobras.empty:
            st.markdown("**Categorias que ainda aparecem mais sobrando para exclusão**")
            sobras_exibicao = tabela_sobras.copy()
            for coluna in [
                "pct_depois",
                "pct_universo",
                "diferenca_depois_pp",
                "excesso_n_aprox",
            ]:
                if coluna in sobras_exibicao.columns:
                    sobras_exibicao[coluna] = pd.to_numeric(
                        sobras_exibicao[coluna], errors="coerce"
                    ).round(2)
            colunas_sobras = [
                c for c in [
                    "variavel", "codigo", "categoria", "n_depois", "pct_depois",
                    "pct_universo", "diferenca_depois_pp", "excesso_n_aprox"
                ] if c in sobras_exibicao.columns
            ]
            st.dataframe(
                sobras_exibicao[colunas_sobras],
                use_container_width=True,
                hide_index=True,
            )

        with st.expander("Ver diagnóstico das margens antes/depois das exclusões sugeridas"):
            diag_sug = sugestao.resumo_margens.copy()
            for coluna in [
                "pct_antes", "pct_universo", "diferenca_antes_pp",
                "pct_depois", "diferenca_depois_pp", "excesso_n_aprox",
            ]:
                if coluna in diag_sug.columns:
                    diag_sug[coluna] = pd.to_numeric(diag_sug[coluna], errors="coerce").round(2)
            st.dataframe(diag_sug, use_container_width=True, hide_index=True)

        if sugestao.ids_excluir:
            tipo_id_ui = st.radio(
                f"Tipo da variável de ID '{id_var_assistente}' no SPSS",
                ["Detectar automaticamente", "Numérica", "Texto"],
                index=0,
                horizontal=True,
                key="ponderacao_assistente_tipo_id",
            )
            tratar_texto = {
                "Detectar automaticamente": None,
                "Numérica": False,
                "Texto": True,
            }[tipo_id_ui]
            modo_ui = st.radio(
                "Código SPSS para os IDs sugeridos",
                [
                    "Marcar IDs para conferir antes de excluir",
                    "Excluir diretamente os IDs sugeridos",
                ],
                index=0,
                key="ponderacao_assistente_modo_syntax",
            )
            modo_syntax = "marcar" if modo_ui.startswith("Marcar") else "excluir"
            assistente_syntax = gerar_syntax_exclusao_ids(
                id_var_assistente,
                sugestao.ids_excluir,
                tratar_como_texto=tratar_texto,
                modo=modo_syntax,
            )
            assistente_ids_csv = pd.DataFrame(
                {id_var_assistente: sugestao.ids_excluir}
            ).to_csv(index=False).encode("utf-8-sig")
            assistente_relatorio = _gerar_relatorio_sugestao_exclusoes_excel(
                sugestao=sugestao,
                id_var=id_var_assistente,
                identificacao=identificacao_final,
                metodo=resultado.metodo,
                tolerancia_pp=float(tolerancia),
                maior_ponderada_antes_pp=float(resultado.maior_diferenca_pp),
                maior_ponderada_depois_pp=float(resultado_pos.maior_diferenca_pp),
                objetivo_exclusao=modo_objetivo_assistente,
            )

            ex1, ex2, ex3 = st.columns(3)
            ex1.download_button(
                "Baixar IDs sugeridos (.csv)",
                data=assistente_ids_csv,
                file_name=f"IDs_Exclusao_{_nome_arquivo_seguro(identificacao_final)}.csv",
                mime="text/csv",
                key="ponderacao_assistente_download_ids",
                use_container_width=True,
            )
            ex2.download_button(
                "Baixar código SPSS (.sps)",
                data=assistente_syntax.encode("utf-8"),
                file_name=f"Exclusoes_{_nome_arquivo_seguro(identificacao_final)}.sps",
                mime="text/plain",
                key="ponderacao_assistente_download_syntax",
                use_container_width=True,
            )
            ex3.download_button(
                "Baixar relatório da sugestão (.xlsx)",
                data=assistente_relatorio,
                file_name=f"Sugestao_Exclusoes_{_nome_arquivo_seguro(identificacao_final)}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="ponderacao_assistente_download_relatorio",
                use_container_width=True,
            )
            with st.expander("Ver código SPSS gerado"):
                st.code(assistente_syntax, language="text")

    if resultado_reutilizado:
        syntax = cache_resultado["syntax"]
        excel_tecnico = cache_resultado["excel_tecnico"]
        excel_calculo = cache_resultado.get("excel_calculo")
        excel_historico = cache_resultado.get("excel_historico")
        erro_excel_calculo = cache_resultado.get("erro_excel_calculo")
        erro_excel_historico = cache_resultado.get("erro_excel_historico")
        nome_registro = cache_resultado["nome_registro"]
    else:
        syntax = gerar_syntax_spss(resultado.fatores)
        excel_tecnico = _gerar_relatorio_excel(resultado, float(tolerancia), len(dados))

        erro_excel_calculo = None
        try:
            excel_calculo = _gerar_calculo_ponderacao_excel(
                resultado=resultado,
                dados=dados,
                meta=meta,
                variaveis=list(variaveis),
                arquivo_origem=getattr(uploaded, "name", "Base SPSS"),
                identificacao=identificacao_final,
            )
        except Exception as exc:
            excel_calculo = None
            erro_excel_calculo = str(exc)

        erro_excel_historico = None
        try:
            excel_historico = _gerar_arquivo_historico_excel(
                resultado=resultado,
                tolerancia_pp=float(tolerancia),
                dados=dados,
                meta=meta,
                variaveis=list(variaveis),
                labels_por_variavel=labels_por_variavel,
                arquivo_origem=getattr(uploaded, "name", "Base SPSS"),
                identificacao=identificacao_final,
            )
        except Exception as exc:
            excel_historico = None
            erro_excel_historico = str(exc)

        nome_registro = _nome_arquivo_seguro(identificacao_final)
        st.session_state[cache_key] = {
            "assinatura": assinatura_atual,
            "resultado": resultado,
            "syntax": syntax,
            "excel_tecnico": excel_tecnico,
            "excel_calculo": excel_calculo,
            "excel_historico": excel_historico,
            "erro_excel_calculo": erro_excel_calculo,
            "erro_excel_historico": erro_excel_historico,
            "nome_registro": nome_registro,
        }
        cache_resultado = st.session_state[cache_key]

    if erro_excel_calculo:
        st.warning(
            "Não foi possível montar a memória de cálculo da ponderação: "
            f"{erro_excel_calculo}"
        )
    if erro_excel_historico:
        st.warning(
            "Não foi possível montar o registro histórico no modelo enviado: "
            f"{erro_excel_historico}"
        )

    st.subheader("Arquivos finais")
    st.caption(
        "O download da base .sav ponderada está temporariamente oculto. "
        "A syntax SPSS continua disponível para aplicação do peso no banco."
    )

    pacote_zip = _gerar_pacote_ponderacao_zip(
        nome_registro=nome_registro,
        excel_historico=excel_historico,
        excel_calculo=excel_calculo,
        syntax=syntax,
        excel_tecnico=excel_tecnico,
        modelo_universo=modelo_universo,
        relatorio_exclusoes=assistente_relatorio,
        ids_exclusoes_csv=assistente_ids_csv,
        syntax_exclusoes=assistente_syntax,
    )
    st.download_button(
        "Baixar todos os arquivos (.zip)",
        data=pacote_zip,
        file_name=f"Ponderacao_{nome_registro}_arquivos.zip",
        mime="application/zip",
        key="ponderacao_download_todos_zip",
        type="primary",
        use_container_width=True,
        help=(
            "O ZIP cria uma pasta com o registro histórico, memória de cálculo, syntax, "
            "relatório técnico e modelo reutilizável do Universo. Se o Assistente de exclusões "
            "tiver uma sugestão ativa, inclui também o relatório, a lista de IDs e a syntax de "
            "exclusão. A base .sav não é incluída."
        ),
    )

    col1, col2, col3 = st.columns(3)
    if excel_historico is not None:
        col1.download_button(
            "Baixar registro histórico (.xlsx)",
            data=excel_historico,
            file_name=f"Ponderacao_{nome_registro}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="ponderacao_download_historico",
            help=(
                "Arquivo preenchido no padrão do modelo histórico: amostra sem peso, universo, "
                "amostra ponderada, diferenças e uma aba de registro técnico."
            ),
            use_container_width=True,
        )
    else:
        col1.warning("Registro histórico indisponível nesta execução.")

    if excel_calculo is not None:
        col2.download_button(
            "Baixar cálculo da ponderação (.xlsx)",
            data=excel_calculo,
            file_name=f"Calculo_da_Ponderacao_{nome_registro}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="ponderacao_download_calculo",
            help=(
                "Memória de cálculo inspirada na planilha usada pela equipe: mostra amostra, "
                "universo, razão, fatores, componentes peso1/peso2 e as linhas da syntax SPSS."
            ),
            use_container_width=True,
        )
    else:
        col2.warning("Memória de cálculo indisponível nesta execução.")

    col3.download_button(
        "Baixar syntax SPSS (.sps)",
        data=syntax.encode("utf-8"),
        file_name=f"Pesos_{nome_registro}.sps",
        mime="text/plain",
        key="ponderacao_download_sps",
        use_container_width=True,
    )

    st.caption(
        "O registro histórico preserva os nomes/rótulos e a ordem dos códigos que vierem do SPSS. "
        "Para aplicar o peso no banco, utilize a syntax gerada no SPSS."
    )

    with st.expander("Relatório técnico detalhado"):
        st.caption(
            "Mantém o relatório técnico da versão anterior, com uma aba por variável, fatores e diagnóstico completo."
        )
        st.download_button(
            "Baixar relatório técnico (.xlsx)",
            data=excel_tecnico,
            file_name=f"Relatorio_tecnico_{nome_registro}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="ponderacao_download_xlsx_tecnico",
        )
