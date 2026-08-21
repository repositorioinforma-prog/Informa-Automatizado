"""
Relatório Automatizado.

Orquestra, em sequência e com as perguntas necessárias em cada etapa, o
fluxo de finalização de um relatório descritivo — equivalente aos 15
códigos VBA usados manualmente hoje (ver conversa de origem). Trabalha
sempre em cima de UM ÚNICO arquivo, que vai sendo atualizado passo a
passo conforme a pessoa avança no assistente.

ESTADO ATUAL: TODOS os 16 códigos estão portados e ligados de verdade —
01, 02, 03, 04, 05 (com exclusão por base reduzida), 06, 07, 08, 09, 10,
11, 12, 14, 16 (core/relatorio_automatizado_math.py), 13
(core/cabecalho_correcao_math.py, precisa de arquivo de referência) e 15
(core/capas_resultados_math.py), além da etapa de Legendas (reaproveita
core/legendas_math.py).

Não depende do `dados` carregado no início do app: tem upload próprio.
"""
import io
import os
import base64

import openpyxl
import pandas as pd
import streamlit as st

from core.legendas_math import parsear_legenda_por_chave, aplicar_legendas_por_chave
from core.base_multiplas_math import (
    parsear_blocos,
    extrair_bases,
    calcular_linhas_base,
    gerar_workbook_com_base,
    bloco_eh_religiao,
)
from core.planilha_utils import worksheet_para_html
from core.pdf_preview import gerar_pdf_preview, wkhtmltopdf_disponivel
from core.cabecalho_correcao_math import parsear_blocos_cabecalho_referencia, aplicar_codigo_13
from core.capas_resultados_math import aplicar_codigo_15
from core.total_automatico_math import (
    encontrar_perguntas_com_base_reduzida,
    detectar_bases_divergentes,
    remover_linha_branca_multiplas_sem_base,
)
from core.divisor_tabelas_math import processar_workbook as processar_divisor_tabelas
from core.relatorio_automatizado_math import (
    aplicar_codigo_01,
    aplicar_codigo_02,
    aplicar_codigo_03,
    aplicar_codigo_04,
    aplicar_codigo_05,
    aplicar_codigo_06,
    aplicar_codigo_07,
    aplicar_codigo_08,
    aplicar_codigo_09,
    aplicar_codigo_10,
    aplicar_codigo_11,
    aplicar_codigo_12,
    aplicar_codigo_14,
    aplicar_codigo_16,
    remover_termo_do_sumario,
)
from core.relatorio_automatizado_math import TERMOS_EXCLUSAO_CODIGO_03


# =========================================================
# Utilidades de estado do assistente
# =========================================================
def _log(msg):
    st.session_state.setdefault("ra_log", []).append(msg)


def _ir_para(etapa):
    st.session_state["ra_step"] = etapa
    st.rerun()


def _reiniciar():
    for chave in list(st.session_state.keys()):
        if chave.startswith("ra_"):
            del st.session_state[chave]
    st.rerun()


def _stub(numero, nome):
    """
    Placeholder para um código ainda não portado para Python: não altera
    o arquivo, só registra no log que essa etapa foi (ou seria) aplicada.
    Será substituído pela lógica real de cada código, um de cada vez.
    """
    _log(f"Código {numero} ({nome}): placeholder — arquivo não alterado ainda.")


def _aplicar_no_arquivo(funcao, numero, nome, aproximado=False, args=(), no_workbook=False):
    """
    Carrega o arquivo em andamento (st.session_state["ra_wb_bytes"]),
    aplica a função no workbook ou na planilha ativa (conforme
    'no_workbook'), salva de volta e registra no log. 'args' passa
    parâmetros extras para a função (ex.: o limite de base reduzida).
    'aproximado=True' sinaliza no log que a etapa envolve uma estimativa
    (ex.: altura de linha) em vez de um cálculo exato.
    """
    wb = openpyxl.load_workbook(io.BytesIO(st.session_state["ra_wb_bytes"]), rich_text=True)
    alvo = wb if no_workbook else wb.active
    resultado = funcao(alvo, *args)

    saida = io.BytesIO()
    wb.save(saida)
    st.session_state["ra_wb_bytes"] = saida.getvalue()

    selo = " (estimativa aproximada — confira no Excel)" if aproximado else ""
    if isinstance(resultado, tuple):
        resultado_txt = ", ".join(str(x) for x in resultado)
    else:
        resultado_txt = f"{resultado} alteração(ões)"
    _log(f"Código {numero} ({nome}): {resultado_txt} aplicada(s){selo}.")


def _aplicar_sequencia(passos):
    """
    Aplica uma sequência de códigos com feedback visual de progresso —
    mostra em qual código está rodando no momento (via st.status), pra
    dar noção de quanto falta em relatórios grandes que demoram mais.

    'passos' é uma lista de dicts no formato:
        {"fn": funcao, "numero": "07", "nome": "...", "aproximado": bool,
         "args": (...), "no_workbook": bool}
    (mesmos parâmetros de `_aplicar_no_arquivo`, exceto fn/numero/nome
    que são obrigatórios).
    """
    with st.status("Aplicando códigos...", expanded=True) as status:
        for p in passos:
            status.update(label=f"Aplicando código {p['numero']} — {p['nome']}...")
            _aplicar_no_arquivo(
                p["fn"], p["numero"], p["nome"],
                aproximado=p.get("aproximado", False),
                args=p.get("args", ()),
                no_workbook=p.get("no_workbook", False),
            )
        status.update(label="Concluído!", state="complete")


# Ordem das etapas do assistente, só pra exibir "Etapa X de Y" no topo —
# não controla o fluxo em si (isso continua sendo feito pelas comparações
# "if etapa == ..." abaixo). Etapas de tela pura (aviso/upload) ficam de
# fora da contagem, já que não envolvem processamento.
_ETAPAS_COM_PROGRESSO = [
    ("limpeza_multiplas", "Limpeza de linha branca (Múltiplas sem base)"),
    ("base_multiplas", "Base nas Múltiplas"),
    ("origem", "Preenche células / bordas (01-02)"),
    ("novos_termos", "Ordenar tabelas (03)"),
    ("codigo_04", "Formatação de rótulos (04)"),
    ("etapa_13", "Corrigir cabeçalhos (13)"),
    ("base_reduzida", "Formata layout / Base reduzida (05)"),
    ("codigo_06_07", "DIN 9 + quebras de página (06-07)"),
    ("legendas", "Legendas"),
    ("codigo_08_12", "Base pequena até Perguntas (08-12)"),
    ("codigo_14", "Cabeçalho/rodapé (14)"),
    ("codigo_15", "Capas de resultados (15)"),
    ("final", "Concluído"),
]


def _mostrar_progresso(etapa_atual):
    nomes = [nome for chave, nome in _ETAPAS_COM_PROGRESSO]
    chaves = [chave for chave, nome in _ETAPAS_COM_PROGRESSO]
    if etapa_atual not in chaves:
        return
    idx = chaves.index(etapa_atual)
    st.progress(
        idx / (len(chaves) - 1),
        text=f"Etapa {idx + 1} de {len(chaves)}: {nomes[idx]}",
    )


# =========================================================
# Assistente
# =========================================================
def modulo_relatorio_automatizado():
    st.header("Relatório Automatizado")

    if "ra_step" not in st.session_state:
        st.session_state["ra_step"] = "aviso"
    if "ra_log" not in st.session_state:
        st.session_state["ra_log"] = []

    if st.session_state["ra_step"] != "aviso":
        if st.button("🔄 Reiniciar do zero", key="ra_reiniciar"):
            _reiniciar()

    etapa = st.session_state["ra_step"]
    _mostrar_progresso(etapa)

    # ---------------------------------------------------------------
    if etapa == "aviso":
        st.warning(
            "Antes de começar: importe aqui o **relatório descritivo já "
            "com as bases nas múltiplas aplicadas** (rode o módulo "
            "\"Base nas Múltiplas\" antes, se ainda não tiver feito isso)."
        )
        if st.button("Entendi, continuar", key="ra_aviso_continuar"):
            _ir_para("upload")
        return

    # ---------------------------------------------------------------
    if etapa == "upload":
        arquivo = st.file_uploader(
            "Relatório descritivo (.xlsx)", type=["xlsx"], key="ra_upload_relatorio"
        )
        if arquivo is not None:
            st.session_state["ra_wb_bytes"] = arquivo.read()
            st.session_state["ra_nome_arquivo"] = arquivo.name
            _log("Relatório carregado.")
            _ir_para("limpeza_multiplas")
        return

    # A partir daqui, sempre existe um arquivo em andamento em
    # st.session_state["ra_wb_bytes"] — mostramos isso pra situar a pessoa
    st.caption(f"Arquivo em andamento: **{st.session_state.get('ra_nome_arquivo', '?')}**")

    # ---------------------------------------------------------------
    if etapa == "limpeza_multiplas":
        st.subheader("Limpeza de linha branca (Múltiplas sem base)")
        st.caption(
            "Tabelas de pergunta do tipo Múltipla às vezes saem do SPSS "
            "com uma linha em branco entre o título e a tabela. Esta "
            "etapa procura essa linha SOMENTE nas tabelas de Múltipla "
            "que ainda não têm uma linha 'Base' por perto, e só remove "
            "quando a linha estiver realmente vazia — sem nenhum valor, "
            "borda ou cor. Se tiver qualquer conteúdo ou formatação, a "
            "linha fica intacta."
        )
        if st.button("Aplicar e continuar", key="ra_limpeza_continuar"):
            with st.spinner("Verificando tabelas de Múltipla..."):
                wb = openpyxl.load_workbook(io.BytesIO(st.session_state["ra_wb_bytes"]), rich_text=True)
                ws = wb.active
                resultado = remover_linha_branca_multiplas_sem_base(ws)
                saida = io.BytesIO()
                wb.save(saida)
                st.session_state["ra_wb_bytes"] = saida.getvalue()
            _log(
                f"Limpeza de linha branca (Múltiplas sem base): "
                f"{resultado['removidas']} linha(s) removida(s), "
                f"{resultado['mantidas']} mantida(s) por terem "
                f"conteúdo/formatação."
            )
            _ir_para("base_multiplas")
        return

    # ---------------------------------------------------------------
    if etapa == "base_multiplas":
        st.subheader("Base nas Múltiplas")
        st.caption(
            "Tabelas de Múltipla Escolha não saem do SPSS com a linha de "
            "Base. Se o relatório carregado tiver alguma, envie um "
            "arquivo com base (qualquer pergunta do mesmo projeto que já "
            "tenha 'Base') pra copiar a linha certa em cada tabela de "
            "múltiplas — casando pelo texto do cabeçalho, não pela "
            "posição da coluna. Se não tiver tabela de múltiplas, pode "
            "pular direto."
        )
        arquivo_bases = st.file_uploader(
            "Relatório com base (opcional)", type=["xlsx"], key="ra_upload_bases_multiplas"
        )

        col1, col2 = st.columns([1, 1])
        with col1:
            aplicar_bm = st.button(
                "Aplicar e continuar", key="ra_bm_aplicar", disabled=arquivo_bases is None
            )
        with col2:
            pular_bm = st.button("Este relatório não tem Múltiplas", key="ra_bm_pular")

        if pular_bm:
            _log("Base nas Múltiplas: etapa pulada.")
            _ir_para("origem")
            return

        if aplicar_bm and arquivo_bases is not None:
            with st.spinner("Lendo e casando as tabelas..."):
                try:
                    blocos_multiplas = parsear_blocos(io.BytesIO(st.session_state["ra_wb_bytes"]))
                    blocos_bases = parsear_blocos(arquivo_bases)
                except Exception as e:
                    st.error(f"Não foi possível ler os arquivos: {e}")
                    return

                if not blocos_multiplas:
                    st.error(
                        "Não encontrei nenhum bloco 'Titulo:' no relatório "
                        "carregado. Confirme se ele segue o formato padrão "
                        "de exportação do SPSS."
                    )
                    return
                if not blocos_bases:
                    st.error(
                        "Não encontrei nenhum bloco 'Titulo:' no relatório "
                        "com base. Confirme se esse é o arquivo certo."
                    )
                    return

                base_total, por_par, por_categoria, por_grupo_sem_categoria = extrair_bases(blocos_bases)
                if base_total is None:
                    st.error(
                        "Não encontrei nenhuma linha 'Base' no relatório "
                        "com base. Confirme se esse é o arquivo certo."
                    )
                    return

                blocos_religiao_bases = [b for b in blocos_bases if bloco_eh_religiao(b)]
                indice_religiao = None
                if blocos_religiao_bases:
                    indice_religiao = extrair_bases(blocos_bases, apenas_blocos=blocos_religiao_bases)[1:]

                linhas_base = calcular_linhas_base(
                    blocos_multiplas, base_total, por_par, por_categoria, por_grupo_sem_categoria,
                    indice_religiao=indice_religiao,
                )

                wb_novo = gerar_workbook_com_base(
                    io.BytesIO(st.session_state["ra_wb_bytes"]), blocos_multiplas, linhas_base
                )
                saida = io.BytesIO()
                wb_novo.save(saida)
                st.session_state["ra_wb_bytes"] = saida.getvalue()

            n_ja_tinham = sum(1 for lb in linhas_base if lb.get("ja_tinha_base"))
            n_religiao = sum(1 for lb in linhas_base if lb.get("eh_religiao"))
            total_nao_encontradas = sum(len(lb["nao_encontradas"]) for lb in linhas_base)
            msg = f"Base nas Múltiplas: {len(blocos_multiplas)} tabela(s) no total"
            if n_ja_tinham:
                msg += f", {n_ja_tinham} já tinha(m) base"
            if n_religiao:
                msg += f", {n_religiao} de religião (casamento restrito)"
            if total_nao_encontradas:
                msg += f", {total_nao_encontradas} coluna(s) sem base correspondente"
            _log(msg + ".")
            _ir_para("origem")
        return

    # ---------------------------------------------------------------
    if etapa == "origem":
        origem = st.radio(
            "O relatório foi gerado no SPSS Relatoria ou no SPSS Inovação?",
            ["SPSS Relatoria", "SPSS Inovação"],
            key="ra_origem"
        )
        if st.button("Continuar", key="ra_origem_continuar"):
            if origem == "SPSS Inovação":
                _aplicar_sequencia([
                    {"fn": aplicar_codigo_01, "numero": "01", "nome": "Preenche células do SPSS pirata"},
                    {"fn": aplicar_codigo_02, "numero": "02", "nome": "Muda Bordas"},
                ])
            else:
                _log("SPSS Relatoria selecionado — códigos 01 e 02 pulados.")
            _ir_para("novos_termos")
        return

    # ---------------------------------------------------------------
    if etapa == "novos_termos":
        st.subheader("Ordenar tabelas (código 03)")
        st.caption(
            f"A lista padrão já tem {len(TERMOS_EXCLUSAO_CODIGO_03)} termos "
            "herdados do VBA original (ex.: 'Base', 'Não sabe', 'Branco/Nulo')."
        )
        termos_novos = st.text_area(
            "Termos extras a incluir na lista de exclusão (um por linha, "
            "opcional) — essas linhas ficam fixas na posição original e não "
            "entram na reordenação, além dos termos padrão:",
            key="ra_termos_novos",
            height=120,
        )
        if st.button("Continuar", key="ra_termos_continuar"):
            termos_extras = [t.strip() for t in termos_novos.splitlines() if t.strip()]
            st.session_state["ra_termos_novos_lista"] = termos_extras
            lista_final = TERMOS_EXCLUSAO_CODIGO_03 + termos_extras
            _aplicar_sequencia([
                {"fn": aplicar_codigo_03, "numero": "03", "nome": "Ordenar tabelas", "args": (lista_final,)},
            ])
            _ir_para("codigo_04")
        return

    # ---------------------------------------------------------------
    if etapa == "codigo_04":
        st.subheader("Formatação de rótulos específicos (código 04)")
        st.caption(
            "Aplica um conjunto de regras fixas de formatação em rótulos "
            "específicos (ex.: Denominações Religiosas, Ensino "
            "Fundamental)."
        )
        if st.button("Aplicar e continuar", key="ra_codigo04_continuar"):
            _aplicar_sequencia([
                {"fn": aplicar_codigo_04, "numero": "04", "nome": "Formatação de rótulos específicos"},
            ])
            _ir_para("etapa_13")
        return

    # ---------------------------------------------------------------
    if etapa == "base_reduzida":
        st.subheader("Formata layout / Base reduzida (código 05)")
        limite = st.number_input(
            "Qual o limite mínimo de Base reduzida? "
            "(colunas com base menor que esse valor serão removidas)",
            min_value=1, value=25, step=1, key="ra_limite_base_reduzida"
        )
        if st.button("Aplicar e continuar", key="ra_base_reduzida_continuar"):
            _aplicar_sequencia([
                {
                    "fn": aplicar_codigo_05, "numero": "05", "nome": "Formata layout / Base reduzida",
                    "args": (limite,),
                },
            ])
            _ir_para("codigo_06_07")
        return

    # ---------------------------------------------------------------
    if etapa == "codigo_06_07":
        st.subheader("DIN 9 + linhas e quebras de página (códigos 06-07)")
        st.caption(
            "Ajusta a fonte das linhas de Pergunta pra tamanho 9 e insere "
            "linhas + quebras de página manuais depois de cada bloco de "
            "pergunta. Em relatórios grandes pode levar alguns segundos."
        )
        if st.button("Aplicar e continuar", key="ra_codigo0607_continuar"):
            _aplicar_sequencia([
                {"fn": aplicar_codigo_06, "numero": "06", "nome": "Ajuste DIN 9 na linha de Pergunta"},
                {"fn": aplicar_codigo_07, "numero": "07", "nome": "Adiciona linhas e quebra de página"},
            ])
            _ir_para("legendas")
        return

    # ---------------------------------------------------------------
    if etapa == "legendas":
        st.subheader("Legendas")
        st.caption(
            "Envie o arquivo com a(s) legenda(s) a serem usadas neste "
            "relatório (mesmo formato do módulo \"Legendas\") — só entram "
            "os itens cujo código/rótulo aparece no cabeçalho de cada "
            "tabela específica, não a legenda inteira em todas."
        )
        arquivo_legenda = st.file_uploader(
            "Arquivo de legendas (.xlsx)", type=["xlsx"], key="ra_upload_legenda"
        )

        col1, col2 = st.columns([1, 1])
        with col1:
            aplicar = st.button("Aplicar legendas e continuar", key="ra_legendas_aplicar")
        with col2:
            pular = st.button("Este relatório não precisa de legenda", key="ra_legendas_pular")

        if pular:
            _log("Legendas: etapa pulada (relatório sem tabelas territoriais).")
            _ir_para("codigo_08_12")
            return

        if aplicar:
            if not arquivo_legenda:
                st.warning("Envie o arquivo de legendas antes de continuar.")
                return
            with st.spinner("Aplicando legendas..."):
                try:
                    dados_referencia = parsear_legenda_por_chave(arquivo_legenda)
                    if not dados_referencia["mapa"]:
                        st.error(
                            "Não encontrei nenhum item de legenda reconhecível "
                            "no arquivo de referência."
                        )
                        return

                    wb = openpyxl.load_workbook(io.BytesIO(st.session_state["ra_wb_bytes"]), rich_text=True)
                    ws = wb.active
                    n_inseridas = aplicar_legendas_por_chave(ws, dados_referencia)

                    saida = io.BytesIO()
                    wb.save(saida)
                    st.session_state["ra_wb_bytes"] = saida.getvalue()
                except Exception as e:
                    st.error(f"Não foi possível aplicar as legendas: {e}")
                    return

            if n_inseridas:
                _log(f"Legendas: {n_inseridas} tabela(s) receberam legenda.")
            else:
                st.info(
                    "Não encontrei nenhuma tabela cujo cabeçalho batesse com "
                    "os códigos/rótulos da referência — seguindo sem aplicar legenda."
                )
                _log("Legendas: nenhuma tabela compatível encontrada, nada a fazer.")
            _ir_para("codigo_08_12")
        return

    # ---------------------------------------------------------------
    if etapa == "codigo_08_12":
        st.subheader("Base pequena até Perguntas (códigos 08-12)")
        st.caption(
            "Marca bases pequenas, ajusta altura de labels/títulos/"
            "perguntas e o título da Renda. Cinco códigos em sequência — "
            "pode levar um tempinho em relatórios grandes."
        )
        if st.button("Aplicar e continuar", key="ra_codigo0812_continuar"):
            _aplicar_sequencia([
                {"fn": aplicar_codigo_08, "numero": "08", "nome": "Base pequena"},
                {
                    "fn": aplicar_codigo_09, "numero": "09", "nome": "Ajusta a altura das labels",
                    "aproximado": True,
                },
                {"fn": aplicar_codigo_10, "numero": "10", "nome": "Ajusta o título da Renda"},
                {
                    "fn": aplicar_codigo_11, "numero": "11", "nome": "Ajusta a altura dos Títulos",
                    "aproximado": True,
                },
                {
                    "fn": aplicar_codigo_12, "numero": "12", "nome": "Ajuste da altura das Perguntas",
                    "aproximado": True,
                },
            ])
            _ir_para("codigo_14")
        return

    # ---------------------------------------------------------------
    if etapa == "etapa_13":
        st.subheader("Corrigir cabeçalhos repetidos (código 13)")
        st.caption(
            "Sobe um arquivo de referência com o cabeçalho \"correto\" de "
            "cada tipo de segmentação (Sexo, Escolaridade, Religião, "
            "Regiões etc. — um bloco \"Titulo: ...\" por tipo, igual ao "
            "relatório, cada um com a linha de grupo + linha de "
            "subcategoria certas). O app identifica o tipo de cada tabela "
            "do relatório pelo conjunto de rótulos do cabeçalho e "
            "substitui pela versão de referência — corrige texto com "
            "artefato de codificação, hifenização diferente e mesclagens "
            "quebradas/duplicadas."
        )
        st.info(
            "**Sexo, Idade, Renda, Escolaridade, Religiões, Voto no 2º "
            "Turno e Avaliações (Aprovação/Reprovação) já vêm "
            "pré-configurados** nos padrões do relatório — não precisa "
            "subir arquivo de referência pra corrigir esses tipos. Só é "
            "necessário subir um arquivo aqui se tiver **outra** "
            "segmentação vindo quebrada."
        )
        arquivo_ref_13 = st.file_uploader(
            "Arquivo de referência dos cabeçalhos (.xlsx)", type=["xlsx"], key="ra_ref_13"
        )
        col1, col2 = st.columns([1, 1])
        with col1:
            aplicar13 = st.button(
                "Aplicar e continuar", key="ra_aplicar_13", disabled=arquivo_ref_13 is None
            )
        with col2:
            pular13 = st.button("Pular esta etapa", key="ra_pular_13")

        if aplicar13 and arquivo_ref_13 is not None:
            with st.spinner("Corrigindo cabeçalhos..."):
                blocos_ref = parsear_blocos_cabecalho_referencia(arquivo_ref_13)
                wb = openpyxl.load_workbook(io.BytesIO(st.session_state["ra_wb_bytes"]), rich_text=True)
                ws = wb.active
                n = aplicar_codigo_13(ws, blocos_ref)
                saida = io.BytesIO()
                wb.save(saida)
                st.session_state["ra_wb_bytes"] = saida.getvalue()
            _log(f"Código 13 (Corrigir cabeçalhos repetidos): {n} bloco(s) corrigido(s).")
            _ir_para("base_reduzida")
        elif pular13:
            _log("Código 13 (Corrigir cabeçalhos repetidos): pulado.")
            _ir_para("base_reduzida")
        return

    # ---------------------------------------------------------------
    if etapa == "codigo_14":
        st.subheader("Cabeçalho/rodapé de impressão fixos (código 14)")
        st.caption("Aplica o aviso legal fixo no cabeçalho e o número de página no rodapé.")
        if st.button("Aplicar e continuar", key="ra_codigo14_continuar"):
            _aplicar_sequencia([
                {
                    "fn": aplicar_codigo_14, "numero": "14", "nome": "Cabeçalho/rodapé de impressão fixos",
                    "no_workbook": True,
                },
            ])
            _ir_para("codigo_15")
        return

    # ---------------------------------------------------------------
    if etapa == "codigo_15":
        st.subheader("Capas de resultados (código 15)")
        st.caption(
            "Insere as capas de separação \"RESULTADOS PELO TOTAL\" (com "
            "6 páginas em branco antes) e \"RESULTADOS PELOS SEGMENTOS\" "
            "logo antes da primeira tabela segmentada. Depende das "
            "quebras de página do código 07 já estarem no arquivo."
        )
        if st.button("Aplicar e continuar", key="ra_codigo15_continuar"):
            with st.spinner("Inserindo capas..."):
                wb = openpyxl.load_workbook(io.BytesIO(st.session_state["ra_wb_bytes"]), rich_text=True)
                ws = wb.active
                resultado = aplicar_codigo_15(ws)
                saida = io.BytesIO()
                wb.save(saida)
                st.session_state["ra_wb_bytes"] = saida.getvalue()
            if resultado["status"] == "ok":
                _log(
                    f"Código 15 (Capas de resultados): capa do Total na linha "
                    f"{resultado['linha_capa_total']}"
                    + (
                        f", capa dos Segmentos na linha {resultado['linha_capa_segmentos']}."
                        if resultado["linha_capa_segmentos"]
                        else " (nenhuma tabela segmentada encontrada)."
                    )
                )
            else:
                st.warning(f"Código 15: {resultado['mensagem']} — etapa pulada.")
                _log(f"Código 15 (Capas de resultados): {resultado['mensagem']}")
            _ir_para("final")
        return

    # ---------------------------------------------------------------
    if etapa == "final":
        st.success("Fluxo concluído!")

        wb_final = openpyxl.load_workbook(io.BytesIO(st.session_state["ra_wb_bytes"]), rich_text=True)
        ws_final = wb_final.active

        perguntas_base_reduzida = encontrar_perguntas_com_base_reduzida(ws_final)
        if perguntas_base_reduzida:
            lista_perguntas = "\n".join(f"- {t}" for t in perguntas_base_reduzida)
            st.warning(
                "Este relatório tem pergunta(s) com **Base reduzida** (feitas só "
                "pra uma parte da amostra). Não esqueça de colocar o \"Para "
                "quem...\" explicando o filtro em cada uma:\n\n" + lista_perguntas
            )

        divergencia = detectar_bases_divergentes(ws_final)
        if divergencia["divergentes"]:
            lista_divergentes = "\n".join(
                f"- {d['titulo']}: base = {d['valor']}" for d in divergencia["divergentes"]
            )
            st.warning(
                f"A maioria das tabelas tem Base = {divergencia['valor_maioria']}, "
                f"mas encontrei tabela(s) com um valor diferente (e sem estar "
                f"marcada(s) como 'Base reduzida'). Vale conferir se é "
                f"intencional:\n\n" + lista_divergentes
            )

        st.warning(
            "Checklist antes de entregar:\n\n"
            "- [ ] Colocar a capa e a contra-capa\n"
            "- [ ] Colocar os cabeçalhos nas perguntas de Múltiplas\n"
            "- [ ] Conferir se a metodologia da capa não mistura tempos "
            "verbais (deve estar tudo no passado)\n"
            "- [ ] Ver se precisa da página do GDS ou se a pesquisa é online\n"
            "- [ ] Conferir o relatório inteiro\n"
            "- [ ] Rodar o Índice (código 16, em Códigos Individuais)"
        )

        st.subheader("Log das etapas")
        for linha in st.session_state.get("ra_log", []):
            st.write(f"- {linha}")

        st.subheader("Logo da Informa")
        st.caption(
            "O passo de aplicar o logo direto no cabeçalho do Excel foi "
            "removido daqui — não funcionava de forma confiável (imagem "
            "em cabeçalho é um recurso legado que o Excel/openpyxl tratam "
            "de forma frágil), então esse passo passou a ser feito "
            "manualmente, direto no Excel. Baixe a imagem abaixo caso não "
            "tenha ela disponível no computador de quem for rodar o "
            "relatório."
        )
        caminho_logo_padrao = os.path.join(
            os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "assets", "logo.jpg"
        )
        if os.path.exists(caminho_logo_padrao):
            with open(caminho_logo_padrao, "rb") as f:
                logo_bytes = f.read()
            st.download_button(
                "🖼️ Baixar logo da Informa",
                data=logo_bytes,
                file_name="logo_informa.jpg",
                mime="image/jpeg",
                key="ra_download_logo",
            )

        st.download_button(
            "Baixar relatório processado",
            data=st.session_state["ra_wb_bytes"],
            file_name=f"processado_{st.session_state.get('ra_nome_arquivo', 'relatorio.xlsx')}",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="ra_download_final"
        )

        st.subheader("Pré-visualização do resultado")

        if wkhtmltopdf_disponivel():
            if st.button("📄 Gerar pré-visualização em PDF (relatório inteiro)", key="ra_gerar_pdf_preview"):
                with st.spinner("Convertendo o relatório inteiro pra PDF — pode levar alguns segundos..."):
                    try:
                        st.session_state["ra_pdf_preview_bytes"] = gerar_pdf_preview(st.session_state["ra_wb_bytes"])
                    except RuntimeError as e:
                        st.session_state.pop("ra_pdf_preview_bytes", None)
                        st.error(f"Não consegui gerar a pré-visualização em PDF: {e}")

            if "ra_pdf_preview_bytes" in st.session_state:
                pdf_bytes = st.session_state["ra_pdf_preview_bytes"]
                base64_pdf = base64.b64encode(pdf_bytes).decode("utf-8")
                st.components.v1.html(
                    f'<iframe src="data:application/pdf;base64,{base64_pdf}" '
                    f'width="100%" height="900" style="border:1px solid #d1d5db;"></iframe>',
                    height=920,
                )
                st.download_button(
                    "Baixar esse PDF de pré-visualização", data=pdf_bytes,
                    file_name="preview_relatorio.pdf", mime="application/pdf",
                    key="ra_download_pdf_preview",
                )
        else:
            st.caption(
                "⚠️ Pré-visualização em PDF não disponível neste ambiente "
                "(wkhtmltopdf não está instalado)."
            )

        with st.expander("Pré-visualização rápida em tabela (mais leve, só valores aproximados)"):
            wb_preview = openpyxl.load_workbook(io.BytesIO(st.session_state["ra_wb_bytes"]))
            aba_escolhida = st.selectbox(
                "Aba", wb_preview.sheetnames, key="ra_preview_aba"
            ) if len(wb_preview.sheetnames) > 1 else wb_preview.sheetnames[0]
            ws_preview = wb_preview[aba_escolhida]

            html_preview, truncou_linhas, truncou_colunas = worksheet_para_html(ws_preview)
            st.components.v1.html(html_preview, height=620, scrolling=True)

            avisos = []
            if truncou_linhas:
                avisos.append(f"mostrando as primeiras 150 linhas de {ws_preview.max_row}")
            if truncou_colunas:
                avisos.append(f"mostrando as primeiras 30 colunas de {ws_preview.max_column}")
            aviso_txt = " e ".join(avisos)
            st.caption(
                (f"⚠️ Pré-visualização parcial ({aviso_txt}) — " if aviso_txt else "")
                + "reproduz mesclagem, cor de fundo, negrito/itálico, alinhamento e bordas, "
                "mas é uma aproximação; o arquivo baixado é a fonte da verdade."
            )

        st.divider()
        st.subheader("Dividir Tabelas + Índice (opcional)")
        st.caption(
            "Continua direto daqui, em cima do arquivo que acabou de ser "
            "gerado — sem precisar baixar e reenviar em outra tela: "
            "primeiro divide as tabelas que passarem do limite de altura "
            "útil por página, depois gera o Índice (que precisa rodar "
            "DEPOIS da divisão, já que dividir tabela desloca a "
            "paginação — gerar o índice antes deixaria os números de "
            "página errados)."
        )
        if st.button("Dividir tabelas e gerar Índice", key="ra_dividir_e_indexar"):
            with st.status("Dividindo tabelas e gerando o Índice...", expanded=True) as status:
                wb_div = openpyxl.load_workbook(io.BytesIO(st.session_state["ra_wb_bytes"]), rich_text=True)
                status.update(label="Analisando e dividindo tabelas...")
                resumo_divisor = processar_divisor_tabelas(wb_div)
                status.update(label="Gerando o Índice...")
                n_titulos = aplicar_codigo_16(wb_div.active)
                saida_div = io.BytesIO()
                wb_div.save(saida_div)
                status.update(label="Concluído!", state="complete")

            st.session_state["ra_wb_bytes_dividido"] = saida_div.getvalue()
            st.session_state["ra_resumo_divisor"] = resumo_divisor
            st.success(f"Tabelas processadas. Índice gerado com {n_titulos} título(s) listado(s).")

        resumo_divisor = st.session_state.get("ra_resumo_divisor")
        wb_bytes_dividido = st.session_state.get("ra_wb_bytes_dividido")
        if resumo_divisor and wb_bytes_dividido is not None:
            c1, c2, c3, c4, c5, c6 = st.columns(6)
            c1.metric("Abas analisadas", resumo_divisor["sheets_analyzed"])
            c2.metric("Tabelas candidatas", resumo_divisor["tables_candidatas"])
            c3.metric("Divididas", resumo_divisor["tables_divididas"])
            c4.metric("Legenda movida", resumo_divisor["legenda_movida"])
            c5.metric("Não corrigidas", resumo_divisor["nao_corrigidas"])
            c6.metric("Partes geradas", resumo_divisor["partes_geradas"])

            if resumo_divisor["nao_corrigidas"]:
                titulos_nao_corrigidas = [
                    a["titulo"] for a in resumo_divisor["acoes"] if a["acao"] == "nao_corrigida"
                ]
                lista_nc = "\n".join(f"- {t}" for t in titulos_nao_corrigidas)
                st.warning(
                    "Tabela(s) protegida(s) sem legenda (ou onde nem sem "
                    "legenda cabe nas linhas úteis) — passam do limite "
                    "mas não têm como ser corrigidas automaticamente sem "
                    "dividir os labels no meio. Precisam de ajuste "
                    "manual:\n\n" + lista_nc
                )

            if resumo_divisor["tables_candidatas"]:
                rows_divisor = [
                    {
                        "Aba": a["table"].sheet,
                        "Tabela": a["titulo"],
                        "Ação": {
                            "dividida": "Dividida em partes",
                            "legenda_movida": "Legenda movida pra página própria",
                            "nao_corrigida": "Não corrigida automaticamente",
                        }.get(a["acao"], a["acao"]),
                        "Partes": a.get("parts", "-"),
                    }
                    for a in resumo_divisor["acoes"]
                ]
                st.dataframe(pd.DataFrame(rows_divisor), use_container_width=True, hide_index=True)

            original_name = st.session_state.get("ra_nome_arquivo", "relatorio.xlsx")
            if "." in original_name:
                stem, ext = original_name.rsplit(".", 1)
                output_name_final = f"{stem}_tabelas_divididas_com_indice.{ext}"
            else:
                output_name_final = original_name + "_tabelas_divididas_com_indice.xlsx"

            st.download_button(
                "Baixar relatório dividido + com Índice",
                data=wb_bytes_dividido,
                file_name=output_name_final,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="ra_download_dividido_indice",
            )

            st.markdown("**Prévia do Sumário**")
            wb_preview_final = openpyxl.load_workbook(io.BytesIO(wb_bytes_dividido), rich_text=True)
            sumario_ws_final = wb_preview_final["Sumário"]
            preview_rows_final = [
                {
                    "Título": sumario_ws_final.cell(row=r, column=1).value,
                    "Página": sumario_ws_final.cell(row=r, column=2).value,
                }
                for r in range(1, sumario_ws_final.max_row + 1)
                if sumario_ws_final.cell(row=r, column=1).value is not None
            ]
            st.dataframe(pd.DataFrame(preview_rows_final), use_container_width=True, hide_index=True)

            st.caption("Remover termo de todos os títulos do Sumário:")
            col_a, col_b = st.columns(2)
            with col_a:
                if st.button("Remover \"(Estimulada e Única)\"", key="ra_remover_unica"):
                    wb_edit = openpyxl.load_workbook(io.BytesIO(wb_bytes_dividido), rich_text=True)
                    n = remover_termo_do_sumario(wb_edit, "(Estimulada e Única)")
                    saida_edit = io.BytesIO()
                    wb_edit.save(saida_edit)
                    st.session_state["ra_wb_bytes_dividido"] = saida_edit.getvalue()
                    st.success(f"Termo removido de {n} título(s).")
                    st.rerun()
            with col_b:
                if st.button("Remover \"(Estimulada e Múltipla)\"", key="ra_remover_multipla"):
                    wb_edit = openpyxl.load_workbook(io.BytesIO(wb_bytes_dividido), rich_text=True)
                    n = remover_termo_do_sumario(wb_edit, "(Estimulada e Múltipla)")
                    saida_edit = io.BytesIO()
                    wb_edit.save(saida_edit)
                    st.session_state["ra_wb_bytes_dividido"] = saida_edit.getvalue()
                    st.success(f"Termo removido de {n} título(s).")
                    st.rerun()
        return
