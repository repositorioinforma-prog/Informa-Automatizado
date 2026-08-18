"""
Total Automático.

Assistente sequencial parecido com o Relatório Automatizado, mas com
menos etapas — pensado pra relatórios "pelo total": limpeza de linha
branca nas Múltiplas sem base, Base nas Múltiplas (por arquivo ou valor
manual), SPSS Relatoria/Inovação (01-02), Ordenar tabelas (03), layout
(sem exclusão por Base reduzida), DIN 9 + linhas/quebras (06-07),
ajuste de altura de labels/títulos/perguntas (09/11/12) e, por fim, uma
capa "Resultados pelo total" (com os campos de bairro/cidade e mês/ano
digitáveis) acima da primeira tabela.

Usa prefixo de chave "ta_" no session_state (Total Automático), bem
diferente de "ra_" (Relatório Automatizado), pra não colidir caso a
pessoa alterne entre os dois módulos na mesma sessão do navegador.

Trabalha sempre em cima de UM ÚNICO arquivo, que vai sendo atualizado
passo a passo conforme a pessoa avança no assistente. Não depende do
`dados` carregado no início do app: tem upload próprio.
"""
import io

import openpyxl
import streamlit as st

from core.base_multiplas_math import (
    parsear_blocos,
    extrair_bases,
    calcular_linhas_base,
    calcular_linhas_base_manual,
    gerar_workbook_com_base,
    bloco_eh_religiao,
)
from core.planilha_utils import worksheet_para_html
from core.relatorio_automatizado_math import (
    aplicar_codigo_01,
    aplicar_codigo_02,
    aplicar_codigo_03,
    aplicar_codigo_06,
    aplicar_codigo_07,
    aplicar_codigo_09,
    aplicar_codigo_11,
    aplicar_codigo_12,
    TERMOS_EXCLUSAO_CODIGO_03,
)
from core.total_automatico_math import (
    remover_linha_branca_multiplas_sem_base,
    aplicar_layout_total_automatico,
    inserir_capa_total_automatico,
    adicionar_linhas_finais_em_branco,
    encontrar_perguntas_com_base_reduzida,
    detectar_bases_divergentes,
)


# =========================================================
# Utilidades de estado do assistente
# =========================================================
def _log(msg):
    st.session_state.setdefault("ta_log", []).append(msg)


def _ir_para(etapa):
    st.session_state["ta_step"] = etapa
    st.rerun()


def _reiniciar():
    for chave in list(st.session_state.keys()):
        if chave.startswith("ta_"):
            del st.session_state[chave]
    st.rerun()


def _aplicar_no_arquivo(funcao, numero, nome, aproximado=False, args=()):
    """
    Carrega o arquivo em andamento (st.session_state["ta_wb_bytes"]),
    aplica a função na planilha ativa, salva de volta e registra no
    log. Mesmo padrão do Relatório Automatizado
    (`analises/relatorio_automatizado.py::_aplicar_no_arquivo`).
    """
    wb = openpyxl.load_workbook(io.BytesIO(st.session_state["ta_wb_bytes"]), rich_text=True)
    ws = wb.active
    resultado = funcao(ws, *args)

    saida = io.BytesIO()
    wb.save(saida)
    st.session_state["ta_wb_bytes"] = saida.getvalue()

    selo = " (estimativa aproximada — confira no Excel)" if aproximado else ""
    if isinstance(resultado, tuple):
        resultado_txt = ", ".join(str(x) for x in resultado)
    else:
        resultado_txt = f"{resultado} alteração(ões)"
    _log(f"Código {numero} ({nome}): {resultado_txt} aplicada(s){selo}.")


def _aplicar_sequencia(passos):
    with st.status("Aplicando códigos...", expanded=True) as status:
        for p in passos:
            status.update(label=f"Aplicando código {p['numero']} — {p['nome']}...")
            _aplicar_no_arquivo(
                p["fn"], p["numero"], p["nome"],
                aproximado=p.get("aproximado", False),
                args=p.get("args", ()),
            )
        status.update(label="Concluído!", state="complete")


_ETAPAS_COM_PROGRESSO = [
    ("limpeza_multiplas", "Limpeza de linha branca (Múltiplas sem base)"),
    ("base_multiplas", "Base nas Múltiplas"),
    ("origem", "Preenche células / bordas (01-02)"),
    ("novos_termos", "Ordenar tabelas (03)"),
    ("layout", "Layout da planilha"),
    ("codigo_06_07", "DIN 9 + quebras de página (06-07)"),
    ("alturas", "Altura de labels/títulos/perguntas (09/11/12)"),
    ("capa", "Capa \"Resultados pelo total\""),
    ("final", "Concluído"),
]


def _mostrar_progresso(etapa_atual):
    nomes = [nome for chave, nome in _ETAPAS_COM_PROGRESSO]
    chaves = [chave for chave, nome in _ETAPAS_COM_PROGRESSO]
    if etapa_atual not in chaves:
        return
    idx = chaves.index(etapa_atual)
    st.progress(
        idx / (len(chaves) - 1),
        text=f"Etapa {idx + 1} de {len(chaves)}: {nomes[idx]}",
    )


# =========================================================
# Assistente
# =========================================================
def modulo_total_automatico():
    st.header("Total Automático")

    if "ta_step" not in st.session_state:
        st.session_state["ta_step"] = "aviso"
    if "ta_log" not in st.session_state:
        st.session_state["ta_log"] = []

    if st.session_state["ta_step"] != "aviso":
        if st.button("🔄 Reiniciar do zero", key="ta_reiniciar"):
            _reiniciar()

    etapa = st.session_state["ta_step"]
    _mostrar_progresso(etapa)

    # ---------------------------------------------------------------
    if etapa == "aviso":
        st.warning(
            "Versão enxuta do Relatório Automatizado, pra relatórios "
            "pelo total: limpeza de linha branca nas Múltiplas sem base "
            "→ Base nas Múltiplas → SPSS Relatoria/Inovação → Ordenar "
            "tabelas → layout → DIN 9 + quebras → altura de "
            "labels/títulos/perguntas → capa \"Resultados pelo total\". "
            "Sem exclusão por Base reduzida, sem cabeçalhos repetidos, "
            "sem Legendas e sem as Capas de Resultados do código 15 — "
            "aqui a capa é outra, mais simples."
        )
        if st.button("Entendi, continuar", key="ta_aviso_continuar"):
            _ir_para("upload")
        return

    # ---------------------------------------------------------------
    if etapa == "upload":
        arquivo = st.file_uploader(
            "Relatório descritivo (.xlsx)", type=["xlsx"], key="ta_upload_relatorio"
        )
        if arquivo is not None:
            st.session_state["ta_wb_bytes"] = arquivo.read()
            st.session_state["ta_nome_arquivo"] = arquivo.name
            _log("Relatório carregado.")
            _ir_para("limpeza_multiplas")
        return

    st.caption(f"Arquivo em andamento: **{st.session_state.get('ta_nome_arquivo', '?')}**")

    # ---------------------------------------------------------------
    if etapa == "limpeza_multiplas":
        st.subheader("Limpeza de linha branca (Múltiplas sem base)")
        st.caption(
            "Tabelas de pergunta do tipo Múltipla às vezes saem do SPSS "
            "com uma linha em branco entre o título e a tabela. Esta "
            "etapa procura essa linha SOMENTE nas tabelas de Múltipla "
            "que ainda não têm uma linha 'Base' por perto, e só remove "
            "quando a linha estiver realmente vazia — sem nenhum valor, "
            "borda ou cor. Se tiver qualquer conteúdo ou formatação, a "
            "linha fica intacta."
        )
        if st.button("Aplicar e continuar", key="ta_limpeza_continuar"):
            with st.spinner("Verificando tabelas de Múltipla..."):
                wb = openpyxl.load_workbook(io.BytesIO(st.session_state["ta_wb_bytes"]), rich_text=True)
                ws = wb.active
                resultado = remover_linha_branca_multiplas_sem_base(ws)
                saida = io.BytesIO()
                wb.save(saida)
                st.session_state["ta_wb_bytes"] = saida.getvalue()
            _log(
                f"Limpeza de linha branca (Múltiplas sem base): "
                f"{resultado['removidas']} linha(s) removida(s), "
                f"{resultado['mantidas']} mantida(s) por terem "
                f"conteúdo/formatação."
            )
            _ir_para("base_multiplas")
        return

    # ---------------------------------------------------------------
    if etapa == "base_multiplas":
        st.subheader("Base nas Múltiplas")
        st.caption(
            "Tabelas de Múltipla Escolha não saem do SPSS com a linha de "
            "Base. Se o relatório carregado tiver alguma, escolha uma "
            "das opções abaixo pra preencher a base. Se não tiver "
            "tabela de múltiplas, pode pular direto."
        )

        modo_base = st.radio(
            "Como você quer preencher a base?",
            ["Enviar arquivo com base", "Digitar um valor manualmente"],
            key="ta_bm_modo",
        )

        pular_bm = st.button("Este relatório não tem Múltiplas", key="ta_bm_pular")
        if pular_bm:
            _log("Base nas Múltiplas: etapa pulada.")
            _ir_para("origem")
            return

        if modo_base == "Digitar um valor manualmente":
            st.caption(
                "Aplica o MESMO valor em toda coluna com dado de cada "
                "tabela de Múltipla que ainda não tem base — útil quando "
                "não tem outro relatório do mesmo projeto pra puxar "
                "bases reais por segmento, só o N total mesmo."
            )
            valor_manual = st.number_input(
                "Valor da base", min_value=0, step=1, key="ta_bm_valor_manual"
            )
            aplicar_manual = st.button("Aplicar valor e continuar", key="ta_bm_aplicar_manual")

            if aplicar_manual:
                with st.spinner("Aplicando o valor em todas as tabelas..."):
                    try:
                        blocos_multiplas = parsear_blocos(io.BytesIO(st.session_state["ta_wb_bytes"]))
                    except Exception as e:
                        st.error(f"Não foi possível ler o relatório: {e}")
                        return

                    if not blocos_multiplas:
                        st.error(
                            "Não encontrei nenhum bloco 'Titulo:' no relatório "
                            "carregado. Confirme se ele segue o formato padrão "
                            "de exportação do SPSS."
                        )
                        return

                    linhas_base = calcular_linhas_base_manual(blocos_multiplas, int(valor_manual))
                    wb_novo = gerar_workbook_com_base(
                        io.BytesIO(st.session_state["ta_wb_bytes"]), blocos_multiplas, linhas_base
                    )
                    saida = io.BytesIO()
                    wb_novo.save(saida)
                    st.session_state["ta_wb_bytes"] = saida.getvalue()

                n_ja_tinham = sum(1 for lb in linhas_base if lb.get("ja_tinha_base"))
                n_aplicadas = sum(1 for lb in linhas_base if lb["valores"])
                msg = (
                    f"Base nas Múltiplas (valor manual = {int(valor_manual)}): "
                    f"{len(blocos_multiplas)} tabela(s) no total, "
                    f"{n_aplicadas} receberam o valor"
                )
                if n_ja_tinham:
                    msg += f", {n_ja_tinham} já tinha(m) base"
                _log(msg + ".")
                _ir_para("origem")
            return

        arquivo_bases = st.file_uploader(
            "Relatório com base (opcional)", type=["xlsx"], key="ta_upload_bases_multiplas"
        )

        aplicar_bm = st.button(
            "Aplicar e continuar", key="ta_bm_aplicar", disabled=arquivo_bases is None
        )

        if aplicar_bm and arquivo_bases is not None:
            with st.spinner("Lendo e casando as tabelas..."):
                try:
                    blocos_multiplas = parsear_blocos(io.BytesIO(st.session_state["ta_wb_bytes"]))
                    blocos_bases = parsear_blocos(arquivo_bases)
                except Exception as e:
                    st.error(f"Não foi possível ler os arquivos: {e}")
                    return

                if not blocos_multiplas:
                    st.error(
                        "Não encontrei nenhum bloco 'Titulo:' no relatório "
                        "carregado. Confirme se ele segue o formato padrão "
                        "de exportação do SPSS."
                    )
                    return
                if not blocos_bases:
                    st.error(
                        "Não encontrei nenhum bloco 'Titulo:' no relatório "
                        "com base. Confirme se esse é o arquivo certo."
                    )
                    return

                base_total, por_par, por_categoria, por_grupo_sem_categoria = extrair_bases(blocos_bases)
                if base_total is None:
                    st.error(
                        "Não encontrei nenhuma linha 'Base' no relatório "
                        "com base. Confirme se esse é o arquivo certo."
                    )
                    return

                blocos_religiao_bases = [b for b in blocos_bases if bloco_eh_religiao(b)]
                indice_religiao = None
                if blocos_religiao_bases:
                    indice_religiao = extrair_bases(blocos_bases, apenas_blocos=blocos_religiao_bases)[1:]

                linhas_base = calcular_linhas_base(
                    blocos_multiplas, base_total, por_par, por_categoria, por_grupo_sem_categoria,
                    indice_religiao=indice_religiao,
                )

                wb_novo = gerar_workbook_com_base(
                    io.BytesIO(st.session_state["ta_wb_bytes"]), blocos_multiplas, linhas_base
                )
                saida = io.BytesIO()
                wb_novo.save(saida)
                st.session_state["ta_wb_bytes"] = saida.getvalue()

            n_ja_tinham = sum(1 for lb in linhas_base if lb.get("ja_tinha_base"))
            n_religiao = sum(1 for lb in linhas_base if lb.get("eh_religiao"))
            total_nao_encontradas = sum(len(lb["nao_encontradas"]) for lb in linhas_base)
            msg = f"Base nas Múltiplas: {len(blocos_multiplas)} tabela(s) no total"
            if n_ja_tinham:
                msg += f", {n_ja_tinham} já tinha(m) base"
            if n_religiao:
                msg += f", {n_religiao} de religião (casamento restrito)"
            if total_nao_encontradas:
                msg += f", {total_nao_encontradas} coluna(s) sem base correspondente"
            _log(msg + ".")
            _ir_para("origem")
        return

    # ---------------------------------------------------------------
    if etapa == "origem":
        origem = st.radio(
            "O relatório foi gerado no SPSS Relatoria ou no SPSS Inovação?",
            ["SPSS Relatoria", "SPSS Inovação"],
            key="ta_origem"
        )
        if st.button("Continuar", key="ta_origem_continuar"):
            if origem == "SPSS Inovação":
                _aplicar_sequencia([
                    {"fn": aplicar_codigo_01, "numero": "01", "nome": "Preenche células do SPSS pirata"},
                    {"fn": aplicar_codigo_02, "numero": "02", "nome": "Muda Bordas"},
                ])
            else:
                _log("SPSS Relatoria selecionado — códigos 01 e 02 pulados.")
            _ir_para("novos_termos")
        return

    # ---------------------------------------------------------------
    if etapa == "novos_termos":
        st.subheader("Ordenar tabelas (código 03)")
        st.caption(
            f"A lista padrão já tem {len(TERMOS_EXCLUSAO_CODIGO_03)} termos "
            "herdados do VBA original (ex.: 'Base', 'Não sabe', 'Branco/Nulo')."
        )
        termos_novos = st.text_area(
            "Termos extras a incluir na lista de exclusão (um por linha, "
            "opcional):",
            key="ta_termos_novos",
            height=120,
        )
        if st.button("Continuar", key="ta_termos_continuar"):
            termos_extras = [t.strip() for t in termos_novos.splitlines() if t.strip()]
            lista_final = TERMOS_EXCLUSAO_CODIGO_03 + termos_extras
            _aplicar_sequencia([
                {"fn": aplicar_codigo_03, "numero": "03", "nome": "Ordenar tabelas", "args": (lista_final,)},
            ])
            _ir_para("layout")
        return

    # ---------------------------------------------------------------
    if etapa == "layout":
        st.subheader("Layout da planilha")
        st.caption(
            "Coluna A = 21, coluna B = 8, colunas C em diante = 8,67; "
            "fonte DIN 10; escala de impressão 100%; linhas de grade "
            "desativadas. (O negrito dos títulos é aplicado no próximo "
            "passo de altura, não aqui.) Sem exclusão de colunas por "
            "Base reduzida."
        )
        if st.button("Aplicar e continuar", key="ta_layout_continuar"):
            _aplicar_sequencia([
                {"fn": aplicar_layout_total_automatico, "numero": "04 (layout)", "nome": "Layout da planilha"},
            ])
            _ir_para("codigo_06_07")
        return

    # ---------------------------------------------------------------
    if etapa == "codigo_06_07":
        st.subheader("DIN 9 + linhas e quebras de página (códigos 06-07)")
        st.caption(
            "Ajusta a fonte das linhas de Pergunta pra tamanho 9 e insere "
            "linhas + quebras de página manuais depois de cada bloco de "
            "pergunta. Em relatórios grandes pode levar alguns segundos."
        )
        if st.button("Aplicar e continuar", key="ta_codigo0607_continuar"):
            _aplicar_sequencia([
                {"fn": aplicar_codigo_06, "numero": "06", "nome": "Ajuste DIN 9 na linha de Pergunta"},
                {"fn": aplicar_codigo_07, "numero": "07", "nome": "Adiciona linhas e quebra de página"},
            ])
            _ir_para("alturas")
        return

    # ---------------------------------------------------------------
    if etapa == "alturas":
        st.subheader("Altura de labels/títulos/perguntas (códigos 09/11/12)")
        st.caption(
            "Ajusta a altura das labels com quebra de linha, remove o "
            "prefixo 'Titulo: ' e deixa os títulos em negrito e "
            "centralizados, e ajusta a altura das linhas de Pergunta. "
            "Não marca base pequena (08) nem ajusta o título da Renda "
            "(10) — esses dois ficam de fora deste fluxo."
        )
        if st.button("Aplicar e continuar", key="ta_alturas_continuar"):
            _aplicar_sequencia([
                {
                    "fn": aplicar_codigo_09, "numero": "09", "nome": "Ajusta a altura das labels",
                    "aproximado": True,
                },
                {
                    "fn": aplicar_codigo_11, "numero": "11", "nome": "Ajusta a altura dos Títulos",
                    "aproximado": True,
                },
                {
                    "fn": aplicar_codigo_12, "numero": "12", "nome": "Ajuste da altura das Perguntas",
                    "aproximado": True,
                },
            ])
            _ir_para("capa")
        return

    # ---------------------------------------------------------------
    if etapa == "capa":
        st.subheader('Capa "Resultados pelo total"')
        st.caption(
            "Insere um bloco de 11 linhas acima da primeira tabela do "
            "relatório (a primeira linha com conteúdo no arquivo), com "
            "uma quebra de página no meio do bloco. Também adiciona 2 "
            "linhas em branco logo depois da última tabela do relatório."
        )
        texto_local = st.text_input(
            "Bairro / cidade / região (linhas 4-5, DIN 28 negrito)",
            key="ta_capa_local",
        )
        texto_mes_ano = st.text_input(
            "Mês e ano (linha 7)",
            key="ta_capa_mes_ano",
        )
        if st.button(
            "Inserir capa e continuar", key="ta_capa_continuar",
            disabled=not (texto_local.strip() and texto_mes_ano.strip()),
        ):
            with st.spinner("Inserindo capa..."):
                wb = openpyxl.load_workbook(io.BytesIO(st.session_state["ta_wb_bytes"]), rich_text=True)
                ws = wb.active
                resultado = inserir_capa_total_automatico(ws, texto_local.strip(), texto_mes_ano.strip())
                if resultado["status"] == "ok":
                    adicionar_linhas_finais_em_branco(ws)
                    saida = io.BytesIO()
                    wb.save(saida)
                    st.session_state["ta_wb_bytes"] = saida.getvalue()
                    _log(
                        f"Capa \"Resultados pelo total\": inserida a partir da "
                        f"linha {resultado['linha_inicio_bloco']}."
                    )
                    _log("Adicionadas 2 linhas em branco após a última tabela do relatório.")
                    _ir_para("final")
                else:
                    st.error(f"Capa: {resultado['mensagem']}")
        return

    # ---------------------------------------------------------------
    if etapa == "final":
        st.success("Fluxo concluído!")

        wb_final = openpyxl.load_workbook(io.BytesIO(st.session_state["ta_wb_bytes"]), rich_text=True)
        ws_final = wb_final.active

        perguntas_base_reduzida = encontrar_perguntas_com_base_reduzida(ws_final)
        if perguntas_base_reduzida:
            lista_perguntas = "\n".join(f"- {t}" for t in perguntas_base_reduzida)
            st.warning(
                "Este relatório tem pergunta(s) com **Base reduzida** (feitas só "
                "pra uma parte da amostra). Não esqueça de colocar o \"Para "
                "quem...\" explicando o filtro em cada uma:\n\n" + lista_perguntas
            )

        divergencia = detectar_bases_divergentes(ws_final)
        if divergencia["divergentes"]:
            lista_divergentes = "\n".join(
                f"- {d['titulo']}: base = {d['valor']}" for d in divergencia["divergentes"]
            )
            st.warning(
                f"A maioria das tabelas tem Base = {divergencia['valor_maioria']}, "
                f"mas encontrei tabela(s) com um valor diferente (e sem estar "
                f"marcada(s) como 'Base reduzida'). Vale conferir se é "
                f"intencional:\n\n" + lista_divergentes
            )

        st.subheader("Log das etapas")
        for linha in st.session_state.get("ta_log", []):
            st.write(f"- {linha}")

        st.download_button(
            "Baixar relatório processado",
            data=st.session_state["ta_wb_bytes"],
            file_name=f"processado_{st.session_state.get('ta_nome_arquivo', 'relatorio.xlsx')}",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="ta_download_final"
        )

        with st.expander("Pré-visualização rápida em tabela (mais leve, só valores aproximados)"):
            wb_preview = openpyxl.load_workbook(io.BytesIO(st.session_state["ta_wb_bytes"]))
            aba_escolhida = st.selectbox(
                "Aba", wb_preview.sheetnames, key="ta_preview_aba"
            ) if len(wb_preview.sheetnames) > 1 else wb_preview.sheetnames[0]
            ws_preview = wb_preview[aba_escolhida]

            html_preview, truncou_linhas, truncou_colunas = worksheet_para_html(ws_preview)
            st.components.v1.html(html_preview, height=620, scrolling=True)

            avisos = []
            if truncou_linhas:
                avisos.append("linhas")
            if truncou_colunas:
                avisos.append("colunas")
            if avisos:
                st.caption(f"⚠️ Pré-visualização truncada ({' e '.join(avisos)}) — o arquivo baixado está completo.")
        return
