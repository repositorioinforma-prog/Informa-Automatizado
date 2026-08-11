"""
Camada 3A - Exportador Excel

Gera o arquivo .xlsx final com:
    - Aba "Amostra Completa" com hierarquia mesclada
      (Meso → Micro → Município → Distrito)
    - Uma aba por campanha, quando houver
    - Colunas: Amostra + Gênero + Idade + Renda

Lógica pura, sem dependência de Streamlit (só openpyxl), seguindo a
convenção de arquitetura do projeto: core/ nunca importa streamlit.

Uso como biblioteca:
    from core.amostra_exportador import exportar_amostra
    exportar_amostra(df_completo, dict_campanhas, caminho_saida)
"""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Estilos (baseados na sua planilha modelo)
# ---------------------------------------------------------------------------
COR_CABECALHO_ESCURO = "1F4E78"   # azul escuro (header de topo)
COR_CABECALHO_VERDE = "70AD47"    # verde (coluna Campanhas)
COR_LINHA_ZEBRA = "F2F2F2"        # cinza claro alternado
COR_BORDA = "808080"

FONTE_HEADER = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
FONTE_CAMPANHA = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
FONTE_CORPO = Font(name="Calibri", size=10)

FILL_HEADER = PatternFill(start_color=COR_CABECALHO_ESCURO, end_color=COR_CABECALHO_ESCURO, fill_type="solid")
FILL_CAMPANHA = PatternFill(start_color=COR_CABECALHO_VERDE, end_color=COR_CABECALHO_VERDE, fill_type="solid")
FILL_ZEBRA = PatternFill(start_color=COR_LINHA_ZEBRA, end_color=COR_LINHA_ZEBRA, fill_type="solid")

ALINHA_CENTRO = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALINHA_ESQUERDA = Alignment(horizontal="left", vertical="center", wrap_text=True)

_borda_fina = Side(border_style="thin", color=COR_BORDA)
BORDA = Border(top=_borda_fina, bottom=_borda_fina, left=_borda_fina, right=_borda_fina)


# ---------------------------------------------------------------------------
# Estrutura de colunas (bate com a imagem 1 que você mandou)
# ---------------------------------------------------------------------------
# Grupos de cabeçalho de 2 níveis: (nome do grupo, [subcolunas])
GRUPOS_HEADER = [
    ("Campanhas", ["Campanhas"]),
    ("Mesorregiões (Intermediárias)", ["Mesorregiões (Intermediárias)"]),
    ("Microrregiões (Imediatas)", ["Microrregiões (Imediatas)"]),
    ("Municípios", ["Municípios"]),
    ("Distritos", ["Distritos"]),
    ("Amostra", ["Amostra"]),
    ("%", ["%"]),
    ("Gênero", ["Masculino", "Feminino"]),
    ("Idade", ["16 a 19", "20 a 29", "30 a 39", "40 a 49", "50 ou mais"]),
    ("Renda Média Domiciliar Familiar", ["Até 2sm", "2 a 5sm", "5 a 10sm", "Mais de 10sm"]),
]


def _colunas_planas():
    """Retorna a lista plana de colunas na ordem final."""
    return [sub for _, subs in GRUPOS_HEADER for sub in subs]


def _mapa_col_para_letra():
    """Mapa nome da coluna → letra do Excel (A, B, C...)"""
    return {c: get_column_letter(i + 1) for i, c in enumerate(_colunas_planas())}


# ---------------------------------------------------------------------------
# Escrita dos cabeçalhos (2 linhas, com merge nos grupos)
# ---------------------------------------------------------------------------
def _escrever_cabecalho(ws):
    col_idx = 1
    for grupo, subs in GRUPOS_HEADER:
        largura = len(subs)
        # Linha 1 (grupo)
        ws.cell(row=1, column=col_idx, value=grupo)
        if largura > 1:
            ws.merge_cells(
                start_row=1, start_column=col_idx,
                end_row=1, end_column=col_idx + largura - 1
            )
        else:
            # grupos de uma coluna só ocupam as duas linhas (merge vertical)
            ws.merge_cells(
                start_row=1, start_column=col_idx,
                end_row=2, end_column=col_idx
            )
        # Linha 2 (subcolunas)
        for j, sub in enumerate(subs):
            if largura > 1:
                ws.cell(row=2, column=col_idx + j, value=sub)
        col_idx += largura

    # Estiliza as 2 linhas de cabeçalho
    for row in ws.iter_rows(min_row=1, max_row=2, min_col=1, max_col=len(_colunas_planas())):
        for cell in row:
            cell.font = FONTE_HEADER
            cell.fill = FILL_HEADER
            cell.alignment = ALINHA_CENTRO
            cell.border = BORDA

    # Altura das linhas de cabeçalho
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 30


# ---------------------------------------------------------------------------
# Escrita das linhas de dados (com merge vertical automático nos agrupamentos)
# ---------------------------------------------------------------------------
def _linha_dados_de_row(row, nome_campanha=""):
    """Converte uma linha do DataFrame do motor em dict {coluna: valor}."""
    return {
        "Campanhas": nome_campanha,
        "Mesorregiões (Intermediárias)": row.get("regiao_intermediaria", ""),
        "Microrregiões (Imediatas)": row.get("regiao_imediata", ""),
        "Municípios": row.get("municipio", ""),
        "Distritos": row.get("distrito", ""),
        "Amostra": int(row["amostra"]),
        "%": round(float(row["percentual"]), 2),
        "Masculino": int(row["cota_masc"]),
        "Feminino": int(row["cota_fem"]),
        "16 a 19": int(row["cota_16_19"]),
        "20 a 29": int(row["cota_20_29"]),
        "30 a 39": int(row["cota_30_39"]),
        "40 a 49": int(row["cota_40_49"]),
        "50 ou mais": int(row["cota_50mais"]),
        "Até 2sm": int(row["cota_renda_ate_2sm"]),
        "2 a 5sm": int(row["cota_renda_2_a_5sm"]),
        "5 a 10sm": int(row["cota_renda_5_a_10sm"]),
        "Mais de 10sm": int(row["cota_renda_mais_10sm"]),
    }


def _escrever_linhas(ws, df, nome_campanha=""):
    """Escreve as linhas de dados a partir da linha 3."""
    colunas = _colunas_planas()
    linha_inicio = 3

    for i, (_, row) in enumerate(df.iterrows()):
        dados = _linha_dados_de_row(row, nome_campanha=nome_campanha)
        linha_excel = linha_inicio + i
        for j, col in enumerate(colunas, start=1):
            cell = ws.cell(row=linha_excel, column=j, value=dados[col])
            cell.font = FONTE_CORPO
            cell.alignment = ALINHA_CENTRO if j >= 6 else ALINHA_ESQUERDA
            cell.border = BORDA
            if i % 2 == 1:
                cell.fill = FILL_ZEBRA


def _mesclar_hierarquia(ws, df):
    """Mescla verticalmente células repetidas nas colunas de hierarquia.

    Ex: se 3 linhas seguidas têm a mesma Mesorregião, mescla essas 3 células
    da coluna Mesorregião numa só.
    """
    if df.empty:
        return

    colunas_hierarquia = ["Campanhas", "Mesorregiões (Intermediárias)",
                          "Microrregiões (Imediatas)", "Municípios"]
    mapa = _mapa_col_para_letra()
    linha_inicio = 3

    for col in colunas_hierarquia:
        letra = mapa[col]
        col_idx = ord(letra) - ord("A") + 1

        # varre as linhas identificando sequências iguais
        inicio = linha_inicio
        atual = ws.cell(row=inicio, column=col_idx).value
        for i in range(1, len(df) + 1):
            linha = linha_inicio + i
            valor = ws.cell(row=linha, column=col_idx).value if i < len(df) else None
            if valor != atual or i == len(df):
                fim = linha - 1
                if fim > inicio:
                    ws.merge_cells(start_row=inicio, start_column=col_idx,
                                   end_row=fim, end_column=col_idx)
                    # Reaplica alinhamento centralizado vertical na célula mesclada
                    ws.cell(row=inicio, column=col_idx).alignment = ALINHA_CENTRO
                inicio = linha
                atual = valor


def _ajustar_larguras(ws):
    """Define larguras de coluna razoáveis."""
    larguras = {
        "Campanhas": 28,
        "Mesorregiões (Intermediárias)": 22,
        "Microrregiões (Imediatas)": 22,
        "Municípios": 22,
        "Distritos": 22,
        "Amostra": 10,
        "%": 8,
        "Masculino": 11, "Feminino": 11,
        "16 a 19": 9, "20 a 29": 9, "30 a 39": 9, "40 a 49": 9, "50 ou mais": 11,
        "Até 2sm": 10, "2 a 5sm": 10, "5 a 10sm": 10, "Mais de 10sm": 12,
    }
    mapa = _mapa_col_para_letra()
    for col, larg in larguras.items():
        ws.column_dimensions[mapa[col]].width = larg


def _estilizar_coluna_campanha(ws, num_linhas):
    """A coluna 'Campanhas' fica com fundo verde e fonte branca, como na imagem 2."""
    if num_linhas == 0:
        return
    for linha in range(3, 3 + num_linhas):
        cell = ws.cell(row=linha, column=1)
        cell.fill = FILL_CAMPANHA
        cell.font = FONTE_CAMPANHA


def _escrever_linhas_totais(ws, df, num_linhas_dados):
    """
    Escreve duas linhas ao final do bloco de dados:
      - TOTAL: soma de cada coluna numérica
      - % DA AMOSTRA: cada coluna como % do total de amostra
    """
    colunas = _colunas_planas()
    linha_total = 3 + num_linhas_dados
    linha_pct = linha_total + 1

    cols_hierarquia = ["Campanhas", "Mesorregiões (Intermediárias)",
                       "Microrregiões (Imediatas)", "Municípios", "Distritos"]
    cols_numericas = [c for c in colunas if c not in cols_hierarquia]

    # Somatórios por coluna, a partir do DataFrame original
    somas = {c: 0 for c in cols_numericas}
    for _, row in df.iterrows():
        dados = _linha_dados_de_row(row)
        for c in cols_numericas:
            somas[c] += dados[c]

    total_amostra = somas.get("Amostra", 0)

    # Label "TOTAL" mesclado nas colunas de hierarquia (A até E)
    ws.cell(row=linha_total, column=1, value="TOTAL")
    ws.merge_cells(start_row=linha_total, start_column=1,
                   end_row=linha_total, end_column=5)
    # Label "% DA AMOSTRA"
    ws.cell(row=linha_pct, column=1, value="% DA AMOSTRA")
    ws.merge_cells(start_row=linha_pct, start_column=1,
                   end_row=linha_pct, end_column=5)

    # Preenche valores nas colunas numéricas
    for c in cols_numericas:
        col_idx = colunas.index(c) + 1
        soma = somas[c]

        # Linha TOTAL
        valor_total = round(soma, 2) if c == "%" else int(soma)
        ws.cell(row=linha_total, column=col_idx, value=valor_total)

        # Linha % DA AMOSTRA
        if c == "%":
            texto_pct = ""  # não faz sentido % do %
        elif c == "Amostra":
            texto_pct = "100%"
        elif total_amostra > 0:
            texto_pct = f"{round(soma / total_amostra * 100, 1)}%"
        else:
            texto_pct = "0%"
        ws.cell(row=linha_pct, column=col_idx, value=texto_pct)

    # Estiliza a linha TOTAL (fundo escuro, fonte branca)
    fill_total = PatternFill(start_color=COR_CABECALHO_ESCURO,
                             end_color=COR_CABECALHO_ESCURO, fill_type="solid")
    fonte_branca = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    for col_idx in range(1, len(colunas) + 1):
        cell = ws.cell(row=linha_total, column=col_idx)
        cell.font = fonte_branca
        cell.fill = fill_total
        cell.alignment = ALINHA_CENTRO
        cell.border = BORDA

    # Estiliza a linha % DA AMOSTRA (fundo cinza, itálico)
    for col_idx in range(1, len(colunas) + 1):
        cell = ws.cell(row=linha_pct, column=col_idx)
        cell.font = Font(name="Calibri", size=10, bold=True, italic=True)
        cell.fill = FILL_ZEBRA
        cell.alignment = ALINHA_CENTRO
        cell.border = BORDA


# ---------------------------------------------------------------------------
# Função principal
# ---------------------------------------------------------------------------
def exportar_amostra(df_completo, dict_campanhas=None, caminho_saida="amostra.xlsx"):
    """
    Gera o arquivo Excel final.

    Args:
        df_completo: DataFrame com a amostra completa (saída de core.amostra_math)
        dict_campanhas: dict {"nome_campanha": df_da_campanha, ...} ou None
            Se None, o arquivo terá apenas a aba "Amostra Completa".
        caminho_saida: caminho do arquivo .xlsx a gerar
    """
    wb = Workbook()

    # Aba 1: Amostra Completa
    ws = wb.active
    ws.title = "Amostra Completa"
    _escrever_cabecalho(ws)
    _escrever_linhas(ws, df_completo, nome_campanha="AMOSTRA COMPLETA")
    _mesclar_hierarquia(ws, df_completo)
    _estilizar_coluna_campanha(ws, len(df_completo))
    _escrever_linhas_totais(ws, df_completo, len(df_completo))
    _ajustar_larguras(ws)
    ws.freeze_panes = "A3"  # trava cabeçalho ao rolar

    # Abas por campanha
    if dict_campanhas:
        for nome, df_camp in dict_campanhas.items():
            # limite do Excel: 31 chars por nome de aba
            nome_aba = nome[:31]
            ws = wb.create_sheet(title=nome_aba)
            _escrever_cabecalho(ws)
            _escrever_linhas(ws, df_camp, nome_campanha=nome)
            _mesclar_hierarquia(ws, df_camp)
            _estilizar_coluna_campanha(ws, len(df_camp))
            _escrever_linhas_totais(ws, df_camp, len(df_camp))
            _ajustar_larguras(ws)
            ws.freeze_panes = "A3"

    wb.save(caminho_saida)
    return caminho_saida


# ---------------------------------------------------------------------------
# Teste: gera um arquivo real usando o motor de cálculo
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    from core.amostra_math import calcular_amostra

    print("Gerando amostra de teste: SP-1000, quebra por Região Imediata...")
    df = calcular_amostra(
        uf="SP",
        amostra_total=1000,
        nivel_quebra="regiao_imediata",
        base_populacional="16_mais",
    )
    print(f"  {len(df)} unidades territoriais")

    caminho = exportar_amostra(
        df_completo=df,
        dict_campanhas=None,
        caminho_saida="teste_amostra_SP_1000.xlsx",
    )
    print(f"OK. Arquivo gerado: {caminho}")