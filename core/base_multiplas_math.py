"""
Motor de "Base nas Múltiplas": lê um relatório de tabelas do tipo Múltipla
(sem linha de Base) e um relatório de tabelas de qualquer outra pergunta do
mesmo projeto (com linha de Base), casa os segmentos pelo texto do
cabeçalho (grupo + categoria) e devolve, para cada tabela de múltiplas, os
valores de base a inserir — sem depender da ordem das colunas, já que a
mesma segmentação pode aparecer em ordens diferentes nos dois arquivos.

Sem dependência do Streamlit — pode ser testado isoladamente.
"""
import re
import unicodedata
from collections import defaultdict

import openpyxl


def normalizar_texto(valor):
    """
    Normaliza um texto de cabeçalho para comparação, de forma tolerante a
    quebras de linha e hífens de quebra de linha do Excel/SPSS — que nem
    sempre caem no mesmo lugar da palavra em arquivos diferentes (ex.:
    "Evangé-\\nlica" num arquivo, "Afro-\\nBrasilei-ras" no outro).

    Em vez de tentar adivinhar onde rejuntar a palavra (frágil — pode
    rejuntar errado), remove TODOS os caracteres que não sejam letra ou
    número (hífen, espaço, quebra de linha, barra etc.), depois de tirar
    acento e caixa. Isso funciona porque o que varia entre os arquivos é
    só a pontuação/quebra de linha, nunca as letras em si.
    """
    if valor is None:
        return ""
    s = str(valor)
    s = s.replace("_x000D_", " ")
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")
    s = s.lower()
    s = re.sub(r"[^a-z0-9]", "", s)
    return s


# Categorias de religião conhecidas. Separadas em duas listas:
# - SINAIS_FORTES: termos que só aparecem em blocos de religião — usados
#   para IDENTIFICAR se um bloco é de religião (basta um bater).
# - as demais (Agnóstico, Ateu, Outras religiões, Não sabe) são ambíguas
#   (podem existir em qualquer outra pergunta) e por isso NÃO servem para
#   identificar o bloco, mas uma vez que o bloco já foi identificado como
#   de religião por um sinal forte, elas também entram na lista completa
#   usada para casar as bases.
RELIGIOES_SINAIS_FORTES = [
    "Denominações Católicas",
    "Denominações Evangélicas",
    "Católica Total",
    "Evangélica Total",
    "Católica de Batismo",
    "Católica Praticante",
    "Evangélicas de Missão",
    "Pentecostal",
    "Testemunha de Jeová",
    "Judaica",
    "Espírita/Kardecista",
    "Afro-Brasileiras",
    "Orientais",
]

RELIGIOES_AMBIGUAS = [
    "Agnóstico",
    "Ateu",
    "Outras religiões",
    "Não sabe",
]

RELIGIOES_SINAIS_FORTES_NORM = {normalizar_texto(r) for r in RELIGIOES_SINAIS_FORTES}
RELIGIOES_TODAS_NORM = RELIGIOES_SINAIS_FORTES_NORM | {normalizar_texto(r) for r in RELIGIOES_AMBIGUAS}


def bloco_eh_religiao(bloco):
    """
    Um bloco é considerado 'de religião' se algum dos seus grupos ou
    categorias bater com um dos SINAIS_FORTES (termos que não têm outro
    uso plausível, ex.: 'Denominações Católicas', 'Pentecostal',
    'Testemunha de Jeová'). Categorias ambíguas sozinhas (ex.: só 'Não
    sabe') não classificam um bloco como de religião, pra não confundir
    com outra pergunta qualquer que também tenha essa categoria.
    """
    textos = set()
    for g in bloco["grupos"]:
        if g is not None:
            textos.add(normalizar_texto(g))
    for c in bloco["categorias"]:
        if c is not None:
            textos.add(normalizar_texto(c))
    return bool(textos & RELIGIOES_SINAIS_FORTES_NORM)


def _eh_linha_vazia(linha):
    return linha is None or all(v is None for v in linha)


def _forward_fill(linha):
    """Preenche células None com o último rótulo de grupo visto à esquerda."""
    resultado = list(linha)
    atual = None
    for i, v in enumerate(resultado):
        if v is not None and str(v).strip() != "":
            atual = v
        resultado[i] = atual
    return resultado


def parsear_blocos(caminho_ou_arquivo, aba=None):
    """
    Parseia um arquivo no formato de tabelas empilhadas com marcadores
    'Titulo:' e 'Pergunta:'. Cada bloco tem uma linha de grupo (pode
    repetir ao longo de várias colunas, ex.: 'Sexo' cobrindo Masculino e
    Feminino) e, se houver segmentação, uma linha de categoria logo
    abaixo; blocos "só Total" (sem segmentação nenhuma) têm só a linha de
    grupo.

    Retorna uma lista de dicts:
      {"titulo", "grupos", "categorias", "dados", "pergunta",
       "linha_pergunta": nº da linha na planilha (1-indexado) onde está
       o marcador 'Pergunta:' — é ali que a linha de Base deve ser
       inserida (uma linha ANTES dela).}
    """
    wb = openpyxl.load_workbook(caminho_ou_arquivo, data_only=True)
    ws = wb[aba] if aba else wb[wb.sheetnames[0]]
    linhas = list(ws.iter_rows(values_only=True))
    n = len(linhas)
    blocos = []
    i = 0
    while i < n:
        linha = linhas[i]
        primeiro = linha[0] if linha else None
        if isinstance(primeiro, str) and primeiro.strip().startswith("Titulo:"):
            titulo = primeiro.strip()
            i += 1
            if i < n and _eh_linha_vazia(linhas[i]):
                i += 1
            linha_grupo = linhas[i]
            i += 1
            grupos = _forward_fill(linha_grupo)

            tem_segmentacao = any(v is not None for v in linha_grupo[2:])
            if tem_segmentacao:
                categorias = list(linhas[i])
                i += 1
            else:
                categorias = [None] * len(linha_grupo)

            dados = []
            pergunta = None
            linha_pergunta_idx = None
            while i < n:
                linha_dado = linhas[i]
                p = linha_dado[0] if linha_dado else None
                if isinstance(p, str) and p.strip().startswith("Pergunta:"):
                    pergunta = p.strip()
                    linha_pergunta_idx = i + 1  # 1-indexado, como no openpyxl
                    i += 1
                    break
                dados.append(linha_dado)
                i += 1
            if i < n and _eh_linha_vazia(linhas[i]):
                i += 1

            blocos.append({
                "titulo": titulo,
                "grupos": grupos,
                "categorias": categorias,
                "dados": dados,
                "pergunta": pergunta,
                "linha_pergunta": linha_pergunta_idx,
            })
        else:
            i += 1
    return blocos


def extrair_bases(blocos_bases, apenas_blocos=None):
    """
    Varre os blocos do arquivo de bases e monta as estruturas de busca:
    - base_total: valor da coluna "Total" (deve ser o mesmo em todos os
      blocos, já que é o n total da amostra)
    - por_par: {(grupo_norm, categoria_norm): valor}
    - por_categoria: {categoria_norm: [(grupo_norm, valor), ...]} — usado
      para casar quando o grupo não bate mas a categoria é inequívoca
    - por_grupo_sem_categoria: {grupo_norm: valor} — casos em que a
      segmentação não tem subcategoria (ex.: "Agnóstico" sozinho), que
      podem aparecer como CATEGORIA no arquivo de múltiplas

    'apenas_blocos': se informado, uma lista de blocos já filtrada (ex.:
    só os blocos de religião) — usado para montar um índice de busca
    restrito, mais seguro para segmentos com categorias ambíguas.
    """
    base_total = None
    por_par = {}
    por_categoria = defaultdict(list)
    por_grupo_sem_categoria = {}

    for b in (apenas_blocos if apenas_blocos is not None else blocos_bases):
        linha_base = next(
            (d for d in b["dados"] if d and isinstance(d[0], str) and d[0].strip() == "Base"),
            None
        )
        if linha_base is None:
            continue
        if base_total is None and len(linha_base) > 1 and linha_base[1] is not None:
            base_total = linha_base[1]

        for col in range(2, len(linha_base)):
            categoria = b["categorias"][col] if col < len(b["categorias"]) else None
            grupo = b["grupos"][col] if col < len(b["grupos"]) else None
            valor = linha_base[col]
            if valor is None or (categoria is None and grupo is None):
                continue

            grp_norm = normalizar_texto(grupo)
            if categoria is None:
                por_grupo_sem_categoria[grp_norm] = valor
                continue

            cat_norm = normalizar_texto(categoria)
            por_par[(grp_norm, cat_norm)] = valor
            por_categoria[cat_norm].append((grp_norm, valor))

    return base_total, por_par, por_categoria, por_grupo_sem_categoria


def casar_base(grupo, categoria, por_par, por_categoria, por_grupo_sem_categoria):
    """
    Casa uma coluna (grupo, categoria) do relatório de múltiplas com o
    valor de base correspondente. Retorna (valor, metodo); metodo é um
    dos: 'exato', 'por_categoria_unica', 'categoria_como_grupo',
    'grupo_sem_categoria', 'nao_encontrado'.

    A coluna "Total" é tratada à parte pelo chamador (é sempre a
    primeira coluna de dados, por posição) — esta função só lida com
    colunas de segmento. Uma categoria em branco aqui significa que o
    GRUPO em si é a segmentação (ex.: "Agnóstico" sem subcategoria), não
    que é a coluna Total — tratar os dois casos como a mesma coisa foi
    justamente o bug que fazia colunas como "Agnóstico"/"Ateu" saírem
    com o valor do total geral em vez do valor certo.
    """
    grp_norm = normalizar_texto(grupo)

    if categoria is None:
        if grp_norm in por_grupo_sem_categoria:
            return por_grupo_sem_categoria[grp_norm], "grupo_sem_categoria"
        return None, "nao_encontrado"

    cat_norm = normalizar_texto(categoria)

    if (grp_norm, cat_norm) in por_par:
        return por_par[(grp_norm, cat_norm)], "exato"

    candidatos = por_categoria.get(cat_norm, [])
    valores_unicos = set(v for _, v in candidatos)
    if len(valores_unicos) == 1:
        return candidatos[0][1], "por_categoria_unica"

    if cat_norm in por_grupo_sem_categoria:
        return por_grupo_sem_categoria[cat_norm], "categoria_como_grupo"

    return None, "nao_encontrado"


def bloco_ja_tem_base(bloco):
    """Verifica se o bloco já tem uma linha 'Base' nos dados (não precisa de nada)."""
    return any(
        d and isinstance(d[0], str) and d[0].strip() == "Base"
        for d in bloco["dados"]
    )


def calcular_linhas_base(
    blocos_multiplas, base_total, por_par, por_categoria, por_grupo_sem_categoria,
    indice_religiao=None,
):
    """
    Para cada bloco do relatório de múltiplas, calcula a linha de Base a
    inserir. Só preenche colunas onde o bloco realmente tem dado (evita
    escrever em colunas em branco no fim de tabelas menores).

    Blocos que JÁ têm uma linha 'Base' (ex.: um relatório misto, com
    tabelas de resposta única no meio das de múltiplas) são pulados por
    completo — não precisam de nada e não devem ser alterados.

    Blocos identificados como de RELIGIÃO (`bloco_eh_religiao`) usam um
    índice de busca separado e restrito só a blocos de religião do
    arquivo de bases (`indice_religiao`), em vez do índice geral — evita
    que uma categoria ambígua (ex.: "Não sabe") case por engano com um
    valor de um bloco sem relação nenhuma com religião. Se
    'indice_religiao' não for informado, cai no índice geral normalmente.

    Retorna uma lista paralela a 'blocos_multiplas', cada item:
      {"valores": {col_idx: valor}, "nao_encontradas": [rotulo, ...],
       "ja_tinha_base": bool, "eh_religiao": bool}
    """
    resultado = []
    for b in blocos_multiplas:
        if bloco_ja_tem_base(b):
            resultado.append({
                "valores": {}, "nao_encontradas": [], "ja_tinha_base": True, "eh_religiao": False
            })
            continue

        eh_religiao = bloco_eh_religiao(b)
        if eh_religiao and indice_religiao is not None:
            pp, pc, pg = indice_religiao
        else:
            pp, pc, pg = por_par, por_categoria, por_grupo_sem_categoria

        valores = {}
        nao_encontradas = []
        n_cols = len(b["grupos"])
        for col in range(1, n_cols):
            categoria = b["categorias"][col] if col < len(b["categorias"]) else None
            grupo = b["grupos"][col] if col < len(b["grupos"]) else None

            tem_dado_na_coluna = any(
                (row[col] if col < len(row) else None) is not None
                for row in b["dados"]
            )
            if not tem_dado_na_coluna:
                continue

            if col == 1:
                # a primeira coluna de dados é sempre a coluna "Total"
                # (por posição, não porque a categoria está em branco —
                # esse era o bug: colunas de grupo sem subcategoria, como
                # "Agnóstico", também têm categoria em branco e NÃO são
                # a coluna Total)
                valores[col] = base_total
                continue

            valor, metodo = casar_base(grupo, categoria, pp, pc, pg)
            if valor is not None:
                valores[col] = valor
            else:
                rotulo = f"{grupo} / {categoria}" if categoria else "Total"
                nao_encontradas.append(rotulo)

        resultado.append({
            "valores": valores, "nao_encontradas": nao_encontradas,
            "ja_tinha_base": False, "eh_religiao": eh_religiao,
        })
    return resultado


def calcular_linhas_base_manual(blocos_multiplas, valor_base):
    """
    Variante de `calcular_linhas_base` pro modo manual: em vez de casar
    contra um arquivo de bases (outra pergunta do mesmo projeto), aplica
    o MESMO valor `valor_base` em toda coluna que o bloco realmente tem
    dado — mesma lógica de detecção de coluna usada no modo com
    arquivo, só que sem nenhum casamento por texto de cabeçalho. Pensado
    pra quando não tem outro relatório do mesmo projeto disponível pra
    puxar bases reais por segmento, só o N total da amostra mesmo (ou
    qualquer valor único que a pessoa queira aplicar em tudo).

    Blocos que já têm 'Base' continuam sendo pulados por completo,
    igual ao modo com arquivo.

    Retorna uma lista no mesmo formato de `calcular_linhas_base`
    (`{"valores", "nao_encontradas", "ja_tinha_base", "eh_religiao"}`),
    já compatível com `gerar_workbook_com_base` sem nenhuma mudança lá.
    """
    resultado = []
    for b in blocos_multiplas:
        if bloco_ja_tem_base(b):
            resultado.append({
                "valores": {}, "nao_encontradas": [], "ja_tinha_base": True, "eh_religiao": False
            })
            continue

        valores = {}
        n_cols = len(b["grupos"])
        for col in range(1, n_cols):
            tem_dado_na_coluna = any(
                (row[col] if col < len(row) else None) is not None
                for row in b["dados"]
            )
            if not tem_dado_na_coluna:
                continue
            valores[col] = valor_base

        resultado.append({
            "valores": valores, "nao_encontradas": [], "ja_tinha_base": False, "eh_religiao": False,
        })
    return resultado


def gerar_workbook_com_base(caminho_multiplas, blocos_multiplas, linhas_base, aba=None):
    """
    Gera um NOVO workbook a partir do relatório de múltiplas original, com
    uma linha "Base" inserida antes da linha 'Pergunta:' de cada bloco.

    Reconstrói a planilha célula a célula (valor + estilo) em vez de usar
    `ws.insert_rows()`, porque essa função do openpyxl é conhecida por
    corromper o conteúdo quando a planilha tem células mescladas — e
    relatórios de tabelas do SPSS/Excel quase sempre têm (os cabeçalhos de
    grupo, ex.: "Sexo" mesclado sobre duas colunas). As mesclagens
    originais são recriadas nas novas posições de linha.

    Tratamento de borda: a última linha de dados de cada tabela tem a
    borda inferior espessa (é o rodapé visual da tabela). Como a linha
    "Base" passa a ser a nova última linha, essa borda espessa é REMOVIDA
    da antiga última linha e movida para a linha "Base" — que herda o
    restante do padrão de borda (esquerda/direita) e a fonte de cada
    coluna da linha de dados acima dela, sem negrito, com formato numérico
    de inteiro (já que são contagens, não percentuais).
    """
    from copy import copy as _copy_style

    from openpyxl.styles import Border, Font, Side

    wb_origem = openpyxl.load_workbook(caminho_multiplas)
    nome_aba = aba or wb_origem.sheetnames[0]
    ws_origem = wb_origem[nome_aba]

    wb_novo = openpyxl.Workbook()
    ws_novo = wb_novo.active
    ws_novo.title = nome_aba

    for letra, dim in ws_origem.column_dimensions.items():
        if dim.width:
            ws_novo.column_dimensions[letra].width = dim.width

    max_col = ws_origem.max_column
    lado_vazio = Side(border_style=None)

    # mapa: linha original (a última linha de dados de um bloco) -> índice do bloco
    # blocos que JÁ têm 'Base' (relatório misto) ficam de fora — não são tocados
    ultima_linha_por_bloco = {
        b["linha_pergunta"] - 1: idx
        for idx, b in enumerate(blocos_multiplas)
        if b["linha_pergunta"] and not linhas_base[idx].get("ja_tinha_base")
    }

    def _copiar_celula(cel_origem, cel_destino, bottom_override=None, negrito=None,
                        formato_numero=None):
        cel_destino.value = cel_origem.value
        if not cel_origem.has_style:
            return
        fonte_origem = cel_origem.font
        if negrito is not None:
            fonte = Font(
                name=fonte_origem.name, size=fonte_origem.size, bold=negrito,
                italic=fonte_origem.italic, vertAlign=fonte_origem.vertAlign,
                underline=fonte_origem.underline, strike=fonte_origem.strike,
                color=_copy_style(fonte_origem.color),
            )
        else:
            fonte = _copy_style(fonte_origem)
        cel_destino.font = fonte
        cel_destino.fill = _copy_style(cel_origem.fill)
        cel_destino.number_format = formato_numero or cel_origem.number_format
        cel_destino.protection = _copy_style(cel_origem.protection)
        cel_destino.alignment = _copy_style(cel_origem.alignment)
        b = cel_origem.border
        nova_bottom = bottom_override if bottom_override is not None else _copy_style(b.bottom)
        cel_destino.border = Border(
            left=_copy_style(b.left), right=_copy_style(b.right),
            top=_copy_style(b.top), bottom=nova_bottom
        )

    linha_destino = 1
    mapa_linha = {}
    bordas_removidas = {}  # idx_bloco -> {coluna: Side da borda inferior original}

    for linha_origem_idx in range(1, ws_origem.max_row + 1):
        idx_bloco_desta_linha = ultima_linha_por_bloco.get(linha_origem_idx)
        mapa_linha[linha_origem_idx] = linha_destino

        for col in range(1, max_col + 1):
            cel_origem = ws_origem.cell(row=linha_origem_idx, column=col)
            cel_destino = ws_novo.cell(row=linha_destino, column=col)
            if idx_bloco_desta_linha is not None:
                # última linha de dados: guarda a borda inferior original e remove
                bordas_removidas.setdefault(idx_bloco_desta_linha, {})[col] = (
                    _copy_style(cel_origem.border.bottom)
                )
                _copiar_celula(cel_origem, cel_destino, bottom_override=lado_vazio)
            else:
                _copiar_celula(cel_origem, cel_destino)
        linha_destino += 1

        if idx_bloco_desta_linha is not None:
            # insere a linha "Base" logo abaixo, reaproveitando a borda
            # espessa que acabou de ser retirada da linha de dados acima
            valores = linhas_base[idx_bloco_desta_linha]["valores"]
            bordas_col = bordas_removidas[idx_bloco_desta_linha]

            cel_ref = ws_origem.cell(row=linha_origem_idx, column=1)
            cel_dest_rotulo = ws_novo.cell(row=linha_destino, column=1, value="Base")
            _copiar_celula(
                cel_ref, cel_dest_rotulo,
                bottom_override=bordas_col.get(1, lado_vazio),
                negrito=False,
            )
            cel_dest_rotulo.value = "Base"

            for col in range(2, max_col + 1):
                valor = valores.get(col - 1)
                cel_ref_col = ws_origem.cell(row=linha_origem_idx, column=col)
                cel_dest = ws_novo.cell(row=linha_destino, column=col)
                _copiar_celula(
                    cel_ref_col, cel_dest,
                    bottom_override=bordas_col.get(col, lado_vazio),
                    negrito=False,
                    formato_numero="0" if valor is not None else None,
                )
                cel_dest.value = valor

            linha_destino += 1

    for rng in ws_origem.merged_cells.ranges:
        nova_min_linha = mapa_linha.get(rng.min_row, rng.min_row)
        nova_max_linha = mapa_linha.get(rng.max_row, rng.max_row)
        ws_novo.merge_cells(
            start_row=nova_min_linha, start_column=rng.min_col,
            end_row=nova_max_linha, end_column=rng.max_col
        )

    return wb_novo
