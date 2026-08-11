"""
Motor de "Corrigir Cabeçalhos" (código 13).

Usa um arquivo de referência (um bloco "Titulo: ..." por tipo de
segmentação — Sexo/Idade/Renda, Escolaridade, Religião, Regiões etc.,
cada um com a linha de grupo + linha de subcategoria "corretas") e
substitui, no relatório principal, o cabeçalho de toda tabela cujo
conjunto de rótulos de grupo bata com um desses tipos.

Corrige na prática:
- Texto com artefato de codificação (ex.: "Ensino_x000D_\\nMédio" em vez
  de "Ensino\\nMédio", herdado de quebras de linha do Windows mal
  interpretadas na exportação do SPSS);
- Rótulos com hifenização/quebra de linha diferente da versão "oficial"
  (ex.: "Ensino Fundamental" vs "Ensino Funda-\\nmental");
- Mesclagens quebradas (ex.: "Denominações Católicas" duplicado em duas
  células adjacentes em vez de uma célula mesclada abrangendo as
  subcategorias).

O "tipo" de cada bloco é identificado pelo CONJUNTO de rótulos de grupo
da própria linha de cabeçalho (comparação normalizada — sem acento,
caixa, hífen/quebra de linha), não por um marcador separado.

Pensado para rodar ANTES do código 05 (que uniformiza fonte/alinhamento
de toda a planilha por cima) — o papel deste código é só corrigir a
ESTRUTURA/CONTEÚDO do cabeçalho, não a aparência visual fina.

Sem dependência do Streamlit — pode ser testado isoladamente.
"""
import openpyxl

from core.base_multiplas_math import normalizar_texto
from core.legendas_math import _eh_linha_vazia, _forward_fill


def _eh_titulo(valor):
    return isinstance(valor, str) and valor.strip().lower().startswith("titulo")


def _extrair_grupos_categorias(ws, linha_titulo, max_col):
    """
    A partir da linha do título (1-indexada), lê a linha de grupo (logo
    abaixo) e, se houver segmentação (algo além da coluna 'Total' em
    B), a linha de subcategoria também.

    Returns:
        (linha_grupo_idx, grupos_forward_filled, linha_categoria_idx ou
        None, categorias ou None)
    """
    linha_grupo_idx = linha_titulo + 1
    linha_grupo = [ws.cell(row=linha_grupo_idx, column=c).value for c in range(1, max_col + 1)]
    grupos = _forward_fill(linha_grupo)

    tem_segmentacao = len(linha_grupo) > 2 and any(v is not None for v in linha_grupo[2:])
    linha_categoria_idx = None
    categorias = None
    if tem_segmentacao:
        candidato_idx = linha_grupo_idx + 1
        candidato = [ws.cell(row=candidato_idx, column=c).value for c in range(1, max_col + 1)]
        if not _eh_titulo(candidato[0]) and not _eh_linha_vazia(candidato):
            linha_categoria_idx = candidato_idx
            categorias = candidato

    return linha_grupo_idx, grupos, linha_categoria_idx, categorias


def _assinatura_tipo(grupos):
    """
    Conjunto normalizado dos rótulos de grupo (a partir da coluna C —
    B costuma ser só 'Total'), usado pra casar o "tipo" de segmentação
    entre o relatório e a referência. frozenset vazio = bloco sem
    segmentação (só "Total"), não é candidato a correção.
    """
    vistos = set()
    for v in grupos[2:]:
        if v is None:
            continue
        norm = normalizar_texto(v)
        if norm:
            vistos.add(norm)
    return frozenset(vistos)


def parsear_blocos_cabecalho_referencia(caminho_arquivo, aba=None):
    """
    Lê o arquivo de referência — um bloco "Titulo: ..." por tipo de
    segmentação, cada um com linha de grupo + linha de subcategoria
    corretas (mais uma linha de exemplo qualquer logo abaixo, que serve
    só pra identificar visualmente o bloco no arquivo — não é usada).

    Returns:
        Lista de dicts: {"assinatura": frozenset(...),
            "valores_grupo": [...], "valores_categoria": [...] ou None,
            "mesclagens_grupo": [(col_ini, col_fim), ...],
            "mesclagens_categoria": [(col_ini, col_fim), ...]}
    """
    wb = openpyxl.load_workbook(caminho_arquivo, rich_text=True)
    ws = wb[aba] if aba else wb[wb.sheetnames[0]]
    max_col = ws.max_column

    linhas_titulo = [
        r for r in range(1, ws.max_row + 1)
        if _eh_titulo(ws.cell(row=r, column=1).value)
    ]

    blocos = []
    for linha_titulo in linhas_titulo:
        linha_grupo_idx, grupos, linha_categoria_idx, categorias = _extrair_grupos_categorias(
            ws, linha_titulo, max_col
        )
        assinatura = _assinatura_tipo(grupos)
        if not assinatura:
            continue  # bloco só "Total", sem segmentação — nada pra este código corrigir

        valores_grupo = [ws.cell(row=linha_grupo_idx, column=c).value for c in range(1, max_col + 1)]
        valores_categoria = None
        mesclagens_categoria = []
        if linha_categoria_idx:
            valores_categoria = [
                ws.cell(row=linha_categoria_idx, column=c).value for c in range(1, max_col + 1)
            ]
            mesclagens_categoria = [
                (rng.min_col, rng.max_col)
                for rng in ws.merged_cells.ranges
                if rng.min_row == linha_categoria_idx == rng.max_row
            ]

        mesclagens_grupo = [
            (rng.min_col, rng.max_col)
            for rng in ws.merged_cells.ranges
            if rng.min_row == linha_grupo_idx == rng.max_row
        ]

        blocos.append({
            "assinatura": assinatura,
            "valores_grupo": valores_grupo,
            "valores_categoria": valores_categoria,
            "mesclagens_grupo": mesclagens_grupo,
            "mesclagens_categoria": mesclagens_categoria,
        })

    return blocos


def _substituir_linha(ws, linha_idx, valores, mesclagens, max_col):
    """Desfaz as mesclagens existentes nessa linha, reescreve os valores
    célula a célula, e reaplica as mesclagens da versão de referência."""
    for rng in list(ws.merged_cells.ranges):
        if rng.min_row == linha_idx == rng.max_row:
            ws.unmerge_cells(str(rng))

    for c in range(1, max_col + 1):
        valor = valores[c - 1] if valores and c - 1 < len(valores) else None
        ws.cell(row=linha_idx, column=c, value=valor)

    for col_inicio, col_fim in mesclagens:
        if col_fim > col_inicio:
            ws.merge_cells(start_row=linha_idx, start_column=col_inicio, end_row=linha_idx, end_column=col_fim)


def aplicar_codigo_13(ws, blocos_referencia):
    """
    Para cada bloco "Titulo: ..." do relatório principal cuja linha de
    grupo tenha o MESMO conjunto de rótulos (normalizado) que algum
    bloco do arquivo de referência, substitui a linha de grupo e a
    linha de subcategoria (valores E mesclagens) pela versão "correta".

    Blocos sem segmentação (só "Total") ou sem correspondência na
    referência ficam intocados.

    Returns:
        Número de blocos corrigidos.
    """
    max_col = ws.max_column
    linhas_titulo = [
        r for r in range(1, ws.max_row + 1)
        if _eh_titulo(ws.cell(row=r, column=1).value)
    ]

    corrigidos = 0
    for linha_titulo in linhas_titulo:
        linha_grupo_idx, grupos, linha_categoria_idx, categorias = _extrair_grupos_categorias(
            ws, linha_titulo, max_col
        )
        assinatura = _assinatura_tipo(grupos)
        if not assinatura:
            continue

        bloco_ref = next((b for b in blocos_referencia if b["assinatura"] == assinatura), None)
        if bloco_ref is None:
            continue

        _substituir_linha(ws, linha_grupo_idx, bloco_ref["valores_grupo"], bloco_ref["mesclagens_grupo"], max_col)
        if linha_categoria_idx and bloco_ref["valores_categoria"]:
            _substituir_linha(
                ws, linha_categoria_idx, bloco_ref["valores_categoria"],
                bloco_ref["mesclagens_categoria"], max_col,
            )
        corrigidos += 1

    return corrigidos
