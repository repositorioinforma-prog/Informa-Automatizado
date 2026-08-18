"""
Motor de "InserirCapasResultados" (código 15) — o mais complexo dos 16
códigos originais do VBA.

Insere duas "capas" de separação no relatório:
    - "RESULTADOS PELOS SEGMENTOS", logo antes da primeira tabela que
      tem alguma segmentação (Sexo, Idade, Renda, Escolaridade,
      Religião, Regiões etc. — identificado pela própria linha de
      cabeçalho conter "Total" + uma palavra de segmento conhecida).
    - "RESULTADOS PELO TOTAL", logo antes da primeira tabela do
      relatório (a primeira ocorrência de "Total" em qualquer célula),
      precedida por um espaço fixo de 140 linhas em branco (com quebra
      de página em posições fixas, não calculadas dinamicamente — ver
      `QUEBRAS_RELATIVAS_ANTES_TOTAL`).

Cada capa é um bloco de 7 linhas (título grande centralizado na 4ª) +
2 linhas em branco depois, com uma célula-marcador invisível (texto
branco, tamanho 1) na primeira linha do bloco — usada só pra
reidentificar o bloco depois (aplicar quebras de página) e pra permitir
rodar o código de novo sem duplicar capas (remove as antigas antes de
inserir as novas).

Sem dependência do Streamlit — pode ser testado isoladamente.
"""
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.pagebreak import Break

from core.planilha_utils import inserir_linhas_seguro, remover_linhas_seguro, linha_vazia_ate_coluna
from core.relatorio_automatizado_math import _ultima_linha_com_conteudo

COL_INICIO = 1  # A
COL_FIM = 13  # M
LINHAS_BLOCO_TITULO = 7
LINHA_DO_TITULO_NO_BLOCO = 4
LINHAS_EM_BRANCO_APOS_CAPA = 2
MARCADOR_TOTAL = "__CAPA_TOTAL__"
MARCADOR_SEGMENTOS = "__CAPA_SEGMENTOS__"

# Espaço ANTES do bloco de título "RESULTADOS PELO TOTAL": 140 linhas em
# branco, com quebra de página nas posições relativas abaixo (medidas
# por Lucas num arquivo real — não são calculadas a partir de
# `linhas_por_pagina`, são fixas). O bloco do título em si já tem 3
# linhas de espaço antes do título por conta própria (ver
# `_inserir_bloco_resultado`/`LINHA_DO_TITULO_NO_BLOCO`); as 140 linhas
# daqui NÃO somam nenhuma linha extra a mais além dessas 140 — não tem
# mais nenhum espaço adicional entre elas e o bloco do título.
# A capa dos Segmentos não usa nada disso (continua sem espaço antes).
LINHAS_BRANCAS_ANTES_TOTAL = 140
ESPACO_TOTAL_ANTES_TITULO = LINHAS_BRANCAS_ANTES_TOTAL  # 140
# Posições relativas (1 = primeira das 140 linhas em branco) onde deve
# haver quebra de página. Página 1 (antes da 1ª quebra) fica com 14
# linhas de verdade no arquivo (offset relativo 12, considerando que o
# início do bloco costuma ficar 2 linhas abaixo do topo do arquivo);
# páginas 2 a 6 têm exatamente 32 linhas cada uma daí em diante.
QUEBRAS_RELATIVAS_ANTES_TOTAL = [12, 44, 76, 108, 140]

PALAVRAS_SEGMENTO = [
    "SEXO", "IDADE", "FAIXA ETARIA", "FAIXA ETÁRIA", "RENDA", "ESCOLARIDADE",
    "RELIGIAO", "RELIGIÃO", "REGIAO", "REGIÃO", "CATOLICAS", "CATÓLICAS",
    "EVANGELICAS", "EVANGÉLICAS", "GOVL", "GOVM", "GOVO",
    "APROVACAO", "APROVAÇÃO", "AVALIACAO", "AVALIAÇÃO",
]


# ---------------------------------------------------------------------------
# Leitura de linhas (equivalentes às funções auxiliares "AM" do VBA)
# ---------------------------------------------------------------------------
def _texto_linha(ws, linha):
    partes = []
    for c in range(COL_INICIO, COL_FIM + 1):
        v = ws.cell(row=linha, column=c).value
        if v is not None and str(v).strip() != "":
            partes.append(str(v).strip().upper())
    return " ".join(partes)


def _linha_tem_total(ws, linha):
    for c in range(COL_INICIO, COL_FIM + 1):
        v = ws.cell(row=linha, column=c).value
        if v is not None and str(v).strip().upper() == "TOTAL":
            return True
    return False


def _linha_tem_conteudo(ws, linha):
    for c in range(COL_INICIO, COL_FIM + 1):
        v = ws.cell(row=linha, column=c).value
        if v is not None and str(v).strip() != "":
            return True
    return False


def _contem_palavra_segmento(texto):
    return any(p in texto for p in PALAVRAS_SEGMENTO)


def _encontrar_linha_cabecalho_total(ws, max_row):
    for r in range(1, max_row + 1):
        if _linha_tem_total(ws, r):
            return r
    return 0


def _encontrar_linha_cabecalho_segmentado(ws, linha_inicial, max_row):
    for r in range(linha_inicial, max_row + 1):
        texto = _texto_linha(ws, r)
        if "TOTAL" in texto and _contem_palavra_segmento(texto):
            return r
    return 0


def _eh_linha_titulo(ws, linha):
    """Uma linha 'Titulo: ...' é o marcador semântico mais confiável do
    início de um bloco de pergunta/tabela — muito mais preciso do que
    só "subir enquanto tiver conteúdo" (que nas versões anteriores
    acabava chegando até a linha 1, incluindo cabeçalhos/títulos
    institucionais que não fazem parte de bloco nenhum)."""
    valor = ws.cell(row=linha, column=1).value
    return isinstance(valor, str) and valor.strip().lower().startswith("titulo")


def _encontrar_inicio_bloco(ws, linha_ref):
    """
    Acha o início "de verdade" do bloco de pergunta/tabela que contém
    `linha_ref` — primeiro tenta achar a linha 'Titulo: ...' mais
    próxima IGUAL ou ACIMA de linha_ref (identificação semântica: o
    título é sempre a primeira linha de um bloco, nesse formato de
    relatório). Só cai pro comportamento conservador antigo ("sobe
    enquanto a linha anterior tiver conteúdo") se não achar nenhum
    'Titulo:' no caminho — o que não deveria acontecer num relatório
    bem formado, mas evita quebrar em arquivos fora do padrão.
    """
    r = linha_ref
    while r >= 1:
        if _eh_linha_titulo(ws, r):
            return r
        r -= 1

    # Fallback conservador (comportamento antigo) — só chega aqui se não
    # existir NENHUMA linha "Titulo:" entre linha_ref e o topo do arquivo
    r = linha_ref
    while r > 1 and _linha_tem_conteudo(ws, r - 1):
        r -= 1
    return r


def _encontrar_linha_marcador(ws, marcador, max_row):
    for r in range(1, max_row + 1):
        if str(ws.cell(row=r, column=1).value or "").strip() == marcador:
            return r
    return 0


# ---------------------------------------------------------------------------
# Configuração de página / quebras
# ---------------------------------------------------------------------------
def _configurar_pagina_am(ws):
    ultima_linha = _ultima_linha_com_conteudo(ws)
    ws.print_area = f"A1:{get_column_letter(COL_FIM)}{ultima_linha}"
    # Escala 100% (não "ajustar largura") — ver mesma explicação no
    # código 07 (core/relatorio_automatizado_math.py), que configura a
    # página do mesmo jeito.
    ws.sheet_properties.pageSetUpPr.fitToPage = False
    ws.page_setup.fitToWidth = None
    ws.page_setup.fitToHeight = None
    ws.page_setup.scale = 100


def _obter_linhas_por_pagina(ws):
    """Usa a 1ª quebra de página manual já existente (normalmente inserida
    pelo código 07) pra descobrir quantas linhas cabem numa página."""
    ids = sorted(b.id for b in ws.row_breaks.brk if b.id is not None)
    if ids:
        return ids[0]
    return 45


def _adicionar_quebra(ws, linha_id):
    if linha_id >= 1 and not any(b.id == linha_id for b in ws.row_breaks.brk):
        ws.row_breaks.append(Break(id=linha_id))


def _limpar_quebras_no_intervalo(ws, linha_inicial, linha_final):
    ini = max(1, linha_inicial)
    if linha_final < 1:
        return
    ws.row_breaks.brk = [b for b in ws.row_breaks.brk if not (ini <= b.id <= linha_final)]


def _aplicar_quebras_bloco_total(ws, linha_inicio_titulo):
    """
    As quebras ANTES do título (dentro das 140 linhas em branco) já
    foram inseridas em posições fixas por `_inserir_espaco_fixo_antes_total`,
    na hora da criação do bloco — não precisam ser recalculadas aqui, e
    a limpeza abaixo NÃO pode alcançar essa área (a última das 140
    quebras fixas fica bem coladinha no início do bloco do título, uma
    margem de limpeza "-2" a partir daqui a apagaria sem querer). Só
    garante a quebra logo DEPOIS do bloco do título (separando a capa
    do que vem a seguir no relatório).
    """
    linha_quebra_depois_titulo = linha_inicio_titulo + LINHAS_BLOCO_TITULO
    _limpar_quebras_no_intervalo(ws, linha_quebra_depois_titulo - 2, linha_quebra_depois_titulo + 2)
    _adicionar_quebra(ws, linha_quebra_depois_titulo - 1)


def _inicio_vao_em_branco_acima(ws, linha_ref, max_col=None):
    """
    Sobe a partir de `linha_ref - 1` enquanto a linha estiver totalmente
    vazia — acha o início do vão em branco imediatamente acima de
    `linha_ref`. Usado pra limpar TODAS as quebras de página que
    sobraram dentro desse vão (normalmente deixadas pelo código 07, de
    quando a tabela anterior ainda não tinha sido empurrada pra cá)
    antes de inserir a quebra nova — uma margem fixa de só 2 linhas
    podia deixar uma quebra antiga "presa" logo fora da margem,
    resultando numa página nova quase vazia (só 1-2 linhas em branco)
    entre a tabela anterior e a capa.
    """
    if max_col is None:
        max_col = ws.max_column
    r = linha_ref - 1
    while r >= 1 and linha_vazia_ate_coluna(ws, r, max_col):
        r -= 1
    return r + 1


def _reduzir_vao_em_branco_acima(ws, linha_ref, linhas_desejadas=2, max_col=None):
    """
    Reduz o vão de linhas em branco imediatamente acima de `linha_ref`
    pra exatamente `linhas_desejadas` linhas. O código 07 normaliza o
    vão pós-Pergunta de toda tabela pra 4 linhas em branco — na
    transição pra "RESULTADOS PELOS SEGMENTOS" isso deixava 4 linhas
    entre a última tabela do Total e a capa, mas só deveriam sobrar 2.

    Se o vão já tiver `linhas_desejadas` ou menos, não faz nada (seguro
    pra rodar de novo sem encolher além da conta numa segunda execução).

    Retorna a nova posição de `linha_ref` depois do ajuste (remover
    linhas de dentro do vão empurra tudo abaixo, incluindo `linha_ref`,
    pra cima).
    """
    inicio_vao = _inicio_vao_em_branco_acima(ws, linha_ref, max_col)
    tamanho_vao = linha_ref - inicio_vao
    if tamanho_vao <= linhas_desejadas:
        return linha_ref
    a_remover = tamanho_vao - linhas_desejadas
    remover_linhas_seguro(ws, inicio_vao, a_remover)
    return linha_ref - a_remover


def _aplicar_quebras_bloco_normal(ws, linha_inicio_titulo):
    linha_quebra_depois_titulo = linha_inicio_titulo + LINHAS_BLOCO_TITULO
    inicio_vao = _inicio_vao_em_branco_acima(ws, linha_inicio_titulo)
    _limpar_quebras_no_intervalo(ws, inicio_vao, linha_quebra_depois_titulo + 2)
    if linha_inicio_titulo > 1:
        _adicionar_quebra(ws, linha_inicio_titulo - 1)
    _adicionar_quebra(ws, linha_quebra_depois_titulo - 1)


def _aplicar_quebras_resultados(ws):
    max_row = _ultima_linha_com_conteudo(ws)
    linha_total = _encontrar_linha_marcador(ws, MARCADOR_TOTAL, max_row)
    if linha_total:
        _aplicar_quebras_bloco_total(ws, linha_total)

    linha_segmentos = _encontrar_linha_marcador(ws, MARCADOR_SEGMENTOS, max_row)
    if linha_segmentos:
        _aplicar_quebras_bloco_normal(ws, linha_segmentos)


# ---------------------------------------------------------------------------
# Inserção / limpeza das capas
# ---------------------------------------------------------------------------
def _inserir_espaco_fixo_antes_total(ws, linha_base):
    """
    Insere o espaço fixo ANTES do bloco de título "RESULTADOS PELO
    TOTAL": 140 linhas em branco + mais 2 linhas em branco logo acima
    de onde o bloco do título vai começar.

    As 140 linhas têm quebra de página nas posições relativas definidas
    em `QUEBRAS_RELATIVAS_ANTES_TOTAL` (12, 33, 68, 104, 140) — valores
    fixos, medidos por Lucas contra um arquivo real, não calculados a
    partir de `linhas_por_pagina`.

    Retorna a linha onde o bloco do título (`_inserir_bloco_resultado`,
    7 linhas) deve começar a partir daqui.
    """
    inserir_linhas_seguro(ws, linha_base, ESPACO_TOTAL_ANTES_TITULO)

    altura_padrao = ws.sheet_format.defaultRowHeight or 15
    for i in range(linha_base, linha_base + ESPACO_TOTAL_ANTES_TITULO):
        ws.row_dimensions[i].height = altura_padrao

    for offset in QUEBRAS_RELATIVAS_ANTES_TOTAL:
        _adicionar_quebra(ws, linha_base + offset - 1)

    return linha_base + ESPACO_TOTAL_ANTES_TITULO


def _inserir_bloco_resultado(ws, linha_base, titulo, marcador, paginas_brancas_antes, linhas_por_pagina):
    linhas_paginas_brancas = paginas_brancas_antes * linhas_por_pagina
    total_linhas_inserir = linhas_paginas_brancas + LINHAS_BLOCO_TITULO + LINHAS_EM_BRANCO_APOS_CAPA

    inserir_linhas_seguro(ws, linha_base, total_linhas_inserir)

    linha_inicio_titulo = linha_base + linhas_paginas_brancas
    altura_padrao = ws.sheet_format.defaultRowHeight or 15

    if linhas_paginas_brancas > 0:
        for i in range(linha_base, linha_inicio_titulo):
            ws.row_dimensions[i].height = altura_padrao

    for i in range(linha_inicio_titulo, linha_inicio_titulo + LINHAS_BLOCO_TITULO):
        ws.row_dimensions[i].height = 18

    linha_titulo = linha_inicio_titulo + LINHA_DO_TITULO_NO_BLOCO - 1
    ws.row_dimensions[linha_titulo].height = 42

    ws.row_dimensions[linha_inicio_titulo + LINHAS_BLOCO_TITULO].height = 18
    ws.row_dimensions[linha_inicio_titulo + LINHAS_BLOCO_TITULO + 1].height = 18

    # marcador invisível (texto branco, tamanho 1) — só pra reidentificar
    # o bloco depois (quebras de página, limpeza numa próxima execução)
    cel_marcador = ws.cell(row=linha_inicio_titulo, column=1, value=marcador)
    cel_marcador.font = Font(size=1, color="FFFFFFFF")

    for rng in list(ws.merged_cells.ranges):
        if rng.min_row == linha_titulo == rng.max_row:
            ws.unmerge_cells(str(rng))
    for c in range(COL_INICIO, COL_FIM + 1):
        ws.cell(row=linha_titulo, column=c, value=None)
    ws.merge_cells(start_row=linha_titulo, start_column=COL_INICIO, end_row=linha_titulo, end_column=COL_FIM)

    cel_titulo = ws.cell(row=linha_titulo, column=COL_INICIO)
    cel_titulo.value = titulo
    cel_titulo.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False, shrink_to_fit=False)
    cel_titulo.font = Font(name="DIN", size=28, bold=True)

    return linha_inicio_titulo


def _limpar_capas_antigas(ws):
    """Remove capas inseridas numa execução anterior deste código (busca
    pelos marcadores invisíveis) — permite rodar de novo sem duplicar."""
    while True:
        max_row = ws.max_row
        encontrou = False
        for i in range(max_row, 0, -1):
            valor = str(ws.cell(row=i, column=1).value or "").strip()
            if valor == MARCADOR_TOTAL:
                linha_inicio = max(1, i - ESPACO_TOTAL_ANTES_TITULO)
                linha_fim = i + LINHAS_BLOCO_TITULO + LINHAS_EM_BRANCO_APOS_CAPA - 1
                remover_linhas_seguro(ws, linha_inicio, linha_fim - linha_inicio + 1)
                encontrou = True
                break
            if valor == MARCADOR_SEGMENTOS:
                linha_inicio = i
                linha_fim = i + LINHAS_BLOCO_TITULO + LINHAS_EM_BRANCO_APOS_CAPA - 1
                remover_linhas_seguro(ws, linha_inicio, linha_fim - linha_inicio + 1)
                encontrou = True
                break
        if not encontrou:
            break


# ---------------------------------------------------------------------------
# Função principal
# ---------------------------------------------------------------------------
def aplicar_codigo_15(ws):
    """
    Insere as capas "RESULTADOS PELO TOTAL" e "RESULTADOS PELOS
    SEGMENTOS" no relatório (ver docstring do módulo). Pode ser
    reaplicado sem duplicar — remove as capas de uma execução anterior
    antes de inserir as novas.

    Depende de quebras de página manuais já existirem no relatório
    (inseridas pelo código 07), usadas pra calcular quantas linhas cabem
    numa página impressa.

    Returns:
        dict com "status" ("ok" ou "erro"), e em caso de sucesso as
        linhas onde cada capa foi inserida.
    """
    _configurar_pagina_am(ws)
    linhas_por_pagina = _obter_linhas_por_pagina(ws)
    if linhas_por_pagina <= 0:
        linhas_por_pagina = 45

    _limpar_capas_antigas(ws)
    _configurar_pagina_am(ws)

    max_row = _ultima_linha_com_conteudo(ws)
    linha_cab_total = _encontrar_linha_cabecalho_total(ws, max_row)
    if linha_cab_total == 0:
        return {"status": "erro", "mensagem": "Não encontrei a primeira tabela do total."}

    linha_inicio_total = _encontrar_inicio_bloco(ws, linha_cab_total)
    linha_cab_seg = _encontrar_linha_cabecalho_segmentado(ws, linha_cab_total + 1, max_row)

    linha_capa_segmentos = None
    if linha_cab_seg:
        linha_inicio_seg = _encontrar_inicio_bloco(ws, linha_cab_seg)
        # o código 07 normaliza o vão pós-Pergunta de toda tabela pra 4
        # linhas em branco — aqui, na transição pro bloco de Segmentos,
        # só devem sobrar 2 antes da capa
        linha_inicio_seg = _reduzir_vao_em_branco_acima(ws, linha_inicio_seg, linhas_desejadas=2)
        # insere primeiro o bloco de baixo (Segmentos) — inserir linhas
        # ali não afeta a posição do bloco do Total, que fica acima
        linha_capa_segmentos = _inserir_bloco_resultado(
            ws, linha_inicio_seg, "RESULTADOS PELOS SEGMENTOS", MARCADOR_SEGMENTOS,
            0, linhas_por_pagina,
        )

    # Espaço fixo (140 linhas em branco com quebras + 2 linhas em branco)
    # ANTES do bloco do título "RESULTADOS PELO TOTAL" — ver
    # `_inserir_espaco_fixo_antes_total`. Substituiu o antigo cálculo
    # dinâmico de "6 páginas em branco" baseado em linhas_por_pagina.
    linha_apos_espaco = _inserir_espaco_fixo_antes_total(ws, linha_inicio_total)
    linha_capa_total = _inserir_bloco_resultado(
        ws, linha_apos_espaco, "RESULTADOS PELO TOTAL", MARCADOR_TOTAL,
        0, linhas_por_pagina,
    )

    _configurar_pagina_am(ws)
    _aplicar_quebras_resultados(ws)
    _configurar_pagina_am(ws)

    # As quebras foram inseridas em pontos diferentes do processo (as 5
    # fixas antes do Total, a de depois do bloco, as dos Segmentos, e
    # quaisquer outras já existentes no arquivo antes do código 15
    # rodar — normalmente do código 07). `.append()` do openpyxl sempre
    # põe no fim da lista Python, independente do valor — então a lista
    # final pode ficar fora de ordem crescente mesmo com cada valor
    # individualmente correto. O Excel descarta a lista INTEIRA de
    # quebras manuais nesse caso (gotcha já mapeado no projeto), o que
    # se manifesta como "só aparecem linhas pontilhadas" (paginação
    # automática) em vez das quebras manuais de verdade. Reordena por
    # id como último passo, depois de todas as quebras já terem sido
    # inseridas.
    ws.row_breaks.brk = sorted(ws.row_breaks.brk, key=lambda b: b.id)

    # a posição da capa de Segmentos pode ter sido deslocada pela
    # inserção da capa do Total (que fica acima dela no arquivo) —
    # busca a posição final de verdade em vez de devolver a de antes
    # desse deslocamento
    if linha_capa_segmentos is not None:
        max_row_final = _ultima_linha_com_conteudo(ws)
        linha_capa_segmentos = _encontrar_linha_marcador(ws, MARCADOR_SEGMENTOS, max_row_final)

    return {
        "status": "ok",
        "linha_capa_total": linha_capa_total,
        "linha_capa_segmentos": linha_capa_segmentos,
    }


def validar_capas(ws, resultado):
    """
    Confere programaticamente que a estrutura das capas ficou correta —
    pra ser chamada DEPOIS de salvar e reabrir o arquivo (não em cima do
    `ws` ainda em memória antes de salvar), pra pegar qualquer coisa que
    o ciclo salvar/reabrir do Excel possa ter derrubado. Levanta
    ValueError com uma mensagem clara se algo estiver errado, em vez de
    devolver silenciosamente um arquivo incorreto.

    Args:
        ws: a planilha já reaberta (via openpyxl.load_workbook, depois
            de salvar) do resultado de `aplicar_codigo_15`.
        resultado: o dict devolvido por `aplicar_codigo_15`.

    Returns:
        True se tudo estiver certo (levanta exceção caso contrário).
    """
    erros = []
    max_row = ws.max_row

    linha_total = _encontrar_linha_marcador(ws, MARCADOR_TOTAL, max_row)
    if not linha_total:
        erros.append(f"Marcador {MARCADOR_TOTAL} não encontrado após salvar/reabrir.")
    else:
        linha_titulo_total = linha_total + LINHA_DO_TITULO_NO_BLOCO - 1
        valor = ws.cell(row=linha_titulo_total, column=1).value
        if valor != "RESULTADOS PELO TOTAL":
            erros.append(
                f"Esperava 'RESULTADOS PELO TOTAL' na linha {linha_titulo_total}, achei {valor!r}."
            )
        tem_mesclagem = any(
            r.min_row == linha_titulo_total == r.max_row and r.min_col == COL_INICIO and r.max_col == COL_FIM
            for r in ws.merged_cells.ranges
        )
        if not tem_mesclagem:
            erros.append(f"Mesclagem A:M não encontrada na linha do título ({linha_titulo_total}).")

        # confere o espaço fixo de 140 linhas em branco: precisa haver
        # pelo menos as 5 quebras manuais fixas (ver
        # QUEBRAS_RELATIVAS_ANTES_TOTAL) entre o início desse espaço e a
        # capa
        ids_quebras = sorted(b.id for b in ws.row_breaks.brk if b.id is not None)
        quebras_antes_do_total = [i for i in ids_quebras if i < linha_total]
        if len(quebras_antes_do_total) < len(QUEBRAS_RELATIVAS_ANTES_TOTAL):
            erros.append(
                f"Esperava pelo menos {len(QUEBRAS_RELATIVAS_ANTES_TOTAL)} quebras de "
                f"página antes da capa do Total, achei {len(quebras_antes_do_total)}."
            )
        linha_depois_bloco_total = linha_total + LINHAS_BLOCO_TITULO
        if (linha_depois_bloco_total - 1) not in ids_quebras:
            erros.append(
                f"Esperava uma quebra de página logo depois do bloco da capa do "
                f"Total (id={linha_depois_bloco_total - 1})."
            )

    linha_seg = _encontrar_linha_marcador(ws, MARCADOR_SEGMENTOS, max_row)
    if resultado.get("linha_capa_segmentos") and not linha_seg:
        erros.append(f"Marcador {MARCADOR_SEGMENTOS} não encontrado após salvar/reabrir.")
    elif linha_seg:
        linha_titulo_seg = linha_seg + LINHA_DO_TITULO_NO_BLOCO - 1
        valor = ws.cell(row=linha_titulo_seg, column=1).value
        if valor != "RESULTADOS PELOS SEGMENTOS":
            erros.append(
                f"Esperava 'RESULTADOS PELOS SEGMENTOS' na linha {linha_titulo_seg}, achei {valor!r}."
            )
        ids_quebras = sorted(b.id for b in ws.row_breaks.brk if b.id is not None)
        if (linha_seg - 1) not in ids_quebras:
            erros.append(f"Esperava uma quebra de página logo antes da capa dos Segmentos (id={linha_seg - 1}).")
        linha_depois_bloco_seg = linha_seg + LINHAS_BLOCO_TITULO
        if (linha_depois_bloco_seg - 1) not in ids_quebras:
            erros.append(
                f"Esperava uma quebra de página logo depois do bloco da capa dos "
                f"Segmentos (id={linha_depois_bloco_seg - 1})."
            )

    area_esperada = f"A1:{get_column_letter(COL_FIM)}{_ultima_linha_com_conteudo(ws)}"
    # openpyxl devolve o print_area formatado como "'NomeDaAba'!$A$1:$M$136"
    # — compara só a parte da faixa de células, sem se importar com o
    # nome da aba nem com "$"
    area_atual_normalizada = str(ws.print_area or "").replace("$", "")
    if area_esperada not in area_atual_normalizada:
        erros.append(f"Área de impressão esperada terminando em '{area_esperada}', achei {ws.print_area!r}.")

    # Nota: o documento de referência original pedia fitToWidth=1 (modo
    # "ajustar largura"), mas isso foi substituído pelo modo "escala
    # 100%" (fitToPage=False + scale=100) num pedido posterior — os dois
    # modos são mutuamente exclusivos no Excel, então validamos o que
    # realmente configuramos agora, não os dois ao mesmo tempo.
    if ws.sheet_properties.pageSetUpPr.fitToPage:
        erros.append("fitToPage deveria estar desligado (modo escala 100%), veio ligado.")
    if ws.page_setup.scale != 100:
        erros.append(f"scale deveria ser 100, veio {ws.page_setup.scale!r}.")

    if erros:
        raise ValueError("Validação das capas falhou depois de salvar/reabrir:\n" + "\n".join(f"- {e}" for e in erros))

    return True
