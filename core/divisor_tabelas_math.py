"""
Divisor de Tabelas de Regiões (v2).

Reescrita completa a pedido do Lucas — a v1 tentava ESTIMAR
matematicamente onde o Excel quebraria a página automaticamente
(altura de linha, papel, margem, escala), e essa estimativa nunca bate
100% com o que o Excel/impressora real decide. Essa incerteza, somada a
bugs na mecânica de remontar mesclagem na hora de dividir (uma
mesclagem que só encostava na borda da região sendo reconstruída era
desfeita mesmo estando majoritariamente FORA dela), foi o motivo da v1
não ter ficado como esperado.

Modelo novo, bem mais simples e prático — critério FIXO, sem
estimativa nenhuma:
    - cada tabela vive num "slot" de página de 32 linhas: 2 em branco
      acima + 28 linhas úteis + 2 em branco abaixo;
    - as 28 linhas úteis contam TUDO: título, cabeçalho, labels, Base,
      Pergunta e legenda (quando houver);
    - se uma tabela ultrapassar as 28 linhas, é candidata a ajuste.

Duas ações possíveis pra quem ultrapassa:
    - tabela NORMAL: divide os labels em partes, repetindo
      cabeçalho+Base+Pergunta(+legenda, se houver) em cada parte, com
      "Continuação" a partir da 2ª;
    - tabela PROTEGIDA (nunca pode dividir os labels no meio — ex.:
      IPV, Avaliação com Aprovação/Regular/Reprovação): se tiver
      legenda, a legenda é movida pra uma página própria, sem tocar nos
      labels. Sem legenda, não tem como corrigir automaticamente — só
      é reportada.

Sem dependência do Streamlit — pode ser testado isoladamente.
"""
from __future__ import annotations

import unicodedata
from copy import copy
from dataclasses import dataclass
from typing import Optional

from openpyxl.formula.translate import Translator
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.pagebreak import Break

from core.planilha_utils import inserir_linhas_seguro, remover_linhas_seguro


# ---------------------------------------------------------------------------
# Constantes do critério fixo
# ---------------------------------------------------------------------------
# O orçamento de página é em ALTURA (pontos), não em contagem de linha —
# nem toda linha tem 15 de altura (código 09 deixa Sexo/Avaliação em 30 e
# Escolaridade em 45, por exemplo), então uma tabela com "28 linhas" só
# cabe de verdade nas 28 se TODAS elas tiverem a altura de referência.
# 32 linhas de referência (15 cada) = 480 de altura total da página; 4
# linhas de referência (2 acima + 2 abaixo, sempre altura padrão de
# verdade, geradas por nós) = 60 reservados; sobram 420 de altura útil.
ALTURA_LINHA_PADRAO = 15.0
LINHAS_POR_PAGINA = 32
LINHAS_RESERVADAS = 4  # 2 acima da tabela + 2 abaixo — nunca contam pro orçamento
ALTURA_TOTAL_PAGINA = LINHAS_POR_PAGINA * ALTURA_LINHA_PADRAO  # 480
ALTURA_RESERVADA = LINHAS_RESERVADAS * ALTURA_LINHA_PADRAO  # 60
ALTURA_UTIL = ALTURA_TOTAL_PAGINA - ALTURA_RESERVADA  # 420

GAP_ROWS = 4  # mesma convenção do código 07: vão padrão entre blocos
BREAK_AFTER_GAP_ROWS = 2  # quebra de página no meio do vão de 4
DEFAULT_ROW_HEIGHT = ALTURA_LINHA_PADRAO
CONTINUATION_TEXT = "Continuação"

# Tabelas que NUNCA podem ter os labels divididos no meio — três formas
# de identificar, combinadas (qualquer uma que bater já protege):
#   1) palavra-chave no TÍTULO da tabela
GATILHOS_TITULO_PROTEGIDO = ("ipv",)
#   2) gatilho no CONTEÚDO das células de label (mesmos gatilhos usados
#      no código 09 pra altura fixa 30 — ver relatorio_automatizado_math.py
#      — mais 'Rejeição', que entra só aqui, não tem altura fixa 30)
GATILHOS_CONTEUDO_PROTEGIDO = ("aprovação", "regular", "reprovação", "rejeição")
#   3) lista manual de títulos exatos (normalizados) — Lucas vai
#      passando casos que não se encaixam nas regras acima
TABELAS_PROTEGIDAS_MANUAL: set[str] = set()


def _norm(value) -> str:
    """Maiúsculo, sem acento, espaços colapsados — usado tanto pra
    identificar marcadores estruturais (TOTAL/BASE/LEGENDA/PERGUNTA)
    quanto pra comparar título/conteúdo contra os gatilhos de tabela
    protegida."""
    if value is None:
        return ""
    text = str(value).strip().upper()
    text = "".join(
        ch for ch in unicodedata.normalize("NFD", text)
        if unicodedata.category(ch) != "Mn"
    )
    return " ".join(text.split())


# ---------------------------------------------------------------------------
# Detecção de tabelas — praticamente inalterada da v1 (Lucas confirmou
# que o problema era de formatação na hora de dividir, não de
# identificar a tabela). Único ajuste: `detectar_tabelas` (ancorada em
# Base, pega TODA tabela, com ou sem legenda) volta a ser o detector
# principal, já que agora o módulo também trata tabelas sem legenda —
# antes só `detectar_tabelas_regionais` (ancorada em LEGENDA) era usada.
# ---------------------------------------------------------------------------
@dataclass
class TableInfo:
    sheet: str
    start_row: int
    title_row: int
    header_start: int
    label_start: int
    label_end: int
    base_row: int
    footer_end: int
    legend_start: Optional[int]
    legend_end: Optional[int]
    title: str

    @property
    def end_row(self) -> int:
        """Última linha do bloco inteiro da tabela — inclui a legenda,
        se houver."""
        return self.legend_end if self.legend_end is not None else self.footer_end

    @property
    def label_count(self) -> int:
        return max(0, self.label_end - self.label_start + 1)

    @property
    def has_legend(self) -> bool:
        return self.legend_start is not None and self.legend_end is not None

    @property
    def tamanho_total(self) -> int:
        """Quantidade de linhas FÍSICAS do bloco inteiro (título até o
        fim, legenda incluída) — só informativo agora (pra exibir na
        tela); a decisão de precisar de ajuste usa altura real
        (`_altura_bloco`), não contagem de linha."""
        return self.end_row - self.start_row + 1

    @property
    def tamanho_sem_legenda(self) -> int:
        """Quantidade de linhas FÍSICAS sem a legenda — só informativo,
        mesma ressalva acima."""
        return self.footer_end - self.start_row + 1


def _row_height(ws, row: int) -> float:
    """Altura de verdade da linha — usa a altura explícita se tiver
    (ex.: 30/45 aplicada pelo código 09), senão cai na altura padrão da
    planilha, senão em `ALTURA_LINHA_PADRAO`."""
    dim = ws.row_dimensions.get(row)
    if dim and dim.height is not None:
        return float(dim.height)
    if ws.sheet_format.defaultRowHeight is not None:
        return float(ws.sheet_format.defaultRowHeight)
    return ALTURA_LINHA_PADRAO


def _altura_bloco(ws, start: int, end: int) -> float:
    """Soma a altura real de cada linha em [start, end] — é isso que é
    comparado contra `ALTURA_UTIL`, não a contagem de linhas."""
    return sum(_row_height(ws, r) for r in range(start, end + 1))


def _row_has_content(ws, row: int, max_col: int) -> bool:
    for c in range(1, max_col + 1):
        v = ws.cell(row=row, column=c).value
        if v is not None and str(v).strip() != "":
            return True
    return False


def _row_has_total(ws, row: int, max_col: int) -> bool:
    return any(_norm(ws.cell(row=row, column=c).value) == "TOTAL" for c in range(1, max_col + 1))


def _row_looks_like_data(ws, row: int, max_col: int) -> bool:
    label = ws.cell(row=row, column=1).value
    if label is None or str(label).strip() == "":
        return False
    if _norm(label) in {"BASE", "BASE REDUZIDA"}:
        return False
    for c in range(2, max_col + 1):
        value = ws.cell(row=row, column=c).value
        if isinstance(value, (int, float)):
            return True
        if isinstance(value, str) and value.startswith("="):
            return True
    return False


def _find_footer_end(ws, base_row: int, max_col: int, max_row: int) -> int:
    """Última linha fixa entre Base e a legenda (Pergunta, e — em
    tabelas de avaliação — uma linha 'Média' entre Base e Pergunta)."""
    end = base_row
    r = base_row + 1
    probe_limit = min(max_row, base_row + 8)
    found_question = False

    while r <= probe_limit:
        a = ws.cell(row=r, column=1).value
        n = _norm(a)

        if n.startswith("LEGENDA"):
            break
        if _row_has_total(ws, r, max_col) or n == "BASE":
            break

        if _row_has_content(ws, r, max_col):
            end = r
            if n.startswith("PERGUNTA:") or str(a or "").lstrip().startswith("*"):
                found_question = True
                r += 1
                while r <= max_row and _row_has_content(ws, r, max_col):
                    nn = _norm(ws.cell(row=r, column=1).value)
                    if nn.startswith("LEGENDA") or _row_has_total(ws, r, max_col):
                        break
                    end = r
                    r += 1
                break
        elif found_question:
            break

        r += 1

    return end


def _find_legend_block(ws, footer_end: int, max_col: int, max_row: int) -> tuple[Optional[int], Optional[int]]:
    """Acha um bloco LEGENDA logo depois do rodapé da tabela (após um
    pequeno vão em branco)."""
    r = footer_end + 1
    max_probe = min(max_row, footer_end + 4)
    while r <= max_probe and not _row_has_content(ws, r, max_col):
        r += 1
    legend_marker = _norm(ws.cell(row=r, column=1).value) if r <= max_row else ""
    if r > max_row or not legend_marker.startswith("LEGENDA"):
        return None, None

    legend_start = r
    legend_end = r
    r += 1
    # a legenda tem uma linha em branco de propósito logo após o
    # título (separador antes do primeiro item, adicionada a pedido do
    # Lucas) — pula ela sem considerar "fim da legenda"; sem isso, o
    # laço abaixo parava bem nessa linha, achando que a legenda não
    # tinha nenhum item.
    if r <= max_row and not _row_has_content(ws, r, max_col):
        r += 1
    while r <= max_row:
        if not _row_has_content(ws, r, max_col):
            break
        if _row_has_total(ws, r, max_col):
            break
        legend_end = r
        r += 1
    return legend_start, legend_end


def _find_title_row(ws, header_start: int, max_col: int) -> int:
    r = header_start - 1
    while r >= 1 and not _row_has_content(ws, r, max_col):
        r -= 1
    if r < 1:
        return header_start
    return r


def detectar_tabelas(ws) -> list[TableInfo]:
    """Detecta toda tabela do relatório (com ou sem legenda), ancorada
    em linhas 'Base'."""
    tables: list[TableInfo] = []
    used_bases: set[int] = set()

    # `ws.max_column`/`ws.max_row` recalculam varrendo a planilha
    # inteira toda vez que são acessados (não são um valor já
    # guardado) — chamar isso dentro de um laço por linha, centenas ou
    # milhares de vezes, é a causa principal do Divisor de Tabelas
    # ficar lento em relatórios grandes. Calcula uma vez só aqui (a
    # análise inteira é só leitura — nada aqui insere linha/coluna) e
    # passa adiante.
    max_col = ws.max_column
    max_row_atual = ws.max_row

    for base_row in range(1, max_row_atual + 1):
        if _norm(ws.cell(row=base_row, column=1).value) != "BASE":
            continue
        if base_row in used_bases:
            continue

        header_start = None
        for r in range(base_row - 1, max(0, base_row - 100), -1):
            if _norm(ws.cell(row=r, column=1).value) == "BASE":
                break
            if _row_has_total(ws, r, max_col):
                header_start = r
                break
        if header_start is None:
            continue

        data_rows = [r for r in range(header_start + 1, base_row) if _row_looks_like_data(ws, r, max_col)]
        if not data_rows:
            continue
        label_start = min(data_rows)
        label_end = max(data_rows)
        if any(not _row_looks_like_data(ws, r, max_col) for r in range(label_start, label_end + 1)):
            continue

        title_row = _find_title_row(ws, header_start, max_col)
        footer_end = _find_footer_end(ws, base_row, max_col, max_row_atual)
        legend_start, legend_end = _find_legend_block(ws, footer_end, max_col, max_row_atual)
        title_value = ws.cell(row=title_row, column=1).value
        title = str(title_value).strip() if title_value is not None else f"Tabela linha {title_row}"

        tables.append(
            TableInfo(
                sheet=ws.title, start_row=title_row, title_row=title_row,
                header_start=header_start, label_start=label_start, label_end=label_end,
                base_row=base_row, footer_end=footer_end, legend_start=legend_start,
                legend_end=legend_end, title=title,
            )
        )
        used_bases.add(base_row)

    out: list[TableInfo] = []
    last_end = 0
    for table in sorted(tables, key=lambda t: t.start_row):
        if table.start_row <= last_end:
            continue
        out.append(table)
        last_end = table.end_row
    return out


# ---------------------------------------------------------------------------
# Tabela protegida
# ---------------------------------------------------------------------------
def eh_tabela_protegida(ws, table: TableInfo, max_col: int) -> bool:
    """Uma tabela é protegida (nunca divide os labels no meio) se
    qualquer uma das 3 regras bater — ver constantes no topo do
    arquivo."""
    titulo_norm = _norm(table.title)
    if any(_norm(g) in titulo_norm for g in GATILHOS_TITULO_PROTEGIDO):
        return True
    if titulo_norm in {_norm(t) for t in TABELAS_PROTEGIDAS_MANUAL}:
        return True
    for r in range(table.label_start, table.label_end + 1):
        for c in range(1, max_col + 1):
            valor = ws.cell(row=r, column=c).value
            if isinstance(valor, str):
                # sem hífen/espaço aqui — o código 04 pode ter
                # reformatado o texto em várias linhas com hífen (ex.:
                # "Apro-\nvação"), que senão não bateria mais com o
                # gatilho "aprovação"
                texto = _norm(valor).replace("-", "").replace(" ", "")
                if any(texto.startswith(_norm(g).replace(" ", "")) for g in GATILHOS_CONTEUDO_PROTEGIDO):
                    return True
    return False


# ---------------------------------------------------------------------------
# Mecânica de captura/cópia/mesclagem — reescrita do zero.
#
# A causa raiz do bug de formatação da v1 estava em `_clear_region`:
# desfazia QUALQUER mesclagem que só encostasse na borda da região
# sendo limpa (comparação por sobreposição), mesmo que a maior parte
# dela estivesse fora — uma mesclagem que ia de A1:E3 era desfeita ao
# limpar a região [3,10], por exemplo, mesmo só tendo 1 linha dentro.
# Aqui só mexe em mesclagem TOTALMENTE contida na região.
# ---------------------------------------------------------------------------
def _capturar_bloco(ws, start: int, end: int, max_col: int):
    """Tira uma 'foto' de todo o conteúdo/formatação/altura/mesclagem
    de [start, end] antes de mexer em qualquer coisa.

    Guarda a REFERÊNCIA do objeto de estilo (Font/Fill/Border/...), não
    uma cópia — nada neste arquivo muda um atributo de estilo "no
    lugar" (é sempre criar/atribuir um objeto novo por completo), então
    compartilhar a referência entre células é seguro: o openpyxl
    deduplica estilos por VALOR na hora de salvar, não por identidade
    do objeto Python. `copy.copy()` num objeto de estilo do openpyxl é
    caro (passa por uma serialização XML interna pra fazer a cópia) —
    evitar isso pra cada célula de cada linha copiada foi o que mais
    pesava no tempo total de divisão de tabelas."""
    cells = {}
    for r in range(start, end + 1):
        for c in range(1, max_col + 1):
            src = ws.cell(row=r, column=c)
            cells[(r, c)] = {
                "value": src.value,
                # `cel.font` (e fill/border/alignment/protection) devolve um
                # StyleProxy, não o objeto Font/... de verdade — precisa de
                # copy() aqui pra materializar o objeto real (é o único
                # jeito de "descolar" do proxy). Isso só é feito UMA VEZ
                # aqui na captura; `_copiar_linha` reaproveita esse mesmo
                # objeto já materializado em cada parte da tabela (uma
                # tabela dividida em 3 partes colava a mesma linha 3
                # vezes — cada colagem copiando de novo era 3x o custo à
                # toa).
                "font": copy(src.font),
                "fill": copy(src.fill),
                "border": copy(src.border),
                "alignment": copy(src.alignment),
                "number_format": src.number_format,
                "protection": copy(src.protection),
            }
    heights = {r: ws.row_dimensions[r].height for r in range(start, end + 1)}
    merges = []
    for rng in list(ws.merged_cells.ranges):
        if rng.min_row >= start and rng.max_row <= end:
            merges.append((rng.min_row, rng.min_col, rng.max_row, rng.max_col))
    return cells, heights, merges


def _limpar_regiao_com_seguranca(ws, start: int, end: int, max_col: int):
    """Desfaz mesclagem e limpa valor/estilo em [start, end] — só mexe
    em mesclagem TOTALMENTE contida na região (esse é o fix do bug da
    v1: uma mesclagem que só encosta na borda, com a maior parte fora
    da região, é preservada intacta).

    Se uma mesclagem preservada (não totalmente contida) tiver alguma
    célula "filha" dentro da região — inevitável quando ela encosta na
    borda —, essas células são somente-leitura no openpyxl
    (`MergedCell`) e não podem ser limpas; são simplesmente puladas.
    """
    from openpyxl.cell.cell import MergedCell

    for rng in list(ws.merged_cells.ranges):
        if rng.min_row >= start and rng.max_row <= end:
            try:
                ws.unmerge_cells(str(rng))
            except KeyError:
                ws.merged_cells.ranges.discard(rng)

    fonte_padrao = Font()
    for r in range(start, end + 1):
        for c in range(1, max_col + 1):
            cel = ws.cell(row=r, column=c)
            if isinstance(cel, MergedCell):
                continue
            cel.value = None
            cel.font = fonte_padrao
            cel.fill = ws.parent._named_styles[0].fill
            cel.border = ws.parent._named_styles[0].border
            cel.alignment = ws.parent._named_styles[0].alignment
            cel.number_format = "General"
            cel.protection = ws.parent._named_styles[0].protection
        ws.row_dimensions[r].height = None


def _copiar_linha(ws, cells, heights, origem: int, destino: int, max_col: int):
    from openpyxl.cell.cell import MergedCell

    for c in range(1, max_col + 1):
        dst = ws.cell(row=destino, column=c)
        if isinstance(dst, MergedCell):
            # célula-filha de uma mesclagem preservada que encosta na
            # região (não totalmente contida, então não foi desfeita) —
            # somente-leitura, não dá pra escrever; pula.
            continue
        dado = cells[(origem, c)]
        valor = dado["value"]
        if isinstance(valor, str) and valor.startswith("="):
            try:
                valor = Translator(
                    valor, origin=f"{get_column_letter(c)}{origem}"
                ).translate_formula(f"{get_column_letter(c)}{destino}")
            except Exception:
                pass
        dst.value = valor
        dst.font = dado["font"]
        dst.fill = dado["fill"]
        dst.border = dado["border"]
        dst.alignment = dado["alignment"]
        dst.number_format = dado["number_format"]
        dst.protection = dado["protection"]
    ws.row_dimensions[destino].height = heights[origem]


def _remontar_mesclagens(ws, merges, mapa_linhas: dict[int, int]):
    """Recria cada mesclagem original na posição de destino — só quando
    TODAS as linhas dela têm mapeamento (ou seja, foram realmente
    copiadas pra esta parte) E o destino continua contíguo (sem buracos
    — uma mesclagem que caía metade numa parte e metade em outra, por
    causa de uma divisão de labels no meio dela, é descartada em vez de
    remontada errada)."""
    for min_r, min_c, max_r, max_c in merges:
        linhas_origem = list(range(min_r, max_r + 1))
        if not all(r in mapa_linhas for r in linhas_origem):
            continue
        linhas_destino = [mapa_linhas[r] for r in linhas_origem]
        if linhas_destino != list(range(linhas_destino[0], linhas_destino[0] + len(linhas_destino))):
            continue
        ws.merge_cells(
            start_row=linhas_destino[0], start_column=min_c,
            end_row=linhas_destino[-1], end_column=max_c,
        )


def _adicionar_quebra(ws, row_id: int):
    if row_id >= 1 and not any(int(b.id) == row_id for b in ws.row_breaks.brk if b.id is not None):
        ws.row_breaks.append(Break(id=row_id))
        ws.row_breaks.brk = sorted(ws.row_breaks.brk, key=lambda b: int(b.id or 0))


# ---------------------------------------------------------------------------
# Ação 1 — dividir os labels em partes (tabela NÃO protegida)
# ---------------------------------------------------------------------------
def _tamanhos_balanceados(total: int, partes: int) -> list[int]:
    """Divide `total` em `partes` grupos o mais parecido possível — os
    primeiros `resto` grupos ficam com 1 a mais (ex.: 13 labels em 2
    partes -> [7, 6], 12 labels em 2 partes -> [6, 6])."""
    base = total // partes
    resto = total % partes
    return [base + (1 if i < resto else 0) for i in range(partes)]


def escolher_grupos_labels(ws, table: TableInfo) -> list[list[int]]:
    """
    Agrupa as linhas de label em partes de tamanho o mais PARECIDO
    possível (proporcional, não "enche a primeira página até estourar")
    — ex.: 12 labels em 2 partes vira 6+6; 13 vira 7+6.

    Ainda respeita a altura real de cada linha: acha o MENOR número de
    partes onde essa divisão balanceada por contagem cabe (em altura)
    em toda parte; só aumenta o número de partes se com aquele número
    alguma parte ainda estourasse o espaço útil. Uma parte de 1 label
    só é sempre aceita mesmo que estoure sozinha — não tem como dividir
    mais que isso.

    Preserva sempre a ORDEM original dos labels — nunca reordena
    categoria de resposta.
    """
    prefixo_altura = _altura_bloco(ws, table.start_row, table.label_start - 1)
    sufixo_altura = _altura_bloco(ws, table.base_row, table.end_row)
    fixo = prefixo_altura + sufixo_altura
    capacidade = max(0.01, ALTURA_UTIL - fixo)

    linhas_labels = list(range(table.label_start, table.label_end + 1))
    total_labels = len(linhas_labels)
    if total_labels <= 1:
        return [linhas_labels]

    for n_partes in range(1, total_labels + 1):
        tamanhos = _tamanhos_balanceados(total_labels, n_partes)
        grupos = []
        cursor = 0
        cabe = True
        for tam in tamanhos:
            grupo = linhas_labels[cursor:cursor + tam]
            cursor += tam
            if len(grupo) > 1:
                altura_grupo = sum(_row_height(ws, r) for r in grupo)
                if altura_grupo > capacidade + 0.01:
                    cabe = False
            grupos.append(grupo)
        if cabe:
            return grupos

    # nunca deveria chegar aqui (n_partes = total_labels sempre cabe,
    # já que grupo de 1 label é sempre aceito) — fallback defensivo
    return [[r] for r in linhas_labels]


def _garantir_duas_linhas_antes_da_quebra(ws, linha_fim_legenda):
    """
    Garante EXATAMENTE 2 linhas em branco entre `linha_fim_legenda` e a
    quebra de página mais próxima logo depois dela — é o padrão do
    relatório.

    Não mexe em nada se não achar nenhuma quebra nas proximidades (não
    força a criação de uma quebra nova, só corrige o espaçamento em
    torno de uma que já existe).

    Insere ou remove linhas ANTES da linha que segura a quebra (nunca a
    própria linha dela) — removê-la faria a quebra desaparecer
    silenciosamente (é assim que `remover_linhas_seguro` funciona; o
    cuidado tem que vir de quem chama).

    Idempotente. Sempre "materializa" as 2 linhas-alvo aplicando um
    estilo (mesmo sem texto) — célula totalmente vazia, sem nenhum
    estilo, some ao salvar no openpyxl.
    """
    candidatos = [
        b.id for b in ws.row_breaks.brk
        if b.id is not None and linha_fim_legenda <= b.id <= linha_fim_legenda + 20
    ]
    if not candidatos:
        return
    quebra_id_atual = min(candidatos)
    quebra_id_alvo = linha_fim_legenda + 2

    if quebra_id_atual < quebra_id_alvo:
        inserir_linhas_seguro(ws, linha_fim_legenda + 1, quebra_id_alvo - quebra_id_atual)
    elif quebra_id_atual > quebra_id_alvo:
        remover_linhas_seguro(ws, linha_fim_legenda + 1, quebra_id_atual - quebra_id_alvo)

    for rr in range(linha_fim_legenda + 1, linha_fim_legenda + 3):
        ws.row_dimensions[rr].height = DEFAULT_ROW_HEIGHT
        ws.cell(row=rr, column=1).font = Font(name="DIN Book", size=9)


def _dividir_tabela(ws, table: TableInfo, max_col: int) -> dict:
    grupos = escolher_grupos_labels(ws, table)
    if len(grupos) <= 1:
        return {"acao": "nao_precisou", "table": table, "parts": 1, "part_sizes": [table.label_count]}

    cells, heights, merges = _capturar_bloco(ws, table.start_row, table.end_row, max_col)
    prefixo = list(range(table.start_row, table.label_start))
    sufixo = list(range(table.base_row, table.end_row + 1))
    part_sizes = [len(g) for g in grupos]

    tamanho_original = table.end_row - table.start_row + 1
    tamanho_novo = sum(len(prefixo) + len(g) + len(sufixo) for g in grupos) + GAP_ROWS * (len(grupos) - 1)
    extra = tamanho_novo - tamanho_original
    if extra > 0:
        inserir_linhas_seguro(ws, table.end_row + 1, extra)

    # qualquer quebra manual pré-existente DENTRO da tabela original é
    # substituída pelas quebras novas que vão ser adicionadas nos vãos
    ws.row_breaks.brk = [
        b for b in ws.row_breaks.brk
        if b.id is None or not (table.start_row <= int(b.id) < table.end_row)
    ]

    _limpar_regiao_com_seguranca(ws, table.start_row, table.start_row + tamanho_novo - 1, max_col)

    cursor = table.start_row
    for parte_idx, grupo in enumerate(grupos):
        linhas_origem = prefixo + grupo + sufixo

        mapa_linhas = {}
        for origem in linhas_origem:
            _copiar_linha(ws, cells, heights, origem, cursor, max_col)
            mapa_linhas[origem] = cursor
            cursor += 1

        if parte_idx > 0:
            offset_cabecalho = prefixo.index(table.header_start)
            linha_continuacao = prefixo[offset_cabecalho]
            linha_continuacao_destino = mapa_linhas[linha_continuacao]
            cel = ws.cell(row=linha_continuacao_destino, column=1)
            cel.value = CONTINUATION_TEXT
            cel.font = Font(name="DIN", size=9, bold=True)

        _remontar_mesclagens(ws, merges, mapa_linhas)

        if parte_idx < len(grupos) - 1:
            inicio_vao = cursor
            for rr in range(inicio_vao, inicio_vao + GAP_ROWS):
                ws.row_dimensions[rr].height = DEFAULT_ROW_HEIGHT
            _adicionar_quebra(ws, inicio_vao + BREAK_AFTER_GAP_ROWS - 1)
            cursor += GAP_ROWS

    # o vão DEPOIS da última parte não é criado por este laço (só as
    # partes intermediárias ganham vão novo) — é herdado de onde a
    # tabela original terminava, deslocado pela inserção lá em cima.
    # Se a tabela tiver legenda, garante que esse vão herdado também
    # tenha as 2 linhas padrão antes da quebra seguinte.
    if table.has_legend:
        _garantir_duas_linhas_antes_da_quebra(ws, cursor - 1)

    return {
        "acao": "dividida", "table": table, "parts": len(grupos), "part_sizes": part_sizes,
        "nova_ultima_linha": cursor - 1,
    }


# ---------------------------------------------------------------------------
# Ação 2 — mover legenda pra página própria (tabela PROTEGIDA, com legenda)
# ---------------------------------------------------------------------------
def _mover_legenda_para_pagina_propria(ws, table: TableInfo) -> dict:
    gap_atual = table.legend_start - table.footer_end - 1
    if gap_atual < GAP_ROWS:
        inserir_linhas_seguro(ws, table.footer_end + 1, GAP_ROWS - gap_atual)
        deslocamento = GAP_ROWS - gap_atual
    elif gap_atual > GAP_ROWS:
        remover_linhas_seguro(ws, table.footer_end + 1, gap_atual - GAP_ROWS)
        deslocamento = -(gap_atual - GAP_ROWS)
    else:
        deslocamento = 0

    inicio_vao = table.footer_end + 1
    for rr in range(inicio_vao, inicio_vao + GAP_ROWS):
        ws.row_dimensions[rr].height = DEFAULT_ROW_HEIGHT

    # limpa qualquer quebra velha dentro/perto do vão antes de adicionar a nova
    limite_inferior = inicio_vao - 1
    limite_superior = inicio_vao + GAP_ROWS + 1
    ws.row_breaks.brk = [
        b for b in ws.row_breaks.brk
        if b.id is None or not (limite_inferior <= int(b.id) <= limite_superior)
    ]
    _adicionar_quebra(ws, inicio_vao + BREAK_AFTER_GAP_ROWS - 1)

    nova_legend_end = table.legend_end + deslocamento
    # o vão DEPOIS da legenda (entre ela e o que vem a seguir no
    # relatório) não é criado por esta função — é herdado de onde a
    # legenda terminava antes de mover, só que agora deslocado. Garante
    # que ele também tenha as 2 linhas padrão antes da próxima quebra.
    _garantir_duas_linhas_antes_da_quebra(ws, nova_legend_end)

    return {
        "acao": "legenda_movida", "table": table,
        "nova_legend_start": table.legend_start + deslocamento,
        "nova_legend_end": nova_legend_end,
    }


# ---------------------------------------------------------------------------
# Orquestração
# ---------------------------------------------------------------------------
def analisar_planilha(ws) -> dict:
    tables = detectar_tabelas(ws)
    max_col = ws.max_column
    candidatas = []
    for table in tables:
        altura_total = _altura_bloco(ws, table.start_row, table.end_row)
        if altura_total <= ALTURA_UTIL + 0.01:
            continue
        protegida = eh_tabela_protegida(ws, table, max_col)
        candidatas.append({
            "table": table,
            "tamanho": table.tamanho_total,  # informativo (contagem de linhas)
            "altura": altura_total,  # é isso que decide, de verdade
            "protegida": protegida,
            "tem_legenda": table.has_legend,
        })
    return {"sheet": ws.title, "tables": tables, "candidatas": candidatas}


def processar_workbook(wb) -> dict:
    """Analisa e ajusta as tabelas do relatório inteiro segundo o
    critério fixo de 28 linhas úteis por página."""
    resultados_por_aba = []

    for ws in wb.worksheets:
        if ws.title.lower() == "sumario":
            continue

        analise = analisar_planilha(ws)
        max_col = ws.max_column
        # processa de baixo pra cima: inserir/remover linha numa tabela
        # não deve invalidar a posição das tabelas ACIMA dela, já
        # calculadas
        candidatas_ordenadas = sorted(
            analise["candidatas"], key=lambda c: c["table"].start_row, reverse=True
        )

        acoes = []
        for item in candidatas_ordenadas:
            table = item["table"]
            protegida = item["protegida"]
            tem_legenda = item["tem_legenda"]

            if not protegida:
                resultado = _dividir_tabela(ws, table, max_col)
            elif tem_legenda:
                altura_sem_legenda = _altura_bloco(ws, table.start_row, table.footer_end)
                if altura_sem_legenda > ALTURA_UTIL + 0.01:
                    resultado = {"acao": "nao_corrigida", "table": table, "motivo": "protegida_sem_espaco_mesmo_sem_legenda"}
                else:
                    resultado = _mover_legenda_para_pagina_propria(ws, table)
            else:
                resultado = {"acao": "nao_corrigida", "table": table, "motivo": "protegida_sem_legenda"}

            acoes.append({
                "titulo": table.title,
                "tamanho": item["tamanho"],
                "altura": item["altura"],
                "tem_legenda": tem_legenda,
                "protegida": protegida,
                **resultado,
            })

        resultados_por_aba.append({
            "sheet": ws.title,
            "tables_detected": len(analise["tables"]),
            "acoes": acoes,
        })

        # A área de impressão fica travada num range fixo (ex.: "A1:M3740")
        # gravado por uma etapa anterior do fluxo — inserir linha aqui não
        # estica esse range sozinho. Sem atualizar, tudo que a divisão de
        # tabelas empurrar pra baixo do fim antigo cai FORA da área de
        # impressão — no "Visualizar Quebra de Página" do Excel aparece
        # como página cinza, mesmo com quebra/formatação corretas por
        # baixo. Reexpande pra cobrir o conteúdo real depois de processar.
        ultima_linha_real = ws.max_row
        max_col_final = ws.max_column
        while ultima_linha_real > 1 and not _row_has_content(ws, ultima_linha_real, max_col_final):
            ultima_linha_real -= 1
        ultima_coluna_letra = get_column_letter(max_col_final)
        ws.print_area = f"A1:{ultima_coluna_letra}{ultima_linha_real}"

        # Linhas de grade desligadas na exportação do Divisor de Tabelas —
        # independente do que o arquivo de entrada já tinha.
        ws.sheet_view.showGridLines = False

    todas_acoes = [a for r in resultados_por_aba for a in r["acoes"]]
    return {
        "sheets_analyzed": len(resultados_por_aba),
        "tables_candidatas": len(todas_acoes),
        "tables_divididas": sum(1 for a in todas_acoes if a["acao"] == "dividida"),
        "legenda_movida": sum(1 for a in todas_acoes if a["acao"] == "legenda_movida"),
        "nao_corrigidas": sum(1 for a in todas_acoes if a["acao"] == "nao_corrigida"),
        "partes_geradas": sum(a.get("parts", 0) for a in todas_acoes if a["acao"] == "dividida"),
        "acoes": todas_acoes,
        "resultados_por_aba": resultados_por_aba,
    }
