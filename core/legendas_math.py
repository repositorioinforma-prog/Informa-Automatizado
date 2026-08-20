"""
Motor de "Legendas": localiza, num relatório de tabelas, os blocos cuja
segmentação é por região/território (Regiões, Bairros, Municípios etc.)
e insere o texto de legenda correspondente (vindo de um segundo arquivo)
logo depois da linha "Pergunta:" de cada bloco.

Sem dependência do Streamlit — pode ser testado isoladamente.
"""
import openpyxl
import unicodedata

from core.base_multiplas_math import normalizar_texto
from core.planilha_utils import inserir_linhas_seguro, remover_linhas_seguro

# Nomes de categoria (grupo da tabela) reconhecidos como "segmentação
# territorial" — qualquer bloco cujo grupo bata com um destes (comparação
# tolerante a acento/caixa/hífen, via normalizar_texto) é candidato a
# receber legenda.
CATEGORIAS_LEGENDA = [
    "Regiões",
    "Mesorregiões",
    "Microrregiões",
    "Regiões Intermediárias",
    "Regiões Imediatas",
    "Municípios",
    "Cidades",
    "Bairros",
    "Regiões Administrativas",
    "RPA",
    "Regiões da Capital",
    "Capital",
    "Sub-Regiões",
    "Sub-regiões",
    "Subregiões",
]
CATEGORIAS_LEGENDA_NORM = {normalizar_texto(c) for c in CATEGORIAS_LEGENDA}


def _forcar_cor_preta(item):
    """
    Reconstrói um item de legenda forçando a cor do texto pra preto
    explícito (RGB), em vez de deixar uma cor referenciada por TEMA
    (theme=N) — cores de tema são relativas ao arquivo: o mesmo índice
    de tema pode ser preto no arquivo de legendas de origem e virar
    outra cor (ex.: laranja) no relatório de destino, porque cada
    workbook tem sua própria paleta de tema. Preserva negrito/itálico
    de cada trecho, só troca a cor. Usada pelos dois motores de
    Legendas (v1 e v2) sempre que um item é transplantado de um
    workbook pra outro.
    """
    from copy import copy as _copy_style

    from openpyxl.cell.rich_text import CellRichText, TextBlock
    from openpyxl.styles.colors import Color

    if not isinstance(item, CellRichText):
        return item
    novos_blocos = []
    for bloco in item:
        if isinstance(bloco, TextBlock):
            nova_fonte = _copy_style(bloco.font) if bloco.font else None
            if nova_fonte is not None:
                nova_fonte.color = Color(rgb="FF000000")
            novos_blocos.append(TextBlock(nova_fonte, bloco.text))
        else:
            novos_blocos.append(bloco)
    return CellRichText(novos_blocos)


def _eh_linha_vazia(linha):
    return linha is None or all(v is None for v in linha)


def _forward_fill(linha):
    resultado = list(linha)
    atual = None
    for i, v in enumerate(resultado):
        if v is not None and str(v).strip() != "":
            atual = v
        resultado[i] = atual
    return resultado


def parsear_blocos_tabela(caminho_ou_arquivo, aba=None):
    """
    Parseia um relatório de tabelas SEM o marcador 'Titulo:' (o título é
    só o texto puro na primeira linha não-vazia de cada bloco) — formato
    usado neste tipo de exportação. Cada bloco é delimitado por linhas em
    branco e termina numa linha 'Pergunta:'.

    Retorna uma lista de dicts:
      {"titulo_idx": linha do título (1-indexada), "titulo": texto,
       "grupos": [...], "categorias": [...], "linha_pergunta": nº da
       linha 'Pergunta:' (1-indexada) ou None}
    """
    wb = openpyxl.load_workbook(caminho_ou_arquivo, data_only=True, rich_text=True)
    ws = wb[aba] if aba else wb[wb.sheetnames[0]]
    linhas = list(ws.iter_rows(values_only=True))
    n = len(linhas)
    blocos = []
    i = 0

    while i < n:
        if _eh_linha_vazia(linhas[i]):
            i += 1
            continue

        titulo_idx = i + 1
        titulo = linhas[i][0]
        i += 1
        if i >= n:
            break

        linha_grupo = linhas[i]
        i += 1
        grupos = _forward_fill(linha_grupo)

        tem_segmentacao = len(linha_grupo) > 2 and any(v is not None for v in linha_grupo[2:])
        if tem_segmentacao and i < n:
            categorias = list(linhas[i])
            i += 1
        else:
            categorias = [None] * len(linha_grupo)

        linha_pergunta_idx = None
        while i < n:
            linha = linhas[i]
            primeiro = linha[0] if linha else None
            if isinstance(primeiro, str) and primeiro.strip().startswith("Pergunta:"):
                linha_pergunta_idx = i + 1
                i += 1
                break
            i += 1

        blocos.append({
            "titulo_idx": titulo_idx,
            "titulo": titulo,
            "grupos": grupos,
            "categorias": categorias,
            "linha_pergunta": linha_pergunta_idx,
        })

    return blocos


def bloco_precisa_legenda(bloco):
    """Bloco é candidato a legenda se algum de seus grupos bater com CATEGORIAS_LEGENDA."""
    textos = {normalizar_texto(g) for g in bloco["grupos"] if g is not None}
    return bool(textos & CATEGORIAS_LEGENDA_NORM)


def parsear_blocos_legenda(caminho_ou_arquivo, aba=None):
    """
    Parseia o arquivo de legendas: blocos que começam com uma linha cujo
    texto COMEÇA com 'LEGENDA' — pode ser só 'LEGENDA', ou com um rótulo
    depois (ex.: 'LEGENDA CAPITAL', 'LEGENDA REGIÕES') — seguidos de
    linhas de texto (uma por item), até uma linha em branco ou o próximo
    cabeçalho 'LEGENDA...'.

    Carrega com rich_text=True para preservar formatação parcial dentro
    da célula (ex.: "Região 1 [9,09%]:" em negrito, o resto normal) — sem
    isso, o openpyxl achata a célula toda para texto simples e o negrito
    se perde.

    Retorna uma lista de dicts: {"itens": [...], "rotulo": str|None,
    "rotulo_original": str|None} — 'rotulo' é o texto normalizado depois
    de 'LEGENDA' (usado só pra comparação/pareamento), 'rotulo_original'
    é o texto exatamente como está escrito no arquivo (usado pra
    reescrever o cabeçalho no relatório de saída). Ambos None se o
    cabeçalho for só 'LEGENDA', sem nada depois. Cada item pode ser uma
    string simples ou um CellRichText (preserva os trechos em negrito).
    """
    wb = openpyxl.load_workbook(caminho_ou_arquivo, rich_text=True)
    ws = wb[aba] if aba else wb[wb.sheetnames[0]]
    linhas = list(ws.iter_rows(values_only=True))
    n = len(linhas)
    blocos = []
    i = 0

    def _rotulo_legenda(texto):
        """Se 'texto' começa com 'LEGENDA', devolve (rótulo_normalizado,
        rótulo_original, True) — ou (None, None, False) se não for uma
        linha de cabeçalho de legenda. O original é preservado (não só o
        normalizado) porque precisamos reescrever exatamente esse texto
        de volta no relatório de saída (ex.: "LEGENDA REGIÕES"), e a
        versão normalizada (sem acento/caixa, só pra comparação) ficaria
        ilegível se usada pra exibição."""
        t = texto.strip().upper()
        if not t.startswith("LEGENDA"):
            return None, None, False
        resto = texto.strip()[len("LEGENDA"):].strip()
        rotulo_original = resto if resto else None
        rotulo_norm = normalizar_texto(resto) if resto else None
        return rotulo_norm, rotulo_original, True

    while i < n:
        primeiro = linhas[i][0] if linhas[i] else None
        rotulo, rotulo_original, eh_legenda = (
            _rotulo_legenda(primeiro) if isinstance(primeiro, str) else (None, None, False)
        )
        if eh_legenda:
            i += 1
            itens = []
            while i < n and not _eh_linha_vazia(linhas[i]):
                v = linhas[i][0]
                _, _, v_eh_legenda = _rotulo_legenda(v) if isinstance(v, str) else (None, None, False)
                if v_eh_legenda:
                    break
                if v is not None:
                    itens.append(v)
                i += 1
            blocos.append({"itens": itens, "rotulo": rotulo, "rotulo_original": rotulo_original})
        else:
            i += 1

    return blocos


def _assinatura_bloco(bloco):
    """
    'Assinatura' de um bloco elegível: (categoria territorial reconhecida,
    valores exatos das categorias). Blocos com a MESMA assinatura usam a
    mesma legenda — é assim que uma segmentação como "Regiões" (que se
    repete idêntica numa tabela por pergunta/candidato) recebe a legenda
    certa em TODAS as ocorrências, não só na primeira.

    Importante usar os valores exatos (não só a quantidade): uma mesma
    categoria territorial pode aparecer dividida em mais de uma tabela
    quando há categorias demais pra caber numa só (ex.: "Regiões 1-7" e
    "Regiões 8-14" — mesma categoria territorial, mesma quantidade de
    colunas, mas valores diferentes e legendas diferentes). Usar só a
    quantidade juntaria essas duas por engano.
    """
    textos = {normalizar_texto(g) for g in bloco["grupos"] if g is not None}
    reconhecidas = sorted(textos & CATEGORIAS_LEGENDA_NORM)
    categoria_territorial = reconhecidas[0] if reconhecidas else None
    valores_categoria = tuple(
        normalizar_texto(c) for c in bloco["categorias"] if c is not None
    )
    return (categoria_territorial, valores_categoria)


def parear_blocos(blocos_tabela_elegiveis, blocos_legenda):
    """
    Agrupa os blocos elegíveis por "tipo" de segmentação territorial
    (assinatura: categoria + valores exatos das categorias) e associa
    cada tipo distinto a UM bloco de legenda — essa mesma legenda é
    reaplicada a TODAS as tabelas daquele tipo ao longo do relatório, não
    só à primeira ocorrência (a mesma segmentação, ex.: "Regiões",
    costuma se repetir em dezenas de tabelas — uma por pergunta/candidato
    — mas o arquivo de legendas normalmente só define a legenda uma vez
    por tipo).

    Se um bloco de legenda tiver um rótulo explícito que bata com uma das
    categorias territoriais reconhecidas (ex.: 'LEGENDA CAPITAL', 'LEGENDA
    REGIÕES'), esse rótulo é usado para casar direto com o tipo certo —
    mais confiável do que depender só da ordem em que os blocos aparecem
    no arquivo. Blocos de legenda sem rótulo (só 'LEGENDA') continuam
    sendo usados na ordem em que aparecem, como reserva para os tipos que
    não tiverem um bloco rotulado correspondente.

    Retorna (pares, avisos), onde pares é uma lista de (bloco_tabela, bloco_legenda) —
    um item por TABELA (podendo repetir o mesmo bloco_legenda várias vezes).
    """
    avisos = []

    # separa os blocos de legenda com rótulo reconhecido dos genéricos
    fila_por_rotulo = {}
    fila_generica = []
    for bl in blocos_legenda:
        rotulo = bl.get("rotulo")
        if rotulo and rotulo in CATEGORIAS_LEGENDA_NORM:
            fila_por_rotulo.setdefault(rotulo, []).append(bl)
        else:
            fila_generica.append(bl)

    ordem_assinaturas = []
    vistos = set()
    for bt in blocos_tabela_elegiveis:
        sig = _assinatura_bloco(bt)
        if sig not in vistos:
            vistos.add(sig)
            ordem_assinaturas.append(sig)

    assinatura_para_legenda = {}
    indice_generico = 0
    for sig in ordem_assinaturas:
        categoria_territorial = sig[0]
        bl = None

        # 1) tenta um bloco de legenda com rótulo explícito batendo com o tipo
        if categoria_territorial and fila_por_rotulo.get(categoria_territorial):
            bl = fila_por_rotulo[categoria_territorial].pop(0)
        # 2) senão, cai pro próximo bloco genérico (sem rótulo), na ordem do arquivo
        elif indice_generico < len(fila_generica):
            bl = fila_generica[indice_generico]
            indice_generico += 1

        if bl is not None:
            assinatura_para_legenda[sig] = bl
        else:
            avisos.append(
                f"Não sobrou nenhum bloco de legenda disponível para o tipo "
                f"'{categoria_territorial or '?'}'."
            )

    sobras_rotulo = sum(len(fila) for fila in fila_por_rotulo.values())
    sobras_generica = len(fila_generica) - indice_generico
    if sobras_rotulo or sobras_generica:
        avisos.append(
            f"{sobras_rotulo + sobras_generica} bloco(s) de legenda sobraram sem "
            "nenhum tipo de segmentação territorial correspondente."
        )

    pares = []
    avisados_tipo = set()
    for bt in blocos_tabela_elegiveis:
        sig = _assinatura_bloco(bt)
        bl = assinatura_para_legenda.get(sig)
        if bl is None:
            continue

        n_categorias = len(sig[1])
        if n_categorias and len(bl["itens"]) != n_categorias and sig not in avisados_tipo:
            avisados_tipo.add(sig)
            avisos.append(
                f"Tipo '{sig[0] or '?'}' ({n_categorias} categoria(s) na tabela): a "
                f"legenda correspondente tem {len(bl['itens'])} item(ns) — confira se "
                "a ordem/quantidade bate."
            )
        pares.append((bt, bl))

    return pares, avisos


def gerar_workbook_com_legendas(caminho_tabela, pares, aba=None):
    """
    Insere, no PRÓPRIO workbook original (carregado a partir de
    'caminho_tabela'), o texto de legenda de cada par (bloco_tabela,
    bloco_legenda) logo depois da linha 'Pergunta:' daquele bloco — mais
    precisamente: mantém a primeira linha em branco após 'Pergunta:' como
    está, insere "LEGENDA" + os itens, e mais uma linha em branco de
    espaçamento antes do restante do arquivo continuar normalmente.

    Diferente de uma versão anterior, NÃO reconstrói a planilha do zero —
    reconstruir descartava configurações do arquivo original que não
    são copiadas por padrão pelo openpyxl (orientação da página, quebras
    de página manuais, área de impressão etc.). Em vez disso, usa
    `ws.insert_rows()` diretamente no arquivo carregado, o que preserva
    tudo isso automaticamente, já que é o mesmo objeto de planilha.

    O único cuidado necessário com `insert_rows()` é com células
    mescladas: o openpyxl não desloca corretamente faixas mescladas que
    estejam na região afetada, então cada faixa mesclada em/abaixo do
    ponto de inserção é desfeita antes de inserir e refeita depois, já
    deslocada.
    """
    from copy import copy as _copy_style

    from openpyxl.cell.rich_text import CellRichText, TextBlock
    from openpyxl.styles import Alignment, Font

    wb = openpyxl.load_workbook(caminho_tabela, rich_text=True)
    nome_aba = aba or wb.sheetnames[0]
    ws = wb[nome_aba]

    font_item = Font(name="DIN Book", size=9, bold=False, color="FF000000")
    alinhamento_esquerda = Alignment(horizontal="left")

    # (linha a manter como está, itens de legenda, linha do título do
    # bloco — usada como referência de estilo do cabeçalho "LEGENDA").
    # Processa do bloco mais abaixo para o mais acima: assim, inserir
    # linhas num bloco não bagunça os índices de linha dos blocos que
    # ainda faltam processar (todos mais acima na planilha).
    insercoes = []
    for bt, bl in pares:
        if bt["linha_pergunta"]:
            insercoes.append((bt["linha_pergunta"] + 1, bl["itens"], bt["titulo_idx"], bl.get("rotulo_original")))
    insercoes.sort(key=lambda x: x[0], reverse=True)

    for linha_manter, itens, titulo_idx, rotulo_original in insercoes:
        linha_insercao = linha_manter + 1
        n_novas_linhas = len(itens) + 2  # "LEGENDA" + itens + 1 linha em branco no final

        # desfaz temporariamente as mesclagens em/abaixo do ponto de
        # inserção (só essas — as de cima não são afetadas e continuam
        # como estavam)
        faixas_a_remesclar = []
        for rng in list(ws.merged_cells.ranges):
            if rng.min_row >= linha_insercao:
                faixas_a_remesclar.append((rng.min_row, rng.min_col, rng.max_row, rng.max_col))
                ws.unmerge_cells(str(rng))

        # `insert_rows` também não desloca a ALTURA customizada de linha
        # (row_dimensions) — ela fica "grudada" no número antigo da linha
        # em vez de acompanhar o conteúdo que se moveu, o que faz uma
        # linha alta e vazia sobrar bem no meio do relatório, longe de
        # onde a formatação original fazia sentido. Guarda e remove antes
        # de inserir, para recolocar já deslocada depois.
        alturas_a_deslocar = {}
        for idx in list(ws.row_dimensions.keys()):
            if idx >= linha_insercao:
                alturas_a_deslocar[idx] = ws.row_dimensions.pop(idx)

        ws.insert_rows(linha_insercao, amount=n_novas_linhas)

        # refaz as mesclagens desfeitas, já na posição deslocada
        for min_row, min_col, max_row, max_col in faixas_a_remesclar:
            ws.merge_cells(
                start_row=min_row + n_novas_linhas, start_column=min_col,
                end_row=max_row + n_novas_linhas, end_column=max_col
            )

        # recoloca as alturas de linha guardadas, já na posição deslocada
        for idx, dim in alturas_a_deslocar.items():
            nova_idx = idx + n_novas_linhas
            dim.index = nova_idx
            ws.row_dimensions[nova_idx] = dim

        # `insert_rows` não desloca quebras de página manuais sozinho —
        # faz isso à mão para toda quebra que estava em/abaixo do ponto
        # de inserção
        for quebra in ws.row_breaks.brk:
            if quebra.id >= linha_insercao:
                quebra.id += n_novas_linhas

        # preenche o conteúdo nas linhas recém-inseridas (title_idx é
        # sempre anterior ao ponto de inserção deste próprio bloco, então
        # continua válido e sem deslocamento neste momento do laço)
        cel_titulo_ref = ws.cell(row=titulo_idx, column=1)

        r = linha_insercao
        texto_cabecalho = f"LEGENDA {rotulo_original}" if rotulo_original else "LEGENDA"
        cel_legenda = ws.cell(row=r, column=1, value=texto_cabecalho)
        cel_legenda.font = _copy_style(cel_titulo_ref.font)
        cel_legenda.alignment = alinhamento_esquerda
        r += 1

        for item in itens:
            item_normalizado = _forcar_cor_preta(item)
            cel_item = ws.cell(row=r, column=1, value=item_normalizado)
            cel_item.font = font_item
            # Linhas de continuação (sem o prefixo em negrito "Região X
            # (%): ", ou seja, texto simples em vez de rich text) ficam
            # visualmente recuadas, igual ao arquivo de legendas de
            # origem — usa o recuo nativo do Excel (não espaços no
            # início do texto, que ficariam ali se alguém copiar/colar).
            eh_continuacao = not isinstance(item_normalizado, CellRichText)
            cel_item.alignment = Alignment(horizontal="left", indent=1) if eh_continuacao else alinhamento_esquerda
            r += 1
        # a última linha inserida (r) fica em branco de propósito —
        # espaçamento antes do restante do arquivo continuar

    _expandir_area_impressao(ws)

    return wb


def _expandir_area_impressao(ws):
    """
    `insert_rows` não expande a área de impressão sozinho — se o arquivo
    tinha uma área de impressão definida, garante que ela cubra até a
    última linha atual da planilha, pra as novas linhas de legenda não
    ficarem de fora na hora de imprimir/exportar em PDF.
    """
    if not ws.print_area:
        return

    from openpyxl.utils.cell import range_boundaries, get_column_letter

    areas = ws.print_area if isinstance(ws.print_area, list) else [ws.print_area]
    novas_areas = []
    for area in areas:
        ref = area.split("!")[-1]
        try:
            min_col, min_row, max_col, max_row = range_boundaries(ref)
        except ValueError:
            novas_areas.append(area)
            continue
        novo_max_row = max(max_row, ws.max_row)
        novas_areas.append(
            f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{novo_max_row}"
        )
    ws.print_area = novas_areas


# ============================================================================
# Motor v2 de Legendas — casamento por CHAVE individual (código/rótulo),
# não por "tipo" de tabela inteiro.
#
# Diferença pro motor original acima: em vez de colar o bloco de legenda
# inteiro sempre que uma tabela é do tipo "Regiões"/"Capital", este motor
# extrai uma CHAVE de cada item da legenda (um código numérico, ex. "1"
# de "Região 1 (9,73%): ...", ou uma chave textual, ex. "CENTRO E OESTE"
# de "Centro e Oeste (25,24%): ...") e só insere, em cada tabela, os itens
# cuja chave realmente aparece no cabeçalho DAQUELA tabela específica —
# então uma tabela que só usa as regiões 1, 3 e 7 recebe só essas três,
# não a legenda completa.
#
# Baseado no algoritmo de referência "InserirLegendasRegioes.bas".
# ============================================================================

TITULOS_LEGENDA_RECONHECIDOS = {
    "LEGENDA", "LEGENDAS",
    "LEGENDA CAPITAL", "LEGENDAS CAPITAL",
    "LEGENDA REGIOES", "LEGENDAS REGIOES",
}

PREFIXOS_CHAVE_NUMERICA = [
    "MACRORREGIOES", "MACRORREGIAO",
    "MESORREGIOES", "MESORREGIAO",
    "REGIOES", "REGIAO",
    "RPA", "ZONA", "MESO",
]


def _normalizar_texto_vba(valor):
    """
    Normalização usada só pelo motor v2 de Legendas — tira acento,
    maiúsculo, colapsa espaço/quebra de linha/tab/NBSP em um espaço só,
    mas MANTÉM pontuação (":", "(", "[" etc.), diferente de
    `normalizar_texto` (base_multiplas_math), que tira toda pontuação —
    aqui a posição de ":" e "(" importa pra extrair a chave.
    """
    if valor is None:
        return ""
    s = str(valor)
    for ch in ("\r", "\n", "\t", "\xa0"):
        s = s.replace(ch, " ")
    s = "".join(c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c))
    s = s.strip()
    while "  " in s:
        s = s.replace("  ", " ")
    return s.upper()


def _eh_titulo_legenda_v2(texto_normalizado):
    return texto_normalizado in TITULOS_LEGENDA_RECONHECIDOS


def _resto_representa_faixa(resto):
    """'2 a 5', '10-15' etc. — não é um código único, é uma faixa; não vira chave."""
    s = " " + _normalizar_texto_vba(resto) + " "
    if not s.strip():
        return False
    if " E " in s or " A " in s or " ATE " in s:
        return True
    t = s.strip()
    return t.startswith("-") or t.startswith("/")


def _extrair_codigo_numerico_inicial(valor):
    """
    Extrai um código numérico de início de célula — direto se a célula já
    for um número inteiro, ou lendo os dígitos iniciais do texto (com ou
    sem prefixo reconhecido como 'Região'/'RPA'/'Zona'/'Meso' na frente).
    Devolve "" se não achar (ou se o resto do texto sugerir uma FAIXA tipo
    '2 a 5', que não é um código único).
    """
    if valor is None:
        return ""
    if isinstance(valor, bool):
        return ""
    if isinstance(valor, (int, float)):
        if abs(valor - round(valor)) < 1e-7:
            return str(int(round(valor)))
        return ""

    s = _normalizar_texto_vba(valor)
    if not s:
        return ""

    prefixo_reconhecido = False
    for prefixo in PREFIXOS_CHAVE_NUMERICA:
        if s == prefixo:
            return ""
        for sep in (" ", "-", ":"):
            if s.startswith(prefixo + sep):
                s = s[len(prefixo) + 1:].strip()
                prefixo_reconhecido = True
                break
        if prefixo_reconhecido:
            break

    digitos = ""
    for ch in s:
        if ch.isdigit():
            digitos += ch
        else:
            break
    if not digitos:
        return ""

    resto = s[len(digitos):].strip()

    if prefixo_reconhecido:
        if _resto_representa_faixa(resto):
            return ""
        return str(int(digitos))

    # sem prefixo reconhecido: só aceita se o "resto" deixar claro que é
    # mesmo um código isolado (nada depois, ou logo um "(" / "[")
    if not resto:
        return str(int(digitos))
    if resto[0] in "([":
        return str(int(digitos))
    return ""


def _extrair_chave_textual_legenda_v2(valor, eh_continuacao):
    """Pra itens sem código numérico (ex.: 'Centro e Oeste (25,24%): ...')
    — usa o texto antes do primeiro '(' ou ':' como chave."""
    if eh_continuacao or valor is None:
        return ""
    s = _normalizar_texto_vba(valor)
    if not s:
        return ""
    if _eh_titulo_legenda_v2(s):
        return ""
    if s in ("TOTAL", "BASE"):
        return ""
    if s.startswith("PERGUNTA"):
        return ""

    p_par = s.find("(")
    p_dois_pontos = s.find(":")
    if p_dois_pontos < 1:
        return ""

    corte = -1
    if p_par >= 1:
        corte = p_par
    if p_dois_pontos >= 1 and (corte == -1 or p_dois_pontos < corte):
        corte = p_dois_pontos
    if corte > 0:
        s = s[:corte].strip()

    if not s or len(s) > 80:
        return ""
    try:
        float(s.replace(",", "."))
        return ""
    except ValueError:
        pass
    if "%" in s:
        return ""
    return s


def _eh_linha_continuacao_v2(cel):
    """Linha de continuação de um item de legenda multi-linha — detecta
    por recuo do Excel (indent) ou por espaço/tab/NBSP no início do texto."""
    try:
        if cel.alignment and cel.alignment.indent and cel.alignment.indent > 0:
            return True
    except AttributeError:
        pass
    valor = cel.value
    if isinstance(valor, str) and valor:
        return valor[0] in (" ", "\t", "\xa0")
    return False


def _linha_totalmente_vazia_v2(ws, linha, max_col):
    for c in range(1, max_col + 1):
        v = ws.cell(row=linha, column=c).value
        if v is not None and str(v).strip() != "":
            return False
    return True


def _area_vazia_v2(ws, linha_ini, linha_fim, max_col):
    for r in range(linha_ini, linha_fim + 1):
        if not _linha_totalmente_vazia_v2(ws, r, max_col):
            return False
    return True


def parsear_legenda_por_chave(caminho_arquivo, aba=None):
    """
    Lê o arquivo de referência da legenda e monta um mapa CHAVE -> bloco
    (linha de início/fim do item + linha do título ao qual pertence),
    junto com um snapshot dos valores e alturas de linha (pra poder
    copiar depois sem precisar manter o arquivo de referência aberto).

    Reconhece vários blocos de título na mesma aba (ex.: um bloco
    "LEGENDA REGIOES" e outro "LEGENDA CAPITAL" empilhados) — cada item
    fica associado ao título mais próximo acima dele.

    Returns:
        dict com "mapa" (chave -> {linha_inicio, linha_fim, linha_titulo}),
        "linhas_titulo" (lista, em ordem), "linhas_valores" (linha ->
        lista de valores), "alturas" (linha -> altura ou None),
        "max_col", e "titulos_aceitos" (derivado automaticamente dos
        próprios títulos encontrados — ex.: "LEGENDA REGIOES" vira
        "REGIOES" — usado depois pra ajudar a achar o cabeçalho certo
        em cada tabela do relatório).
    """
    wb = openpyxl.load_workbook(caminho_arquivo, rich_text=True)
    ws = wb[aba] if aba else wb[wb.sheetnames[0]]
    max_row = ws.max_row
    max_col = ws.max_column

    linhas_valores = {
        r: [ws.cell(row=r, column=c).value for c in range(1, max_col + 1)]
        for r in range(1, max_row + 1)
    }
    alturas = {}
    for r in range(1, max_row + 1):
        dim = ws.row_dimensions.get(r)
        alturas[r] = dim.height if (dim and dim.height) else None

    # linhas de título (varre todas as colunas, igual ao original)
    linhas_titulo = []
    titulos_aceitos = set()
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            norm = _normalizar_texto_vba(ws.cell(row=r, column=c).value)
            if _eh_titulo_legenda_v2(norm):
                linhas_titulo.append(r)
                for prefixo in ("LEGENDAS ", "LEGENDA "):
                    if norm.startswith(prefixo):
                        resto = norm[len(prefixo):].strip()
                        if resto:
                            titulos_aceitos.add(resto)
                        break
                break

    linhas_titulo_set = set(linhas_titulo)

    titulo_anterior = {}
    atual = 0
    for r in range(1, max_row + 1):
        titulo_anterior[r] = atual
        if r in linhas_titulo_set:
            atual = r

    proximo_titulo = {}
    atual = 0
    for r in range(max_row, 0, -1):
        proximo_titulo[r] = atual
        if r in linhas_titulo_set:
            atual = r

    # linhas de item (têm uma chave extraível) — só coluna A, igual ao original
    linhas_item = []
    for r in range(1, max_row + 1):
        if r in linhas_titulo_set:
            continue
        cel = ws.cell(row=r, column=1)
        continuacao = _eh_linha_continuacao_v2(cel)

        chave = ""
        for c in range(1, max_col + 1):
            codigo = _extrair_codigo_numerico_inicial(ws.cell(row=r, column=c).value)
            if codigo:
                chave = codigo
                break
        if not chave and not continuacao:
            chave = _extrair_chave_textual_legenda_v2(cel.value, continuacao)

        if chave:
            linhas_item.append((r, chave))

    quebras_legenda = sorted(b.id for b in ws.row_breaks.brk if b.id is not None)

    def _fim_pagina(linha_inicio):
        fim = max_row
        for quebra_id in quebras_legenda:
            inicio_pagina = quebra_id + 1
            if inicio_pagina > linha_inicio:
                fim = min(fim, inicio_pagina - 1)
                break
        return fim

    mapa = {}
    for idx, (linha_inicio, chave) in enumerate(linhas_item):
        linha_fim = max_row
        if idx + 1 < len(linhas_item):
            linha_fim = min(linha_fim, linhas_item[idx + 1][0] - 1)
        pt = proximo_titulo.get(linha_inicio, 0)
        if pt:
            linha_fim = min(linha_fim, pt - 1)
        linha_fim = min(linha_fim, _fim_pagina(linha_inicio))
        if linha_fim < linha_inicio:
            linha_fim = linha_inicio

        while linha_fim > linha_inicio and _linha_totalmente_vazia_v2(ws, linha_fim, max_col):
            linha_fim -= 1

        linha_titulo = titulo_anterior.get(linha_inicio, 0) or (linhas_titulo[0] if linhas_titulo else 0)

        if chave not in mapa:  # chaves repetidas: preserva a primeira ocorrência
            mapa[chave] = {
                "linha_inicio": linha_inicio,
                "linha_fim": linha_fim,
                "linha_titulo": linha_titulo,
            }

    return {
        "mapa": mapa,
        "linhas_titulo": linhas_titulo,
        "linhas_valores": linhas_valores,
        "alturas": alturas,
        "max_col": max_col,
        "titulos_aceitos": titulos_aceitos,
    }


def _normalizar_fonte_corpo_legenda(item):
    """
    Força fonte DIN Book tamanho 9 preta em TODOS os trechos de um item
    de legenda (rich text) — inclusive no prefixo em negrito antes do
    ":" — preservando negrito/itálico de cada trecho, só uniformizando
    família/tamanho/cor. Sem isso, o prefixo em negrito ficava com a
    fonte que o arquivo de referência original usava (nem sempre DIN
    Book), inconsistente com o resto do relatório.
    """
    from copy import copy as _copy_style

    from openpyxl.cell.rich_text import CellRichText, TextBlock
    from openpyxl.styles.colors import Color

    if not isinstance(item, CellRichText):
        return item
    novos_blocos = []
    for bloco in item:
        if isinstance(bloco, TextBlock):
            nova_fonte = _copy_style(bloco.font) if bloco.font else None
            if nova_fonte is not None:
                nova_fonte.rFont = "DIN Book"
                nova_fonte.sz = 9
                nova_fonte.color = Color(rgb="FF000000")
            novos_blocos.append(TextBlock(nova_fonte, bloco.text))
        else:
            novos_blocos.append(bloco)
    return CellRichText(novos_blocos)


def _normalizar_fonte_titulo_legenda(item):
    """Igual a `_normalizar_fonte_corpo_legenda`, mas pro título (DIN 10,
    sempre negrito) — cobre o caso do título vir como rich text (com
    trechos sem fonte própria explícita), que senão caía no Calibri
    padrão do Excel em vez de herdar a fonte da célula."""
    from copy import copy as _copy_style

    from openpyxl.cell.rich_text import CellRichText, TextBlock
    from openpyxl.styles.colors import Color

    if not isinstance(item, CellRichText):
        return item
    novos_blocos = []
    for bloco in item:
        if isinstance(bloco, TextBlock):
            nova_fonte = _copy_style(bloco.font) if bloco.font else None
            if nova_fonte is not None:
                nova_fonte.rFont = "DIN"
                nova_fonte.sz = 10
                nova_fonte.b = True
                nova_fonte.color = Color(rgb="FF000000")
            novos_blocos.append(TextBlock(nova_fonte, bloco.text))
        else:
            novos_blocos.append(bloco)
    return CellRichText(novos_blocos)


def _copiar_linha_legenda(ws_destino, linha_destino, dados_referencia, linha_ref, max_col_destino, tipo="item"):
    """
    Copia uma linha da referência pro relatório, já reaplicando a
    formatação que se perde numa cópia de valor puro:
      - linha de título (tipo="titulo", ex.: "LEGENDA" ou "LEGENDA
        REGIÕES") fica em DIN tamanho 10, negrito;
      - corpo da legenda (tipo="item") fica em DIN Book tamanho 9,
        preservando negrito nos trechos até ":" e cor forçada pra preto
        (evita o problema de cor de tema divergindo entre o arquivo de
        legenda e o relatório);
      - linhas de continuação (texto simples, sem prefixo em negrito)
        ganham o mesmo DIN Book 9 e recuo.

    A fonte da CÉLULA (não só dos trechos de rich text) é sempre
    definida também — sem isso, qualquer trecho de um item de legenda
    sem formatação própria explícita cai no Calibri padrão do Excel em
    vez de herdar DIN/DIN Book.
    """
    from openpyxl.cell.rich_text import CellRichText
    from openpyxl.styles import Alignment, Font

    valores = dados_referencia["linhas_valores"].get(linha_ref, [])
    for c in range(1, max_col_destino + 1):
        valor = valores[c - 1] if c - 1 < len(valores) else None

        if tipo == "titulo":
            valor_normalizado = _normalizar_fonte_titulo_legenda(valor)
            cel = ws_destino.cell(row=linha_destino, column=c, value=valor_normalizado)
            if c == 1 and valor is not None:
                cel.font = Font(name="DIN", size=10, bold=True, color="FF000000")
            continue

        valor_normalizado = _normalizar_fonte_corpo_legenda(_forcar_cor_preta(valor))
        cel = ws_destino.cell(row=linha_destino, column=c, value=valor_normalizado)
        if c == 1 and valor is not None:
            # fonte de base da célula — cobre tanto o texto simples de
            # continuação quanto qualquer trecho "solto" sem formatação
            # própria dentro de um item em rich text
            cel.font = Font(name="DIN Book", size=9, bold=False, color="FF000000")
            if not isinstance(valor_normalizado, CellRichText):
                cel.alignment = Alignment(horizontal="left", indent=1)

    # Altura fixa 13 em toda linha de legenda colada (título e item) — a
    # pedido do Lucas, substitui a altura copiada do arquivo de
    # referência (que variava conforme como a legenda estava formatada
    # lá, sem relação com a altura desejada aqui no relatório).
    ws_destino.row_dimensions[linha_destino].height = 13


def _colar_legenda_filtrada(ws, dados_referencia, chaves, linha_destino):
    mapa = dados_referencia["mapa"]
    max_col = dados_referencia["max_col"]

    linha_titulo_ref = None
    for chave in chaves:
        if chave in mapa and mapa[chave]["linha_titulo"]:
            linha_titulo_ref = mapa[chave]["linha_titulo"]
            break
    if not linha_titulo_ref and dados_referencia["linhas_titulo"]:
        linha_titulo_ref = dados_referencia["linhas_titulo"][0]
    if not linha_titulo_ref:
        return

    _copiar_linha_legenda(ws, linha_destino, dados_referencia, linha_titulo_ref, max_col, tipo="titulo")

    # a legenda em si começa IMEDIATAMENTE na linha seguinte ao título —
    # sem linha em branco de separação entre as duas
    linha_atual = linha_destino + 1
    for chave in chaves:
        if chave not in mapa:
            continue
        ini, fim = mapa[chave]["linha_inicio"], mapa[chave]["linha_fim"]
        for offset, linha_ref in enumerate(range(ini, fim + 1)):
            _copiar_linha_legenda(ws, linha_atual + offset, dados_referencia, linha_ref, max_col, tipo="item")
        linha_atual += (fim - ini + 1)


def _ja_tem_legenda_aqui(ws, linha_destino):
    """Confere se já existe um título de legenda reconhecido bem nessa
    posição — sinal de que uma execução anterior já inseriu a legenda
    ali, então não insere de novo (evita duplicar ao rodar 2x)."""
    norm = _normalizar_texto_vba(ws.cell(row=linha_destino, column=1).value)
    return _eh_titulo_legenda_v2(norm)


def _garantir_duas_linhas_apos_legenda(ws, linha_fim_legenda):
    """
    Garante EXATAMENTE 2 linhas em branco entre a última linha da
    legenda recém-colada e a quebra de página que já existia logo
    depois dela (deixada pelo código 07) — é o padrão do relatório.

    Sem isso, sobra só 1 linha em branco ali: o código 07 reserva 4
    linhas em branco depois da Pergunta (quebra de página no meio,
    depois da 2ª), mas o ponto onde a legenda é colada
    (`linha_pergunta + 2`) cai bem em cima da 2ª dessas 4 linhas — a
    legenda "engole" uma das duas linhas que ficariam antes da quebra,
    sobrando só a outra.

    Não mexe em nada se não achar nenhuma quebra de página nas
    proximidades (não força a criação de uma quebra nova, só corrige o
    espaçamento em torno de uma que já existe).

    Insere ou remove linhas ANTES da linha que segura a quebra (nunca a
    própria linha da quebra) — removê-la faria a quebra desaparecer
    (`remover_linhas_seguro` descarta silenciosamente uma quebra cuja
    linha é removida; é assim que a função de segurança do projeto
    funciona, então o cuidado tem que vir de quem chama).

    Idempotente: rodar de novo não fica ajustando à toa se já estiver
    certo. Sempre "materializa" as 2 linhas-alvo aplicando um estilo
    (mesmo sem texto) — uma célula totalmente vazia, sem nenhum estilo,
    some ao salvar no openpyxl.
    """
    from openpyxl.styles import Font

    candidatos = [
        b.id for b in ws.row_breaks.brk
        if b.id is not None and linha_fim_legenda <= b.id <= linha_fim_legenda + 20
    ]
    if not candidatos:
        return
    quebra_id_atual = min(candidatos)
    quebra_id_alvo = linha_fim_legenda + 2

    if quebra_id_atual < quebra_id_alvo:
        inserir_linhas_seguro(ws, linha_fim_legenda + 1, quebra_id_alvo - quebra_id_atual)
    elif quebra_id_atual > quebra_id_alvo:
        remover_linhas_seguro(ws, linha_fim_legenda + 1, quebra_id_atual - quebra_id_alvo)

    for rr in range(linha_fim_legenda + 1, linha_fim_legenda + 3):
        ws.row_dimensions[rr].height = 15
        ws.cell(row=rr, column=1).font = Font(name="DIN Book", size=9)


def aplicar_legendas_por_chave(ws, dados_referencia, titulos_aceitos=None):
    """
    Para cada linha 'Pergunta:' do relatório, procura — de baixo pra
    cima, dentro dos limites da mesma página impressa (usa as quebras de
    página já existentes, inseridas pelo código 07) — um cabeçalho de
    tabela cujas células tenham códigos/chaves batendo com itens da
    legenda de referência. Insere a legenda logo abaixo da Pergunta
    (com uma linha em branco de separação), com o título correspondente
    e SÓ os itens cujas chaves realmente aparecem naquele cabeçalho
    específico — uma tabela que só usa as regiões 1, 3 e 7 recebe só
    essas três, não a lista completa.

    Idempotente: confere se o espaço já está ocupado antes de inserir,
    então rodar de novo não duplica.

    Args:
        ws: planilha do relatório.
        dados_referencia: dict devolvido por `parsear_legenda_por_chave`.
        titulos_aceitos: conjunto de rótulos de título aceitos (ex.:
            {"REGIOES", "CAPITAL"}) usado como uma das regras de
            confiança pra achar o cabeçalho certo. Se None, usa o que
            foi derivado automaticamente da própria referência.

    Returns:
        Número de tabelas que receberam legenda.
    """
    mapa = dados_referencia["mapa"]
    if not mapa:
        return 0

    max_col = ws.max_column
    max_row = ws.max_row

    titulos_aceitos_norm = titulos_aceitos or dados_referencia.get("titulos_aceitos") or set()

    linhas_pergunta = [
        r for r in range(1, max_row + 1)
        if _normalizar_texto_vba(ws.cell(row=r, column=1).value).startswith("PERGUNTA")
    ]
    if not linhas_pergunta:
        return 0

    quebras_relatorio = sorted(b.id for b in ws.row_breaks.brk if b.id is not None)

    def _primeira_linha_pagina(linha_ref):
        melhor = 1
        for quebra_id in quebras_relatorio:
            inicio_pagina = quebra_id + 1
            if inicio_pagina <= linha_ref:
                melhor = inicio_pagina
            else:
                break
        return melhor

    tem_total = {}
    tem_titulo_aceito = {}
    for r in range(1, max_row + 1):
        tt, tta = False, False
        for c in range(1, max_col + 1):
            norm = _normalizar_texto_vba(ws.cell(row=r, column=c).value)
            if norm == "TOTAL":
                tt = True
            if norm and any(titulo in norm for titulo in titulos_aceitos_norm):
                tta = True
        tem_total[r] = tt
        tem_titulo_aceito[r] = tta

    def _existe_true(flags, ini, fim):
        ini, fim = max(1, ini), min(max_row, fim)
        return any(flags.get(r, False) for r in range(ini, fim + 1))

    def _chaves_na_linha(r):
        encontradas = []
        for c in range(1, max_col + 1):
            valor = ws.cell(row=r, column=c).value
            codigo = _extrair_codigo_numerico_inicial(valor)
            chave = None
            if codigo and codigo in mapa:
                chave = codigo
            else:
                norm = _normalizar_texto_vba(valor)
                if norm and norm in mapa:
                    chave = norm
            if chave and chave not in encontradas:
                encontradas.append(chave)
        return encontradas

    # FASE 1: planeja de baixo pra cima (números de linha continuam válidos)
    planos = []
    for i in range(len(linhas_pergunta) - 1, -1, -1):
        linha_pergunta = linhas_pergunta[i]
        linha_pergunta_anterior = linhas_pergunta[i - 1] if i > 0 else 0

        primeira_linha_pagina = _primeira_linha_pagina(linha_pergunta)
        if linha_pergunta_anterior > 0:
            primeira_linha_pagina = max(primeira_linha_pagina, linha_pergunta_anterior + 1)
        primeira_linha_pagina = max(1, primeira_linha_pagina)

        linha_cabecalho, chaves_melhor = None, None
        for r in range(linha_pergunta - 1, primeira_linha_pagina - 1, -1):
            chaves = _chaves_na_linha(r)
            qtd = len(chaves)
            if qtd > 0:
                achou_titulo = _existe_true(tem_titulo_aceito, max(primeira_linha_pagina, r - 3), r)
                achou_total = _existe_true(tem_total, max(primeira_linha_pagina, r - 2), r)
                if (achou_titulo and qtd >= 1) or (achou_total and qtd >= 2) or (qtd >= 3):
                    linha_cabecalho, chaves_melhor = r, chaves
                    break

        if linha_cabecalho:
            qtd_linhas = sum(
                mapa[k]["linha_fim"] - mapa[k]["linha_inicio"] + 1 for k in chaves_melhor if k in mapa
            ) + 1  # +1 = só a linha do título; a legenda começa logo em seguida, sem linha em branco
            if qtd_linhas > 1:
                planos.append((linha_pergunta, chaves_melhor, qtd_linhas))

    if not planos:
        return 0

    # FASE 2: aplica (já em ordem de baixo pra cima)
    inseridas = 0
    for linha_pergunta, chaves, qtd_linhas in planos:
        linha_destino_prevista = linha_pergunta + 2
        if _ja_tem_legenda_aqui(ws, linha_destino_prevista):
            # execução anterior já inseriu a legenda aqui — não duplica o
            # conteúdo, mas AINDA ASSIM confere o espaçamento antes da
            # quebra: uma legenda inserida por uma versão anterior do
            # código (ou por qualquer outro motivo) pode ter ficado com
            # esse espaçamento errado, e rodar de novo é a chance de
            # autocorrigir isso sem duplicar nada.
            fim_legenda_existente = linha_destino_prevista
            r = linha_destino_prevista + 1
            while r <= max_row and not _linha_totalmente_vazia_v2(ws, r, max_col):
                fim_legenda_existente = r
                r += 1
            _garantir_duas_linhas_apos_legenda(ws, fim_legenda_existente)
            continue

        linha_espaco = linha_pergunta + 1
        if not _area_vazia_v2(ws, linha_espaco, linha_espaco, max_col):
            inserir_linhas_seguro(ws, linha_espaco, 1)

        linha_destino = linha_pergunta + 2
        inserir_linhas_seguro(ws, linha_destino, qtd_linhas)

        _colar_legenda_filtrada(ws, dados_referencia, chaves, linha_destino)
        _garantir_duas_linhas_apos_legenda(ws, linha_destino + qtd_linhas - 1)
        inseridas += 1

    return inseridas
