"""
Utilidades genéricas para manipular planilhas .xlsx, reaproveitadas por
vários dos códigos portados do "Relatório Automatizado".

Sem dependência do Streamlit — pode ser testado isoladamente.
"""
import math

import pandas as pd
from openpyxl.utils import get_column_letter


def inserir_linhas_seguro(ws, linha_insercao, quantidade):
    """
    Insere 'quantidade' linhas em branco em 'linha_insercao', deslocando
    corretamente mesclagens, alturas de linha customizadas e quebras de
    página manuais que estejam em/abaixo do ponto de inserção — coisas
    que o `ws.insert_rows()` sozinho não desloca (comportamento já
    mapeado e corrigido no módulo de Legendas).
    """
    if quantidade <= 0:
        return

    faixas_a_remesclar = []
    for rng in list(ws.merged_cells.ranges):
        if rng.min_row >= linha_insercao:
            faixas_a_remesclar.append((rng.min_row, rng.min_col, rng.max_row, rng.max_col))
            try:
                ws.unmerge_cells(str(rng))
            except KeyError:
                # openpyxl pode falhar tentando apagar células "filhas" da
                # mesclagem que nunca foram individualmente instanciadas
                # (acontece com mesclagens já existentes de uma execução
                # anterior, ex.: título de uma capa criada pelo código 15).
                # Remove a mesclagem direto do conjunto interno nesse caso,
                # sem passar pelo método que quebra.
                ws.merged_cells.ranges.discard(rng)

    alturas_a_deslocar = {}
    for idx in list(ws.row_dimensions.keys()):
        if idx >= linha_insercao:
            alturas_a_deslocar[idx] = ws.row_dimensions.pop(idx)

    ws.insert_rows(linha_insercao, amount=quantidade)

    for min_row, min_col, max_row, max_col in faixas_a_remesclar:
        ws.merge_cells(
            start_row=min_row + quantidade, start_column=min_col,
            end_row=max_row + quantidade, end_column=max_col
        )

    for idx, dim in alturas_a_deslocar.items():
        nova_idx = idx + quantidade
        dim.index = nova_idx
        ws.row_dimensions[nova_idx] = dim

    for quebra in ws.row_breaks.brk:
        if quebra.id >= linha_insercao:
            quebra.id += quantidade


def remover_linhas_seguro(ws, linha_remocao, quantidade):
    """
    Remove 'quantidade' linhas a partir de 'linha_remocao' (inclusive),
    deslocando corretamente mesclagens, alturas de linha customizadas e
    quebras de página manuais que estejam abaixo da faixa removida —
    espelha `inserir_linhas_seguro`, na direção contrária. Só deve ser
    usada em linhas que já se sabe estarem vazias (o chamador é
    responsável por confirmar isso antes) — não existe checagem de
    conteúdo aqui.
    """
    if quantidade <= 0:
        return

    linha_fim_removida = linha_remocao + quantidade - 1

    # Desfaz mesclagens que tocam a faixa removida (não deveria acontecer
    # se o chamador só remove linhas vazias, mas por segurança) e as que
    # estão abaixo dela (serão remescladas na posição nova)
    faixas_a_remesclar = []
    for rng in list(ws.merged_cells.ranges):
        if rng.min_row > linha_fim_removida:
            faixas_a_remesclar.append((rng.min_row, rng.min_col, rng.max_row, rng.max_col))
            try:
                ws.unmerge_cells(str(rng))
            except KeyError:
                ws.merged_cells.ranges.discard(rng)

    alturas_a_deslocar = {}
    for idx in list(ws.row_dimensions.keys()):
        if idx >= linha_remocao:
            dim = ws.row_dimensions.pop(idx)
            if idx > linha_fim_removida:
                alturas_a_deslocar[idx] = dim
            # alturas dentro da própria faixa removida são descartadas

    ws.delete_rows(linha_remocao, amount=quantidade)

    for min_row, min_col, max_row, max_col in faixas_a_remesclar:
        ws.merge_cells(
            start_row=min_row - quantidade, start_column=min_col,
            end_row=max_row - quantidade, end_column=max_col
        )

    for idx, dim in alturas_a_deslocar.items():
        nova_idx = idx - quantidade
        dim.index = nova_idx
        ws.row_dimensions[nova_idx] = dim

    quebras_restantes = []
    for quebra in ws.row_breaks.brk:
        if quebra.id >= linha_remocao and quebra.id <= linha_fim_removida:
            continue  # quebra estava dentro da faixa removida, descarta
        if quebra.id > linha_fim_removida:
            quebra.id -= quantidade
        quebras_restantes.append(quebra)
    ws.row_breaks.brk = quebras_restantes


def estimar_altura_calculada(texto, largura_total_colunas, tamanho_fonte):
    """
    Estimativa (aproximada) da altura que uma célula com quebra de linha
    precisaria para caber todo o texto, dada a largura total das colunas
    envolvidas e o tamanho da fonte.

    IMPORTANTE: isto é uma APROXIMAÇÃO por contagem de caracteres — não é
    o cálculo real do Excel (que depende da métrica exata da fonte, algo
    que só o próprio Excel consegue calcular ao renderizar). Serve como
    um bom ponto de partida; ajustes finos de altura continuam sendo
    melhor feitos no Excel (duplo clique na borda da linha).

    Calibrado com margem de segurança pra fontes largas como a "DIN" (a
    fonte usada nos rótulos do relatório) — a estimativa original (fator
    de caractere 0.52, sem margem) subestimava o número de linhas em
    textos como "Não aprova nenhuma área" numa coluna de 20.67 de
    largura: previa 1 linha quando na real precisava de 2, cortando o
    texto visualmente. O fator de caractere maior (0.62) e a margem de
    10% a menos na largura útil (0.90) tornam a estimativa mais
    conservadora — prefere estourar um pouco a altura a arriscar cortar
    texto, o que é bem pior visualmente.
    """
    if not texto:
        return tamanho_fonte * 1.6

    largura_pt = (largura_total_colunas + 0.62) * 7 * 0.90
    largura_char = tamanho_fonte * 0.62
    caracteres_por_linha = max(1, int(largura_pt / largura_char))
    n_linhas = max(1, math.ceil(len(str(texto)) / caracteres_por_linha))
    altura_por_linha = tamanho_fonte * 1.6
    return n_linhas * altura_por_linha


def largura_total_mesclagem(ws, celula):
    """Largura total (em unidades de coluna do Excel) da célula ou da sua faixa mesclada."""
    faixa = None
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= celula.row <= rng.max_row and rng.min_col <= celula.column <= rng.max_col:
            faixa = rng
            break

    if faixa is None:
        colunas = [celula.column]
    else:
        colunas = range(faixa.min_col, faixa.max_col + 1)

    total = 0.0
    largura_padrao = ws.sheet_format.defaultColWidth or 8.43
    for c in colunas:
        letra = get_column_letter(c)
        dim = ws.column_dimensions.get(letra)
        total += dim.width if (dim and dim.width) else largura_padrao
    return total


def worksheet_para_dataframe(ws, max_linhas=300, max_colunas=40):
    """
    Converte uma planilha num pandas.DataFrame só de VALORES (sem
    formatação — fonte, cor, borda etc. não têm como aparecer numa
    pré-visualização em tabela HTML/Streamlit), pra uso em pré-
    visualização dentro do app antes do download.

    Células dentro de uma faixa mesclada (fora a célula-âncora, que já
    tem o valor) recebem o mesmo valor da âncora — sem isso, a pré-
    visualização mostraria um monte de células vazias onde na verdade
    o Excel mostra um bloco mesclado com texto, o que confundiria mais
    do que ajudaria a pessoa a validar o resultado.

    Trunca em `max_linhas`/`max_colunas` por performance — planilhas de
    relatório real passam de 3000 linhas, e renderizar tudo isso como
    HTML deixaria a pré-visualização lenta sem ganho real (a pessoa
    consegue ver o começo e decidir se baixa o arquivo completo).

    Returns:
        (df, truncado_linhas, truncado_colunas) — o DataFrame já com as
        mesclagens preenchidas, e dois booleanos indicando se cada eixo
        foi cortado (pra o chamador avisar a pessoa, se for o caso).
    """
    total_linhas = ws.max_row
    total_colunas = ws.max_column

    limite_linhas = min(total_linhas, max_linhas)
    limite_colunas = min(total_colunas, max_colunas)

    linhas = []
    for r in range(1, limite_linhas + 1):
        linha = [ws.cell(row=r, column=c).value for c in range(1, limite_colunas + 1)]
        linhas.append(linha)

    colunas_nomes = [get_column_letter(c) for c in range(1, limite_colunas + 1)]
    df = pd.DataFrame(linhas, columns=colunas_nomes)

    # Preenche as mesclagens que caem dentro da área exibida
    for rng in ws.merged_cells.ranges:
        if rng.min_row > limite_linhas or rng.min_col > limite_colunas:
            continue
        valor_ancora = ws.cell(row=rng.min_row, column=rng.min_col).value
        if valor_ancora is None:
            continue
        max_r = min(rng.max_row, limite_linhas)
        max_c = min(rng.max_col, limite_colunas)
        for r in range(rng.min_row, max_r + 1):
            for c in range(rng.min_col, max_c + 1):
                df.iat[r - 1, c - 1] = valor_ancora

    truncado_linhas = total_linhas > max_linhas
    truncado_colunas = total_colunas > max_colunas
    return df, truncado_linhas, truncado_colunas


# ---------------------------------------------------------------------------
# Pré-visualização "de verdade" (HTML que imita o Excel: mesclagem, cor de
# fundo, negrito/itálico, bordas, alinhamento) — worksheet_para_dataframe()
# acima vira uma tabela genérica sem nada disso, o que fica muito diferente
# do arquivo real. Esta função gera HTML pra renderizar num componente
# (st.components.v1.html), não um DataFrame.
# ---------------------------------------------------------------------------
_MAPA_BORDA_PX = {
    "thin": 1, "hair": 1, "dotted": 1, "dashed": 1,
    "medium": 2, "mediumDashed": 2, "slantDashDot": 2,
    "thick": 3, "double": 3,
}


def _cor_hex(cor_openpyxl, padrao=None):
    """Extrai um #RRGGBB de um objeto Color do openpyxl, se der pra saber
    a cor real (RGB explícito) — cores indexadas/de tema não têm mapeamento
    simples fora do Excel, então caem no padrão."""
    if cor_openpyxl is None:
        return padrao
    try:
        if cor_openpyxl.type == "rgb" and cor_openpyxl.rgb and len(cor_openpyxl.rgb) == 8:
            rgb = cor_openpyxl.rgb[2:]  # descarta o alpha (AARRGGBB -> RRGGBB)
            if rgb != "000000" or padrao is None:
                return f"#{rgb}"
    except AttributeError:
        pass
    return padrao


def _estilo_borda_css(borda):
    """Monta o CSS de borda (um lado por vez) a partir de um Border do openpyxl."""
    partes = {}
    for lado_nome, lado in (
        ("top", borda.top), ("bottom", borda.bottom),
        ("left", borda.left), ("right", borda.right),
    ):
        if lado is not None and lado.style:
            espessura = _MAPA_BORDA_PX.get(lado.style, 1)
            cor = _cor_hex(lado.color, "#000000")
            estilo = "double" if lado.style == "double" else "solid"
            partes[lado_nome] = f"border-{lado_nome}: {espessura}px {estilo} {cor};"
    return "".join(partes.values())


def _estilo_celula_css(cel):
    """Monta o CSS inline de uma célula (fonte, preenchimento, alinhamento, borda)."""
    partes = []

    fonte = cel.font
    if fonte:
        if fonte.bold:
            partes.append("font-weight:bold;")
        if fonte.italic:
            partes.append("font-style:italic;")
        if fonte.size:
            partes.append(f"font-size:{fonte.size * 1.05:.0f}px;")
        cor_fonte = _cor_hex(fonte.color)
        if cor_fonte:
            partes.append(f"color:{cor_fonte};")

    fill = cel.fill
    if fill and fill.patternType == "solid":
        cor_fundo = _cor_hex(fill.fgColor)
        if cor_fundo:
            partes.append(f"background-color:{cor_fundo};")

    alinh = cel.alignment
    if alinh:
        mapa_h = {"left": "left", "center": "center", "right": "right", "general": None}
        h = mapa_h.get(alinh.horizontal) if alinh.horizontal else None
        if h:
            partes.append(f"text-align:{h};")
        mapa_v = {"top": "top", "center": "middle", "bottom": "bottom"}
        v = mapa_v.get(alinh.vertical) if alinh.vertical else None
        if v:
            partes.append(f"vertical-align:{v};")
        if alinh.wrap_text:
            partes.append("white-space:normal;word-break:break-word;")
        else:
            partes.append("white-space:nowrap;")

    partes.append(_estilo_borda_css(cel.border))
    return "".join(partes)


def _formatar_valor_preview(valor):
    """
    Formata o valor de uma célula pra exibição na pré-visualização: número
    com casa decimal (float) mostra só 1 casa — a planilha real guarda a
    precisão completa, isso é só pra não poluir a leitura na tela (ex.:
    "38.58568887059666" vira "38.6"). Inteiros e texto ficam como estão.
    """
    if valor is None:
        return ""
    if isinstance(valor, bool):
        return str(valor)
    if isinstance(valor, float):
        return f"{valor:.1f}"
    return str(valor)


def worksheet_para_html_paginado(ws):
    """
    Como `worksheet_para_html`, mas sem truncar linhas/colunas (o
    relatório inteiro) e dividido em UMA TABELA HTML POR PÁGINA (nas
    mesmas linhas onde a planilha tem uma quebra de página manual),
    cada uma dentro de uma <div> com quebra de página — pensado pra
    virar um PDF paginado de verdade (ver core/pdf_preview.py).

    Importante: a quebra de página é aplicada na <div> que envolve cada
    tabela, NÃO num <tr> dentro de uma tabela única — "page-break-before"
    em linha de tabela não é respeitado de forma confiável por praticamente
    nenhum motor de renderização HTML->PDF (é uma limitação conhecida de
    CSS, não só do wkhtmltopdf), por isso cada página vira sua própria
    tabela HTML independente.

    Returns:
        str de HTML completo (com <html>/<head>/<body>, já pronto pra
        passar direto pro conversor de PDF).
    """
    total_linhas = ws.max_row
    total_colunas = ws.max_column

    celulas_cobertas = set()
    spans = {}
    for rng in ws.merged_cells.ranges:
        spans[(rng.min_row, rng.min_col)] = (
            rng.max_row - rng.min_row + 1, rng.max_col - rng.min_col + 1
        )
        for r in range(rng.min_row, rng.max_row + 1):
            for c in range(rng.min_col, rng.max_col + 1):
                if (r, c) != (rng.min_row, rng.min_col):
                    celulas_cobertas.add((r, c))

    larguras_px = []
    largura_padrao = ws.sheet_format.defaultColWidth or 8.43
    for c in range(1, total_colunas + 1):
        letra = get_column_letter(c)
        dim = ws.column_dimensions.get(letra)
        largura = dim.width if (dim and dim.width) else largura_padrao
        larguras_px.append(max(20, round(largura * 6.2)))
    colgroup = "".join(f'<col style="width:{w}px;">' for w in larguras_px)

    # pontos onde uma nova página começa (linha 1-based) — a partir das
    # quebras de página manuais já presentes na planilha
    inicios_pagina = sorted({b.id + 1 for b in ws.row_breaks.brk if b.id is not None})
    limites = [1] + [i for i in inicios_pagina if i > 1] + [total_linhas + 1]
    limites = sorted(set(limites))

    def _renderizar_linha(r):
        celulas_html = []
        for c in range(1, total_colunas + 1):
            if (r, c) in celulas_cobertas:
                continue
            cel = ws.cell(row=r, column=c)
            texto = _formatar_valor_preview(cel.value)
            texto = (
                texto.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
                .replace("\n", "<br>")
            )
            attrs = ""
            span_r, span_c = spans.get((r, c), (1, 1))
            if span_r > 1:
                attrs += f' rowspan="{span_r}"'
            if span_c > 1:
                attrs += f' colspan="{span_c}"'
            estilo = _estilo_celula_css(cel)
            celulas_html.append(f'<td style="{estilo}"{attrs}>{texto}</td>')
        dim_linha = ws.row_dimensions.get(r)
        altura_px = round((dim_linha.height or 15) * 1.1) if dim_linha else 17
        return f'<tr style="height:{altura_px}px;">{"".join(celulas_html)}</tr>'

    paginas_html = []
    for i in range(len(limites) - 1):
        inicio_pag, fim_pag = limites[i], limites[i + 1] - 1
        linhas_da_pagina = [
            _renderizar_linha(r) for r in range(inicio_pag, fim_pag + 1)
            if any((r, c) not in celulas_cobertas for c in range(1, total_colunas + 1))
            or True  # inclui mesmo linhas "vazias" (mantém espaçamento do layout original)
        ]
        if not any(ws.cell(row=r, column=c).value not in (None, "") for r in range(inicio_pag, fim_pag + 1) for c in range(1, total_colunas + 1)):
            # página inteiramente vazia (ex.: páginas em branco antes de uma capa) —
            # ainda assim gera uma <div> própria, só que sem tabela dentro,
            # pra manter a contagem/posição de páginas do PDF fiel ao original
            paginas_html.append('<div class="pagina"></div>')
            continue
        tabela = (
            f'<table><colgroup>{colgroup}</colgroup>'
            f'<tbody>{"".join(linhas_da_pagina)}</tbody></table>'
        )
        paginas_html.append(f'<div class="pagina">{tabela}</div>')

    orientacao = (ws.page_setup.orientation or "portrait") if ws.page_setup else "portrait"

    return f"""<!DOCTYPE html>
<html>
<head>
<meta charset="utf-8">
<style>
  @page {{ size: A4 {orientacao}; margin: 10mm; }}
  body {{ margin: 0; }}
  .pagina {{ page-break-after: always; }}
  .pagina:last-child {{ page-break-after: avoid; }}
  table {{ border-collapse: collapse; font-family: Calibri, Arial, sans-serif;
           font-size: 10px; table-layout: fixed; width: 100%; }}
  td {{ overflow: hidden; }}
</style>
</head>
<body>
  {"".join(paginas_html)}
</body>
</html>"""


def worksheet_para_html(ws, max_linhas=150, max_colunas=30):
    """
    Gera uma tabela HTML que imita a aparência real da planilha no Excel:
    mesclagem de células (colspan/rowspan), cor de fundo, negrito/itálico,
    tamanho e cor de fonte, alinhamento, quebra de texto e bordas — tudo
    lido direto dos estilos do openpyxl.

    Diferente de `worksheet_para_dataframe`, isso é HTML pronto pra
    renderizar (ex.: com `st.components.v1.html`), não um DataFrame — não
    dá pra ordenar/filtrar como uma tabela de dados, é uma pré-visualização
    visual mesmo.

    Trunca em `max_linhas`/`max_colunas` por performance (ver docstring de
    `worksheet_para_dataframe` — mesmo motivo).

    Returns:
        (html, truncado_linhas, truncado_colunas)
    """
    total_linhas = ws.max_row
    total_colunas = ws.max_column
    limite_linhas = min(total_linhas, max_linhas)
    limite_colunas = min(total_colunas, max_colunas)

    # Mapeia cada célula coberta por uma mesclagem (exceto a âncora) pra
    # não renderizá-la — e guarda o rowspan/colspan da âncora.
    celulas_cobertas = set()
    spans = {}
    for rng in ws.merged_cells.ranges:
        if rng.min_row > limite_linhas or rng.min_col > limite_colunas:
            continue
        max_r = min(rng.max_row, limite_linhas)
        max_c = min(rng.max_col, limite_colunas)
        spans[(rng.min_row, rng.min_col)] = (max_r - rng.min_row + 1, max_c - rng.min_col + 1)
        for r in range(rng.min_row, max_r + 1):
            for c in range(rng.min_col, max_c + 1):
                if (r, c) != (rng.min_row, rng.min_col):
                    celulas_cobertas.add((r, c))

    # Larguras de coluna (aproximação: unidade de largura do Excel -> px)
    larguras_px = []
    largura_padrao = ws.sheet_format.defaultColWidth or 8.43
    for c in range(1, limite_colunas + 1):
        letra = get_column_letter(c)
        dim = ws.column_dimensions.get(letra)
        largura = dim.width if (dim and dim.width) else largura_padrao
        larguras_px.append(max(24, round(largura * 7 + 5)))

    linhas_html = []
    for r in range(1, limite_linhas + 1):
        dim_linha = ws.row_dimensions.get(r)
        altura_px = round((dim_linha.height or 15) * 1.33) if dim_linha else 20
        celulas_html = []
        for c in range(1, limite_colunas + 1):
            if (r, c) in celulas_cobertas:
                continue
            cel = ws.cell(row=r, column=c)
            texto = _formatar_valor_preview(cel.value)
            texto = (
                texto.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
                .replace("\n", "<br>")
            )
            attrs = ""
            span_r, span_c = spans.get((r, c), (1, 1))
            if span_r > 1:
                attrs += f' rowspan="{span_r}"'
            if span_c > 1:
                attrs += f' colspan="{span_c}"'
            estilo = _estilo_celula_css(cel)
            celulas_html.append(f'<td style="{estilo}"{attrs}>{texto}</td>')
        linhas_html.append(f'<tr style="height:{altura_px}px;">{"".join(celulas_html)}</tr>')

    colgroup = "".join(f'<col style="width:{w}px;">' for w in larguras_px)

    html = f"""
    <div style="overflow:auto; max-height:600px; border:1px solid #d1d5db;">
      <table style="border-collapse:collapse; font-family:Calibri,Arial,sans-serif;
                     font-size:13px; table-layout:fixed;">
        <colgroup>{colgroup}</colgroup>
        <tbody>{"".join(linhas_html)}</tbody>
      </table>
    </div>
    """

    truncado_linhas = total_linhas > max_linhas
    truncado_colunas = total_colunas > max_colunas
    return html, truncado_linhas, truncado_colunas


# ---------------------------------------------------------------------------
# Detecção de blocos de tabela e exclusão de colunas por "Base reduzida"
# abaixo de um limite — usado tanto pelo código 05 do Relatório
# Automatizado quanto pelo Processador de Base Reduzida standalone (as
# duas telas compartilham este mesmo motor, testado uma vez só).
# ---------------------------------------------------------------------------
def normalizar_texto_maiusculo(valor):
    return str(valor).strip().upper() if valor is not None else ""


def linha_vazia_ate_coluna(ws, r, max_col):
    return all(ws.cell(row=r, column=c).value is None for c in range(1, max_col + 1))


def encontrar_inicio_bloco(ws, linha_ref, max_col):
    """Sobe a partir de `linha_ref` enquanto a linha anterior não estiver
    totalmente vazia — acha o início do bloco/tabela contíguo."""
    r = linha_ref
    while r > 1 and not linha_vazia_ate_coluna(ws, r - 1, max_col):
        r -= 1
    return r


def encontrar_fim_bloco(ws, linha_ref, max_col):
    """Desce a partir de `linha_ref` enquanto a linha seguinte não estiver
    totalmente vazia — acha o fim do bloco/tabela contíguo."""
    r = linha_ref
    max_row = ws.max_row
    while r < max_row and not linha_vazia_ate_coluna(ws, r + 1, max_col):
        r += 1
    return r


def _excluir_coluna_no_bloco(ws, linha_inicio, linha_fim, coluna, max_col):
    """Remove uma coluna dentro de [linha_inicio, linha_fim], deslocando
    tudo à direita dela uma posição pra esquerda (valor + estilo), e
    limpa a última coluna do bloco (que sobra "duplicada" depois do
    deslocamento)."""
    from copy import copy as _copy_style
    from openpyxl.styles import Border

    for r in range(linha_inicio, linha_fim + 1):
        for c in range(coluna, max_col):
            origem = ws.cell(row=r, column=c + 1)
            destino = ws.cell(row=r, column=c)
            destino.value = origem.value
            if origem.has_style:
                destino.font = _copy_style(origem.font)
                destino.border = _copy_style(origem.border)
                destino.fill = _copy_style(origem.fill)
                destino.number_format = origem.number_format
                destino.protection = _copy_style(origem.protection)
                destino.alignment = _copy_style(origem.alignment)
        ultima = ws.cell(row=r, column=max_col)
        ultima.value = None
        ultima.border = Border()


def excluir_colunas_base_reduzida(ws, limite=25, texto_alvo="BASE REDUZIDA"):
    """
    Para toda linha cuja coluna A seja exatamente `texto_alvo` (comparação
    exata, sem diferenciar maiúsc./minúsc.), exclui as colunas desse
    bloco de tabela cujo valor numérico seja menor que `limite` —
    deslocando o resto pra esquerda e reconstruindo corretamente
    qualquer mesclagem de cabeçalho que abranja essas colunas (ex.: um
    título "Regiões" mesclado sobre várias colunas de região encolhe
    para cobrir só as colunas que sobraram, em vez de quebrar).

    Detecta os limites do bloco/tabela automaticamente (sobe/desce até
    achar uma linha totalmente vazia), então funciona tanto num
    relatório grande com várias tabelas quanto num arquivo com uma
    tabela só.

    Returns:
        Número de colunas excluídas no total (soma de todos os blocos).
    """
    max_col_geral = ws.max_column
    linhas_alvo = [
        r for r in range(1, ws.max_row + 1)
        if normalizar_texto_maiusculo(ws.cell(row=r, column=1).value) == texto_alvo
    ]

    colunas_excluidas_total = 0

    for linha_base in linhas_alvo:
        inicio = encontrar_inicio_bloco(ws, linha_base, max_col_geral)
        fim = encontrar_fim_bloco(ws, linha_base, max_col_geral)

        ultima_col_linha = 1
        for c in range(1, max_col_geral + 1):
            if ws.cell(row=linha_base, column=c).value is not None:
                ultima_col_linha = c

        colunas_para_excluir = []
        for c in range(ultima_col_linha, 1, -1):
            valor = ws.cell(row=linha_base, column=c).value
            numero = None
            if isinstance(valor, (int, float)):
                numero = valor
            elif isinstance(valor, str):
                limpo = "".join(ch for ch in valor.replace(",", ".") if ch.isdigit() or ch in ".-")
                if limpo:
                    try:
                        numero = float(limpo)
                    except ValueError:
                        numero = None
            if numero is not None and numero < limite:
                colunas_para_excluir.append(c)

        if not colunas_para_excluir:
            continue

        faixas_capturadas = []
        for rng in list(ws.merged_cells.ranges):
            if rng.min_row <= fim and rng.max_row >= inicio:
                # guarda o valor de verdade (só a célula do canto superior
                # esquerdo de uma mesclagem tem o texto) para restaurar
                # depois — sem isso, o deslocamento de coluna pode
                # sobrescrever esse valor com uma célula vizinha vazia
                valor_original = ws.cell(row=rng.min_row, column=rng.min_col).value
                faixas_capturadas.append((rng.min_row, rng.min_col, rng.max_row, rng.max_col, valor_original))
                try:
                    ws.unmerge_cells(str(rng))
                except KeyError:
                    ws.merged_cells.ranges.discard(rng)

        for col_excluir in colunas_para_excluir:  # já em ordem decrescente
            _excluir_coluna_no_bloco(ws, inicio, fim, col_excluir, max_col_geral)
            colunas_excluidas_total += 1

        for min_row, min_col, max_row, max_col, valor_original in faixas_capturadas:
            excluidas_antes = sum(1 for c in colunas_para_excluir if c < min_col)
            excluidas_dentro = sum(1 for c in colunas_para_excluir if min_col <= c <= max_col)
            novo_min_col = min_col - excluidas_antes
            novo_max_col = max_col - excluidas_antes - excluidas_dentro
            if novo_max_col >= novo_min_col:
                ws.cell(row=min_row, column=novo_min_col).value = valor_original
                # remescla se sobrou mais de 1 COLUNA (novo_max_col >
                # novo_min_col) OU se a mesclagem já era vertical desde o
                # início (max_row > min_row, ex.: "Total" ocupando as 2
                # linhas de cabeçalho) — checar só a largura ignorava
                # mesclagens puramente verticais, perdendo elas mesmo
                # quando nenhuma coluna daquele intervalo foi excluída
                if novo_max_col > novo_min_col or max_row > min_row:
                    ws.merge_cells(
                        start_row=min_row, start_column=novo_min_col,
                        end_row=max_row, end_column=novo_max_col
                    )

    return colunas_excluidas_total
