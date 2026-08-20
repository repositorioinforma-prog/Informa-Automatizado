"""
Códigos do "Relatório Automatizado" portados de VBA para Python/openpyxl.

Cada função aqui corresponde a um dos códigos originais (numeração igual
à da conversa de origem). Onde o código VBA original depende do AutoFit
real do Excel (algo que não tem como ser replicado com exatidão fora do
Excel, já que depende do motor de renderização de fonte), a altura é
estimada por contagem de caracteres — ver `core.planilha_utils` — e isso
fica marcado explicitamente no retorno de cada função, para o app avisar
a pessoa.

Sem dependência do Streamlit — pode ser testado isoladamente.
"""
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter

from core.planilha_utils import (
    estimar_altura_calculada,
    estimar_numero_linhas,
    inserir_linhas_seguro,
    largura_total_mesclagem,
    remover_linhas_seguro,
)

THICK = Side(style="thick")
MEDIUM = Side(style="medium")


def _remover_prefixo_case_insensitive(texto, prefixo):
    if texto[:len(prefixo)].lower() == prefixo.lower():
        return texto[len(prefixo):]
    return texto


# =========================================================
# Código 04 — Formatação de rótulos específicos (quebra de linha manual)
# =========================================================
# Cada regra: texto EXATO a buscar -> texto formatado (com \n),
# altura de linha (None = não mexe), largura de coluna (None = não mexe).
# Lista extraída fielmente do VBA original — específica de projeto, então
# é normal precisar adicionar/editar linhas aqui conforme o projeto mudar.
REGRAS_CODIGO_04_EXATAS = [
    ("Ensino Funda- mental", "Ensino\nFunda-\nMental", 45, 15),
    ("Aluno em domicílio (Educação Pública)", " Aluno em domicílio \n(Educação Pública)", 35, 15),
    ("Uso recente do Sistema de Saúde Pública", "Uso recente do Sistema \nde Saúde Pública", 35, 15),
    ("Participação em Tratamento Contínuo da Prefeitura",
     " Participação em Tratamento \nContínuo da Prefeitura ", 35, 15),
    ("Benefício da Cidade recebido por alguém no domicílio",
     "Benefício da Cidade recebido \n por alguém no domicílio", 35, 15),
    ("Renda Média Mensal Domiciliar (em Salário Minímo)",
     " Renda Média \n Mensal Domiciliar\n(em Salário Minímo)", 60, 15),
    ("Ensino Médio", "Ensino\nMédio", 45, 15),
    ("Ensino Superior", "Ensino\nSuperior", 45, 15),
    ("Católica Total", "Católica\nTotal", 45, 15),
    ("Católica de Batismo", "Católica de\nBatismo", 45, 15),
    ("Católica Praticante", "Católica\nPraticante", 45, 15),
    ("Evangélica Total", "Evangé-\nlica Total", 45, 15),
    ("Pente- costal", "Pente-\ncostal", 45, 15),
    (" Católica de Batismo", " Católica de\nBatismo", None, None),
    ("Espírita/Kardecista", "Espírita/\nKardecista", None, None),
    ("Afro-Brasileiras", "Afro-\nBrasileiras", None, None),
    ("Evangélicas de Missão", "Evangé-\nlicas de\nMissão", None, None),
]

# Regras de "Avaliação do Governo X": coluna 1 mantém o texto original
# (altura 18); demais colunas quebram em 2 linhas (altura 34).
GOVERNADORES_CODIGO_04 = [
    "Avaliação do Governo Tarcísio de Freitas",
    "Avaliação do Governo Romeu Zema",
    "Avaliação do Governo Eduardo Riedel",
    "Avaliação do Governo Cláudio Castro",
    "Avaliação do Governo Renato Casagrande",
    "Avaliação do Governo do ex-Presidente Bolsonaro",
]

# Prefixos que, quando a célula COMEÇA com um deles, recebem quebra de
# linha logo depois do prefixo (mantendo o resto do texto na linha de baixo)
PREFIXOS_CODIGO_04 = [
    "Avaliação do Governo Municipal",
    "Avaliação da gestão do",
    "Avaliação da gestão da",
    "Avaliação da gestão de w",
    "Avaliação da gestão atual com",
]


def aplicar_codigo_04(ws):
    """
    Aplica quebra de linha manual e ajustes de altura/largura em rótulos
    específicos (ver REGRAS_CODIGO_04_EXATAS, GOVERNADORES_CODIGO_04,
    PREFIXOS_CODIGO_04) — lista de correções fixas usada há tempos nesse
    tipo de relatório. "Denominações Católicas"/"Denominações
    Evangélicas" também são mescladas com a célula à direita, como no
    VBA original.
    """
    alterados = 0
    centralizado = Alignment(horizontal="center", vertical="center")

    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cel = ws.cell(row=r, column=c)
            texto = cel.value
            if not isinstance(texto, str):
                continue

            if texto in ("Denominações Católicas", "Denominações Evangélicas"):
                ja_mesclada = any(
                    rng.min_row <= r <= rng.max_row and rng.min_col <= c <= rng.max_col
                    for rng in ws.merged_cells.ranges
                )
                if not ja_mesclada:
                    ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c + 1)
                # aplica a formatação sempre, mesmo quando a célula já veio
                # mesclada de fábrica (o VBA original só formatava quando
                # TAMBÉM precisava mesclar — em arquivos que já chegam
                # pré-mesclados isso deixava a formatação de fora)
                cel.alignment = centralizado
                ws.row_dimensions[r].height = 18
                alterados += 1
                continue

            casou = False
            for original, formatado, altura, largura in REGRAS_CODIGO_04_EXATAS:
                if texto == original:
                    cel.value = formatado
                    cel.alignment = centralizado
                    if altura is not None:
                        ws.row_dimensions[r].height = altura
                    if largura is not None:
                        ws.column_dimensions[get_column_letter(c)].width = largura
                    alterados += 1
                    casou = True
                    break
            if casou:
                continue

            for nome_governo in GOVERNADORES_CODIGO_04:
                if texto == nome_governo:
                    if c == 1:
                        ws.row_dimensions[r].height = 18
                    else:
                        resto = nome_governo[len("Avaliação do Governo"):].strip()
                        cel.value = "Avaliação do Governo\n" + resto
                        ws.row_dimensions[r].height = 34
                    cel.alignment = centralizado
                    alterados += 1
                    casou = True
                    break
            if casou:
                continue

            if texto == "Repro- vação":
                if c == 1:
                    cel.value = "Reprovação"
                    cel.alignment = Alignment(horizontal="left", vertical="center")
                else:
                    cel.value = "Repro-\nvação"
                    ws.row_dimensions[r].height = 34
                    cel.alignment = centralizado
                alterados += 1
                continue

            for prefixo in PREFIXOS_CODIGO_04:
                if texto.startswith(prefixo):
                    resto = texto[len(prefixo):]
                    if resto:
                        cel.value = prefixo + "\n" + resto
                        cel.alignment = Alignment(wrap_text=True)
                        ws.row_dimensions[r].height = 34
                        alterados += 1
                    break

    return alterados




# =========================================================
# Código 01 — Preenche células do SPSS pirata
# =========================================================
def aplicar_codigo_01(ws):
    """
    Para cada linha onde a coluna B é 'Total': se não está mesclada
    (tabela de 2 colunas, sem segmento), aplica borda espessa na célula
    A correspondente. Se está mesclada em 2 linhas (tabela com
    segmento), mescla as 2 células correspondentes na coluna A (se
    ambas vazias e ainda não mescladas) e aplica borda espessa ao redor.
    """
    alterados = 0
    for r in range(1, ws.max_row + 1):
        cel_total = ws.cell(row=r, column=2)
        valor = str(cel_total.value).strip() if cel_total.value is not None else ""
        if valor != "Total":
            continue

        faixa_mesclada = None
        for rng in ws.merged_cells.ranges:
            if rng.min_row <= r <= rng.max_row and rng.min_col <= 2 <= rng.max_col:
                faixa_mesclada = rng
                break

        if faixa_mesclada is None:
            a1 = ws.cell(row=r, column=1)
            a1.border = Border(top=THICK, bottom=THICK, left=THICK, right=THICK)
            alterados += 1
        elif faixa_mesclada.max_row - faixa_mesclada.min_row + 1 == 2:
            linha1 = faixa_mesclada.min_row
            linha2 = linha1 + 1
            a1 = ws.cell(row=linha1, column=1)
            a2 = ws.cell(row=linha2, column=1)

            ja_mesclada_a = any(
                rng.min_row <= linha1 <= rng.max_row and rng.min_col <= 1 <= rng.max_col
                for rng in ws.merged_cells.ranges
            )
            v1 = str(a1.value).strip() if a1.value is not None else ""
            v2 = str(a2.value).strip() if a2.value is not None else ""

            if not ja_mesclada_a and v1 == "" and v2 == "":
                # aplica a borda ANTES de mesclar: o merge_cells troca o
                # objeto interno das células que não são a "mestre", então
                # qualquer alteração feita DEPOIS do merge nessas células
                # se perde
                a1.border = Border(top=THICK, left=THICK, right=THICK)
                a2.border = Border(bottom=THICK, left=THICK, right=THICK)
                ws.merge_cells(start_row=linha1, start_column=1, end_row=linha2, end_column=1)
                alterados += 1

    return alterados


# =========================================================
# Código 02 — Muda Bordas (espessa -> média)
# =========================================================
def aplicar_codigo_02(ws):
    """Converte toda borda 'espessa' (thick) da planilha para 'média' (medium)."""
    alterados = 0
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cel = ws.cell(row=r, column=c)
            b = cel.border
            if not any([b.left, b.top, b.right, b.bottom]):
                continue

            mudou = False

            def _rebaixar(lado):
                nonlocal mudou
                if lado is not None and lado.style == "thick":
                    mudou = True
                    return Side(style="medium", color=lado.color)
                return lado

            novo = Border(
                left=_rebaixar(b.left), right=_rebaixar(b.right),
                top=_rebaixar(b.top), bottom=_rebaixar(b.bottom),
            )
            if mudou:
                cel.border = novo
                alterados += 1

    return alterados


# =========================================================
# Código 03 — Ordenar tabelas (por valor da coluna B, descendente)
# =========================================================
# Lista de rótulos que NUNCA entram na ordenação — ficam fixos na posição
# original (linha) onde já estão. É extensa e específica de projeto
# (categorias de resposta padrão, "Não sabe", "Branco/Nulo", opções fixas
# de escala etc.) — o VBA original tinha essa lista hardcoded; aqui vira
# parâmetro com esse valor como padrão, editável pela tela de Códigos
# Individuais / Relatório Automatizado sem precisar mexer no código.
# Comparação sempre case-insensitive (equivalente ao
# `d.CompareMode = vbTextCompare` do VBA original).
TERMOS_EXCLUSAO_CODIGO_03 = [
    "Base",
    "Base reduzida",
    "Outros",
    "Direita",
    "Centro",
    "Esquerda",
    "Branco/Nulo",
    "Votaria em todos",
    "Nenhum",
    "Aumenta",
    "Diminui",
    "Não sabe",
    "Não sabe avaliar",
    "Não conhece",
    "Conhece um pouco ",
    "Conhece de ouvir falar ",
    "Conhece bem ",
    "Positiva",
    "Negativa",
    "Regular",
    "Não rejeito nenhum ",
    "Não sabe/Indeciso",
    "Não irá votar ",
    "Não foi votar ",
    "Concordo totalmente",
    "Nem concordo nem discordo",
    "Discordo totalmente",
    "Um candidato opositor a Família Reis ",
    "Um candidato da Família Reis ",
    "Um candidato aliado aos Reis, mas que não seja da família; ",
    "Depende do candidato",
    "Conhece pela foto",
    "Conhece muito",
    "Conhece de nome",
    "Nunca vi",
    "Nunca ouvi falar",
    "Votarei Nulo ou em Branco ",
    "Concordo plenamente",
    "Concordo",
    "Discordo",
    "Discordo plenamente",
    "Nenhuma palavra",
    "É conhecedor de Brasília e das prioridades das pessoas ",
    "Sempre",
    "Frequentemente",
    "Eventualmente",
    "Raramente",
    "Nunca",
    "Afeta muito",
    "Afeta um pouco",
    "Não afeta",
    "Atualmente participa",
    "Já participou",
    "Nunca participou",
    "Algumas vezes por semana",
    "Algumas vezes no mês",
    "Grande interesse",
    "Interesse moderado",
    "Baixo interesse",
    "Diariamente",
    "Certamente votaria",
    "Provavelmente votaria",
    "Provavelmente não votaria",
    "Certamente não votaria",
    "Certamente voto",
    "Certamente não voto",
    "Não conheço suficiente para votar",
    "Provavelmente voto",
    "Provavelmente não voto",
    "Todas as alternativas",
    "Nenhum desses",
    "Branco / Nulo",
    "Ajuda muito na decisão",
    "Ajuda na decisão",
    "Atrapalha na decisão",
    "Atrapalha muito na decisão",
    "Preparado",
    "Despreparado",
    "Experiente",
    "Inexperiente",
    "Tem influência política",
    "Não tem influência política",
    "Conhece a cidade",
    "Não sei",
    "Não votarei",
    "Aprovação",
    "Reprovação",
    "Nenhum deles",
    "Melhorou",
    "Piorou",
    "Continua da mesma forma",
    "Nada a melhorar",
    "Vai melhorar",
    "Ficar como está",
    "Vai piorar",
    "Confia muito",
    "Confia",
    "Confia mais ou menos",
    "Confia pouco",
    "Não confia",
    "Trocaria",
    "Não trocaria",
    "Outras cidades",
    "Não faz nada",
    "Nada contra a gestão atual",
    "Votou branco/nulo",
    "Está preparado",
    "Não está preparado",
    "Tem capacidade",
    "Não tem capacidade",
    "Tem força política",
    "Não tem força política",
    "Tem autonomia",
    "Não tem autonomia",
    "Ajudaria",
    "Não ajudaria",
    "Demais, deveria diminuir",
    "Na medida certa",
    "Pouca, deveria aumentar",
    "Lembra",
    "Não lembra",
    "Desde que nasceu",
    "Mora há mais de 20 anos",
    "Mora entre 20 e 10 anos",
    "Mora entre 10 e 5 anos",
    "Mora há menos de 5 anos",
    "Irá cumprir todas as promessas que faz",
    "Irá cumprir somente algumas",
    "Não irá cumprir nenhuma promessa que faz",
    "Promete muito mais do que conseguirá fazer em Maricá",
    "Promete na medida certa",
    "Promete pouco para Maricá, pois conseguiria fazer mais",
    "Sim",
    "Não",
    "Indiferente",
    "Outros aspectos",
    "Nada de bom",
    "Não vê nada de ruim",
    "Não sente falta de nada",
    "Muito interesse",
    "Interesse regular",
    "Nenhum interesse",
    "Excelente",
    "Boa",
    "Ruim",
    "Péssima",
    "Votou Branco ou Nulo",
    "Repro- vação",
    "Média",
    "Otimista",
    "Pessimista",
    "Nada importante",
    "Pouco importante",
    "Importante",
    "Muito importante",
    "NS",
    "Evoluindo",
    "Mesma coisa / parada",
    "Regredindo",
    "Da esquerda",
    "Da direita",
    "Do centro",
    "Conheço bem",
    "Conheço de ouvir falar",
    "Conheço das redes sociais",
    "Não conheço",
    "Indeciso",
    "Aumenta minha chance de votar nesse candidato",
    "É indiferente o apoio",
    "Diminui a minha chance de votar nesse candidato",
    "Não lembra/ Não tem candidato",
    "Não sabe em quem votar",
    "Outros candidatos",
    "Nulo / Branco",
    "Outras áreas",
    "Nenhuma área",
    "Todas as áreas",
    "Outras necessidades",
    "Nenhuma necessidade",
    "Influencia meu voto no candidato que apoiar",
    "Posso considerar o voto no candidato apoiado",
    "Posso considerar não votar no candidato apoiado",
    "Influencia a não votar no candidato apoiado",
    "Conheço muito",
    "Conheço um pouco",
    "Conheço somente de foto",
    "Conheço de foto, mas, não sabia que era Rodrigo Bacellar",
    "Nunca ouvi falar e nunca vi",
    "Tenho informações positivas",
    "Tenho informações negativas",
    "Não tenho informações",
    "Já viu",
    "Nunca viu",
    "Nenhum dos dois governos está bem",
    "Concorda totalmente",
    "Concorda em parte",
    "Não concorda e nem discorda",
    "Discorda em parte",
    "Discorda totalmente",
    "Melhor",
    "Igual",
    "Pior",
    "Todos devem permanecer presos e ser condenados",
    "A prisão deve ser apenas para quem danificou prédios",
    "Todos devem receber anistia",
    "Maricá oferece vagas de emprego até demais para os moradores",
    "Maricá oferece emprego na medida certa, todos poderiam estar empregados",
    "Maricá oferece poucas vagas de emprego, e vai oferecer cada vez menos",
    "Maricá oferece poucas vagas de emprego, mas, começará a oferecer mais",
    "O turismo está crescendo, mas falta gerar emprego e renda aos moradores",
    "O turismo está crescendo e tem ajudado a gerar emprego e renda aos moradores",
    "O turismo está estagnado",
    "O turismo está diminuindo",
    "Maricá está crescendo, mas sem planejamento e ficará pior",
    "Maricá está crescendo, sem planejamento, mas haverá uma organização",
    "Maricá já está crescendo de forma organizada",
    "Maricá está estagnada",
    "Maricá está regredindo",
]


def aplicar_codigo_03(ws, termos_exclusao=None):
    """
    Localiza cada célula com o texto exato "Total" (mesclada ou não,
    comparação sem diferenciar maiúsc/minúsc). Abaixo de cada uma,
    considera como "bloco de categorias" as linhas seguintes enquanto a
    coluna B não estiver vazia e não for outro "Total". Dentro desse
    bloco, reordena as linhas cujo rótulo (coluna A) NÃO está na lista de
    exclusão, por valor numérico da coluna B em ordem DECRESCENTE — as
    linhas da lista de exclusão (ex.: "Não sabe", "Branco/Nulo", "Base")
    ficam paradas na posição em que já estavam, servindo de âncora fixa
    entre os grupos que são de fato reordenados.

    Só troca os VALORES das células (todas as colunas usadas), igual ao
    VBA original — não mexe em formatação, então cada linha mantém a
    aparência (fonte, borda, preenchimento) que já tinha na posição onde
    ela está, só o conteúdo é que muda de lugar.

    Args:
        ws: planilha a processar.
        termos_exclusao: lista de rótulos a manter fixos (case-insensitive).
            Se None, usa `TERMOS_EXCLUSAO_CODIGO_03` (lista padrão herdada
            do VBA original).

    Returns:
        Número de tabelas (blocos "Total") que tiveram linhas reordenadas.
    """
    termos = {t.strip().lower() for t in (termos_exclusao or TERMOS_EXCLUSAO_CODIGO_03)}
    max_col = ws.max_column
    max_row = ws.max_row

    # Acha todas as células "Total" (exato, sem diferenciar maiúsc/minúsc),
    # em ordem de leitura (linha por linha, como o Find/FindNext padrão do Excel)
    celulas_total = [
        cel
        for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col)
        for cel in row
        if cel.value is not None and str(cel.value).strip().lower() == "total"
    ]

    tabelas_reordenadas = 0

    for cel in celulas_total:
        # Se a célula "Total" está mesclada, o bloco de categorias começa
        # logo após o fim da mesclagem; senão, na linha seguinte.
        rng_merge = None
        for rng in ws.merged_cells.ranges:
            if rng.min_row <= cel.row <= rng.max_row and rng.min_col <= cel.column <= rng.max_col:
                rng_merge = rng
                break
        linha_inicio = (rng_merge.max_row + 1) if rng_merge is not None else (cel.row + 1)

        # Bloco vai até a coluna B ficar vazia ou repetir "Total"
        linha_fim = linha_inicio
        while linha_fim <= max_row:
            valor_b = ws.cell(row=linha_fim, column=2).value
            valor_b_str = "" if valor_b is None else str(valor_b).strip()
            if valor_b_str == "" or valor_b_str.lower() == "total":
                break
            linha_fim += 1
        linha_fim -= 1

        if linha_fim < linha_inicio:
            continue

        linhas_para_ordenar = [
            r for r in range(linha_inicio, linha_fim + 1)
            if str(ws.cell(row=r, column=1).value or "").strip().lower() not in termos
        ]
        if len(linhas_para_ordenar) < 2:
            continue

        def _chave_ordenacao(r):
            v = ws.cell(row=r, column=2).value
            try:
                return float(str(v).replace(",", ".")) if v is not None else float("-inf")
            except (TypeError, ValueError):
                return float("-inf")

        # Snapshot dos valores ANTES de escrever nada (senão sobrescreveria
        # linhas que ainda vamos ler, já que a origem/destino se cruzam)
        valores_por_linha = {
            r: [ws.cell(row=r, column=c).value for c in range(1, max_col + 1)]
            for r in linhas_para_ordenar
        }
        linhas_ordenadas = sorted(linhas_para_ordenar, key=_chave_ordenacao, reverse=True)

        # Reescreve nas mesmas posições originais (linhas_para_ordenar, em
        # ordem), usando os valores na ordem já ordenada — exatamente como
        # o VBA faz ao reler tempArray (ordenado) de volta nas linhas
        # originais na sequência em que apareciam.
        for linha_destino, linha_origem in zip(linhas_para_ordenar, linhas_ordenadas):
            for c, valor in enumerate(valores_por_linha[linha_origem], start=1):
                ws.cell(row=linha_destino, column=c, value=valor)

        tabelas_reordenadas += 1

    return tabelas_reordenadas


# =========================================================
# Código 06 — Ajuste DIN 9 na linha de Pergunta (+ autofit aproximado)
# =========================================================
def aplicar_codigo_06(ws):
    """
    Reduz para tamanho 9 a fonte de qualquer célula contendo 'Pergunta:'.
    NÃO faz o autofit real (não é possível fora do Excel) — só ajusta a
    fonte, que é a parte 100% confiável deste código.
    """
    alterados = 0
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cel = ws.cell(row=r, column=c)
            if isinstance(cel.value, str) and "pergunta:" in cel.value.lower():
                fonte = cel.font
                cel.font = Font(
                    name=fonte.name, size=9, bold=fonte.bold, italic=fonte.italic,
                    vertAlign=fonte.vertAlign, underline=fonte.underline,
                    strike=fonte.strike, color=fonte.color,
                )
                alterados += 1
    return alterados


# =========================================================
# Código 08 — Base pequena
# =========================================================
def aplicar_codigo_08(ws):
    """
    Em toda linha 'Base'/'Base reduzida', marca com '*' os valores entre
    20 e 30 (exclusivos) e insere uma linha de nota "*Base pequena para
    análise estatística" logo abaixo (pulando a linha 'Média', se ela
    estiver a até 3 linhas de distância) — sempre em fonte DIN tamanho 9.
    """
    linhas_marcadas = 0
    r = 1
    while r <= ws.max_row:
        texto = str(ws.cell(row=r, column=1).value or "").strip().upper()
        if texto in ("BASE", "BASE REDUZIDA"):
            deslocado = _analisar_linha_codigo_08(ws, r)
            if deslocado:
                linhas_marcadas += 1
                r += deslocado  # o resto do arquivo desceu — pula a nota recém-inserida
        r += 1
    return linhas_marcadas


def _analisar_linha_codigo_08(ws, linha):
    col_max = 1
    for c in range(1, ws.max_column + 1):
        if ws.cell(row=linha, column=c).value is not None:
            col_max = c

    mensagem_adicionada = False
    deslocamento_total = 0

    for coluna in range(1, col_max + 1):
        cel = ws.cell(row=linha, column=coluna)
        valor = cel.value
        if isinstance(valor, (int, float)) and 20 < valor < 30:
            cel.value = f"{valor}*"
            cel.alignment = _com_wrap_text(cel.alignment, False)

            if not mensagem_adicionada:
                tem_media = any(
                    str(ws.cell(row=linha + i, column=1).value or "").strip().upper() == "MÉDIA"
                    for i in range(1, 4)
                )
                offset = 3 if tem_media else 2
                inserir_linhas_seguro(ws, linha + offset, 1)
                nota = ws.cell(row=linha + offset, column=1, value="*Base pequena para análise estatística")
                nota.alignment = _com_wrap_text(nota.alignment, False)
                nota.font = Font(name="DIN", size=9)
                mensagem_adicionada = True
                deslocamento_total = offset

    return deslocamento_total


def _com_wrap_text(alinhamento_atual, valor):
    """Retorna uma cópia do alinhamento existente, só trocando wrap_text — preserva
    horizontal/vertical/indent/etc. (o VBA original só mexe no WrapText)."""
    a = alinhamento_atual
    return Alignment(
        horizontal=a.horizontal, vertical=a.vertical, text_rotation=a.text_rotation,
        wrap_text=valor, shrink_to_fit=a.shrink_to_fit, indent=a.indent,
        justifyLastLine=a.justifyLastLine, readingOrder=a.readingOrder,
    )


def _com_horizontal_vertical(alinhamento_atual, horizontal, vertical):
    """Retorna uma cópia do alinhamento existente, só trocando horizontal/vertical —
    preserva wrap_text/indent/etc. (o VBA original só mexe nesses dois quando
    define HorizontalAlignment/VerticalAlignment, nunca no WrapText)."""
    a = alinhamento_atual
    return Alignment(
        horizontal=horizontal, vertical=vertical, text_rotation=a.text_rotation,
        wrap_text=a.wrap_text, shrink_to_fit=a.shrink_to_fit, indent=a.indent,
        justifyLastLine=a.justifyLastLine, readingOrder=a.readingOrder,
    )


# =========================================================
# Código 09 — Ajusta a altura das labels (autofit aproximado)
# =========================================================
# Cada grupo: (gatilhos de texto, altura fixa, coluna mínima onde vale
# — None = qualquer coluna). Comparação por "começa com",
# case-insensitive. Escolaridade ('Ensino...') foi pra 45 a pedido do
# Lucas (era 30, junto com Sexo, até essa mudança); Sexo
# ('Masculino'/'Feminino') continua em 30, em qualquer coluna. Os
# segmentos de avaliação ('Aprovação'/'Regular'/'Reprovação') também
# ficam em 30, mas SÓ quando aparecem da coluna C em diante — se
# aparecerem na coluna A (rótulo da linha, não valor de segmento), não
# deve ser alterado.
_GATILHOS_ALTURA_FIXA = (
    (("masculino", "feminino"), 30, None),
    (("ensino",), 45, None),
    (("aprovação", "regular", "reprovação"), 30, 3),
)


def _altura_fixa_para_linha(ws, r):
    """
    Se a linha tiver algum gatilho de altura fixa (ver
    `_GATILHOS_ALTURA_FIXA`), retorna a altura correspondente — essas
    linhas não passam pela estimativa normal do código 09. Senão,
    retorna None.
    """
    for gatilhos, altura, coluna_minima in _GATILHOS_ALTURA_FIXA:
        col_inicio = coluna_minima or 1
        for c in range(col_inicio, ws.max_column + 1):
            valor = ws.cell(row=r, column=c).value
            if isinstance(valor, str):
                texto = valor.strip().lower()
                if any(texto.startswith(gatilho) for gatilho in gatilhos):
                    return altura
    return None


def aplicar_codigo_09(ws):
    """
    Para toda célula com quebra de linha (WrapText) ativada, estima a
    altura necessária (aproximação por contagem de caracteres — ver
    aviso em `core.planilha_utils.estimar_altura_calculada`) e aplica na
    linha. O AutoFit real do Excel ajusta a altura pra caber exatamente
    (pode até diminuir uma linha que estava alta demais) — então aqui
    também aplica sempre a estimativa, não só quando ela é maior que a
    altura atual.

    Exceção: linhas com algum gatilho de `_GATILHOS_ALTURA_FIXA`
    (subcategoria de Sexo, Escolaridade, ou segmento de avaliação)
    sempre ficam com a altura fixa correspondente, em vez da
    estimativa — um valor pré-definido, não calculado.
    """
    alterados = 0
    for r in range(1, ws.max_row + 1):
        altura_fixa = _altura_fixa_para_linha(ws, r)
        if altura_fixa is not None:
            if ws.row_dimensions[r].height != float(altura_fixa):
                ws.row_dimensions[r].height = float(altura_fixa)
                alterados += 1
            continue

        maior_estim = None
        for c in range(1, ws.max_column + 1):
            cel = ws.cell(row=r, column=c)
            if not (cel.alignment and cel.alignment.wrap_text):
                continue
            if not cel.value:
                continue

            largura = largura_total_mesclagem(ws, cel)
            tamanho_fonte = cel.font.size or 10
            estim = estimar_altura_calculada(cel.value, largura, tamanho_fonte)
            if maior_estim is None or estim > maior_estim:
                maior_estim = estim

        if maior_estim is None:
            continue

        altura_atual = ws.row_dimensions[r].height
        nova_altura = round(maior_estim, 1)
        if altura_atual != nova_altura:
            ws.row_dimensions[r].height = nova_altura
            alterados += 1
    return alterados


# =========================================================
# Código 11 — Ajusta a altura dos Títulos (+ remove "Titulo: ")
# =========================================================
def aplicar_codigo_11(ws):
    """
    Para toda célula contendo 'Titulo: ': estima a altura (aproximada,
    igual ao código 09) e aplica na "escada" 30/45/60/75... Remove o
    prefixo 'Titulo: ' do texto, mantendo só o título em si, centraliza
    o texto horizontal e verticalmente, e deixa em negrito.
    """
    alterados = 0
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cel = ws.cell(row=r, column=c)
            if not (isinstance(cel.value, str) and "titulo: " in cel.value.lower()):
                continue

            largura = largura_total_mesclagem(ws, cel)
            tamanho_fonte = cel.font.size or 10
            altura_calc = estimar_altura_calculada(cel.value, largura, tamanho_fonte) + 2
            altura_final = _altura_staircase_titulo(altura_calc)
            ws.row_dimensions[r].height = altura_final

            texto_novo = _remover_prefixo_case_insensitive(cel.value, "Titulo: ")
            cel.value = texto_novo
            cel.alignment = _com_horizontal_vertical(cel.alignment, "center", "center")

            fonte_atual = cel.font
            cel.font = Font(
                name=fonte_atual.name, size=fonte_atual.size, bold=True,
                italic=fonte_atual.italic, color=fonte_atual.color,
            )
            alterados += 1
    return alterados


def _altura_staircase_titulo(altura_calculada):
    altura = 30
    while altura < altura_calculada:
        altura += 15
    return altura


# =========================================================
# Código 12 — Ajuste da altura das Perguntas
# =========================================================
def aplicar_codigo_12(ws):
    """
    Para toda célula contendo 'Pergunta: ': estima quantas linhas o
    texto vai ocupar (quebra automática, mesma aproximação usada nos
    demais códigos de altura — ver `estimar_numero_linhas`) e aplica
    altura = número de linhas × 15 (1 linha -> 15, 2 linhas -> 30, 3
    linhas -> 45...). Fórmula fixa a pedido do Lucas — substituiu a
    "escada" antiga (20 -> 30 -> +15/+15...). Não remove o texto (esse
    código, diferente do 11, mantém 'Pergunta: ' no texto).
    """
    alterados = 0
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cel = ws.cell(row=r, column=c)
            if not (isinstance(cel.value, str) and "pergunta: " in cel.value.lower()):
                continue

            largura = largura_total_mesclagem(ws, cel)
            tamanho_fonte = cel.font.size or 10
            n_linhas = estimar_numero_linhas(cel.value, largura, tamanho_fonte)
            altura_final = n_linhas * 15
            if ws.row_dimensions[r].height != altura_final:
                ws.row_dimensions[r].height = altura_final
                alterados += 1
    return alterados


# =========================================================
# Código 05 — Base reduzida (exclusão de colunas, direto na planilha)
# =========================================================
# A lógica de detectar blocos de tabela e excluir colunas de "Base
# reduzida" abaixo de um limite (com mesclagem preservada) mora em
# core/planilha_utils.py — compartilhada com o Processador de Base
# Reduzida standalone (analises/base_reduzida.py), pra não duplicar essa
# lógica (já testada com arquivo real, inclusive com cabeçalho mesclado
# tipo "Regiões") em dois lugares.
from core.planilha_utils import (
    normalizar_texto_maiusculo as _normalizar_texto_base,
    linha_vazia_ate_coluna as _linha_vazia_ate_coluna,
    encontrar_inicio_bloco as _encontrar_inicio_bloco,
    encontrar_fim_bloco as _encontrar_fim_bloco,
    excluir_colunas_base_reduzida as _excluir_colunas_base_reduzida,
)

def _remover_linha_extra_multipla(ws):
    """
    Logo abaixo de um título mesclado de pergunta do tipo Múltipla
    ('Múltipla', 'Multipla' ou 'Estimulada e Múltipla'), se a linha
    imediatamente seguinte estiver totalmente em branco, remove essa
    linha (efeito de uma linha extra que sobra nesse tipo de tabela).
    Processa de baixo para cima.

    A comparação ignora acento (usa `unicodedata` pra tirar diacríticos
    antes de comparar) — sem isso, o texto real do relatório ('Múltipla',
    com acento) nunca batia contra os literais 'MULTIPLA'/'ESTIMULADA E
    MULTIPLA' (sem acento), e a linha extra nunca era removida na
    prática. Bug antigo, corrigido aqui — não é específico do Total
    Automático, também afeta o código 05 do Relatório Automatizado.
    """
    import unicodedata

    def _sem_acento(texto):
        return unicodedata.normalize("NFKD", texto).encode("ascii", "ignore").decode("ascii")

    max_col = ws.max_column
    removidas = 0
    r = ws.max_row
    while r >= 1:
        cel = ws.cell(row=r, column=1)
        faixa = None
        for rng in ws.merged_cells.ranges:
            if rng.min_row <= r <= rng.max_row and rng.min_col <= 1 <= rng.max_col:
                faixa = rng
                break

        if faixa is not None and faixa.min_row == r:
            texto = _sem_acento(_normalizar_texto_base(ws.cell(row=faixa.min_row, column=1).value))
            eh_multipla = "ESTIMULADA E MULTIPLA" in texto or "MULTIPLA" in texto
            if eh_multipla:
                linha_abaixo = faixa.max_row + 1
                if linha_abaixo <= ws.max_row and _linha_vazia_ate_coluna(ws, linha_abaixo, max_col):
                    ws.delete_rows(linha_abaixo, 1)
                    removidas += 1
        elif faixa is None:
            texto = _sem_acento(_normalizar_texto_base(cel.value))
            eh_multipla = "ESTIMULADA E MULTIPLA" in texto or "MULTIPLA" in texto
            if eh_multipla:
                linha_abaixo = r + 1
                if linha_abaixo <= ws.max_row and _linha_vazia_ate_coluna(ws, linha_abaixo, max_col):
                    ws.delete_rows(linha_abaixo, 1)
                    removidas += 1
        r -= 1

    return removidas


def aplicar_layout_basico_planilha(ws):
    """
    Parte "layout" do VBA 'ConfigurarUmaPlanilha' — tudo, EXCETO a
    exclusão de colunas de Base reduzida (essa fica isolada em
    `aplicar_codigo_05`, que chama esta função por baixo). Extraída à
    parte pra ser reaproveitada por outros fluxos que precisam do mesmo
    layout sem a exclusão por limite (ex.: Total Automático).

    - fonte DIN 10 e alinhamento centralizado em toda a planilha;
    - coluna A alinhada à esquerda (horizontal) e centralizada (vertical);
    - margens, orientação PAISAGEM e papel A4;
    - largura da coluna A = 21, coluna B = 8, demais colunas = 8.67
      (+ compensação empírica AJUSTE_LARGURA_COLUNA_DIN);
    - escala de impressão 100% (não "ajustar à largura");
    - toda célula mesclada na coluna A contendo 'Pergunta' ou '*' alinhada
      à esquerda;
    - linhas de grade ATIVADAS (a planilha do SPSS às vezes já vem com
      elas desligadas — força ligado em vez de só herdar o que quer que
      o arquivo de origem tiver).

    (Não remove mais a linha extra abaixo de títulos 'Múltipla' — tinha
    uma tentativa de correção aqui, mas não funcionou como esperado num
    arquivo real e foi desativada a pedido do Lucas. A função
    `_remover_linha_extra_multipla` continua no arquivo, só não é mais
    chamada; ver comentário logo abaixo.)
    """
    from openpyxl.worksheet.page import PageMargins

    ws.sheet_view.showGridLines = True

    fonte_padrao = Font(name="DIN", size=10)

    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cel = ws.cell(row=r, column=c)
            cel.font = fonte_padrao
            horizontal = "left" if c == 1 else "center"
            cel.alignment = _com_horizontal_vertical(cel.alignment, horizontal, "center")

    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_margins = PageMargins(
        left=1.5 / 2.54, right=1.3 / 2.54, top=2.5 / 2.54, bottom=2 / 2.54,
        header=1.3 / 2.54, footer=1.3 / 2.54,
    )
    # Escala de impressão em 100% (Layout de Página > Escala), não o
    # modo "ajustar a X página(s) de largura" — os dois são mutuamente
    # exclusivos no Excel via pageSetUpPr.fitToPage.
    ws.sheet_properties.pageSetUpPr.fitToPage = False
    ws.page_setup.fitToWidth = None
    ws.page_setup.fitToHeight = None
    ws.page_setup.scale = 100

    # O Excel parece recalcular a largura "em caracteres" exibida usando
    # a métrica da fonte DIN (aplicada nas células logo abaixo) em vez da
    # fonte "Normal" do arquivo (Calibri, que é o que o openpyxl usa como
    # referência ao gravar o número puro) — resultado: gravar 20.67 faz
    # o Excel mostrar ~19.89 na caixa de diálogo de largura de coluna, um
    # desvio de ~0.78 constante (não proporcional) nos testes feitos.
    # Compensa gravando o valor-alvo + 0.78, calibrado empiricamente
    # contra esse comportamento — não é uma unidade "oficial" documentada
    # do Excel, então se um dia isso mudar de comportamento (nova versão
    # do Excel, fonte diferente etc.), essa constante pode precisar ser
    # recalibrada.
    AJUSTE_LARGURA_COLUNA_DIN = 0.78
    ws.column_dimensions["A"].width = 21 + AJUSTE_LARGURA_COLUNA_DIN
    if ws.max_column >= 2:
        ws.column_dimensions["B"].width = 8 + AJUSTE_LARGURA_COLUNA_DIN
    for c in range(3, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(c)].width = 8.67 + AJUSTE_LARGURA_COLUNA_DIN

    for r in range(1, ws.max_row + 1):
        cel = ws.cell(row=r, column=1)
        faixa = None
        for rng in ws.merged_cells.ranges:
            if rng.min_row <= r <= rng.max_row and rng.min_col <= 1 <= rng.max_col:
                faixa = rng
                break
        if faixa is not None:
            texto = str(ws.cell(row=faixa.min_row, column=1).value or "")
            if "pergunta" in texto.lower() or "*" in texto:
                cel_ref = ws.cell(row=faixa.min_row, column=1)
                a = cel_ref.alignment
                cel_ref.alignment = Alignment(
                    horizontal="left", vertical=a.vertical, text_rotation=a.text_rotation,
                    wrap_text=a.wrap_text, shrink_to_fit=a.shrink_to_fit, indent=a.indent,
                    justifyLastLine=a.justifyLastLine, readingOrder=a.readingOrder,
                )

    # Removida a chamada de `_remover_linha_extra_multipla(ws)` — a
    # correção do bug de acentuação (comparava com 'MULTIPLA' sem
    # acento contra texto real 'MÚLTIPLA' com acento) não resolveu o
    # problema na prática num arquivo real; desativada a pedido do
    # Lucas até investigar com mais calma. A função continua definida
    # acima, só não é mais chamada daqui.


def aplicar_codigo_05(ws, limite=25):
    """
    Porta o VBA 'ConfigurarUmaPlanilha' por completo: aplica o layout
    básico (ver `aplicar_layout_basico_planilha`) e, por fim, exclui as
    colunas de Base reduzida abaixo do limite escolhido.
    """
    aplicar_layout_basico_planilha(ws)
    excluidas = _excluir_colunas_base_reduzida(ws, limite)

    return excluidas



def aplicar_codigo_06_autofit(ws):
    """
    Aproxima o `usedRange.Rows.AutoFit` do VBA original: para toda célula
    com texto, estima a altura necessária (mesma aproximação usada nos
    códigos 09/11/12) e aplica na linha, se maior que a altura atual.
    Não sobrescreve pra baixo — só aumenta quando o texto realmente
    precisar de mais espaço.
    """
    alterados = 0
    for r in range(1, ws.max_row + 1):
        maior_estim = None
        for c in range(1, ws.max_column + 1):
            cel = ws.cell(row=r, column=c)
            if not cel.value or not isinstance(cel.value, str):
                continue

            largura = largura_total_mesclagem(ws, cel)
            tamanho_fonte = cel.font.size or 10
            estim = estimar_altura_calculada(cel.value, largura, tamanho_fonte)
            if maior_estim is None or estim > maior_estim:
                maior_estim = estim

        if maior_estim is None:
            continue

        altura_atual = ws.row_dimensions[r].height or 15
        if maior_estim > altura_atual:
            ws.row_dimensions[r].height = round(maior_estim, 1)
            alterados += 1

    return alterados


# =========================================================
# Código 07 — Adiciona linhas + quebra de página
# =========================================================
def aplicar_codigo_07(ws):
    """
    Para toda célula na coluna A contendo 'Pergunta' ou '*' (mesclada ou
    não — o VBA original só olha células mescladas, mas isso deixa o
    código mais tolerante a variações do arquivo real): normaliza o vão
    logo abaixo pra ficar com exatamente 4 linhas em branco (insere a
    mais se faltar, remove o excedente se já houver linhas em branco ali
    de sobra) e adiciona uma quebra de página manual bem no meio dessas
    4 linhas (depois da 2ª, antes da 3ª). Separadamente, sempre que
    encontrar um trecho de 4 ou mais linhas em branco seguidas que NÃO
    esteja logo depois de um bloco "Pergunta"/"*" (esse já ganhou sua
    própria quebra ali do lado, então outra aqui seria redundante),
    insere uma quebra de página no meio dele — no máximo uma por trecho,
    não uma a cada 4 linhas dentro do mesmo vão vazio longo. Processa de
    baixo para cima, como o original, para as inserções/remoções não
    bagunçarem as linhas ainda não visitadas.
    """
    from openpyxl.worksheet.pagebreak import Break

    def _adicionar_quebra_sem_duplicar(linha_id):
        """Só adiciona a quebra se ainda não existir uma nesse id — evita
        duplicar se a função rodar mais de uma vez sobre o mesmo arquivo
        (ex.: reaplicação acidental do mesmo código)."""
        if not any(b.id == linha_id for b in ws.row_breaks.brk):
            ws.row_breaks.append(Break(id=linha_id))
            return True
        return False

    # Fixa a impressão em escala 100% (modo "Ajustar a: 100% do tamanho
    # normal", não o modo "Ajustar a N página(s) de largura") — o Excel
    # trata os dois como mutuamente exclusivos via pageSetUpPr.fitToPage.
    # Com fitToPage=False + scale=100, o Excel usa a paginação natural da
    # planilha (incluindo as quebras manuais inseridas abaixo) sem
    # recalcular nada pra "caber" num número de páginas — mais previsível
    # do que o modo "ajustar largura" que usávamos antes (que também
    # evitava o bug de quebra manual sendo ignorada, mas por um caminho
    # mais indireto).
    ws.sheet_properties.pageSetUpPr.fitToPage = False
    ws.page_setup.fitToWidth = None
    ws.page_setup.fitToHeight = None
    ws.page_setup.scale = 100

    def _e_gatilho(valor):
        texto = str(valor) if valor is not None else ""
        return bool(texto) and ("pergunta" in texto.lower() or "*" in texto)

    quebras_adicionadas = 0
    linhas_inseridas = 0

    ultima_linha_original = ws.max_row  # antes de qualquer inserção/remoção nesta função

    r = ws.max_row
    while r >= 1:
        cel = ws.cell(row=r, column=1)

        if _e_gatilho(cel.value):
            # Conta quantas linhas já vazias existem logo abaixo, pra
            # normalizar o vão pra exatamente 4 (em vez de sempre somar
            # +3 por cima do que já existe, o que gerava vãos de tamanho
            # variável e às vezes até uma quebra automática extra do
            # próprio Excel no meio de um vão maior que uma página).
            gap_existente = 0
            rr = r + 1
            while rr <= ws.max_row and ws.cell(row=rr, column=1).value is None:
                gap_existente += 1
                rr += 1

            if gap_existente < 4:
                faltam = 4 - gap_existente
                inserir_linhas_seguro(ws, r + 1, faltam)
                linhas_inseridas += faltam
            elif gap_existente > 4:
                excesso = gap_existente - 4
                remover_linhas_seguro(ws, r + 5, excesso)

            # Só adiciona a quebra se sobrar conteúdo real depois dela —
            # quando a "Pergunta" já é a última coisa do arquivo, o
            # Excel usa a tag <dimension> (calculada a partir das
            # células com valor, não da altura de linha) pra saber até
            # onde a planilha "existe de verdade": uma quebra apontando
            # pra além disso fica pendurada e é descartada
            # silenciosamente ao reabrir o arquivo. Nesse caso a quebra
            # também não serviria pra nada mesmo — não há mais conteúdo
            # pra paginar depois dela.
            if r < ultima_linha_original:
                if _adicionar_quebra_sem_duplicar(r + 2):  # quebra no meio das 4 (depois da 2ª)
                    quebras_adicionadas += 1
            r -= 1
            continue

        if cel.value is None:
            # Acha o trecho contínuo de linhas vazias inteiro (varrendo
            # pra cima) antes de decidir se ele merece uma quebra — em
            # vez de contar de 4 em 4 e potencialmente disparar várias
            # vezes dentro do mesmo vão vazio longo.
            fim_trecho = r
            inicio_trecho = r
            rr = r - 1
            while rr >= 1 and ws.cell(row=rr, column=1).value is None:
                inicio_trecho = rr
                rr -= 1

            tamanho_trecho = fim_trecho - inicio_trecho + 1
            # A linha logo depois do trecho (rumo ao topo do arquivo) —
            # se for um gatilho de Pergunta/*, esse vão vazio é só o
            # respiro que o código já inseriu ao lado dele, e a quebra
            # daquele bloco já resolve a paginação ali; não duplica.
            logo_apos_pergunta = rr >= 1 and _e_gatilho(ws.cell(row=rr, column=1).value)

            if tamanho_trecho >= 4 and not logo_apos_pergunta:
                meio = (inicio_trecho + fim_trecho) // 2
                if meio - 1 >= 1:
                    if _adicionar_quebra_sem_duplicar(meio - 1):
                        quebras_adicionadas += 1

            r = inicio_trecho - 1
            continue

        r -= 1

    # O Excel espera a lista de quebras em ordem CRESCENTE de linha — como
    # este código varre o arquivo de baixo pra cima, elas são adicionadas
    # em ordem decrescente, e o Excel parece descartar a lista inteira
    # silenciosamente (sem erro, sem quebra nenhuma aparecendo como
    # manual) quando ela não está ordenada. Reordena antes de devolver.
    ws.row_breaks.brk = sorted(ws.row_breaks.brk, key=lambda b: b.id)

    return quebras_adicionadas, linhas_inseridas


# =========================================================
# Código 10 — Ajusta o título da Renda
# =========================================================
def aplicar_codigo_10(ws):
    """Quebra 'Renda Média Mensal Domiciliar (em Salário Mínimo)' em 2 linhas."""
    alvo = "Renda Média Mensal Domiciliar (em Salário Mínimo)"
    formatado = "Renda Média Mensal Domiciliar\n(em Salário Mínimo)"
    alterados = 0
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cel = ws.cell(row=r, column=c)
            if cel.value == alvo:
                cel.value = formatado
                cel.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                ws.row_dimensions[r].height = 30
                alterados += 1
    return alterados


# =========================================================
# Código 14 — Cabeçalho/rodapé de impressão fixos
# =========================================================
def aplicar_codigo_14(wb):
    """Aplica cabeçalho fixo (aviso legal) e rodapé (nº de página) em todas as abas."""
    texto_cabecalho = (
        "NÃO ATUAMOS COM DIVULGAÇÃO DE PESQUISAS ELEITORAIS DESDE 2010.\n"
        "A PRESENTE PESQUISA NÃO ESTÁ REGISTRADA, SUA DIVULGAÇÃO ESTÁ SUJEITA A\n"
        "PENALIDADES PREVISTAS NA LEI ELEITORAL."
    )
    for ws in wb.worksheets:
        ws.oddHeader.left.text = f'&"DIN,Regular"&9{texto_cabecalho}'
        ws.oddHeader.center.text = ""
        ws.oddHeader.right.text = ""
        ws.oddFooter.left.text = ""
        ws.oddFooter.center.text = ""
        ws.oddFooter.right.text = '&"DIN,Regular"&9&P'
        ws.page_margins.header = 1.8 / 2.54
        ws.page_margins.top = 2.4 / 2.54
        ws.page_margins.footer = 0.5 / 2.54
        ws.page_margins.bottom = 1.5 / 2.54
    return len(wb.worksheets)


# =========================================================
# Código 16 — Índice / Sumário (páginas reais pelas quebras)
# =========================================================
def _ultima_linha_com_conteudo(ws):
    """Última linha com algum valor em qualquer coluna, varrendo de baixo pra cima."""
    for r in range(ws.max_row, 0, -1):
        for c in range(1, ws.max_column + 1):
            if ws.cell(row=r, column=c).value not in (None, ""):
                return r
    return 1


def _primeira_pagina(ws):
    """Espelha GetFirstPageNumber do VBA: só usa o número customizado se
    'usar número de primeira página' estiver realmente marcado; senão
    (Automático), a primeira página é 1."""
    try:
        if ws.page_setup.useFirstPageNumber and ws.page_setup.firstPageNumber:
            return int(ws.page_setup.firstPageNumber)
    except (TypeError, ValueError, AttributeError):
        pass
    return 1


def _pagina_da_linha(linha, quebras_ordenadas, primeira_pagina):
    """Conta quantas quebras de página já começaram até (e incluindo) `linha`."""
    cnt = sum(1 for inicio in quebras_ordenadas if inicio <= linha)
    return primeira_pagina + cnt


TERMOS_EXCLUIR_SUMARIO = ("CONTINUACAO", "PARA QUEM")


def _sumario_norm(texto):
    """Maiúsculo, sem acento — usado só pra comparar contra TOTAL/
    SEGMENTO/os termos excluídos do sumário, sem depender de acentuação
    exata."""
    import unicodedata
    t = str(texto or "").strip().upper()
    return "".join(ch for ch in unicodedata.normalize("NFD", t) if unicodedata.category(ch) != "Mn")


def aplicar_codigo_16(ws):
    """
    Cria (recriando se já existir) a aba "Sumário", listando o título de
    cada bloco (mesclagem na coluna A, exceto blocos "Total"/dentro de um
    "Total" até aparecer um "Segmento", "Pergunta..." e "*...") junto com
    o número de página real — calculado a partir das quebras de página
    manuais já presentes na planilha (ex.: inseridas pelo código 07 ou
    pelo código 15/InserirCapasResultados), não da posição na tela.

    Diferente do VBA original (que precisava alternar pra Page Break
    Preview e capturar `ws.HPageBreaks`), aqui as quebras já estão
    disponíveis diretamente em `ws.row_breaks` (mesmo mecanismo que o
    código 07 usa para inseri-las), então não há nada "visual" a simular.

    Formatação (a pedido do Lucas):
        - tudo em DIN Book 9;
        - sem linha de cabeçalho "Título"/"Página";
        - "RESULTADOS PELO TOTAL" vira a primeira linha, em negrito, com
          o número de página real dela (antes era só um texto solto sem
          página nenhuma);
        - "RESULTADOS PELOS SEGMENTOS" (já listada como qualquer outro
          título) fica em negrito;
        - linhas "Continuação" ou "Para quem..." (que às vezes acabam
          mescladas na coluna A perto de uma tabela dividida ou de uma
          Base reduzida) nunca entram na lista.

    Retorna a quantidade de linhas listadas no sumário.
    """
    wb = ws.parent

    if "Sumário" in wb.sheetnames:
        del wb["Sumário"]
    resumo = wb.create_sheet("Sumário")

    fonte_padrao = Font(name="DIN Book", size=9)
    fonte_negrito = Font(name="DIN Book", size=9, bold=True)

    ultima_linha = _ultima_linha_com_conteudo(ws)

    # Quebras de página já presentes: Break(id=n) => nova página começa na linha n+1
    quebras_inicio = sorted({b.id + 1 for b in ws.row_breaks.brk if b.id is not None})
    primeira_pagina = _primeira_pagina(ws)

    # Mescalgens da coluna A, indexadas pela linha do topo (só interessa
    # onde a mesclagem começa exatamente na linha i, igual ao VBA)
    merges_col_a = {
        rng.min_row: rng
        for rng in ws.merged_cells.ranges
        if rng.min_col <= 1 <= rng.max_col
    }

    entradas = []  # (titulo, pagina, negrito)
    titulos_vistos = set()
    ignorar = False
    pagina_total_capa = None

    i = 1
    while i <= ultima_linha:
        rng = merges_col_a.get(i)
        if rng is not None:
            altura_merge = rng.max_row - rng.min_row + 1
            titulo = str(ws.cell(row=rng.min_row, column=rng.min_col).value or "").strip()
            titulo_norm = _sumario_norm(titulo)

            if "TOTAL" in titulo_norm:
                ignorar = True
                if pagina_total_capa is None:
                    pagina_total_capa = _pagina_da_linha(i, quebras_inicio, primeira_pagina)
            elif "SEGMENTO" in titulo_norm:
                ignorar = False

            eh_termo_excluido = any(titulo_norm.startswith(t) for t in TERMOS_EXCLUIR_SUMARIO)

            if not ignorar and not eh_termo_excluido and titulo:
                if not titulo.startswith("Pergunta") and not titulo.startswith("*"):
                    if titulo not in titulos_vistos:
                        negrito = "SEGMENTO" in titulo_norm
                        pagina = _pagina_da_linha(i, quebras_inicio, primeira_pagina)
                        entradas.append((titulo, pagina, negrito))
                        titulos_vistos.add(titulo)

            i += altura_merge
        else:
            i += 1

    linhas_finais = []
    if pagina_total_capa is not None:
        linhas_finais.append(("RESULTADOS PELO TOTAL", pagina_total_capa, True))
    linhas_finais.extend(entradas)

    for idx, (titulo, pagina, negrito) in enumerate(linhas_finais, start=1):
        fonte = fonte_negrito if negrito else fonte_padrao
        c1 = resumo.cell(row=idx, column=1, value=titulo)
        c2 = resumo.cell(row=idx, column=2, value=pagina)
        c1.font = fonte
        c2.font = fonte

    # Larguras aproximadas (sem AutoFit real disponível fora do Excel)
    maior_titulo = max((len(t) for t, _, _ in linhas_finais), default=10)
    resumo.column_dimensions["A"].width = max(15, min(60, maior_titulo + 2))
    resumo.column_dimensions["B"].width = 10

    return len(linhas_finais)


def remover_termo_do_sumario(wb, termo):
    """
    Remove `termo` (ex.: '(Estimulada e Única)') do texto de cada linha
    da coluna A da aba "Sumário", colapsando o espaço duplo que sobra
    no lugar. Não mexe em mais nada (número de página, negrito etc.
    ficam intactos). Retorna quantas linhas foram alteradas.

    Não faz nada (retorna 0) se a aba "Sumário" ainda não existir —
    precisa rodar o código 16 primeiro.
    """
    if "Sumário" not in wb.sheetnames:
        return 0
    resumo = wb["Sumário"]
    alterados = 0
    for r in range(1, resumo.max_row + 1):
        cel = resumo.cell(row=r, column=1)
        if isinstance(cel.value, str) and termo in cel.value:
            novo = cel.value.replace(termo, "")
            novo = " ".join(novo.split())
            cel.value = novo
            alterados += 1
    return alterados
