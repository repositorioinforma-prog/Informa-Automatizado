"""
Total Automático.

Motor de um fluxo mais enxuto que o Relatório Automatizado, pensado pra
relatórios "pelo total" — sem exclusão de colunas por Base reduzida,
sem cabeçalhos repetidos (13), sem Legendas, sem cabeçalho/rodapé fixo
(14) e com uma capa própria (mais simples que a do código 15) no lugar
das "Capas de Resultados".

Reaproveita o máximo possível do que já existe e já foi validado:
    - layout de planilha (fonte DIN 10, larguras de coluna, escala
      100% etc.) vem de `core.relatorio_automatizado_math.
      aplicar_layout_basico_planilha` — a mesma função usada por baixo
      do código 05 no Relatório Automatizado, só que sem a parte de
      exclusão por limite de Base reduzida.
    - inserção segura de linhas (preserva mesclagem/altura/quebras) vem
      de `core.planilha_utils.inserir_linhas_seguro` /
      `remover_linhas_seguro`, o mesmo motor usado pelo código 15.

Sem dependência do Streamlit — pode ser testado isoladamente.
"""
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.pagebreak import Break

from core.planilha_utils import (
    inserir_linhas_seguro,
    remover_linhas_seguro,
    linha_vazia_ate_coluna,
)
from core.relatorio_automatizado_math import (
    aplicar_layout_basico_planilha,
    _ultima_linha_com_conteudo,
)


# ---------------------------------------------------------------------------
# Limpeza de linha branca nas Múltiplas sem base (passo novo, ANTES de
# Base nas Múltiplas)
# ---------------------------------------------------------------------------
def _celula_totalmente_vazia(cel):
    """
    Uma célula só é considerada 'vazia' pra esta limpeza se não tiver
    valor, nem borda em nenhum lado, nem preenchimento — critério
    deliberadamente rígido (ver `remover_linha_branca_multiplas_sem_base`).
    """
    if cel.value not in (None, ""):
        return False
    borda = cel.border
    if borda is not None:
        lados = [borda.left, borda.right, borda.top, borda.bottom, borda.diagonal]
        if any(lado is not None and lado.style for lado in lados):
            return False
    preenchimento = cel.fill
    if preenchimento is not None and getattr(preenchimento, "fill_type", None) not in (None, "none"):
        return False
    return True


def _linha_totalmente_vazia(ws, r, col_ini, col_fim):
    return all(_celula_totalmente_vazia(ws.cell(row=r, column=c)) for c in range(col_ini, col_fim + 1))


def _bloco_ja_tem_base_proximo(ws, linha_inicio, limite_linhas=60):
    """
    Olha pra frente a partir de `linha_inicio` (sem passar do próximo
    'Titulo:') procurando uma linha 'Base' — se achar, o bloco já tem
    base e não deve ser mexido por esta limpeza.
    """
    limite = min(ws.max_row, linha_inicio + limite_linhas)
    for r in range(linha_inicio, limite + 1):
        valor = ws.cell(row=r, column=1).value
        if valor is None:
            continue
        texto = str(valor).strip().upper()
        if texto.startswith("TITULO"):
            return False
        if texto == "BASE" or texto.startswith("BASE "):
            return True
    return False


def remover_linha_branca_multiplas_sem_base(ws):
    """
    Antes de aplicar a Base nas Múltiplas: pra cada tabela de pergunta
    do tipo Múltipla ('Múltipla', 'Multipla' ou 'Estimulada e
    Múltipla', comparação sem acento) que AINDA não tem uma linha
    'Base' nas proximidades, verifica se a linha logo abaixo do título
    está REALMENTE em branco — sem nenhum valor, borda ou preenchimento
    em nenhuma célula, em nenhuma coluna. Só remove nesse caso; se tiver
    qualquer conteúdo, borda ou cor, não mexe.

    Reaproveita a mesma detecção de título (accent-insensitive) da
    tentativa anterior (`_remover_linha_extra_multipla`, em
    `core.relatorio_automatizado_math`, hoje desativada), mas com um
    critério de "vazio" bem mais rígido — aquela só olhava o valor da
    célula, não borda/preenchimento, e não funcionou num teste real.
    Reescrita do zero aqui (função nova, independente), usando
    `remover_linhas_seguro` em vez de `ws.delete_rows()` cru.

    Retorna {"removidas": int, "mantidas": int} — 'mantidas' é quantas
    linhas candidatas foram encontradas mas NÃO removidas por terem
    algum conteúdo/formatação.
    """
    import unicodedata

    def _sem_acento(texto):
        return unicodedata.normalize("NFKD", texto).encode("ascii", "ignore").decode("ascii")

    max_col = ws.max_column
    removidas = 0
    mantidas = 0
    r = ws.max_row
    while r >= 1:
        faixa = None
        for rng in ws.merged_cells.ranges:
            if rng.min_row <= r <= rng.max_row and rng.min_col <= 1 <= rng.max_col:
                faixa = rng
                break

        if faixa is not None:
            eh_inicio_da_faixa = faixa.min_row == r
            linha_titulo_valor = ws.cell(row=faixa.min_row, column=1).value if eh_inicio_da_faixa else None
            linha_titulo_fim = faixa.max_row
        else:
            eh_inicio_da_faixa = True
            linha_titulo_valor = ws.cell(row=r, column=1).value
            linha_titulo_fim = r

        if eh_inicio_da_faixa and linha_titulo_valor is not None:
            texto = _sem_acento(str(linha_titulo_valor).strip().upper())
            eh_titulo = texto.startswith("TITULO")
            eh_multipla = "ESTIMULADA E MULTIPLA" in texto or "MULTIPLA" in texto
            if eh_titulo and eh_multipla:
                linha_abaixo = linha_titulo_fim + 1
                if linha_abaixo <= ws.max_row and not _bloco_ja_tem_base_proximo(ws, linha_abaixo):
                    if _linha_totalmente_vazia(ws, linha_abaixo, 1, max_col):
                        remover_linhas_seguro(ws, linha_abaixo, 1)
                        removidas += 1
                    else:
                        mantidas += 1
        r -= 1

    return {"removidas": removidas, "mantidas": mantidas}


# ---------------------------------------------------------------------------
# Layout (passo 4 do fluxo)
# ---------------------------------------------------------------------------
def aplicar_layout_total_automatico(ws):
    """
    Aplica o mesmo layout básico do código 05 do Relatório Automatizado
    (fonte DIN 10, larguras de coluna A=21/B=8/C+=8.67, escala 100% etc,
    sem a exclusão de colunas por Base reduzida) e, além disso, desativa
    as linhas de grade da planilha — comportamento novo, específico do
    Total Automático (o Relatório Automatizado não desativa gridlines,
    esta função não altera aquele fluxo).

    O negrito dos títulos NÃO é aplicado aqui — continua vindo do
    código 11 (parte do ajuste de altura de labels/títulos/perguntas,
    passo seguinte), igual ao Relatório Automatizado, pra não duplicar
    essa lógica em dois lugares.
    """
    aplicar_layout_basico_planilha(ws)
    ws.sheet_view.showGridLines = False


# ---------------------------------------------------------------------------
# Capa "Resultados pelo total" (passo 7 do fluxo)
# ---------------------------------------------------------------------------
COL_INICIO = 1  # A
COL_FIM_CAPA = 8  # H — largura da capa (mesclagens), NÃO da tabela em si
COL_FIM_BUSCA = 13  # M — mesma faixa usada na detecção de tabelas do código 15

LINHAS_BLOCO_CAPA = 11
MARCADOR_CAPA = "__CAPA_TOTAL_AUTOMATICO__"

# Linhas do bloco (1-indexadas dentro do bloco de 11):
#   1-2: vazias
#   3:   "Resultados pelo total" — DIN Book 11, altura 13.8
#   4-5: placeholder (bairro/cidade etc.) — mesclagem ÚNICA A4:H5, DIN 28
#        negrito, altura 25.1 em cada uma das duas linhas
#   6:   "Diagnóstico Político e Eleitoral" — DIN 12 negrito
#   7:   placeholder (mês/ano) — DIN 12 (tamanho/negrito não especificados
#        por você; usei o mesmo tamanho da linha 6, sem negrito, pra
#        diferenciar do cabeçalho — ajuste fácil se não for isso)
#   8-9: vazias — respiro logo abaixo do mês/ano, ANTES da quebra
#        — quebra de página logo ABAIXO da linha 9
#   10:  vazia, fonte DIN Book 10
#   11:  vazia
# Altura da linha 6 também não foi especificada — fica com a altura
# padrão da planilha (não seto nada nela de propósito).
LINHA_TITULO_RESULTADOS = 3
LINHA_INICIO_PLACEHOLDER_LOCAL = 4
LINHA_FIM_PLACEHOLDER_LOCAL = 5
LINHA_SUBTITULO_FIXO = 6
LINHA_PLACEHOLDER_MES_ANO = 7
LINHA_BRANCO_PRE_QUEBRA = 9  # última linha em branco antes da quebra de página
LINHA_BRANCO_DIN_BOOK = 10

ALTURA_TITULO_RESULTADOS = 13.8
ALTURA_PLACEHOLDER_LOCAL = 25.1

LINHAS_BRANCO_FINAL_RELATORIO = 2  # linhas em branco após a última tabela do relatório


def _encontrar_primeira_linha_total(ws, max_row=None):
    """
    Acha a primeira linha que tem uma célula com o texto exato 'Total'
    (case-insensitive) — é como o código 15 identifica onde a primeira
    tabela do relatório começa. Reimplementado aqui (em vez de importado
    de `core.capas_resultados_math`) pra não criar acoplamento entre os
    dois módulos de capa e não arriscar mexer no código 15, já validado.
    """
    if max_row is None:
        max_row = ws.max_row
    for r in range(1, max_row + 1):
        for c in range(COL_INICIO, COL_FIM_BUSCA + 1):
            v = ws.cell(row=r, column=c).value
            if v is not None and str(v).strip().upper() == "TOTAL":
                return r
    return 0


def _encontrar_primeira_linha_com_conteudo(ws, max_row=None, max_col=None):
    """
    Acha a primeira linha do arquivo inteiro que tem algum conteúdo em
    qualquer coluna até `max_col`. Como a capa do Total Automático só
    vai acima da PRIMEIRA tabela do arquivo (não de uma tabela no meio
    do documento, como no código 15), não existe nada antes dela — a
    primeira linha não-vazia É o início do bloco.

    Preferido a "subir a partir da linha do Total enquanto a linha
    anterior não estiver vazia" porque, nesse fluxo, o código 07 (que
    roda antes da etapa da capa) já normalizou o vão entre 'Pergunta:'
    e 'Total' pra um bloco fixo de linhas em branco — subir parando na
    primeira linha vazia pararia bem no meio desse vão, sem nunca
    chegar ao título de verdade. Também não dá pra caçar pelo texto
    'Titulo: ' (como o código 15 faz) porque o código 11 já roda antes
    da capa nesse fluxo e remove esse prefixo.
    """
    if max_row is None:
        max_row = ws.max_row
    if max_col is None:
        max_col = COL_FIM_BUSCA
    for r in range(1, max_row + 1):
        if not linha_vazia_ate_coluna(ws, r, max_col):
            return r
    return 0


def _mesclar_limpo(ws, linha_ini, linha_fim, col_ini, col_fim):
    """Desfaz qualquer mesclagem que já exista tocando o intervalo, limpa
    os valores e mescla de novo do zero — evita erro de 'mesclagem já
    existe' em reexecuções."""
    for rng in list(ws.merged_cells.ranges):
        if not (rng.max_row < linha_ini or rng.min_row > linha_fim
                or rng.max_col < col_ini or rng.min_col > col_fim):
            try:
                ws.unmerge_cells(str(rng))
            except KeyError:
                ws.merged_cells.ranges.discard(rng)
    for r in range(linha_ini, linha_fim + 1):
        for c in range(col_ini, col_fim + 1):
            ws.cell(row=r, column=c, value=None)
    ws.merge_cells(start_row=linha_ini, start_column=col_ini, end_row=linha_fim, end_column=col_fim)


def remover_capa_antiga_total_automatico(ws):
    """
    Remove uma capa inserida numa execução anterior deste código (busca
    pelo marcador invisível) — permite rodar de novo sem duplicar,
    mesmo padrão usado no código 15.
    """
    max_row = ws.max_row
    for i in range(max_row, 0, -1):
        valor = str(ws.cell(row=i, column=1).value or "").strip()
        if valor == MARCADOR_CAPA:
            remover_linhas_seguro(ws, i, LINHAS_BLOCO_CAPA)
            return True
    return False


def inserir_capa_total_automatico(ws, texto_local, texto_mes_ano):
    """
    Insere, logo acima da primeira tabela do relatório (a primeira
    linha com qualquer conteúdo no arquivo), um bloco de 11 linhas:
        3: "Resultados pelo total"
        4-5: `texto_local` (nome do bairro/cidade etc. — mesclagem
             única A:H ocupando as duas linhas)
        6: "Diagnóstico Político e Eleitoral"
        7: `texto_mes_ano`
        8-9: vazias (respiro logo abaixo do mês/ano)
        10: vazia, fonte DIN Book 10
        11: vazia
    com uma quebra de página manual entre a linha 9 e a linha 10 — a
    página atual termina com o respiro logo abaixo do mês/ano (8-9), e a
    página nova começa com mais duas linhas em branco (10-11) antes do
    título da tabela original (agora deslocado pra depois do bloco).

    Se já existir uma capa de uma execução anterior (marcador
    invisível), remove antes de inserir a nova — pode rodar de novo
    sem duplicar.

    Retorna um dict com a linha onde o bloco começou, ou
    {"status": "erro", "mensagem": ...} se não achar nenhuma tabela.
    """
    remover_capa_antiga_total_automatico(ws)

    linha_total = _encontrar_primeira_linha_total(ws)
    if not linha_total:
        return {
            "status": "erro",
            "mensagem": (
                "Não encontrei nenhuma célula 'Total' no arquivo — não "
                "deu pra identificar onde a primeira tabela começa."
            ),
        }

    # A capa vai acima da primeira tabela do arquivo — como não existe
    # nada antes dela, a primeira linha com qualquer conteúdo É o início
    # do bloco (ver docstring de `_encontrar_primeira_linha_com_conteudo`
    # pra entender por que não dá pra usar "subir a partir do Total").
    linha_base = _encontrar_primeira_linha_com_conteudo(ws)
    if not linha_base:
        linha_base = linha_total

    inserir_linhas_seguro(ws, linha_base, LINHAS_BLOCO_CAPA)

    # marcador invisível (texto branco, tamanho 1) na 1ª linha do bloco —
    # só pra reidentificar o bloco numa reexecução futura.
    cel_marcador = ws.cell(row=linha_base, column=1, value=MARCADOR_CAPA)
    cel_marcador.font = Font(size=1, color="FFFFFFFF")

    linha_titulo = linha_base + (LINHA_TITULO_RESULTADOS - 1)
    linha_placeholder_ini = linha_base + (LINHA_INICIO_PLACEHOLDER_LOCAL - 1)
    linha_placeholder_fim = linha_base + (LINHA_FIM_PLACEHOLDER_LOCAL - 1)
    linha_subtitulo = linha_base + (LINHA_SUBTITULO_FIXO - 1)
    linha_mes_ano = linha_base + (LINHA_PLACEHOLDER_MES_ANO - 1)
    linha_branco_pre_quebra = linha_base + (LINHA_BRANCO_PRE_QUEBRA - 1)
    linha_branco_din_book = linha_base + (LINHA_BRANCO_DIN_BOOK - 1)

    centralizado = Alignment(horizontal="center", vertical="center", wrap_text=False)
    centralizado_quebra = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Linha 3 — "Resultados pelo total"
    _mesclar_limpo(ws, linha_titulo, linha_titulo, COL_INICIO, COL_FIM_CAPA)
    cel_titulo = ws.cell(row=linha_titulo, column=COL_INICIO, value="Resultados pelo total")
    cel_titulo.font = Font(name="DIN Book", size=11)
    cel_titulo.alignment = centralizado
    ws.row_dimensions[linha_titulo].height = ALTURA_TITULO_RESULTADOS

    # Linhas 4-5 — placeholder do bairro/cidade (mesclagem única A:H,
    # ocupando as duas linhas)
    _mesclar_limpo(ws, linha_placeholder_ini, linha_placeholder_fim, COL_INICIO, COL_FIM_CAPA)
    cel_local = ws.cell(row=linha_placeholder_ini, column=COL_INICIO, value=texto_local)
    cel_local.font = Font(name="DIN", size=28, bold=True)
    cel_local.alignment = centralizado_quebra
    ws.row_dimensions[linha_placeholder_ini].height = ALTURA_PLACEHOLDER_LOCAL
    ws.row_dimensions[linha_placeholder_fim].height = ALTURA_PLACEHOLDER_LOCAL

    # Linha 6 — "Diagnóstico Político e Eleitoral"
    _mesclar_limpo(ws, linha_subtitulo, linha_subtitulo, COL_INICIO, COL_FIM_CAPA)
    cel_subtitulo = ws.cell(row=linha_subtitulo, column=COL_INICIO, value="Diagnóstico Político e Eleitoral")
    cel_subtitulo.font = Font(name="DIN", size=12, bold=True)
    cel_subtitulo.alignment = centralizado

    # Linha 7 — placeholder do mês/ano
    _mesclar_limpo(ws, linha_mes_ano, linha_mes_ano, COL_INICIO, COL_FIM_CAPA)
    cel_mes_ano = ws.cell(row=linha_mes_ano, column=COL_INICIO, value=texto_mes_ano)
    cel_mes_ano.font = Font(name="DIN", size=12, bold=False)
    cel_mes_ano.alignment = centralizado

    # Linhas 8-9 — vazias, respiro logo abaixo do mês/ano (nada a fazer,
    # já vêm em branco da inserção)

    # Linha 10 — vazia, mas com fonte DIN Book 10 já aplicada (pra quem
    # for editar à mão depois herdar a fonte certa em vez da padrão)
    for c in range(COL_INICIO, COL_FIM_CAPA + 1):
        ws.cell(row=linha_branco_din_book, column=c).font = Font(name="DIN Book", size=10)

    # Linha 11 fica vazia (nada a fazer — já vem em branco da inserção)

    # Quebra de página ENTRE a linha 9 (branco) e a linha 10 (branco,
    # DIN Book 10) — a página nova começa com essas duas linhas em
    # branco (10-11) logo antes do título da tabela original (linha
    # 12), e a página anterior termina com as linhas 8-9 (branco) logo
    # abaixo do mês/ano.
    # (Break(id=X) = quebra depois da linha X, nova página começa em
    # X+1.) Nesse fluxo o código 07 já rodou antes (já existem quebras
    # mais abaixo no arquivo), e `.append()` do openpyxl sempre põe no
    # FIM da lista Python independente do valor — como a capa fica
    # perto do topo do arquivo, isso deixaria a lista fora de ordem
    # crescente, e o Excel descarta a lista inteira de quebras nesse
    # caso (gotcha já mapeado no projeto). Por isso reordena a lista
    # inteira por id logo depois de adicionar a nova quebra.
    ws.row_breaks.append(Break(id=linha_branco_pre_quebra))
    ws.row_breaks.brk = sorted(ws.row_breaks.brk, key=lambda b: b.id)

    return {
        "status": "ok",
        "linha_inicio_bloco": linha_base,
        "linha_titulo": linha_titulo,
        "linha_placeholder_local": linha_placeholder_ini,
        "linha_subtitulo": linha_subtitulo,
        "linha_placeholder_mes_ano": linha_mes_ano,
    }


def _eh_linha_titulo(ws, r):
    """
    Identifica se a linha `r` é um título de pergunta ('Titulo: ...').
    Combina DOIS sinais porque, dependendo de QUANDO essas funções de
    análise rodam no fluxo, o texto pode já não ter mais o prefixo
    'Titulo: ' — o código 11 remove esse prefixo (parte do ajuste de
    altura dos títulos), e essas análises rodam depois dele (na etapa
    final):
      1) o texto começa com 'Titulo:' (funciona ANTES do código 11); ou
      2) a célula é o início de uma mesclagem de uma linha só, com mais
         de uma coluna, na coluna A (sobrevive ao código 11 — só a
         mesclagem em si não muda).
    Retorna (True, texto_do_titulo) ou (False, None).
    """
    valor = ws.cell(row=r, column=1).value
    if valor is not None and str(valor).strip().upper().startswith("TITULO"):
        return True, str(valor).strip()
    for rng in ws.merged_cells.ranges:
        if rng.min_row == r and rng.max_row == r and rng.min_col == 1 and rng.max_col > 1:
            v = ws.cell(row=r, column=1).value
            if v is not None and str(v).strip():
                return True, str(v).strip()
    return False, None


def encontrar_perguntas_com_base_reduzida(ws):
    """
    Varre o arquivo procurando linhas 'Base reduzida' (rótulo que o
    próprio SPSS já traz quando uma pergunta foi feita só pra uma parte
    da amostra — filtrada por uma pergunta anterior) e retorna os
    títulos das perguntas correspondentes, na ordem em que aparecem.

    Pra cada 'Base reduzida' encontrada, associa o título da pergunta
    mais próxima ACIMA dela no arquivo (ver `_eh_linha_titulo`).

    Retorna uma lista de títulos (sem duplicar o mesmo título mais de
    uma vez, mesmo que ele tenha mais de uma linha 'Base reduzida' —
    o que não deveria acontecer, mas por segurança).
    """
    titulos = []
    titulo_atual = None
    vistos = set()
    for r in range(1, ws.max_row + 1):
        eh_titulo, texto_titulo = _eh_linha_titulo(ws, r)
        if eh_titulo:
            titulo_atual = texto_titulo
            continue
        valor = ws.cell(row=r, column=1).value
        if valor is None:
            continue
        texto_upper = str(valor).strip().upper()
        if texto_upper == "BASE REDUZIDA":
            if titulo_atual is not None and titulo_atual not in vistos:
                titulos.append(titulo_atual)
                vistos.add(titulo_atual)
    return titulos


def detectar_bases_divergentes(ws):
    """
    Varre todas as linhas 'Base' (NÃO 'Base reduzida' — essas têm um
    valor menor por design, filtrada de propósito, não é uma
    divergência) e compara o valor da coluna Total (coluna B, sempre a
    primeira coluna de dado — mesma convenção usada em
    `core.base_multiplas_math.calcular_linhas_base`) entre todas elas.

    Se a maioria das tabelas compartilhar o mesmo valor de Base mas uma
    ou mais tiverem um valor diferente, retorna essas tabelas
    divergentes — pode ser um problema real nos dados (base errada) ou
    uma pergunta que deveria ter sido marcada 'Base reduzida' e não foi.

    Retorna {"valor_maioria": valor ou None, "divergentes": [{"titulo",
    "valor"}, ...]}. Se tiver menos de 2 linhas 'Base' no arquivo pra
    comparar, retorna valor_maioria=None e divergentes=[] (não dá pra
    falar em "maioria" com um dado só).
    """
    from collections import Counter

    bases = []
    titulo_atual = None
    for r in range(1, ws.max_row + 1):
        eh_titulo, texto_titulo = _eh_linha_titulo(ws, r)
        if eh_titulo:
            titulo_atual = texto_titulo
            continue
        valor = ws.cell(row=r, column=1).value
        if valor is None:
            continue
        texto_upper = str(valor).strip().upper()
        if texto_upper == "BASE":
            valor_total = ws.cell(row=r, column=2).value
            if isinstance(valor_total, (int, float)):
                bases.append((titulo_atual, valor_total))

    if len(bases) < 2:
        return {"valor_maioria": None, "divergentes": []}

    contagem = Counter(valor for _, valor in bases)
    valor_maioria, _ = contagem.most_common(1)[0]

    divergentes = [
        {"titulo": titulo, "valor": valor}
        for titulo, valor in bases
        if valor != valor_maioria
    ]

    return {"valor_maioria": valor_maioria, "divergentes": divergentes}


def _encontrar_ultima_linha_pergunta(ws):
    """
    Acha a última linha com uma célula começando com 'Pergunta:' — é a
    linha da PERGUNTA (o rodapé de cada bloco, com o texto exato da
    pergunta do SPSS) da ÚLTIMA tabela do relatório.
    """
    for r in range(ws.max_row, 0, -1):
        valor = ws.cell(row=r, column=1).value
        if valor is not None and str(valor).strip().upper().startswith("PERGUNTA"):
            return r
    return 0


def adicionar_linhas_finais_em_branco(ws, quantidade=LINHAS_BRANCO_FINAL_RELATORIO):
    """
    Garante `quantidade` linhas em branco, altura 15, logo abaixo da
    linha 'Pergunta:' da última tabela do relatório. Se por algum
    motivo não achar nenhuma linha 'Pergunta:' no arquivo, cai pra
    última linha com conteúdo (qualquer que seja) como referência.

    "Materializa" as células aplicando a fonte padrão da planilha (DIN
    10) mesmo sem nenhum valor — uma célula com `value=None` some no
    arquivo salvo (o openpyxl não grava células totalmente vazias),
    então sem um estilo aplicado essas linhas não fariam parte da área
    usada/impressa e o "respiro" desapareceria ao salvar.

    Idempotente na prática: reaplicar não duplica nada (só re-garante o
    estilo/altura nas mesmas linhas), já que sempre calcula a partir da
    mesma linha de referência.

    Retorna a linha de referência (a última 'Pergunta:', ou a última
    linha com conteúdo se não achar nenhuma) ANTES de adicionar o respiro.
    """
    ALTURA_LINHA_BRANCO_FINAL = 15

    ultima = _encontrar_ultima_linha_pergunta(ws)
    if not ultima:
        ultima = _ultima_linha_com_conteudo(ws)
    fonte_padrao = Font(name="DIN", size=10)
    for i in range(1, quantidade + 1):
        linha = ultima + i
        ws.cell(row=linha, column=1).font = fonte_padrao
        ws.row_dimensions[linha].height = ALTURA_LINHA_BRANCO_FINAL
    return ultima
