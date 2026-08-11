"""
Camada 1 - Consolidação da Base Mestre

Junta as fontes brutas do IBGE em duas tabelas mestre:
    - master_municipios.csv   (5.570 linhas)
    - master_distritos.csv    (10.694 linhas)

Cada linha traz: hierarquia territorial completa (Região Intermediária,
Região Imediata, Município, Distrito) + população total + população por
sexo/idade nas faixas da pesquisa (16-19, 20-29, 30-39, 40-49, 50+) +
renda domiciliar (Até 2SM, 2-5SM, 5-10SM, Mais de 10SM).

Fontes:
    - DTB 2025 (IBGE): hierarquia territorial oficial por código
    - IBGE Censo 2022: sexo/idade por município e distrito
    - SIDRA tabela 9514: idade exata (16, 17, 18, 19 anos separados) por município
    - Censo 2010: renda domiciliar por município e distrito

Uso (script utilitário, roda fora do fluxo do app — só quando os dados
do IBGE precisarem ser atualizados/reconstruídos):
    1. Coloque os 6 arquivos brutos IBGE em ./dados_brutos_gerador_amostra/
       na raiz do projeto (não incluídos no repositório por serem grandes;
       baixe novamente do IBGE se precisar reconstruir).
    2. python scripts/amostrador_build_master.py
    3. Os CSVs serão gerados em ./dados/gerador_amostra/
       (sobrescrevendo os que já estão versionados no repo)
"""
import openpyxl
import csv
import re
import unicodedata
from pathlib import Path

# ---------------------------------------------------------------------------
# Configuração de caminhos
# ---------------------------------------------------------------------------
RAIZ_PROJETO = Path(__file__).parent.parent
BRUTOS = RAIZ_PROJETO / "dados_brutos_gerador_amostra"
SAIDA = RAIZ_PROJETO / "dados" / "gerador_amostra"
SAIDA.mkdir(parents=True, exist_ok=True)

ARQ_DTB_MUN = BRUTOS / "RELATORIO_DTB_BRASIL_2025_MUNICIPIOS.xlsx"
ARQ_DTB_DIST = BRUTOS / "RELATORIO_DTB_BRASIL_2025_DISTRITOS.xlsx"
ARQ_SEXO_IDADE_MUN = BRUTOS / "Dados_Sexo_e_Idade_por_Municipio_-_IBGE_2022.xlsx"
ARQ_SEXO_IDADE_DIST = BRUTOS / "Dados_Sexo_e_Idade_por_Distrito_-_IBGE_2022.xlsx"
ARQ_IDADE_EXATA = BRUTOS / "Cidades_-_Idade.xlsx"
ARQ_RENDA = BRUTOS / "Renda_2010_-_Por_municipios_e_distritos.xlsx"

# Código UF (2 dígitos) → sigla, para compatibilizar com base de Renda (que usa sigla)
UF_SIGLA = {
    "11": "RO", "12": "AC", "13": "AM", "14": "RR", "15": "PA", "16": "AP", "17": "TO",
    "21": "MA", "22": "PI", "23": "CE", "24": "RN", "25": "PB", "26": "PE", "27": "AL",
    "28": "SE", "29": "BA", "31": "MG", "32": "ES", "33": "RJ", "35": "SP", "41": "PR",
    "42": "SC", "43": "RS", "50": "MS", "51": "MT", "52": "GO", "53": "DF",
}


# ---------------------------------------------------------------------------
# Utilitários
# ---------------------------------------------------------------------------
def norm(s):
    """Normaliza texto para join por nome: remove acento, pontuação, caixa."""
    if s is None:
        return ""
    s = str(s).strip().lower()
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode()
    s = re.sub(r"[^a-z0-9 ]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def g(d, key, default=0):
    """Pega valor numérico de um dict, retorna default se ausente ou não-número."""
    v = d.get(key, default)
    return v if isinstance(v, (int, float)) else default


# ---------------------------------------------------------------------------
# 1. DTB Municípios — hierarquia por código de município
# ---------------------------------------------------------------------------
print("Carregando DTB - Municípios...")
wb = openpyxl.load_workbook(ARQ_DTB_MUN, read_only=True, data_only=True)
ws = wb["DTB_Municípios"]
rows = list(ws.iter_rows(values_only=True))
header_idx = next(i for i, r in enumerate(rows) if r[0] == "UF")

hier_mun = {}
for r in rows[header_idx + 1:]:
    if not r[0]:
        continue
    uf, nome_uf, cod_ri, nome_ri, cod_rim, nome_rim, cod_mun_curto, cod_mun, nome_mun = r[:9]
    hier_mun[str(cod_mun)] = dict(
        uf=uf, nome_uf=nome_uf,
        cod_regiao_intermediaria=cod_ri, regiao_intermediaria=nome_ri,
        cod_regiao_imediata=cod_rim, regiao_imediata=nome_rim,
        cod_municipio=str(cod_mun), municipio=nome_mun,
    )
print(f"  {len(hier_mun)} municípios carregados")


# ---------------------------------------------------------------------------
# 2. DTB Distritos — hierarquia + código do município pai
# ---------------------------------------------------------------------------
print("Carregando DTB - Distritos...")
wb = openpyxl.load_workbook(ARQ_DTB_DIST, read_only=True, data_only=True)
ws = wb["DTB_Distritos"]
rows = list(ws.iter_rows(values_only=True))
header_idx = next(i for i, r in enumerate(rows) if r[0] == "UF")

hier_dist = {}
for r in rows[header_idx + 1:]:
    if not r[0]:
        continue
    (uf, nome_uf, cod_ri, nome_ri, cod_rim, nome_rim,
     cod_mun_curto, cod_mun, nome_mun,
     cod_dist_curto, cod_dist, nome_dist) = r[:12]
    hier_dist[str(cod_dist)] = dict(
        uf=uf, nome_uf=nome_uf,
        cod_regiao_intermediaria=cod_ri, regiao_intermediaria=nome_ri,
        cod_regiao_imediata=cod_rim, regiao_imediata=nome_rim,
        cod_municipio=str(cod_mun), municipio=nome_mun,
        cod_distrito=str(cod_dist), distrito=nome_dist,
    )
print(f"  {len(hier_dist)} distritos carregados")


# ---------------------------------------------------------------------------
# 3. Sexo/Idade IBGE 2022 (município e distrito) — indexado por código IBGE
# ---------------------------------------------------------------------------
def load_sexo_idade(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    header_idx = next(i for i, r in enumerate(rows) if r[0] in ("CD_MUN", "CD_DIST"))
    cols = rows[0]  # linha 0 tem os nomes legíveis das colunas
    data = {}
    for r in rows[header_idx + 1:]:
        if r[0] is None:
            continue
        data[str(r[0])] = dict(zip(cols, r))
    return data

print("Carregando Sexo/Idade IBGE 2022...")
sexo_idade_mun = load_sexo_idade(ARQ_SEXO_IDADE_MUN)
sexo_idade_dist = load_sexo_idade(ARQ_SEXO_IDADE_DIST)
print(f"  {len(sexo_idade_mun)} municípios | {len(sexo_idade_dist)} distritos")


# ---------------------------------------------------------------------------
# 4. Idade exata 16-19 (SIDRA 9514) — só município, indexado por nome+UF
# ---------------------------------------------------------------------------
print("Carregando Idade exata 16-19 (SIDRA 9514)...")
wb = openpyxl.load_workbook(ARQ_IDADE_EXATA, read_only=True, data_only=True)
ws = wb["População residente (Pessoas)"]
rows = list(ws.iter_rows(values_only=True))

idade_exata = {}
current_mun = None
for r in rows[4:]:  # cabeçalho ocupa as primeiras 4 linhas
    if r[0]:
        current_mun = norm(r[0])
        idade_exata[current_mun] = {}
    if current_mun and r[3] in ("16 anos", "17 anos", "18 anos", "19 anos"):
        idade = int(r[3].split()[0])
        val = r[4] if isinstance(r[4], (int, float)) else 0
        idade_exata[current_mun][idade] = val
print(f"  {len(idade_exata)} municípios com idade exata")


# ---------------------------------------------------------------------------
# 5. Renda 2010 (Censo) — indexado por nome+UF normalizado
# ---------------------------------------------------------------------------
print("Carregando Renda 2010...")
wb = openpyxl.load_workbook(ARQ_RENDA, read_only=True, data_only=True)

renda_mun = {}
ws = wb["Renda 2010 - Municípios"]
for i, r in enumerate(ws.iter_rows(values_only=True)):
    if i == 0:
        continue
    nome, ate2, de2a5, de5a10, mais10 = r
    renda_mun[norm(nome)] = dict(
        ate_2sm=ate2, de_2_a_5sm=de2a5, de_5_a_10sm=de5a10, mais_10sm=mais10
    )

renda_dist = {}
ws = wb["Renda 2010 - Dsitritos"]  # sic — o typo está no arquivo original do IBGE
for i, r in enumerate(ws.iter_rows(values_only=True)):
    if i == 0:
        continue
    nome, ate2, de2a5, de5a10, mais10 = r
    renda_dist[norm(nome)] = dict(
        ate_2sm=ate2, de_2_a_5sm=de2a5, de_5_a_10sm=de5a10, mais_10sm=mais10
    )
print(f"  {len(renda_mun)} municípios | {len(renda_dist)} distritos com renda")


# ---------------------------------------------------------------------------
# 6. Monta master_municipios.csv
# ---------------------------------------------------------------------------
print("\nConsolidando municípios...")
mun_rows = []
renda_miss_mun = []

for cod, hier in hier_mun.items():
    si = sexo_idade_mun.get(cod)
    if not si:
        continue
    sigla = UF_SIGLA.get(str(hier["uf"]), "")
    nome_norm = norm(hier["municipio"]) + " " + norm(sigla)
    nome_mun_uf = norm(f'{hier["municipio"]} ({sigla})')
    exact = idade_exata.get(nome_mun_uf)

    # 16-19: usa idade exata (SIDRA) para o total, divide masc/fem pela proporção
    # observada na faixa 15-19 do arquivo Sexo/Idade
    masc_15_19 = g(si, "Sexo masculino, 15 a 19 anos")
    fem_15_19 = g(si, "Sexo feminino, 15 a 19 anos")
    tot_15_19 = masc_15_19 + fem_15_19

    if exact:
        exact_total = sum(exact.values())
        if tot_15_19 > 0:
            masc_16_19 = round(exact_total * (masc_15_19 / tot_15_19))
            fem_16_19 = exact_total - masc_16_19
        else:
            masc_16_19 = fem_16_19 = 0
    else:
        # fallback improvável (base é 100% coberta)
        masc_16_19, fem_16_19 = masc_15_19, fem_15_19

    # Demais faixas: soma direta das faixas quinquenais do IBGE
    masc_20_29 = g(si, "Sexo masculino, 20 a 24 anos") + g(si, "Sexo masculino, 25 a 29 anos")
    fem_20_29 = g(si, "Sexo feminino, 20 a 24 anos") + g(si, "Sexo feminino, 25 a 29 anos")
    masc_30_39 = g(si, "Sexo masculino, 30 a 39 anos")
    fem_30_39 = g(si, "Sexo feminino, 30 a 39 anos")
    masc_40_49 = g(si, "Sexo masculino, 40 a 49 anos")
    fem_40_49 = g(si, "Sexo feminino, 40 a 49 anos")
    masc_50mais = (
        g(si, "Sexo masculino, 50 a 59 anos")
        + g(si, "Sexo masculino, 60 a 69 anos")
        + g(si, "Sexo masculino, 70 anos ou mais")
    )
    fem_50mais = (
        g(si, "Sexo feminino, 50 a 59 anos")
        + g(si, "Sexo feminino, 60 a 69 anos")
        + g(si, "Sexo feminino, 70 anos ou mais")
    )

    renda = renda_mun.get(nome_norm)
    if not renda:
        renda_miss_mun.append(f'{hier["municipio"]} ({sigla})')
        renda = dict(ate_2sm=0, de_2_a_5sm=0, de_5_a_10sm=0, mais_10sm=0)

    mun_rows.append(dict(
        cod_municipio=cod, municipio=hier["municipio"], uf=sigla, nome_uf=hier["nome_uf"],
        cod_regiao_intermediaria=hier["cod_regiao_intermediaria"],
        regiao_intermediaria=hier["regiao_intermediaria"],
        cod_regiao_imediata=hier["cod_regiao_imediata"],
        regiao_imediata=hier["regiao_imediata"],
        populacao_total=g(si, "Quantidade de moradores"),
        masc_total=g(si, "Sexo masculino"), fem_total=g(si, "Sexo feminino"),
        masc_16_19=masc_16_19, fem_16_19=fem_16_19,
        masc_20_29=masc_20_29, fem_20_29=fem_20_29,
        masc_30_39=masc_30_39, fem_30_39=fem_30_39,
        masc_40_49=masc_40_49, fem_40_49=fem_40_49,
        masc_50mais=masc_50mais, fem_50mais=fem_50mais,
        renda_ate_2sm=renda["ate_2sm"], renda_2_a_5sm=renda["de_2_a_5sm"],
        renda_5_a_10sm=renda["de_5_a_10sm"], renda_mais_10sm=renda["mais_10sm"],
        idade_exata_disponivel=bool(exact),
    ))

with open(SAIDA / "master_municipios.csv", "w", newline="", encoding="utf-8") as f:
    writer = csv.DictWriter(f, fieldnames=list(mun_rows[0].keys()))
    writer.writeheader()
    writer.writerows(mun_rows)

print(f"  master_municipios.csv: {len(mun_rows)} linhas")
print(f"  Municípios sem match de renda (novos, criados pós-2010): {len(renda_miss_mun)}")
if renda_miss_mun:
    print(f"    Exemplos: {renda_miss_mun[:5]}")


# ---------------------------------------------------------------------------
# 7. Monta master_distritos.csv
# ---------------------------------------------------------------------------
print("\nConsolidando distritos...")
dist_rows = []
renda_miss_dist = []

for cod, hier in hier_dist.items():
    si = sexo_idade_dist.get(cod)
    if not si:
        continue
    sigla = UF_SIGLA.get(str(hier["uf"]), "")
    nome_key = norm(f'{hier["distrito"]} - {hier["municipio"]} ({sigla})')

    # Distritos: 15-19 direto do IBGE (não temos idade exata em nível de distrito)
    masc_16_19 = g(si, "Sexo masculino, 15 a 19 anos")
    fem_16_19 = g(si, "Sexo feminino, 15 a 19 anos")
    masc_20_29 = g(si, "Sexo masculino, 20 a 24 anos") + g(si, "Sexo masculino, 25 a 29 anos")
    fem_20_29 = g(si, "Sexo feminino, 20 a 24 anos") + g(si, "Sexo feminino, 25 a 29 anos")
    masc_30_39 = g(si, "Sexo masculino, 30 a 39 anos")
    fem_30_39 = g(si, "Sexo feminino, 30 a 39 anos")
    masc_40_49 = g(si, "Sexo masculino, 40 a 49 anos")
    fem_40_49 = g(si, "Sexo feminino, 40 a 49 anos")
    masc_50mais = (
        g(si, "Sexo masculino, 50 a 59 anos")
        + g(si, "Sexo masculino, 60 a 69 anos")
        + g(si, "Sexo masculino, 70 anos ou mais")
    )
    fem_50mais = (
        g(si, "Sexo feminino, 50 a 59 anos")
        + g(si, "Sexo feminino, 60 a 69 anos")
        + g(si, "Sexo feminino, 70 anos ou mais")
    )

    renda = renda_dist.get(nome_key)
    if not renda:
        renda_miss_dist.append(f'{hier["distrito"]} - {hier["municipio"]} ({sigla})')
        renda = dict(ate_2sm=0, de_2_a_5sm=0, de_5_a_10sm=0, mais_10sm=0)

    dist_rows.append(dict(
        cod_distrito=cod, distrito=hier["distrito"],
        cod_municipio=hier["cod_municipio"], municipio=hier["municipio"],
        uf=sigla, nome_uf=hier["nome_uf"],
        cod_regiao_intermediaria=hier["cod_regiao_intermediaria"],
        regiao_intermediaria=hier["regiao_intermediaria"],
        cod_regiao_imediata=hier["cod_regiao_imediata"],
        regiao_imediata=hier["regiao_imediata"],
        populacao_total=g(si, "Quantidade de moradores"),
        masc_total=g(si, "Sexo masculino"), fem_total=g(si, "Sexo feminino"),
        masc_16_19=masc_16_19, fem_16_19=fem_16_19,
        masc_20_29=masc_20_29, fem_20_29=fem_20_29,
        masc_30_39=masc_30_39, fem_30_39=fem_30_39,
        masc_40_49=masc_40_49, fem_40_49=fem_40_49,
        masc_50mais=masc_50mais, fem_50mais=fem_50mais,
        renda_ate_2sm=renda["ate_2sm"], renda_2_a_5sm=renda["de_2_a_5sm"],
        renda_5_a_10sm=renda["de_5_a_10sm"], renda_mais_10sm=renda["mais_10sm"],
    ))

with open(SAIDA / "master_distritos.csv", "w", newline="", encoding="utf-8") as f:
    writer = csv.DictWriter(f, fieldnames=list(dist_rows[0].keys()))
    writer.writeheader()
    writer.writerows(dist_rows)

print(f"  master_distritos.csv: {len(dist_rows)} linhas")
print(f"  Distritos sem match de renda (novos, criados pós-2010): {len(renda_miss_dist)}")
if renda_miss_dist:
    print(f"    Exemplos: {renda_miss_dist[:5]}")

print("\nOK. Base mestre consolidada em:", SAIDA)